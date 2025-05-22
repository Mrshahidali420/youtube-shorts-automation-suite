# --- START OF FILE performance_tracker.py ---

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YouTube Performance Tracker

This script fetches performance metrics for uploaded videos using the YouTube Data API.
It reads the YouTube Video IDs from the Excel file and updates the metrics,
focusing only on videos scheduled within the last 7 days (or those missing a valid schedule time).
"""

import os
import json
import time
import pickle
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple, Union
import traceback # For detailed error logging

# Import excel_utils module
try:
    # First try to import from the utils package
    import sys
    import importlib.util

    # Get the script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Try to import excel_utils from utils directory
    excel_utils_path = os.path.join(script_dir, 'utils', 'excel_utils.py')
    if os.path.exists(excel_utils_path):
        # Use importlib to load the module from the file path
        spec = importlib.util.spec_from_file_location("excel_utils", excel_utils_path)
        excel_utils = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(excel_utils)

        # Get the required functions
        load_workbook_safely = excel_utils.load_workbook_safely
        save_workbook_safely = excel_utils.save_workbook_safely
        create_excel_backup = excel_utils.create_excel_backup
        find_column_indices = excel_utils.find_column_indices
        get_cell_value = excel_utils.get_cell_value
        parse_date_value = excel_utils.parse_date_value

        # Try to get logging functions
        try:
            log_info = excel_utils.log_info
            log_success = excel_utils.log_success
            log_warning = excel_utils.log_warning
            log_error = excel_utils.log_error
        except AttributeError:
            # Define basic loggers if not in excel_utils
            def log_info(msg, indent=0): print(f"{'  '*indent}INFO: {msg}")
            def log_success(msg, indent=0): print(f"{'  '*indent}SUCCESS: {msg}")
            def log_warning(msg, indent=0): print(f"{'  '*indent}WARNING: {msg}")
            def log_error(msg, indent=0, include_traceback=False): print(f"{'  '*indent}ERROR: {msg}")

        EXCEL_UTILS_AVAILABLE = True
        print("INFO: Using excel_utils from utils directory.")

    # If not found in utils, try to import directly from the root
    else:
        excel_utils_path = os.path.join(script_dir, 'excel_utils.py')
        if os.path.exists(excel_utils_path):
            spec = importlib.util.spec_from_file_location("excel_utils", excel_utils_path)
            excel_utils = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(excel_utils)

            # Get the required functions
            load_workbook_safely = excel_utils.load_workbook_safely
            save_workbook_safely = excel_utils.save_workbook_safely
            create_excel_backup = excel_utils.create_excel_backup
            find_column_indices = excel_utils.find_column_indices
            get_cell_value = excel_utils.get_cell_value
            parse_date_value = excel_utils.parse_date_value

            # Try to get logging functions
            try:
                log_info = excel_utils.log_info
                log_success = excel_utils.log_success
                log_warning = excel_utils.log_warning
                log_error = excel_utils.log_error
            except AttributeError:
                # Define basic loggers if not in excel_utils
                def log_info(msg, indent=0): print(f"{'  '*indent}INFO: {msg}")
                def log_success(msg, indent=0): print(f"{'  '*indent}SUCCESS: {msg}")
                def log_warning(msg, indent=0): print(f"{'  '*indent}WARNING: {msg}")
                def log_error(msg, indent=0, include_traceback=False): print(f"{'  '*indent}ERROR: {msg}")

            EXCEL_UTILS_AVAILABLE = True
            print("INFO: Using excel_utils from root directory.")
        else:
            raise ImportError("excel_utils not found")

except ImportError:
    # Fall back to direct openpyxl usage if excel_utils is not available
    print("WARNING: excel_utils not found or import failed. Using direct openpyxl.")
    from openpyxl import load_workbook
    EXCEL_UTILS_AVAILABLE = False
    # Define basic logging helpers if excel_utils failed completely
    def log_info(msg, indent=0): print(f"{'  '*indent}INFO: {msg}")
    def log_success(msg, indent=0): print(f"{'  '*indent}SUCCESS: {msg}")
    def log_warning(msg, indent=0): print(f"{'  '*indent}WARNING: {msg}")
    def log_error(msg, indent=0, include_traceback=False): print(f"{'  '*indent}ERROR: {msg}")

# Try to import dateutil for better date parsing
try:
    import dateutil.parser as date_parser
    DATEUTIL_AVAILABLE = True
except ImportError:
    DATEUTIL_AVAILABLE = False
    log_warning("dateutil library not found (pip install python-dateutil). Robust date parsing disabled.")


# Google API imports
try:
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    GOOGLE_API_AVAILABLE = True
except ImportError:
    print("ERROR: Google API libraries not found. Install with:")
    print("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
    GOOGLE_API_AVAILABLE = False

# Colorama import
try:
    import colorama
    from colorama import Fore, Style, init
    init(autoreset=True)
    COLOR_ENABLED = True
    # print(f"{Fore.GREEN}Colorama loaded successfully. Colored output enabled.{Style.RESET_ALL}") # Optional: Keep this line if you like the confirmation
except ImportError:
    print("Warning: 'colorama' not found. Install it for colored output (`pip install colorama`). Output will be monochrome.")
    class DummyColor:
        def __getattr__(self, name): return ""
    Fore = DummyColor(); Style = DummyColor()
    COLOR_ENABLED = False


# Import constants module for path definitions
try:
    # Try importing from the package structure first
    from .utils import constants
except ImportError:
    try:
        # Try importing directly if run as a script
        from utils import constants
    except ImportError:
        # Last resort: try absolute import
        try:
            import sys
            sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            from youtube_shorts.utils import constants
        except ImportError:
            print("ERROR: Could not import constants module. Path-related functionality may not work correctly.")

# Configuration constants
EXCEL_FILE_PATH = constants.EXCEL_FILE_PATH
UPLOADED_SHEET_NAME = constants.UPLOADED_SHEET_NAME
CLIENT_SECRETS_FILE = constants.CLIENT_SECRETS_FILE
TOKEN_FILE = constants.TOKEN_FILE
ERROR_LOG_FILE = constants.PERFORMANCE_TRACKER_LOG_FILE
SCOPES = constants.SCOPES[0:1]  # Just use the readonly scope from constants.SCOPES

# --- Logging helper functions ---
def sanitize_message(message: str) -> str:
    """Redacts potentially sensitive information from log messages."""
    import re
    # Patterns for API keys, secrets, URLs with keys, file paths with sensitive extensions
    patterns = [
        (r'AIza[0-9A-Za-z\-_]{35}', 'API_KEY_REDACTED'),
        (r'(["\'])?(api[_-]?k[e]y|t[o]ken|s[e]cret|p[a]ssword|a[u]th|cr[e]dential)["\']?\s*[:=]\s*["\']?([^"\',\s]{8,})["\']?', r'\1\2\3=REDACTED'),
        (r'(https?://[^\s]+[?&][^\s]*(?:k[e]y|t[o]ken|s[e]cret|p[a]ssword|a[u]th)=[^\s&"]+)', r'URL_WITH_SENSITIVE_PARAMS_REDACTED'),
        (r'([\w\-]+\.)(k[e]y|p[e]m|c[e]rt|p12|pfx|p[a]ssword|t[o]ken|s[e]cret)', r'\1REDACTED'),
    ]
    sanitized = message
    for pattern, replacement in patterns:
        sanitized = re.sub(pattern, replacement, sanitized, flags=re.IGNORECASE)
    return sanitized

def log_error_to_file(message: str, include_traceback: bool = False):
    """Logs a detailed error message to the error log file."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sanitized_message = sanitize_message(message) # Sanitize before writing
    full_message = f"[{timestamp}] {sanitized_message}\n"
    if include_traceback:
        try:
            exc_info = traceback.format_exc()
            if exc_info and exc_info.strip() != 'NoneType: None':
                sanitized_traceback = sanitize_message(exc_info) # Sanitize traceback too
                full_message += sanitized_traceback + "\n"
        except Exception as e:
            # Log the error but continue
            full_message += f"[Error getting traceback: {e}]\n"

    # Ensure the logs directory exists
    os.makedirs(os.path.dirname(ERROR_LOG_FILE), exist_ok=True)

    try:
        with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f: f.write(full_message)
    except (IOError, PermissionError) as e:
        print(f"CRITICAL: Failed to write to error log file '{ERROR_LOG_FILE}': {e}")
    except Exception as e:
        print(f"CRITICAL: Unexpected error writing to error log file '{ERROR_LOG_FILE}': {e}")

# --- Print Helper Functions (using Colorama if available) ---
def print_section_header(title: str): print(f"\n{Style.BRIGHT}{Fore.CYAN}--- {title} ---{Style.RESET_ALL}")
# Redefine print helpers to use log_error_to_file correctly
def print_info(msg: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.DIM}{Fore.BLUE}i INFO:{Style.RESET_ALL} {msg}")
def print_success(msg: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.GREEN}OK SUCCESS:{Style.RESET_ALL} {Fore.GREEN}{msg}{Style.RESET_ALL}")
def print_warning(msg: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.YELLOW}WARN WARNING:{Style.RESET_ALL} {Fore.YELLOW}{msg}{Style.RESET_ALL}")
def print_error(msg: str, indent: int = 0, log_to_file: bool = True, include_traceback: bool = False):
    prefix = "  " * indent
    print(f"{prefix}{Style.BRIGHT}{Fore.RED}ERR ERROR:{Style.RESET_ALL} {Fore.RED}{msg}{Style.RESET_ALL}")
    # Log the error using the dedicated function
    if log_to_file: log_error_to_file(f"ERROR: {msg}", include_traceback=include_traceback)
# --- End Print Helper Functions ---

def get_authenticated_service():
    """Handles OAuth 2.0 authentication and returns the YouTube API service."""
    if not GOOGLE_API_AVAILABLE: print_error("Google API libraries not available."); return None
    creds = None

    # Ensure the data directory exists
    os.makedirs(os.path.dirname(TOKEN_FILE), exist_ok=True)

    # Load cached token
    if os.path.exists(TOKEN_FILE):
        try:
            with open(TOKEN_FILE, 'rb') as token: creds = pickle.load(token)
            # Check if credentials were loaded successfully (basic check)
            if not hasattr(creds, 'valid'):
                print_warning(f"Loaded token file '{TOKEN_FILE}' seems corrupted or not credentials. Re-authenticating.")
                creds = None # Force re-auth
            else:
                print_success("Cached credentials loaded.")
        except (EOFError, pickle.UnpicklingError, TypeError, AttributeError, ValueError) as e:
            print_warning(f"Failed to load cached credentials: {e}. Will re-authenticate.")
            creds = None # Force re-auth
        except Exception as e:
            print_warning(f"Unexpected error loading cached credentials: {e}. Will re-authenticate.")
            creds = None # Force re-auth

    # Refresh or get new credentials
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print_info("Cached credentials expired. Attempting to refresh...")
            try:
                creds.refresh(Request())
                print_success("Credentials refreshed successfully.")
            except Exception as e:
                print_warning(f"Failed to refresh credentials: {e}. Will perform new auth.")
                creds = None # Reset if refresh fails
        else:
            print_info("No valid cached credentials. Starting new authentication flow.")
            if not os.path.exists(CLIENT_SECRETS_FILE):
                 print_error(f"FATAL: Client secrets file not found: {CLIENT_SECRETS_FILE}")
                 log_error_to_file(f"FATAL: Client secrets file not found: {CLIENT_SECRETS_FILE}")
                 return None
            try:
                flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
                creds = flow.run_local_server(port=0)
                print_success("Authentication flow completed.")
            except Exception as e:
                print_error(f"Auth flow error: {e}", include_traceback=True)
                log_error_to_file(f"Authentication flow failed: {e}", include_traceback=True)
                return None
        # Save credentials if newly obtained or refreshed
        if creds and creds.valid:
            try:
                # Ensure the directory exists before saving
                os.makedirs(os.path.dirname(TOKEN_FILE), exist_ok=True)
                with open(TOKEN_FILE, 'wb') as token: pickle.dump(creds, token)
                print_success(f"Credentials saved to: {TOKEN_FILE}")
            except Exception as e: print_warning(f"Failed to save credentials: {e}")

    # Build and return service
    if creds and creds.valid:
        try:
            service = build('youtube', 'v3', credentials=creds)
            print_success("YouTube Data API service built.")
            return service
        except Exception as e:
            print_error(f"API service build error: {e}", include_traceback=True)
            log_error_to_file(f"Failed to build YouTube API service: {e}", include_traceback=True)
            return None
    else:
        print_error("Authentication failed.")
        return None


def get_video_stats(service, video_id: str) -> Optional[Dict[str, int]]:
    """Fetches view, like, and comment counts for a single video ID."""
    if not service or not video_id:
        return None
    try:
        response = service.videos().list(
            part="statistics",
            id=video_id
        ).execute()
        if response and response.get('items'):
            stats = response['items'][0]['statistics']
            # Use .get with default 0 and handle potential non-integer values
            views = 0; likes = 0; comments = 0
            try: views = int(stats.get('viewCount', 0))
            except (ValueError, TypeError): pass
            try: likes = int(stats.get('likeCount', 0))
            except (ValueError, TypeError): pass
            try: comments = int(stats.get('commentCount', 0))
            except (ValueError, TypeError): pass
            return {'viewCount': views, 'likeCount': likes, 'commentCount': comments}
        else:
            # Don't print warning for every possibly deleted/private video, just log it
            log_info(f"Video ID {video_id} not found or no statistics returned by API.")
            return None # Return None to indicate not found or no stats
    except HttpError as e:
        print_error(f"API error fetching stats for {video_id}: {e}")
        log_error_to_file(f"API HttpError fetching stats for {video_id}: {e}")
        return None # Indicate error
    except Exception as e:
        print_error(f"Unexpected error fetching stats for {video_id}: {e}")
        log_error_to_file(f"Unexpected error fetching stats for {video_id}: {e}", include_traceback=True)
        return None # Indicate error


def update_excel_with_stats(excel_path: str, sheet_name: str, stats_data: Dict[str, Dict]) -> bool:
    """Updates the Excel file with fetched statistics using excel_utils if available."""
    # Use excel_utils if available, otherwise fall back to direct openpyxl
    if EXCEL_UTILS_AVAILABLE:
        try:
            log_info(f"Updating Excel using excel_utils for file: {excel_path}")
            # Create backup before modifying
            backup_path = create_excel_backup(excel_path)
            if backup_path: print_success(f"Created backup at: {backup_path}", indent=1)
            else: print_warning("Could not create backup before updating stats", indent=1)

            # Load workbook safely
            wb = load_workbook_safely(excel_path)
            if wb is None: print_error(f"Failed to load workbook: {excel_path}"); return False

            if sheet_name not in wb.sheetnames: print_error(f"Sheet '{sheet_name}' not found."); wb.close(); return False
            sheet = wb[sheet_name]

            # Define target column names (lowercase for flexible matching)
            column_names = {
                'id': ['youtube video id'], # Must find this
                'views': ['views (yt)'], # Must find or create
                'likes': ['likes (yt)'], # Must find or create
                'comments': ['comments (yt)'], # Must find or create
                'last_updated': ['last updated'] # Must find or create
            }
            # Find column indices using excel_utils (case-insensitive)
            columns = find_column_indices(sheet, column_names)

            # Check if ID column was found
            if 'id' not in columns or columns['id'] is None:
                print_error("'YouTube Video ID' column not found.")
                wb.close()
                return False

            # Add missing columns if needed
            required_columns = {
                'views': 'Views (YT)',
                'likes': 'Likes (YT)',
                'comments': 'Comments (YT)',
                'last_updated': 'Last Updated'
            }
            next_col = sheet.max_column + 1
            headers_updated = False
            for key, header_text in required_columns.items():
                if key not in columns or columns[key] is None:
                    print_info(f"Adding missing header: '{header_text}'", indent=1)
                    sheet.cell(row=1, column=next_col, value=header_text)
                    columns[key] = next_col
                    next_col += 1
                    headers_updated = True

            # Verify all required stat columns are now available
            if not all(key in columns and columns[key] is not None for key in required_columns.keys()):
                print_error("Could not find or create necessary stat columns.")
                wb.close()
                return False

            # --- Update stats ---
            updated_count = 0
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S") # Timestamp for update

            for row_idx in range(2, sheet.max_row + 1): # Iterate through data rows
                try:
                    # Get video ID using excel_utils (handle potential non-string values)
                    youtube_id = get_cell_value(sheet, row_idx, columns['id'])
                    youtube_id = str(youtube_id).strip() if youtube_id else None

                    if youtube_id and youtube_id != "N/A" and youtube_id in stats_data:
                        stats = stats_data[youtube_id]
                        # Check if stats are valid (not None)
                        if stats is None:
                            log_warning(f"No valid stats fetched for {youtube_id}. Skipping update for this row.", indent=1)
                            continue

                        # Use cell coordinates for direct update
                        sheet.cell(row=row_idx, column=columns['views'], value=stats.get('viewCount'))
                        sheet.cell(row=row_idx, column=columns['likes'], value=stats.get('likeCount'))
                        sheet.cell(row=row_idx, column=columns['comments'], value=stats.get('commentCount'))
                        sheet.cell(row=row_idx, column=columns['last_updated'], value=now_str)
                        updated_count += 1
                        print_info(f"Updated stats for {youtube_id}: V={stats.get('viewCount')}, L={stats.get('likeCount')}, C={stats.get('commentCount')}", indent=1)
                except KeyError as ke: print_error(f"Missing key when processing row {row_idx} for ID '{youtube_id}': {ke}", indent=1); continue
                except Exception as e: print_error(f"Error updating row {row_idx} for ID '{youtube_id}': {e}", indent=1); continue

            # Save workbook only if updates were made
            if updated_count > 0:
                print_success(f"Updated stats for {updated_count} videos.")
                if save_workbook_safely(wb, excel_path): # Use safe save
                    print_success(f"Excel file saved: {excel_path}")
                    return True
                else:
                    print_error(f"Failed to save Excel file: {excel_path}")
                    return False
            else:
                print_info("No videos found in sheet requiring stat updates.")
                # Close workbook even if no changes were made (or headers added)
                if headers_updated:
                     if save_workbook_safely(wb, excel_path):
                         print_success(f"Excel file saved (headers added): {excel_path}")
                         return False # Technically no *data* updates, but file changed
                     else:
                         print_error(f"Failed to save Excel file after adding headers: {excel_path}")
                         return False
                else:
                     try: wb.close();
                     except: pass
                     return False

        except Exception as e:
            print_error(f"Unexpected error updating Excel with excel_utils: {e}", include_traceback=True)
            log_error_to_file(f"Unexpected error updating Excel with excel_utils: {e}", include_traceback=True)
            if 'wb' in locals() and wb:
                try: wb.close();
                except: pass
            return False
    else:
        # Fall back to original openpyxl implementation if excel_utils is not available
        try:
            wb = load_workbook(excel_path)
            if sheet_name not in wb.sheetnames: print_error(f"Sheet '{sheet_name}' not found."); return False
            sheet = wb[sheet_name]
            header = [str(cell.value).lower().strip() if cell.value is not None else '' for cell in sheet[1]] # Lowercase header

            # --- Find columns (case-insensitive) ---
            try:
                id_col_idx = header.index('youtube video id') + 1 if 'youtube video id' in header else None
                views_col_idx = header.index('views (yt)') + 1 if 'views (yt)' in header else None
                likes_col_idx = header.index('likes (yt)') + 1 if 'likes (yt)' in header else None
                comments_col_idx = header.index('comments (yt)') + 1 if 'comments (yt)' in header else None
                last_updated_col_idx = header.index('last updated') + 1 if 'last updated' in header else None

                if id_col_idx is None:
                    print_error("'YouTube Video ID' column not found.")
                    wb.close()
                    return False

                # Add missing headers if needed (less robust than excel_utils)
                needs_save = False
                if views_col_idx is None:
                    views_col_idx = len(header) + 1
                    sheet.cell(row=1, column=views_col_idx, value="Views (YT)")
                    needs_save = True

                if likes_col_idx is None:
                    likes_col_idx = len(header) + (2 if views_col_idx > len(header) else 1)
                    sheet.cell(row=1, column=likes_col_idx, value="Likes (YT)")
                    needs_save = True

                if comments_col_idx is None:
                    comments_col_idx = len(header) + (3 if likes_col_idx > len(header) else (2 if views_col_idx > len(header) else 1))
                    sheet.cell(row=1, column=comments_col_idx, value="Comments (YT)")
                    needs_save = True

                if last_updated_col_idx is None:
                    last_updated_col_idx = len(header) + (4 if comments_col_idx > len(header) else (3 if likes_col_idx > len(header) else (2 if views_col_idx > len(header) else 1)))
                    sheet.cell(row=1, column=last_updated_col_idx, value="Last Updated")
                    needs_save = True

            except ValueError as e:
                print_error(f"Error finding required columns: {e}")
                wb.close()
                return False
            except Exception as e:
                print_error(f"Unexpected error finding columns: {e}", include_traceback=True)
                wb.close()
                return False
            # --- End Find Columns ---

            updated_count = 0
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for row_idx in range(2, sheet.max_row + 1):
                try:
                    video_id_cell = sheet.cell(row=row_idx, column=id_col_idx)
                    youtube_id = str(video_id_cell.value).strip() if video_id_cell.value else None

                    if youtube_id and youtube_id != "N/A" and youtube_id in stats_data:
                        stats = stats_data[youtube_id]
                        if stats is None: # Skip if API call failed for this ID
                            log_warning(f"No stats fetched for {youtube_id}. Skipping row {row_idx}.", indent=1)
                            continue
                        # Update values directly
                        sheet.cell(row=row_idx, column=views_col_idx, value=stats.get('viewCount'))
                        sheet.cell(row=row_idx, column=likes_col_idx, value=stats.get('likeCount'))
                        sheet.cell(row=row_idx, column=comments_col_idx, value=stats.get('commentCount'))
                        sheet.cell(row=row_idx, column=last_updated_col_idx, value=now_str)
                        updated_count += 1
                        print_info(f"Updated stats for {youtube_id}: V={stats.get('viewCount')}, L={stats.get('likeCount')}, C={stats.get('commentCount')}", indent=1)
                except Exception as e: print_error(f"Error updating row {row_idx} for ID '{youtube_id}': {e}", indent=1); continue

            if updated_count > 0 or needs_save:
                print_success(f"Updated stats for {updated_count} videos.") if updated_count > 0 else print_info("Saving Excel file due to header changes.")
                try:
                    wb.save(excel_path)
                    print_success(f"Excel file saved: {excel_path}")
                    return True
                except PermissionError: print_error(f"PermissionError saving '{excel_path}'. Is it open?"); wb.close(); return False
                except Exception as e: print_error(f"Error saving Excel: {e}"); wb.close(); return False
            else:
                print_info("No videos found in sheet requiring stat updates.")
                wb.close()
                return False

        except FileNotFoundError: print_error(f"Excel file not found: {excel_path}"); return False
        except Exception as e: print_error(f"Unexpected error updating Excel: {e}", include_traceback=True); return False

# --- Main Function ---
def main():
    """Main function to track performance."""
    print_section_header("Starting YouTube Performance Tracker")

    if not GOOGLE_API_AVAILABLE:
        print_error("Google API libraries not installed. Please install required packages:")
        print_info("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
        return

    service = get_authenticated_service()
    if not service:
        print_error("Could not authenticate with YouTube Data API. Exiting.")
        return

    print_info(f"Loading Excel file: {EXCEL_FILE_PATH}")
    videos_to_fetch: List[str] = [] # List of YouTube IDs to fetch stats for

    # Initialize variables
    wb = None
    sheet = None
    id_col_idx = None
    schedule_time_col_idx = None
    use_excel_utils_local = EXCEL_UTILS_AVAILABLE # Local flag for this run

    try:
        # Use excel_utils if available for loading
        if use_excel_utils_local:
            try:
                wb = load_workbook_safely(EXCEL_FILE_PATH, read_only=True, data_only=True)
                if wb is None: print_error(f"Failed to load workbook: {EXCEL_FILE_PATH}. Exiting."); return
                if UPLOADED_SHEET_NAME not in wb.sheetnames: print_error(f"Sheet '{UPLOADED_SHEET_NAME}' not found in '{EXCEL_FILE_PATH}'. Exiting."); wb.close(); return
                sheet = wb[UPLOADED_SHEET_NAME]
                print_info(f"Successfully loaded sheet '{UPLOADED_SHEET_NAME}' using excel_utils")

                # Find column indices using excel_utils
                column_names = {
                    'id': ['youtube video id'], # Case-insensitive by default
                    'schedule_time': ['schedule time', 'upload timestamp', 'upload time', 'upload date']
                }
                columns = find_column_indices(sheet, column_names)
                id_col_idx = columns.get('id')
                schedule_time_col_idx = columns.get('schedule_time')

                if id_col_idx is None: print_error("'YouTube Video ID' column not found. Cannot fetch stats."); wb.close(); return

            except Exception as e:
                print_error(f"Error using excel_utils to load workbook: {e}", include_traceback=True)
                use_excel_utils_local = False # Fallback to direct openpyxl

        # Fallback or primary loading using openpyxl
        if not use_excel_utils_local:
            wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
            if UPLOADED_SHEET_NAME not in wb.sheetnames: print_error(f"Sheet '{UPLOADED_SHEET_NAME}' not found. Exiting."); wb.close(); return
            sheet = wb[UPLOADED_SHEET_NAME]
            header = [str(cell.value).lower().strip() if cell.value else '' for cell in sheet[1]] # Lowercase header
            print_info(f"Loaded sheet '{UPLOADED_SHEET_NAME}'. Header: {header}")

            # Find columns (case-insensitive)
            id_col_idx = header.index('youtube video id') + 1 if 'youtube video id' in header else None
            schedule_time_col_idx = None
            possible_time_cols = ['schedule time', 'upload timestamp', 'upload time', 'upload date']
            for col_name in possible_time_cols:
                if col_name in header: schedule_time_col_idx = header.index(col_name) + 1; break

            if id_col_idx is None: print_error("'YouTube Video ID' column not found. Cannot fetch stats."); wb.close(); return

        # Handle case where schedule time column is not found
        if schedule_time_col_idx is None:
            print_warning("'Schedule Time' (or equivalent) column not found. Cannot filter by date, will check ALL videos.")

        # --- Calculate the date 7 days ago ---
        now = datetime.now()
        seven_days_ago = now - timedelta(days=7)
        print_info(f"Checking for videos scheduled on or after: {seven_days_ago.strftime('%Y-%m-%d')}")

        print_info("Scanning for recent videos to update stats...")
        skipped_old = 0

        # Iterate through rows to find relevant video IDs
        for row_idx in range(2, sheet.max_row + 1):
            youtube_id = None
            reference_datetime = None # The date/time used for filtering

            try:
                # Get YouTube ID (using helper or direct access)
                if use_excel_utils_local: youtube_id = get_cell_value(sheet, row_idx, id_col_idx)
                else: youtube_id = sheet.cell(row=row_idx, column=id_col_idx).value
                youtube_id = str(youtube_id).strip() if youtube_id else None

                if not youtube_id or youtube_id == "N/A": continue

                # --- Attempt to Get & Parse "Schedule Time" for Filtering (only if column exists) ---
                should_fetch = True # Assume we should fetch unless filtered out
                if schedule_time_col_idx:
                    # Get the value
                    if use_excel_utils_local: schedule_time_value = get_cell_value(sheet, row_idx, schedule_time_col_idx)
                    else: schedule_time_value = sheet.cell(row=row_idx, column=schedule_time_col_idx).value

                    # Try to parse the value
                    if schedule_time_value and str(schedule_time_value).strip().upper() not in ["N/A", "NA", "NONE", "", "NULL", "UNDEFINED"]:
                        try:
                            # --- Use the most robust parsing available ---
                            if use_excel_utils_local and EXCEL_UTILS_AVAILABLE:
                                reference_datetime = parse_date_value(schedule_time_value) # excel_utils' robust parser
                            elif DATEUTIL_AVAILABLE:
                                # Use dateutil parser as fallback
                                if isinstance(schedule_time_value, datetime): reference_datetime = schedule_time_value
                                else: reference_datetime = date_parser.parse(str(schedule_time_value), fuzzy=True) # Use fuzzy for flexibility
                            else:
                                # Basic parsing if no utils available
                                if isinstance(schedule_time_value, datetime): reference_datetime = schedule_time_value
                                elif isinstance(schedule_time_value, float): # Basic Excel number check
                                     reference_datetime = datetime.fromtimestamp(time.mktime(time.gmtime((schedule_time_value - 25569) * 86400.0)))
                                else: # Try basic string format
                                     reference_datetime = datetime.strptime(str(schedule_time_value), "%Y-%m-%d %H:%M:%S")

                            # Apply the date filter if parsing succeeded
                            if reference_datetime:
                                if reference_datetime < seven_days_ago:
                                    skipped_old += 1
                                    should_fetch = False # Don't fetch this one
                            else:
                                # Parsing failed, but value exists - include it?
                                print_warning(f"Row {row_idx} (ID: {youtube_id}) - Could not parse Schedule Time '{schedule_time_value}'. Including for stats check.", indent=1)

                        except (ValueError, TypeError) as e:
                            # Parsing failed for this row's date
                            print_warning(f"Row {row_idx} (ID: {youtube_id}) - Could not parse Schedule Time '{schedule_time_value}': {e}. Including for stats check.", indent=1)
                    else:
                        # Schedule time is missing or explicitly N/A - include it
                        print_info(f"Row {row_idx} (ID: {youtube_id}) - Missing or N/A schedule time. Including for stats check.", indent=1)
                # --- End Schedule Time Parsing and Filtering ---

                # Add to fetch list if not filtered out
                if should_fetch and youtube_id not in videos_to_fetch:
                    videos_to_fetch.append(youtube_id)

            except Exception as row_err:
                print_error(f"Error processing row {row_idx}: {row_err}", indent=1, include_traceback=True)
                log_error_to_file(f"Error processing Excel row {row_idx}: {row_err}", include_traceback=True)
                continue
        # End of row iteration

        wb.close() # Close the read-only workbook

        if schedule_time_col_idx and skipped_old > 0: print_info(f"Skipped {skipped_old} videos scheduled older than 7 days.")
        elif not schedule_time_col_idx: print_warning("Could not filter by date (missing 'Schedule Time' column).")

    except FileNotFoundError: print_error(f"Excel file not found at: {EXCEL_FILE_PATH}. Exiting."); return
    except Exception as e: print_error(f"Error reading Excel file for IDs: {e}", include_traceback=True); log_error_to_file(f"Error reading Excel file for IDs: {e}", include_traceback=True); return

    if not videos_to_fetch:
        print_info("No videos found scheduled within the last 7 days (or requiring update) requiring stat updates.")
        return

    print_info(f"Found {len(videos_to_fetch)} videos scheduled within the last 7 days (or requiring update) to fetch/update stats for.")

    # --- Fetch stats in batches ---
    batch_size = 50 # YouTube API limit per request
    all_fetched_stats: Dict[str, Optional[Dict[str, int]]] = {} # Store fetched stats, including None for errors/not found

    for i in range(0, len(videos_to_fetch), batch_size):
        batch_ids = videos_to_fetch[i:i + batch_size]
        print_info(f"Fetching stats for batch {i//batch_size + 1}: {len(batch_ids)} videos.")

        try:
            response = service.videos().list(
                part="statistics",
                id=",".join(batch_ids) # Comma-separated IDs
            ).execute()

            found_ids_in_response = set()
            if response and response.get('items'):
                for item in response['items']:
                    video_id = item.get('id')
                    stats = item.get('statistics')
                    if video_id and stats:
                        # Use .get with default 0 and handle potential non-integer values
                        views = 0; likes = 0; comments = 0
                        try: views = int(stats.get('viewCount', 0))
                        except (ValueError, TypeError): pass
                        try: likes = int(stats.get('likeCount', 0))
                        except (ValueError, TypeError): pass
                        try: comments = int(stats.get('commentCount', 0))
                        except (ValueError, TypeError): pass

                        all_fetched_stats[video_id] = {'viewCount': views, 'likeCount': likes, 'commentCount': comments}
                        found_ids_in_response.add(video_id)

                print_success(f"Successfully fetched stats for {len(response['items'])} videos in batch.")
            elif response and 'items' in response and not response['items']:
                print_warning(f"API returned empty items list for batch {i//batch_size + 1}. IDs might be invalid/private.")
            else:
                print_error(f"API call for batch {i//batch_size + 1} failed or returned unexpected format.")

            # Mark IDs in the batch that were *not* found in the response as None (likely deleted/private)
            for batch_id in batch_ids:
                if batch_id not in found_ids_in_response:
                    all_fetched_stats[batch_id] = None # Mark as not found

        except HttpError as e:
            print_error(f"API error fetching batch {i//batch_size + 1}: {e}")
            log_error_to_file(f"API HttpError fetching batch {i//batch_size + 1}: {e}")
            # Mark all IDs in this batch as errored (None)
            for batch_id in batch_ids: all_fetched_stats[batch_id] = None
        except Exception as e:
            print_error(f"Unexpected error fetching batch {i//batch_size + 1}: {e}")
            log_error_to_file(f"Unexpected error fetching batch {i//batch_size + 1}: {e}", include_traceback=True)
            # Mark all IDs in this batch as errored (None)
            for batch_id in batch_ids: all_fetched_stats[batch_id] = None

        # Add a small delay between batches to be kind to the API
        if i + batch_size < len(videos_to_fetch):
            time.sleep(1)

    # --- Update Excel with Fetched Stats ---
    successful_fetches = sum(1 for stats in all_fetched_stats.values() if stats is not None)
    if successful_fetches > 0:
        print_info(f"Total stats fetched for {successful_fetches} unique videos.")
        update_excel_with_stats(EXCEL_FILE_PATH, UPLOADED_SHEET_NAME, all_fetched_stats)
    else:
        print_warning("No stats were successfully fetched for the recent videos identified.")

if __name__ == "__main__":
    main()

# --- END OF FILE ---