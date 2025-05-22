# -*- coding: utf-8 -*-
import os
import json
import yt_dlp
import google.generativeai as genai
import re
import time
import concurrent.futures
import math # Needed for scoring
from datetime import datetime, timedelta # Keep timedelta for cache cleanup
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import traceback
import random # Needed for weighted selection if we adapt quota later
from typing import List, Dict, Set, Any, Optional, Tuple
import collections # For metadata validation
# For YouTube Data API
try:
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    YOUTUBE_API_AVAILABLE = True
except ImportError:
    print("Warning: Google API libraries not found. Install with:")
    print("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
    YOUTUBE_API_AVAILABLE = False

# --- Colorama Setup ---
try:
    import colorama
    from colorama import Fore, Style, Back, init
    init(autoreset=True)
    COLOR_ENABLED = True
except ImportError:
    print("Warning: 'colorama' not found. Install (`pip install colorama`). Monochrome output.")
    class DummyColor:
        def __getattr__(self, name): return ""
    Fore = DummyColor(); Style = DummyColor(); Back = DummyColor()
    COLOR_ENABLED = False
# --- End Colorama Setup ---

# --- NEW: Import constants from the new location ---
try:
    from .utils import constants # If downloader_channel.py is a module in youtube_shorts
except ImportError:
    # Fallback if run as script directly from youtube_shorts folder
    try:
        from utils import constants
    except ImportError:
        # Absolute fallback if utils isn't directly importable (e.g. running from project root)
        # This might indicate a PYTHONPATH issue or that the script isn't being run as part of the package.
        # For now, assume it can find 'utils' if it's in the same directory level.
        print("CRITICAL: Could not import constants.py. Ensure it's in utils/ and runnable.")
        # Define a minimal set of constants here to avoid crashing, but this is not ideal.
        class MinimalConstants:
            CONFIG_FILE_PATH = "config/config.txt" # Will be relative to where script is run
            CHANNELS_FILE_PATH = "config/channels.txt"
            SHORTS_DOWNLOADS_DIR = "output/shorts_downloads"
            SHORTS_METADATA_DIR = "output/shorts_metadata"
            EXCEL_FILE_PATH = "data/shorts_data.xlsx"
            DOWNLOADER_CHANNEL_LOG_FILE = "logs/downloader_channel_log.txt"
            CHANNEL_PROCESSED_IDS_CACHE = "data/channel_processed_ids_cache.json"
            CHANNEL_LISTING_CACHE = "data/channel_listing_cache.json"
            FFMPEG_PATH = "drivers/ffmpeg.exe"
            SEO_METADATA_PROMPT_FILE = "config/seo_metadata_prompt.txt" # Corrected name
            METADATA_METRICS_FILE = "data/metadata_metrics_channel.json" # Specific
            PERFORMANCE_METRICS_FILE = "data/performance_metrics_channel.json" # Specific
            TUNING_SUGGESTIONS_LOG_FILE = "logs/tuning_suggestions_channel.log" # Specific
            UPLOAD_CORRELATION_CACHE = "data/upload_correlation_cache.json"
            CHANNEL_PERFORMANCE_CACHE_FILE = "data/channel_performance_cache.json" # New cache for channel scores
            SUGGESTED_CHANNELS_LOG_FILE = "logs/suggested_channels.log"
            KEYWORDS_CACHE_FILE = "data/generated_keywords_cache.json"
            PLAYLIST_DATA_CACHE_FILE = "data/playlists_data_cache.json"
            DOWNLOADED_SHEET_NAME = "Downloaded"
            UPLOADED_SHEET_NAME = "Uploaded"
            EXPECTED_DOWNLOADED_HEADERS = ["Video Index", "Optimized Title", "Downloaded Date", "Views", "Uploader", "Original Title"]
            EXPECTED_UPLOADED_HEADERS = ["Video Index", "Optimized Title", "YouTube Video ID", "Upload Timestamp", "Scheduled Time", "Publish Status", "Views (YT)", "Likes (YT)", "Comments (YT)"]
            UPLOADED_YT_ID_COL_NAME = 'youtube video id'
            UPLOADED_VIEWS_COL_NAME = 'views (yt)'
            UPLOADED_LIKES_COL_NAME = 'likes (yt)'
            UPLOADED_COMMENTS_COL_NAME = 'comments (yt)'
            DOWNLOADED_ORIGINAL_TITLE_COL_IDX = 6
            DOWNLOADED_UPLOADER_COL_IDX = 5
            YT_PLAYLIST_FETCH_LIMIT = 50
            MAX_SHORT_DURATION = 61
            METADATA_TIMEOUT_SECONDS = 30
            DEFAULT_UPLOADER_NAME = "Unknown Uploader"
            METADATA_ERROR_THRESHOLD = 0.15
            METADATA_TIMEOUT_THRESHOLD = 0.10
            VALIDATION_WARNING_THRESHOLD = 0.20
            MAX_ERROR_SAMPLES = 5
            MIN_RUNS_BEFORE_TUNING = 3
            BASE_WEIGHT_FOR_CHANNELS = 1
            MAX_RUNS_TO_KEEP = 20
            NICHE_FILE_PATH = "config/niche.txt" # Added for consistency
            # For excel_utils fallback
            BACKUP_FOLDER = "backups/excel"
        constants = MinimalConstants()
        print("WARNING: Using minimal fallback constants due to import error.")
# --- End NEW Import ---

# --- Runtime Configuration Constants ---
# These are runtime settings that are not path-related
# and are still needed locally in this file

# Excel Sheet Names / Headers / Indices
DOWNLOADED_SHEET_NAME = constants.DOWNLOADED_SHEET_NAME
UPLOADED_SHEET_NAME = constants.UPLOADED_SHEET_NAME
EXPECTED_DOWNLOADED_HEADERS = constants.EXPECTED_DOWNLOADED_HEADERS
EXPECTED_UPLOADED_HEADERS = constants.EXPECTED_UPLOADED_HEADERS

# Column indices for reading performance data
UPLOADED_YT_ID_COL_NAME = constants.UPLOADED_YT_ID_COL_NAME
UPLOADED_VIEWS_COL_NAME = constants.UPLOADED_VIEWS_COL_NAME
UPLOADED_LIKES_COL_NAME = constants.UPLOADED_LIKES_COL_NAME
UPLOADED_COMMENTS_COL_NAME = constants.UPLOADED_COMMENTS_COL_NAME

# Column indices for reading previously downloaded
DOWNLOADED_ORIGINAL_TITLE_COL_IDX = constants.DOWNLOADED_ORIGINAL_TITLE_COL_IDX
DOWNLOADED_UPLOADER_COL_IDX = constants.DOWNLOADED_UPLOADER_COL_IDX

# Define these for backward compatibility
ORIGINAL_TITLE_COL_IDX = DOWNLOADED_ORIGINAL_TITLE_COL_IDX
UPLOADER_COL_IDX = DOWNLOADED_UPLOADER_COL_IDX

# yt-dlp Settings
YT_PLAYLIST_FETCH_LIMIT = constants.YT_PLAYLIST_FETCH_LIMIT
MAX_SHORT_DURATION = constants.MAX_SHORT_DURATION

# Gemini Settings
METADATA_TIMEOUT_SECONDS = constants.METADATA_TIMEOUT_SECONDS
DEFAULT_UPLOADER_NAME = constants.DEFAULT_UPLOADER_NAME

# Metadata improvement settings
METADATA_ERROR_THRESHOLD = constants.METADATA_ERROR_THRESHOLD
METADATA_TIMEOUT_THRESHOLD = constants.METADATA_TIMEOUT_THRESHOLD
VALIDATION_WARNING_THRESHOLD = constants.VALIDATION_WARNING_THRESHOLD
MAX_ERROR_SAMPLES = constants.MAX_ERROR_SAMPLES

# Parameter tuning settings
MIN_RUNS_BEFORE_TUNING = constants.MIN_RUNS_BEFORE_TUNING
BASE_WEIGHT_FOR_CHANNELS = constants.BASE_WEIGHT_FOR_CHANNELS
MAX_RUNS_TO_KEEP = constants.MAX_RUNS_TO_KEEP

# Other settings
MAX_TITLE_LENGTH = 100 # YouTube's recommended max title length
TARGET_TITLE_LENGTH = 90 # Target length before adding #Shorts
SHORTS_SUFFIX = " #Shorts"

# --- Global Path Definitions (now derived from constants) ---
# Using constants directly instead of building paths from script_directory
config_file_path = constants.CONFIG_FILE_PATH
channels_file_path = constants.CHANNELS_FILE_PATH
download_folder = constants.SHORTS_DOWNLOADS_DIR
metadata_folder = constants.SHORTS_METADATA_DIR
excel_file = constants.EXCEL_FILE_PATH
ERROR_LOG_FILE = constants.DOWNLOADER_CHANNEL_LOG_FILE # Use specific log
channel_processed_ids_cache_file = constants.CHANNEL_PROCESSED_IDS_CACHE
channel_listing_cache_file = constants.CHANNEL_LISTING_CACHE
ffmpeg_path = constants.FFMPEG_PATH # From constants

# --- Self-Improvement File Paths (now derived from constants) ---
seo_metadata_prompt_cache_path = constants.SEO_METADATA_PROMPT_FILE
metadata_metrics_file_path = os.path.join(constants.DATA_DIR, "metadata_metrics_channel.json") # Specific metrics
performance_metrics_file_path = os.path.join(constants.DATA_DIR, "performance_metrics_channel.json") # Specific
tuning_suggestions_file_path = os.path.join(constants.LOGS_DIR, "tuning_suggestions_channel.log") # Specific
upload_correlation_cache_path = constants.UPLOAD_CORRELATION_CACHE
channel_performance_cache_path = os.path.join(constants.DATA_DIR, "channel_performance_cache.json") # Specific
suggested_channels_file_path = constants.SUGGESTED_CHANNELS_LOG_FILE
keywords_cache_file_path = constants.GENERATED_KEYWORDS_CACHE_FILE
playlist_data_cache_path = constants.PLAYLIST_DATA_CACHE_FILE

# --- Global Cache for SEO Prompt ---
_current_seo_prompt_template = None

# --- Print Helper Functions ---
def print_info(msg, indent=0):
    prefix = "  " * indent
    print(f"{prefix}{Fore.BLUE}i INFO:{Style.RESET_ALL} {msg}")

def print_success(msg, indent=0):
    prefix = "  " * indent
    print(f"{prefix}{Fore.GREEN}OK SUCCESS:{Style.RESET_ALL} {msg}")

def print_warning(msg, indent=0):
    prefix = "  " * indent
    print(f"{prefix}{Fore.YELLOW}WARN WARNING:{Style.RESET_ALL} {msg}")

def print_error(msg, indent=0, include_traceback=False):
    prefix = "  " * indent
    print(f"{prefix}{Fore.RED}ERR ERROR:{Style.RESET_ALL} {msg}")
    if include_traceback:
        traceback.print_exc()
    log_error(f"ERROR: {msg}" + (f"\n{traceback.format_exc()}" if include_traceback else ""))

def print_fatal(msg, indent=0):
    prefix = "  " * indent
    full_msg = f"FATAL: {msg}"
    print(f"{prefix}{Back.RED}{Fore.WHITE}{Style.BRIGHT} {full_msg} {Style.RESET_ALL}")
    log_error(full_msg)
    exit(1)

def print_section_header(title):
    """Prints a formatted section header."""
    print(f"\n{Fore.CYAN}{Style.BRIGHT}--- {title} ---{Style.RESET_ALL}")
# --- End Print Helper Functions ---

# --- Logging Functions (Modified to use print_error internally for console) ---
def log_error(message: str):
    """Logs an error message to the download error log file."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    full_message = f"[{timestamp}] {message}\n"
    try:
        # --- Use constant for log file path ---
        with open(constants.DOWNLOADER_CHANNEL_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(full_message)
    except Exception as e:
        # Use direct print as colored print might fail if colorama failed
        print(f"CRITICAL: Failed write to log '{constants.DOWNLOADER_CHANNEL_LOG_FILE}': {e}\nOriginal: {full_message}")

def log_error_to_file(message: str, include_traceback: bool = False):
    """Logs a detailed error message to the error log file with optional traceback."""
    error_msg = message
    if include_traceback:
        error_msg += f"\n{traceback.format_exc()}"
    log_error(error_msg)

# --- Configuration Loading ---
config = {}
try:
    with open(config_file_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and "=" in line:
                key, value = line.split("=", 1)
                config[key.strip()] = value.strip()
except FileNotFoundError:
    print_fatal(f"Config file '{config_file_path}' not found.")
except Exception as e:
    print_fatal(f"Error reading config '{config_file_path}': {e}")

# --- Get Configurable Settings ---
API_KEY = config.get("API_KEY") or config.get("GEMINI_API_KEY") # Allow either name
if not API_KEY:
    print_fatal("API_KEY or GEMINI_API_KEY not found in 'config.txt'.")

_DEFAULT_MAX_DOWNLOADS = 24
try:
    max_downloads = int(config.get("MAX_DOWNLOADS", _DEFAULT_MAX_DOWNLOADS))
    if max_downloads <= 0:
        print_warning(f"MAX_DOWNLOADS invalid. Using default: {_DEFAULT_MAX_DOWNLOADS}")
        max_downloads = _DEFAULT_MAX_DOWNLOADS
except (ValueError, TypeError):
    print_warning(f"Invalid MAX_DOWNLOADS. Using default: {_DEFAULT_MAX_DOWNLOADS}")
    max_downloads = _DEFAULT_MAX_DOWNLOADS

# SEO Config
seo_channel_name = config.get("SEO_CHANNEL_NAME", "Our Awesome Channel")
seo_channel_topic = config.get("SEO_CHANNEL_TOPIC", "interesting videos")
seo_example_tags_raw = config.get("SEO_EXAMPLE_TAGS", "tag1, tag2, youtube, video, shorts")
seo_example_tags = [tag.strip() for tag in seo_example_tags_raw.split(',') if tag.strip()]
seo_example_hashtags_raw = config.get("SEO_EXAMPLE_HASHTAGS", "#shorts #video #youtube")
seo_example_hashtags = [ht.strip() for ht in seo_example_hashtags_raw.split() if ht.strip().startswith('#')]

# FFMPEG Path
ffmpeg_path_config = config.get("FFMPEG_PATH")
if ffmpeg_path_config and os.path.exists(ffmpeg_path_config):
    ffmpeg_path = ffmpeg_path_config
    print_info(f"Using FFmpeg path from config: {ffmpeg_path}")
elif not os.path.exists(ffmpeg_path):
    print_warning(f"FFmpeg not found at default path '{ffmpeg_path}'. Downloads requiring merging might fail.")
    print_warning("Ensure ffmpeg.exe is in the script directory or set FFMPEG_PATH in config.txt")
    ffmpeg_path = "ffmpeg" # Fallback to hoping it's in system PATH

# Other settings
MAX_TITLE_LENGTH = 100 # From keyword downloader
TARGET_TITLE_LENGTH = 90 # From keyword downloader
SHORTS_SUFFIX = " #Shorts" # From keyword downloader

print_success("Configuration loaded.")
print(f"Settings: Max Downloads={max_downloads}, SEO Channel='{seo_channel_name}', Topic='{seo_channel_topic}'")

# Configure Gemini API
try:
    genai.configure(api_key=API_KEY)
    print_success("Gemini API configured.")
except Exception as e:
    print_fatal(f"Failed to configure Gemini API: {e}")

# --- Function Definitions ---

# --- Performance/Metrics Functions ---
def load_performance_metrics():
    """Loads overall performance metrics from the JSON file."""
    default_metrics = {
        "runs": [], "total_shorts_found": 0, "total_suitable_shorts": 0,
        "total_downloads_attempted": 0, "total_successful_downloads": 0,
        "total_metadata_api_calls": 0, "total_metadata_errors": 0,
        "total_download_errors": 0, "channel_performance": {},
        "last_parameter_update": ""
    }
    try:
        if os.path.exists(performance_metrics_file_path):
            with open(performance_metrics_file_path, "r", encoding="utf-8") as f:
                metrics = json.load(f)
            for key, value in default_metrics.items():
                metrics.setdefault(key, value) # Ensure all keys exist
            return metrics
        else:
            return default_metrics
    except Exception as e:
        print_warning(f"Error loading performance metrics: {e}. Using default values.")
        return default_metrics

def save_performance_metrics(metrics):
    """Saves overall performance metrics to the JSON file."""
    try:
        with open(performance_metrics_file_path, "w", encoding="utf-8") as f:
            json.dump(metrics, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print_error(f"Error saving performance metrics: {e}", 1)

def load_metadata_metrics():
    """Loads metadata generation metrics from the JSON file."""
    default_metrics = {
        "total_api_calls": 0, "parse_failures": 0, "timeouts": 0,
        "empty_title_errors": 0, "empty_description_errors": 0, "empty_tags_errors": 0,
        # Validation warning metrics
        "validation_title_mismatches": 0,
        "validation_tag_list_errors": 0,
        "validation_keyword_stuffing": 0,
        "last_run_date": datetime.now().isoformat(), "error_samples": [],
        "total_api_calls_previous": 0, "total_errors_previous": 0
    }
    try:
        if os.path.exists(metadata_metrics_file_path):
            with open(metadata_metrics_file_path, "r", encoding="utf-8") as f:
                metrics = json.load(f)
            for key, value in default_metrics.items():
                metrics.setdefault(key, value)
            return metrics
        else:
            return default_metrics
    except Exception as e:
        print_warning(f"Error loading metadata metrics: {e}. Using default values.")
        return default_metrics

def save_metadata_metrics(metrics):
    """Saves metadata generation metrics to the JSON file."""
    try:
        metrics["last_run_date"] = datetime.now().isoformat()
        with open(metadata_metrics_file_path, "w", encoding="utf-8") as f:
            json.dump(metrics, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print_error(f"Error saving metadata metrics: {e}", 1)

def add_error_sample(metrics, error_type, error_details, video_title):
    """Adds an error sample to the metadata metrics."""
    if "error_samples" not in metrics:
        metrics["error_samples"] = []
    error_sample = {
        "type": error_type,
        "details": error_details,
        "video_title": video_title,
        "timestamp": datetime.now().isoformat()
    }
    metrics["error_samples"].append(error_sample)
    metrics["error_samples"] = metrics["error_samples"][-MAX_ERROR_SAMPLES:] # Keep only last N
# --- End Performance/Metrics Functions ---

# --- Correlation Cache Functions ---
def load_correlation_cache():
    """Loads the upload correlation cache that maps video index to YouTube ID."""
    default_cache = []
    if not os.path.exists(upload_correlation_cache_path):
        return default_cache
    try:
        with open(upload_correlation_cache_path, "r", encoding="utf-8") as f:
            content = f.read()
            if not content:
                return default_cache
            cache = json.loads(content)
            if not isinstance(cache, list):
                print_warning(f"Correlation cache file '{os.path.basename(upload_correlation_cache_path)}' invalid format.")
                return default_cache
            return cache
    except json.JSONDecodeError:
        print_error(f"Error decoding correlation cache '{os.path.basename(upload_correlation_cache_path)}'.")
        return default_cache
    except Exception as e:
        print_error(f"Error loading correlation cache: {e}", include_traceback=True)
        return default_cache

def save_correlation_cache(cache_data):
    """Saves the upload correlation cache."""
    try:
        with open(upload_correlation_cache_path, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print_error(f"Error saving correlation cache: {e}", include_traceback=True)

def cleanup_correlation_cache(days_to_keep=7):
    """Cleans up old entries from the correlation cache."""
    cache = load_correlation_cache()
    if not cache:
        return

    now = datetime.now()
    cutoff_date = now - timedelta(days=days_to_keep)
    original_count = len(cache)
    cleaned_cache = []
    removed_count = 0

    for entry in cache:
        try:
            added_time = datetime.fromisoformat(entry.get("added_timestamp"))
            if added_time >= cutoff_date:
                cleaned_cache.append(entry)
            else:
                removed_count += 1
        except (ValueError, TypeError, KeyError):
            cleaned_cache.append(entry)
            print_warning(f"Invalid timestamp in correlation cache entry: {entry.get('video_index', 'Unknown')}. Keeping.")

    if removed_count > 0:
        save_correlation_cache(cleaned_cache)
        print_info(f"Cleaned up {removed_count} old entries (>{days_to_keep} days) from correlation cache. {len(cleaned_cache)} remaining.")
    else:
        print_info("No old entries found in correlation cache to cleanup.")
# --- End Correlation Cache Functions ---

# --- Metadata Validation ---
def validate_generated_metadata(metadata, video_title, metadata_metrics=None):
    """Performs cross-validation checks on generated metadata and updates metrics."""
    warnings = []
    title = metadata.get("title", "")
    description = metadata.get("description", "")
    tags = metadata.get("tags", [])

    # Flags to track specific validation issues (to only increment metrics once per validation)
    title_mismatch_detected = False
    tag_list_error_detected = False
    stuffing_detected = False

    if not title: warnings.append("Generated title is empty.")
    if not description: warnings.append("Generated description is empty.")
    if not tags: warnings.append("Generated tags list is empty.")

    # 1. Title in Description (improved check)
    # Remove #Shorts for check, compare lowercase
    title_base = title.replace("#Shorts", "").strip().lower()

    # Check if title is at the beginning of the description (first 200 chars)
    # This is more lenient than requiring exact match
    desc_start = description[:200].lower()

    if title_base and len(title_base) > 5:  # Only check if title is meaningful
        # Try different approaches to find the title in the description
        title_words = set(re.findall(r'\b\w+\b', title_base))
        significant_words = [w for w in title_words if len(w) > 3 and w not in ['this', 'that', 'with', 'from', 'have', 'what', 'when', 'where', 'will', 'your']]

        # Check if at least 70% of significant words from title appear in the first 200 chars of description
        if len(significant_words) > 0:
            words_found = sum(1 for word in significant_words if word in desc_start)
            match_percentage = words_found / len(significant_words)

            if match_percentage < 0.7:  # Less than 70% of significant words found
                warnings.append(f"Generated title base ('{title_base[:50]}...') not found in description.")
                if metadata_metrics and not title_mismatch_detected:
                    metadata_metrics["validation_title_mismatches"] += 1
                    title_mismatch_detected = True

    # 2. Tags mentioned in Description (improved check)
    tags_in_desc_heading = "Tags Used in Video :-"
    tags_listed_incorrectly = False

    if tags_in_desc_heading.lower() in description.lower():
        try:
            # Find text after heading, split by comma/newline, strip whitespace
            desc_parts = description.lower().split(tags_in_desc_heading.lower(), 1)
            if len(desc_parts) > 1:
                # Take text after heading, stop at next potential heading or end
                tags_text_area = desc_parts[1].split("\n\n")[0].split("ignored hashtags :-")[0]

                # Improved tag extraction - handle asterisk prefixes and other formatting
                raw_tags = re.split(r'[\n,]', tags_text_area)
                tags_mentioned_in_desc = set()

                for tag in raw_tags:
                    # Clean up the tag - remove asterisks, hashes, and other formatting
                    cleaned_tag = re.sub(r'^\s*[\*\-•#]\s*', '', tag).strip().lower()
                    if cleaned_tag:
                        tags_mentioned_in_desc.add(cleaned_tag)

                # Clean up the generated tags for comparison
                tags_generated_lower = set()
                for tag in tags:
                    # Remove special characters and normalize
                    cleaned_tag = re.sub(r'[^\w\s]', '', tag.lower()).strip()
                    if cleaned_tag:
                        tags_generated_lower.add(cleaned_tag)

                # Compare the cleaned sets
                # Allow for some flexibility - check if at least 80% of tags are present
                if tags_generated_lower and tags_mentioned_in_desc:
                    common_tags = tags_generated_lower.intersection(tags_mentioned_in_desc)
                    match_percentage = len(common_tags) / len(tags_generated_lower)

                    if match_percentage < 0.8:  # Less than 80% of tags found
                        missing_in_desc = tags_generated_lower - tags_mentioned_in_desc
                        extra_in_desc = tags_mentioned_in_desc - tags_generated_lower

                        if missing_in_desc:
                            warnings.append(f"Tags listed in <tags> but missing under description heading: {missing_in_desc}")
                            tags_listed_incorrectly = True
                        if extra_in_desc:
                            warnings.append(f"Tags mentioned under description heading but not in <tags>: {extra_in_desc}")
                            tags_listed_incorrectly = True

        except Exception as e:
            warnings.append(f"Error parsing tags from description: {e}")
            tags_listed_incorrectly = True
    else:
        warnings.append(f"'{tags_in_desc_heading}' heading not found in description.")
        tags_listed_incorrectly = True

    if tags_listed_incorrectly and metadata_metrics and not tag_list_error_detected:
        metadata_metrics["validation_tag_list_errors"] += 1
        tag_list_error_detected = True

    # 3. Keyword Stuffing (Simple Heuristic)
    words_in_desc = description.lower().split()
    if tags and words_in_desc:
        max_tag_word_occurrences = 7 # Configurable threshold
        # Use Counter for more efficient word counting
        word_counts = collections.Counter(words_in_desc)

        for tag in tags:
            tag_words = tag.lower().split()
            for tag_word in tag_words:
                if len(tag_word) > 3: # Only check meaningful words
                    count = word_counts.get(tag_word, 0)
                    if count > max_tag_word_occurrences:
                        warnings.append(f"Potential keyword stuffing: Word '{tag_word}' (from tag '{tag}') appears {count} times in description.")
                        if metadata_metrics and not stuffing_detected:
                            metadata_metrics["validation_keyword_stuffing"] += 1
                            stuffing_detected = True
                        break # Only report once per tag

    if warnings:
        print_warning(f"Metadata validation warnings for '{video_title}':", 3)
        for warn in warnings:
            print_warning(f"  - {warn}", 4)
    else:
        print_success(f"Metadata validation passed for '{video_title}'.", 3)

# --- SEO Prompt Handling ---
def load_or_get_seo_prompt_template():
    """Loads or gets the SEO metadata prompt template."""
    global _current_seo_prompt_template
    if _current_seo_prompt_template:
        return _current_seo_prompt_template

    prompt_loaded = False
    if os.path.exists(seo_metadata_prompt_cache_path):
        try:
            with open(seo_metadata_prompt_cache_path, "r", encoding="utf-8") as f:
                _current_seo_prompt_template = f.read()
            # Check if the prompt has the required placeholders
            if "{video_topic}" in _current_seo_prompt_template and "{uploader_name}" in _current_seo_prompt_template and "{channel_name}" in _current_seo_prompt_template and "{channel_topic}" in _current_seo_prompt_template:
                print_info(f"Loaded SEO metadata prompt from cache: {os.path.basename(seo_metadata_prompt_cache_path)}")
                prompt_loaded = True
            else:
                print_warning(f"Cached SEO prompt invalid (missing placeholders). Using default.")
                _current_seo_prompt_template = None
        except Exception as e:
            print_warning(f"Error loading cached SEO prompt: {e}. Using default.")
            _current_seo_prompt_template = None

    if not prompt_loaded:
        print_info("Using default inline SEO metadata prompt.")
        # Use the prompt definition from the channel script, ensuring placeholders match
        _current_seo_prompt_template = f"""
        do not include any explanation or any other text. just give me the metadata in below format.
        only apply the below format. do not include any other text or explanation.
        Generate SEO-optimized metadata for a YouTube Shorts video in the following structured format:
        You are a YOUTUBE SEO EXPERT A GURU one in million. you have insight knowledge of youtube shorts.
        you know how the ranking algorithm works and how to get more views and engagement.
        you know how creator like mrbeast, tseries, and other top creators get more views and engagement.
        your master of youtube shorts. you have worked with big creator know all secrets of youtube shorts.
        you have worked in google youtube team and you know all secrets of youtube shorts.
        Our Channel Name is "{{channel_name}}" and we are a channel about {{channel_topic}}. <-- Use channel_name/topic placeholders
        include a copyright fair use disclaimer in the description.
        APPLY ALL OF THE ABOVE KNOWLEDGE AND SECRETS TO BELOW metadata.

        <metadata>
            <title>
                Create an engaging, fast-paced, and action-driven title (max 100 chars incl. #Shorts) with a high CTR based on the video topic: '{{video_topic}}'.
                Use keywords for '{{channel_topic}}'. Use relevant emojis (🔥, 💪, 👀), numbers, power words (BEST, HOT, ULTIMATE, SECRET, TRY THIS). Add "#Shorts" at the end.
            </title>
            <description>
                Write an SEO-optimized description (max 4500 chars):
                    * Start with the optimized video title.
                    * 2-3 sentence engaging summary about '{{video_topic}}' and '{{channel_topic}}', using keywords/LSI naturally.
                    * **Include credit: "Credit: {{uploader_name}}"** <-- Use uploader_name placeholder
                    * Include copyright disclaimer:
                      --------------【Copyright Disclaimer】-------------
                      All the videos, songs, images, and graphics used in the video belong to
                      their respective owners and I or this channel "{{channel_name}}" does not claim any right over them.
                      Copyright Disclaimer under section 107 of the Copyright Act of 1976, allowance is made for "fair use" for purposes such as criticism, comment, news reporting, teaching, scholarship, education and research. Fair use is a use permitted by copyright statute that might otherwise be infringing.
                    * After disclaimer, add 10-15 relevant hashtags (inspired by: {seo_example_hashtags_raw}). <-- Insert example hashtags from config
                    * Add heading "Tags Used in Video :-" and list all tags from <tags> section below, comma-separated.
                    * End with a Call to Action (e.g., "👍 Like & Subscribe to {{channel_name}}!").
                    * Add heading "Ignored Hashtags :-" followed by a diverse list of relevant hashtags.
            </description>
            <tags>
                Suggest 15-25 SEO-friendly tags (comma-separated, max 500 chars total).
                * Start with keywords for '{{video_topic}}'. Include tags for '{{channel_topic}}' and channel name '{{channel_name}}'.
                * Use mix of general/specific tags. Inspire from: {seo_example_tags_raw} <-- Insert example tags from config
            </tags>
        </metadata>

        **Video Topic**: {{video_topic}}
        """
    return _current_seo_prompt_template

def save_seo_prompt_template(prompt_text):
    """Saves the SEO metadata prompt template to the cache file."""
    global _current_seo_prompt_template
    try:
        with open(seo_metadata_prompt_cache_path, "w", encoding="utf-8") as f:
            f.write(prompt_text)
        _current_seo_prompt_template = prompt_text
        print_success(f"Saved updated SEO metadata prompt to cache: {os.path.basename(seo_metadata_prompt_cache_path)}")
    except Exception as e:
        print_error(f"Error saving updated SEO metadata prompt to cache: {e}")

def improve_metadata_prompt(error_metrics):
    """Uses Gemini to improve the SEO metadata prompt based on error patterns."""
    current_prompt = load_or_get_seo_prompt_template()
    if not current_prompt:
        print_error("Cannot improve prompt: Failed to load current SEO prompt.")
        return None

    total_calls = error_metrics.get("total_api_calls", 0)
    if total_calls == 0:
        print_warning("No API calls recorded yet. Cannot improve prompt.")
        return None

    error_summary = [f"Total API calls: {total_calls}"]
    # Calculate error rates for relevant metrics
    for error_type in ["parse_failures", "timeouts", "empty_description_errors", "empty_tags_errors"]:
        count = error_metrics.get(error_type, 0)
        rate = count / total_calls if total_calls > 0 else 0
        # Clarify metric name in summary
        metric_name = error_type.replace("_", " ").title()
        error_summary.append(f"{metric_name}: {count} ({rate:.1%})")

    # Add validation warning metrics
    validation_summary = []
    for warning_type in ["validation_title_mismatches", "validation_tag_list_errors", "validation_keyword_stuffing"]:
        count = error_metrics.get(warning_type, 0)
        rate = count / total_calls if total_calls > 0 else 0
        # Clarify metric name in summary
        metric_name = warning_type.replace("validation_", "").replace("_", " ").title()
        validation_summary.append(f"{metric_name}: {count} ({rate:.1%})")

    error_summary.append("\n=== Validation Warnings ===")
    error_summary.extend(validation_summary)

    error_samples = error_metrics.get("error_samples", [])
    if error_samples:
        error_summary.append("\nRecent error samples:")
        for i, sample in enumerate(error_samples[-5:], 1):
            error_summary.append(f"Sample {i}: {sample.get('type')} - {sample.get('details')}")
    error_summary_text = "\n".join(error_summary)

    meta_prompt = f"""
    Review the following prompt used to generate SEO-optimized YouTube Shorts metadata for a channel-based downloader:

    ```
    {current_prompt}
    ```

    Based on the following issues observed (metrics over multiple runs):

    {error_summary_text}

    Please provide an improved version of the SEO metadata generation prompt. Focus on clarity, robustness, and addressing the specific issues indicated by the metrics. The improved prompt should instruct the metadata generation model to:

    1.  Adhere strictly to the requested XML structure (`<metadata>`, `<title>`, `<description>`, `<tags>`) and output *only* this structure.
    2.  Generate SEO-focused titles, descriptions, and tags.
    3.  Use keywords **naturally** within the description. **Explicitly instruct the model to avoid excessive repetition of the same keywords or phrases.** Aim for readability and value over simple repetition.
    4.  Ensure the title is properly reflected in the description content (to prevent title mismatches).
    5.  Provide clear instructions for including all tags in the "Tags Used in Video :-" section of the description, ensuring the tag list in the description matches exactly with the tags in the <tags> section.
    6.  Ensure all specific elements requested within the description (e.g., title inclusion, credit, disclaimer, tags list, CTA) are present and correctly formatted.
    7.  Generate comprehensive and relevant tag lists.
    8.  The prompt uses placeholders: {{channel_name}}, {{channel_topic}}, {{video_topic}}, {{uploader_name}}. Ensure these remain intact and are properly used.
    9.  Add a final self-check step within the prompt for the model to verify all elements are present and constraints (like keyword usage) are met before outputting.

    Provide ONLY the improved prompt text, without any explanations or additional text.
    """
    try:
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(meta_prompt)
        improved_prompt = response.text.strip()

        # Basic validation
        if "<metadata>" not in improved_prompt or "<title>" not in improved_prompt or "<description>" not in improved_prompt or "<tags>" not in improved_prompt:
            print_error("Generated prompt missing required XML tags. Keeping current.", 1)
            return None

        # Check channel downloader specific placeholders
        if "{video_topic}" not in improved_prompt or "{uploader_name}" not in improved_prompt or "{channel_name}" not in improved_prompt or "{channel_topic}" not in improved_prompt:
             print_error("Generated prompt missing required placeholders. Keeping current.", 1)
             return None

        return improved_prompt
    except Exception as e:
        print_error(f"Error generating improved SEO metadata prompt: {e}", 1, include_traceback=True)
        return None
# --- End SEO Prompt Handling ---

# --- Category Suggestion ---
def get_suggested_category(title: str, description: str):
    """Uses Gemini to suggest the most appropriate YouTube category based on video content."""
    if not title or not description:
        print_warning("Cannot suggest category: Title or Description empty.", 2)
        return None

    # --- Keep the list here for validation ---
    KNOWN_CATEGORIES = [
        "Film & Animation", "Autos & Vehicles", "Music", "Pets & Animals",
        "Sports", "Travel & Events", "Gaming", "People & Blogs",
        "Comedy", "Entertainment", "News & Politics", "Howto & Style",
        "Education", "Science & Technology", "Nonprofits & Activism"
        # Removed "Movies", "Shows" as they are less common for Shorts uploads via Studio
    ]
    KNOWN_CATEGORIES_LOWER = {cat.lower() for cat in KNOWN_CATEGORIES}

    # --- MODIFY THE PROMPT ---
    valid_categories_string = ", ".join([f'"{cat}"' for cat in KNOWN_CATEGORIES]) # Create quoted list for prompt

    prompt = f"""
    Analyze the following YouTube Shorts video Title and Description:

    Title: {title}
    Description: {description[:1000]} # Limit description length

    Select the single BEST matching official YouTube Video Category for this content.
    You MUST choose EXACTLY ONE category name from this official list:
    {valid_categories_string}

    Output ONLY the chosen category name from the list, with the exact capitalization shown in the list, and nothing else.
    For example, if the content is about gaming, output: Gaming
    If it's about fitness exercises, the best fit from the list might be "Howto & Style" or "Sports". Choose the single most appropriate one FROM THE LIST PROVIDED.
    """

    try:
        print_info("Requesting category suggestion (with explicit list)...", 3)
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        suggested_cat_raw = response.text.strip()

        # --- Keep the validation logic, but it should pass more often now ---
        if suggested_cat_raw and suggested_cat_raw in KNOWN_CATEGORIES: # Check against the original list directly now
            print_success(f"Suggested category: {suggested_cat_raw}", 3)
            return suggested_cat_raw
        elif suggested_cat_raw:
            # This might still happen if Gemini hallucinates or ignores instructions
            print_warning(f"Gemini suggested a category NOT in the provided list: '{suggested_cat_raw}'. Ignoring.", 3)
            # Log the failed suggestion for debugging
            log_error(f"Category Suggestion Mismatch: Gemini suggested '{suggested_cat_raw}' which is not in the allowed list based on Title: '{title}'")
            return None
        else:
            print_warning("Gemini returned an empty category suggestion.", 3)
            return None
    except Exception as e:
        print_error(f"Error getting category suggestion: {e}", 3)
        log_error(f"Gemini Category Suggestion Error: {e} for Title: '{title}'") # Log error
        return None
# --- End Category Suggestion ---

# --- Tuning Suggestions ---
def generate_performance_summary(metrics):
    """Generates a summary of performance metrics for tuning suggestions."""
    summary = ["=== Channel Downloader Performance Summary ==="]
    runs_data = metrics.get('runs', [])
    summary.append(f"Total runs recorded: {len(runs_data)}")

    # Overall metrics
    summary.append(f"Total videos found (all runs): {metrics.get('total_shorts_found', 0)}")
    summary.append(f"Total suitable videos (all runs): {metrics.get('total_suitable_shorts', 0)}")
    summary.append(f"Total downloads attempted (all runs): {metrics.get('total_downloads_attempted', 0)}")
    summary.append(f"Total successful downloads (all runs): {metrics.get('total_successful_downloads', 0)}")
    summary.append(f"Overall Download success rate: {metrics.get('total_successful_downloads', 0) / max(1, metrics.get('total_downloads_attempted', 1)):.1%}")

    # Metadata performance
    summary.append(f"\n=== Metadata Performance (All Runs) ===")
    summary.append(f"Total metadata API calls: {metrics.get('total_metadata_api_calls', 0)}")
    summary.append(f"Total metadata errors: {metrics.get('total_metadata_errors', 0)}")
    summary.append(f"Overall Metadata error rate: {metrics.get('total_metadata_errors', 0) / max(1, metrics.get('total_metadata_api_calls', 1)):.1%}")

    # Add specific metadata errors/warnings
    metadata_metrics = load_metadata_metrics() # Load the detailed metrics
    summary.append(f"  Parse Failures: {metadata_metrics.get('parse_failures', 0)}")
    summary.append(f"  Timeouts: {metadata_metrics.get('timeouts', 0)}")
    summary.append(f"  Empty Descriptions: {metadata_metrics.get('empty_description_errors', 0)}")
    summary.append(f"  Empty Tags: {metadata_metrics.get('empty_tags_errors', 0)}")
    # Add specific validation warnings
    summary.append(f"  Validation - Title Mismatches: {metadata_metrics.get('validation_title_mismatches', 0)}")
    summary.append(f"  Validation - Tag List Errors: {metadata_metrics.get('validation_tag_list_errors', 0)}")
    summary.append(f"  Validation - Keyword Stuffing: {metadata_metrics.get('validation_keyword_stuffing', 0)}")

    # Recent runs
    if runs_data:
        summary.append(f"\n=== Recent Runs ({min(5, len(runs_data))}) ===")
        for run in runs_data[-5:]:
            summary.append(f" - Run {run.get('date', 'Unknown')[:10]}: Suitable={run.get('suitable_shorts',0)}, Attempted={run.get('downloads_attempted', 0)}, Succeeded={run.get('downloads_successful', 0)}, MetaErrors={run.get('metadata_errors', 0)}")

    # Channel performance
    channel_performance = metrics.get('channel_performance', {})
    if channel_performance:
         sorted_channels = sorted(channel_performance.items(), key=lambda item: item[1], reverse=True)
         summary.append(f"\n=== Top 10 Performing Channels (by Score) ===")
         for chan, score in sorted_channels[:10]:
             summary.append(f" - {chan}: {score:.2f}")

    return "\n".join(summary)

def generate_tuning_suggestions(metrics, config):
    """Generates parameter tuning suggestions using Gemini."""
    performance_summary = generate_performance_summary(metrics)
    config_text = "\n".join([f"{key}={value}" for key, value in config.items() if 'API_KEY' not in key])

    prompt = f"""
    Analyze the following performance report from a YouTube Shorts CHANNEL downloader script.
    Suggest config.txt adjustments and explain *why* based ONLY on the provided data.

    Performance Report:
    {performance_summary}

    Current Config (excluding API keys):
    {config_text}

    Focus on:
    - MAX_DOWNLOADS: Total downloads per run.
    - METADATA_TIMEOUT_SECONDS: Metadata generation timeout.
    - YT_PLAYLIST_FETCH_LIMIT: Initial videos checked per channel. (Is it finding enough suitable ones?)
    - SEO_... parameters: Are metadata errors high? Maybe review prompt indirectly via these.

    Provide specific, data-driven recommendations ONLY IF the data suggests a change. If performance looks good, say so.
    Format clearly.
    """

    try:
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        suggestions = response.text.strip()

        with open(tuning_suggestions_file_path, "a", encoding="utf-8") as f:
            f.write(f"\n\n=== Tuning Suggestions ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ===\nConfig:\n{config_text}\n\nSummary:\n{performance_summary}\n\nSuggestions:\n{suggestions}\n")

        return suggestions
    except Exception as e:
        print_error(f"Error generating tuning suggestions: {e}", 1)
        return None
# --- End Tuning Suggestions ---

def generate_seo_metadata_v2(video_topic, uploader_name=DEFAULT_UPLOADER_NAME, original_title="Unknown Title"):
    """Generates SEO-optimized metadata using Gemini API, including uploader credit."""
    # Use SEO example tags and hashtags from config
    example_tags_str = ", ".join(seo_example_tags)
    example_hashtags_str = " ".join(seo_example_hashtags)
    # Updated prompt to include uploader_name placeholder
    prompt = f"""
    do not include any explanation or any other text. just give me the metadata in below format.
    only apply the below format. do not include any other text or explanation.
    Generate SEO-optimized metadata for a YouTube Shorts video in the following structured format:
    You are a YOUTUBE SEO EXPERT A GURU one in million. you have insight knowledge of youtube shorts.
    you know how the ranking algorithm works and how to get more views and engagement.
    you know how creator like mrbeast, tseries, and other top creators get more views and engagement.
    your master of youtube shorts. you have worked with big creator know all secrets of youtube shorts.
    you have worked in google youtube team and you know all secrets of youtube shorts.
    Our Channel Name is "{seo_channel_name}" and we are a channel about {seo_channel_topic}.
    include a copyright fair use disclaimer in the description.
    APPLY ALL OF THE ABOVE KNOWLEDGE AND SECRETS TO BELOW metadata.

    <metadata>
        <title>
            Create an engaging, fast-paced, and action-driven title (max 100 chars incl. #Shorts) with a high CTR based on the video topic: '{video_topic}'.
            Use keywords for '{seo_channel_topic}'. Use relevant emojis (🔥, 💪, 👀), numbers, power words (BEST, HOT, ULTIMATE, SECRET, TRY THIS). Add "#Shorts" at the end.
        </title>
        <description>
            Write an SEO-optimized description (max 4500 chars):
                * Start with the optimized video title.
                * 2-3 sentence engaging summary about '{video_topic}' and '{seo_channel_topic}', using keywords/LSI naturally.
                * **Include credit: "Credit: {uploader_name}"** <-- UPLOADER CREDIT HERE
                * Include copyright disclaimer:
                  --------------【Copyright Disclaimer】-------------
                  All the videos, songs, images, and graphics used in the video belong to
                  their respective owners and I or this channel "{seo_channel_name}" does not claim any right over them.
                  Copyright Disclaimer under section 107 of the Copyright Act of 1976, allowance is made for “fair use” for purposes such as criticism, comment, news reporting, teaching, scholarship, education and research. Fair use is a use permitted by copyright statute that might otherwise be infringing.
                * After disclaimer, add 10-15 relevant hashtags (inspired by: {example_hashtags_str}).
                * Add heading "Tags Used in Video :-" and list all tags from <tags> section below, comma-separated.
                * End with a Call to Action (e.g., "👍 Like & Subscribe to {seo_channel_name}!").
                * Add heading "Ignored Hashtags :-" followed by a diverse list of relevant hashtags.
        </description>
        <tags>
            Suggest 15-25 SEO-friendly tags (comma-separated, max 500 chars total).
            * Start with keywords for '{video_topic}'. Include tags for '{seo_channel_topic}' and channel name '{seo_channel_name}'.
            * Use mix of general/specific tags. Inspire from: {example_tags_str}
        </tags>
    </metadata>

    **Video Topic**: {video_topic}
    """
    # Default includes credit
    default_metadata = { "title": f"{video_topic[:80]} #Shorts", "description": f"Desc: {video_topic}.\n\nCredit: {uploader_name}\n\n[Disclaimer]", "tags": ["default"] }
    try:
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        raw_text = ""
        try: raw_text = "".join(part.text for part in response.parts)
        except Exception: raw_text = response.text # Fallback
        if not raw_text: log_error(f"Gemini response empty for '{video_topic}'."); return default_metadata

        title_match = re.search(r"<title>(.*?)</title>", raw_text, re.DOTALL | re.IGNORECASE)
        desc_match = re.search(r"<description>(.*?)</description>", raw_text, re.DOTALL | re.IGNORECASE)
        tags_match = re.search(r"<tags>(.*?)</tags>", raw_text, re.DOTALL | re.IGNORECASE)
        metadata = {}
        # Title processing
        if title_match:
            metadata["title"] = title_match.group(1).strip()
            metadata["title"] = metadata["title"][:90].strip() # Limit length before adding #Shorts
            if not metadata["title"].lower().endswith("#shorts"): metadata["title"] += " #Shorts"
        else: metadata["title"] = default_metadata["title"]
        # Description processing
        if desc_match:
            metadata["description"] = desc_match.group(1).strip()
            credit_line = f"Credit: {uploader_name}"
            if credit_line not in metadata["description"]: metadata["description"] += f"\n\n{credit_line}" # Ensure credit
        else: metadata["description"] = default_metadata["description"]
        # Tags processing
        if tags_match:
            tags_raw = tags_match.group(1).strip()
            all_tags = [tag.strip() for tag in tags_raw.split(',') if tag.strip()]
            current_len = 0; final_tags = []
            for tag in all_tags:
                tag_len_with_comma = len(tag) + (1 if final_tags else 0)
                if current_len + tag_len_with_comma <= 495: final_tags.append(tag); current_len += tag_len_with_comma
                else: break # Stop if limit exceeded
            metadata["tags"] = final_tags
        else: metadata["tags"] = default_metadata["tags"]
        # Basic validation
        return metadata if metadata.get("title") and metadata.get("description") else default_metadata
    except Exception as e:
        log_error(f"Error in generate_seo_metadata for '{video_topic}': {e}")
        return default_metadata


def generate_metadata_with_timeout_v2(video_title, uploader_name, original_title="Unknown Title", timeout=METADATA_TIMEOUT_SECONDS):
    """Generates metadata with timeout, includes category suggestion."""
    metadata_metrics = load_metadata_metrics()
    metadata_metrics["total_api_calls"] += 1
    error_type = None
    error_details = None
    result_metadata = None

    try:
        with concurrent.futures.ThreadPoolExecutor() as executor:
            print_info("Generating primary SEO metadata (Title/Desc/Tags)...", 2)
            # Use the load_or_get_seo_prompt_template function directly instead of generate_seo_metadata_v2
            prompt_template = load_or_get_seo_prompt_template()
            if not prompt_template:
                print_error("Failed to load SEO prompt template. Using minimal fallback.")
                prompt_template = """<metadata><title>{video_topic} #Shorts</title><description>Credit: {uploader_name}</description><tags>shorts</tags></metadata>"""

            # Format the prompt with actual data
            prompt = prompt_template.replace("{video_topic}", video_title)\
                                    .replace("{uploader_name}", uploader_name)\
                                    .replace("{channel_name}", seo_channel_name)\
                                    .replace("{channel_topic}", seo_channel_topic)\
                                    .replace("{original_title}", original_title)

            # Create a function to generate metadata that can be executed with timeout
            def generate_metadata_internal():
                try:
                    model = genai.GenerativeModel("gemini-2.0-flash")
                    response = model.generate_content(prompt)
                    raw_text = response.text

                    # Default fallback structure
                    metadata = {
                        "title": f"{video_title} #Shorts",
                        "description": f"Default SEO description.\nCredit: {uploader_name}\nOriginal Title: {original_title}",
                        "tags": []
                    }

                    # --- Parsing (Adjusted for SEO prompt structure) ---
                    title_match = re.search(r"<title>(.*?)</title>", raw_text, re.DOTALL | re.IGNORECASE)
                    desc_match = re.search(r"<description>(.*?)</description>", raw_text, re.DOTALL | re.IGNORECASE)
                    tags_match = re.search(r"<tags>(.*?)</tags>", raw_text, re.DOTALL | re.IGNORECASE)

                    parsing_warnings = []
                    if title_match and title_match.group(1).strip():
                        metadata["title"] = title_match.group(1).strip()
                    else:
                        parsing_warnings.append("title")
                        metadata["title"] = f"{video_title} #Shorts"

                    if desc_match and desc_match.group(1).strip():
                        metadata["description"] = desc_match.group(1).strip()
                        # Ensure credit is present (as per prompt structure)
                        if f"Credit: {uploader_name}" not in metadata["description"]:
                            metadata["description"] += f"\n\nCredit: {uploader_name}"
                        if f"Original Title: {original_title}" not in metadata["description"] and original_title != "Unknown Title":
                            metadata["description"] += f"\nOriginal Title: {original_title}"
                    else:
                        parsing_warnings.append("description")
                        metadata["description"] = f"Default desc.\nCredit: {uploader_name}\nOriginal Title: {original_title}"

                    if tags_match and tags_match.group(1).strip():
                        tags_raw = tags_match.group(1).strip()
                        metadata["tags"] = [tag.strip() for tag in re.split(r'[,\n]', tags_raw) if tag.strip()] # Handle comma or newline
                    else:
                        parsing_warnings.append("tags")
                        metadata["tags"] = [seo_channel_topic.lower(), "shorts"]

                    if parsing_warnings:
                        print_warning(f"Could not parse <{'>, <'.join(parsing_warnings)}> for topic: {video_title}. Used fallbacks.", 1)

                    # --- Post-Processing (Title Length/Suffix) ---
                    temp_title = metadata.get("title", video_title)
                    if len(temp_title) > TARGET_TITLE_LENGTH: # Use TARGET_TITLE_LENGTH (90)
                        truncated = temp_title[:TARGET_TITLE_LENGTH]
                        last_space = truncated.rfind(' ')
                        temp_title = truncated[:last_space].strip() if last_space > 0 else truncated.strip()
                    if not temp_title.lower().endswith(SHORTS_SUFFIX.lower()):
                        if len(temp_title) + len(SHORTS_SUFFIX) <= MAX_TITLE_LENGTH:
                            temp_title += SHORTS_SUFFIX
                    metadata["title"] = temp_title

                    return metadata
                except Exception as e:
                    print_error(f"Error during metadata generation/processing for '{video_title}': {e}", 1, include_traceback=True)
                    return {
                        "title": f"{video_title[:80]} #Shorts",
                        "description": f"Default desc. Error: {e}\nCredit: {uploader_name}\nOriginal Title: {original_title}",
                        "tags": ["error"]
                    }

            # Execute with timeout
            future = executor.submit(generate_metadata_internal)
            result_metadata = future.result(timeout=timeout)

            # Final check
            if not isinstance(result_metadata, dict) or not result_metadata.get("title") or not result_metadata.get("description") or not result_metadata.get("tags"):
                print_error(f"Critical Warning: Metadata invalid/empty for '{video_title}'. Final fallback.", 1)
                result_metadata = {
                    "title": f"{video_title[:80]} #Shorts",
                    "description": f"Final fallback.\nCredit: {uploader_name}\nOriginal Title: {original_title}",
                    "tags": ["fallback"]
                }

            # Add category suggestion
            suggested_category = None
            if result_metadata:
                gen_title = result_metadata.get("title", video_title)
                gen_desc = result_metadata.get("description", "")
                if gen_title and gen_desc:
                    try:
                        suggested_category = get_suggested_category(gen_title, gen_desc)
                        if suggested_category:
                            result_metadata['suggested_category'] = suggested_category
                    except Exception as cat_err:
                        print_error(f"Error during category suggestion call: {cat_err}", 2)
                else:
                    print_warning("Skipping category suggestion due to empty title/description.", 2)

            # Check for errors (parsing/empty)
            if "Default SEO description" in result_metadata.get("description", ""):
                metadata_metrics["empty_description_errors"] += 1
                error_type = "empty_description"
                error_details = f"Failed description: {video_title}"

            if not result_metadata.get("tags") or "error" in result_metadata.get("tags",[]):
                metadata_metrics["empty_tags_errors"] += 1
                error_type = error_type or "empty_tags"
                error_details = error_details or f"Failed tags: {video_title}"

            if error_type and error_details:
                add_error_sample(metadata_metrics, error_type, error_details, video_title)

            save_metadata_metrics(metadata_metrics)
            return result_metadata

    except concurrent.futures.TimeoutError:
        print_warning(f"Primary metadata generation timed out for: {video_title}", 1)
        metadata_metrics["timeouts"] += 1
        add_error_sample(metadata_metrics, "timeout", f"Timeout for: {video_title}", video_title)
        save_metadata_metrics(metadata_metrics)
        return {
            "title": f"{video_title[:80]} #Shorts",
            "description": f"Timeout.\nCredit: {uploader_name}\nOriginal Title: {original_title}",
            "tags": ["timeout"]
        }
    except Exception as e:
        print_error(f"Error submitting metadata generation job for '{video_title}': {e}", 1, include_traceback=True)
        metadata_metrics["parse_failures"] += 1
        add_error_sample(metadata_metrics, "exception", f"Error: {str(e)} for: {video_title}", video_title)
        save_metadata_metrics(metadata_metrics)
        return {
            "title": f"{video_title[:80]} #Shorts",
            "description": f"Error: {e}\nCredit: {uploader_name}\nOriginal Title: {original_title}",
            "tags": ["error"]
        }


def create_folders():
    """Creates necessary folders if they don't exist."""
    for folder in [download_folder, metadata_folder]:
        try: os.makedirs(folder, exist_ok=True)
        except OSError as e: log_error(f"FATAL: Error creating folder '{folder}': {e}"); raise


def load_or_create_excel():
    """Loads Excel, checks/creates sheets, corrects headers, loads previous videos."""
    previously_downloaded_videos = set()
    wb = None; downloaded_sheet = None; uploaded_sheet = None # Initialize

    # Try to import excel_utils module
    try:
        # First try to import from the current directory
        import excel_utils
        excel_utils_available = True
        print_info("Using excel_utils module for robust Excel handling with auto-closing")
    except ImportError:
        # If that fails, try to import using the full path
        try:
            import sys
            import os
            script_directory = os.path.dirname(os.path.abspath(__file__))
            excel_utils_path = os.path.join(script_directory, "excel_utils.py")

            if os.path.exists(excel_utils_path):
                import importlib.util
                spec = importlib.util.spec_from_file_location("excel_utils", excel_utils_path)
                excel_utils = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(excel_utils)
                excel_utils_available = True
                print_info(f"Using excel_utils module from {excel_utils_path}")
            else:
                excel_utils_available = False
                print_warning(f"excel_utils.py not found at {excel_utils_path}")
        except Exception as e:
            excel_utils_available = False
            print_warning(f"excel_utils module not available: {e}. Using fallback Excel handling.")

    # Define sheet configuration
    sheets_config = {
        DOWNLOADED_SHEET_NAME: EXPECTED_DOWNLOADED_HEADERS,
        UPLOADED_SHEET_NAME: EXPECTED_UPLOADED_HEADERS
    }

    if excel_utils_available:
        # Use the excel_utils module for robust Excel handling
        wb, sheets, save_needed = excel_utils.load_or_create_excel(excel_file, sheets_config)

        if not wb:
            log_error(f"FATAL: Failed to load or create Excel file: {excel_file}")
            raise Exception(f"Failed to load or create Excel file: {excel_file}")

        downloaded_sheet = sheets.get(DOWNLOADED_SHEET_NAME)
        uploaded_sheet = sheets.get(UPLOADED_SHEET_NAME)

        # Check/Correct headers for Downloaded Sheet
        if downloaded_sheet:
            current_headers = [str(c.value) if c.value is not None else '' for c in downloaded_sheet[1]]
            if tuple(current_headers) != tuple(EXPECTED_DOWNLOADED_HEADERS):
                log_error(f"Correcting headers in '{DOWNLOADED_SHEET_NAME}'.")
                print(f"Warning: Correcting headers in '{DOWNLOADED_SHEET_NAME}'.")
                for i, h in enumerate(EXPECTED_DOWNLOADED_HEADERS):
                    downloaded_sheet.cell(1, i + 1, h)
                save_needed = True

        # Check/Correct headers for Uploaded Sheet
        if uploaded_sheet:
            current_headers = [str(c.value) if c.value is not None else '' for c in uploaded_sheet[1]]
            if tuple(current_headers) != tuple(EXPECTED_UPLOADED_HEADERS):
                log_error(f"Correcting headers in '{UPLOADED_SHEET_NAME}'.")
                print(f"Warning: Correcting headers in '{UPLOADED_SHEET_NAME}'.")
                for i, h in enumerate(EXPECTED_UPLOADED_HEADERS):
                    uploaded_sheet.cell(1, i + 1, h)
                save_needed = True

        # Load previous videos (only if downloaded_sheet is valid)
        if downloaded_sheet:
            print("Loading previous Title/Uploader pairs...")
            max_col = max(ORIGINAL_TITLE_COL_IDX, UPLOADER_COL_IDX)
            for row in downloaded_sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
                if len(row) >= max_col:
                    title, uploader = row[ORIGINAL_TITLE_COL_IDX - 1], row[UPLOADER_COL_IDX - 1]
                    if isinstance(title, str) and title.strip() and isinstance(uploader, str) and uploader.strip():
                         previously_downloaded_videos.add((title.strip(), uploader.strip()))
            print(f"Loaded {len(previously_downloaded_videos)} previous Title/Uploader pairs.")
        else:
            log_error("Could not load previous videos: Downloaded sheet object invalid.")

        # Save if needed using robust save mechanism with Excel auto-closing
        if save_needed:
            if not excel_utils.safe_save_workbook(wb, excel_file, close_excel=True, create_backup=True):
                print_warning(f"Could not save structural changes to Excel. Will try again later.")
                log_error(f"Could not save structural changes to Excel. Will try again later.")

        print_success("Excel loaded successfully.")
        return wb, downloaded_sheet, uploaded_sheet, previously_downloaded_videos

    else:
        # Fallback to original implementation if excel_utils is not available
        save_needed = False # Flag to save only if changes were made

        if not os.path.exists(excel_file):
            print(f"Excel file not found '{excel_file}'. Creating.")
            wb = Workbook(); save_needed = True
            downloaded_sheet = wb.active; downloaded_sheet.title = DOWNLOADED_SHEET_NAME
            downloaded_sheet.append(EXPECTED_DOWNLOADED_HEADERS)
            uploaded_sheet = wb.create_sheet(title=UPLOADED_SHEET_NAME)
            uploaded_sheet.append(EXPECTED_UPLOADED_HEADERS)
        else:
            print(f"Loading existing Excel: {excel_file}")
            try:
                wb = load_workbook(excel_file)
                # Check/Correct Downloaded Sheet
                if DOWNLOADED_SHEET_NAME not in wb.sheetnames:
                    log_error(f"Sheet '{DOWNLOADED_SHEET_NAME}' missing. Creating."); downloaded_sheet = wb.create_sheet(DOWNLOADED_SHEET_NAME, 0); downloaded_sheet.append(EXPECTED_DOWNLOADED_HEADERS); save_needed = True
                else:
                    downloaded_sheet = wb[DOWNLOADED_SHEET_NAME]; current_headers = [str(c.value) if c.value is not None else '' for c in downloaded_sheet[1]]
                    if tuple(current_headers) != tuple(EXPECTED_DOWNLOADED_HEADERS):
                        log_error(f"Correcting headers in '{DOWNLOADED_SHEET_NAME}'."); print(f"Warning: Correcting headers in '{DOWNLOADED_SHEET_NAME}'.")
                        for i, h in enumerate(EXPECTED_DOWNLOADED_HEADERS): downloaded_sheet.cell(1, i + 1, h); save_needed = True
                # Check/Correct Uploaded Sheet
                if UPLOADED_SHEET_NAME not in wb.sheetnames:
                    log_error(f"Sheet '{UPLOADED_SHEET_NAME}' missing. Creating."); uploaded_sheet = wb.create_sheet(UPLOADED_SHEET_NAME); uploaded_sheet.append(EXPECTED_UPLOADED_HEADERS); save_needed = True
                else:
                    uploaded_sheet = wb[UPLOADED_SHEET_NAME]; current_headers = [str(c.value) if c.value is not None else '' for c in uploaded_sheet[1]]
                    if tuple(current_headers) != tuple(EXPECTED_UPLOADED_HEADERS):
                        log_error(f"Correcting headers in '{UPLOADED_SHEET_NAME}'."); print(f"Warning: Correcting headers in '{UPLOADED_SHEET_NAME}'.")
                        for i, h in enumerate(EXPECTED_UPLOADED_HEADERS): uploaded_sheet.cell(1, i + 1, h); save_needed = True

                # Load previous videos (only if downloaded_sheet is valid)
                if downloaded_sheet:
                    print("Loading previous Title/Uploader pairs...")
                    max_col = max(ORIGINAL_TITLE_COL_IDX, UPLOADER_COL_IDX)
                    for row in downloaded_sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
                        if len(row) >= max_col:
                            title, uploader = row[ORIGINAL_TITLE_COL_IDX - 1], row[UPLOADER_COL_IDX - 1]
                            if isinstance(title, str) and title.strip() and isinstance(uploader, str) and uploader.strip():
                                 previously_downloaded_videos.add((title.strip(), uploader.strip()))
                    print(f"Loaded {len(previously_downloaded_videos)} previous Title/Uploader pairs.")
                else: log_error("Could not load previous videos: Downloaded sheet object invalid.")

                print("Excel loaded.")
            except Exception as e:
                log_error(f"FATAL: Error loading/validating Excel '{excel_file}': {e}"); traceback.print_exc(); raise

        # Save workbook only if changes were made during loading/creation/correction
        if save_needed:
            try:
                # Try to save with a simple retry mechanism
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        wb.save(excel_file)
                        print(f"Saved structural changes to Excel file: {excel_file}")
                        break
                    except PermissionError as pe:
                        if attempt < max_retries - 1:
                            print(f"PermissionError saving Excel (attempt {attempt+1}/{max_retries}): {pe}")

                            # Try to close Excel if excel_utils is available
                            try:
                                import excel_utils
                                if excel_utils.close_excel_processes_with_file(excel_file):
                                    print("Successfully closed Excel processes. Retrying immediately...")
                                    continue
                            except ImportError:
                                pass  # excel_utils not available

                            print(f"Retrying in 2 seconds...")
                            time.sleep(2)
                        else:
                            raise
            except Exception as e:
                log_error(f"Error saving structural changes to Excel: {e}")
                print(f"Error saving structural changes to Excel: {e}")
                # Try to save to a backup file
                try:
                    backup_path = f"{excel_file}.backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
                    wb.save(backup_path)
                    print(f"Saved backup Excel file: {backup_path}")
                except Exception as be:
                    log_error(f"Failed to save backup Excel file: {be}")

        return wb, downloaded_sheet, uploaded_sheet, previously_downloaded_videos


def get_last_video_index(downloaded_sheet: Worksheet) -> int:
    """Gets the next video index based on the 'Downloaded' sheet."""
    max_index = 0
    if not downloaded_sheet: log_error("Cannot get last video index: Sheet is None."); return 1
    try:
        for row in downloaded_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row and row[0] and isinstance(row[0], str) and row[0].lower().startswith("video"):
                try: max_index = max(max_index, int(row[0][len("video"):]))
                except (ValueError, IndexError): pass # Ignore parsing errors
    except Exception as e: log_error(f"Error get_last_video_index: {e}"); return 1
    next_index = max_index + 1; print(f"Next video index: {next_index}"); return next_index


def load_channels() -> list:
    """Loads channel URLs from the channels file."""
    print(f"Loading channels: {channels_file_path}")
    if not os.path.exists(channels_file_path): log_error(f"FATAL: Channels file '{channels_file_path}' not found."); raise FileNotFoundError(f"{channels_file_path}")
    try:
        with open(channels_file_path, "r", encoding="utf-8") as file:
            channels = [line.strip() for line in file if line.strip() and not line.startswith('#')]
    except Exception as e: log_error(f"FATAL: Error reading channels file '{channels_file_path}': {e}"); raise
    if not channels: log_error("FATAL: Channels file empty."); raise ValueError("Channels file is empty.")
    print(f"Loaded {len(channels)} channels."); return channels


# This function has been moved to the Playlist Management Functions section


# This function has been moved to the Playlist Management Functions section

def save_metadata_file(entry, video_index, seo_metadata, channel_url=None):
    """Saves the combined metadata to a JSON file, including channel URL as discovery_keyword."""
    original_title = entry.get("title", f"unk_{video_index}")
    uploader = entry.get("uploader", DEFAULT_UPLOADER_NAME)
    view_count = entry.get('view_count', 0);
    try:
        view_count = int(view_count) if view_count is not None else 0
    except: view_count = 0
    credit_uploader_name = uploader if uploader != DEFAULT_UPLOADER_NAME else "Original Uploader"
    optimized_description = seo_metadata.get("description", ""); credit_line = f"Credit: {credit_uploader_name}"
    if optimized_description and credit_line not in optimized_description: optimized_description += f"\n\n{credit_line}"
    elif not optimized_description: optimized_description = f"Default description.\n\n{credit_line}"
    metadata = {
        "video_index": str(video_index),
        "id": entry.get("id"),
        "original_title": original_title,
        "uploader": uploader,
        "view_count": view_count,
        "download_timestamp": datetime.now().isoformat(), # Use ISO format
        "optimized_title": seo_metadata.get("title", f"{original_title[:80]} #Shorts"),
        "optimized_description": optimized_description,
        "optimized_tags": seo_metadata.get("tags", []),
        "suggested_category": seo_metadata.get("suggested_category"),
        "target_playlist": seo_metadata.get("target_playlist"), # Added for playlist management
        "discovery_keyword": channel_url, # Store channel URL as discovery_keyword for correlation cache
        "metadata_strategy": "B: Channel-Based" # Strategy marker to differentiate from keyword-based downloads
    }
    metadata_file_path = os.path.join(metadata_folder, f"video{video_index}.json")
    try:
        with open(metadata_file_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, ensure_ascii=False, indent=4)
        return metadata
    except Exception as e:
        log_error(f"Error saving metadata {metadata_file_path}: {e}")
        return metadata # Return potentially incomplete data


# --- Main Execution Logic ---
def main():
    wb = None; downloaded_sheet = None; uploaded_sheet = None
    downloaded_video_data_for_excel = [] # Store tuples for final Excel append
    previously_downloaded_videos = set()
    channel_processed_ids = {} # Processed video ID cache {channel_url: [ids]}
    channel_listing_cache = {} # Channel entry list cache {channel_url: [entries]}
    channel_scores = {} # Scores based on performance {channel_url: float_score}

    try:
        print(f"{Fore.CYAN}--- Starting Channel Downloader Script (with Self-Improvement) ---{Style.RESET_ALL}")
        create_folders()
        performance_metrics = load_performance_metrics() # Load overall performance metrics
        metadata_metrics = load_metadata_metrics() # Load metadata error metrics
        wb, downloaded_sheet, uploaded_sheet, previously_downloaded_videos = load_or_create_excel()
        channels = load_channels()
        channel_processed_ids = load_cache(channel_processed_ids_cache_file, "Channel Processed IDs")
        channel_listing_cache = load_cache(channel_listing_cache_file, "Channel Listing") # Load permanent channel list cache
        channel_scores = load_cache(channel_performance_cache_path, "Channel Performance") # Load channel scores

        # --- Load/Fetch Playlist Data ---
        print_section_header("Loading/Fetching Playlist Data")
        playlist_data_cache = load_cache(playlist_data_cache_path, "Playlist Data")
        # Simple cache expiry - refetch all if cache is old (could be more granular)
        refetch_playlists = True
        if "timestamp" in playlist_data_cache:
            try:
                cache_time = datetime.fromisoformat(playlist_data_cache["timestamp"])
                if datetime.now() - cache_time < timedelta(hours=24): # 24hr cache
                    refetch_playlists = False
                    print_info("Using cached playlist data.")
            except:
                print_warning("Error reading playlist cache timestamp. Refetching.")

        # Initialize service variable for potential API calls
        service = None
        existing_playlists_api = {} # Will store API results if fetched

        if refetch_playlists:
            print_info("Fetching fresh playlist data from YouTube API...")
            service = get_authenticated_service() # Get service if needed
            if service:
                existing_playlists_api = fetch_channel_playlists(service)
                # Update cache: Keep existing suggestions, update API ones
                new_cache_data = {"timestamp": datetime.now().isoformat()}
                for p_id, p_title in existing_playlists_api.items():
                    new_cache_data[p_id] = p_title # Add/update fetched playlists
                # Keep previous NEW suggestions if they don't conflict with fetched IDs/Titles
                existing_titles_lower = {t.lower() for t in existing_playlists_api.values()}
                for key, val in playlist_data_cache.items():
                    if key.startswith("NEW: ") and key not in new_cache_data:
                        # Avoid adding if an existing playlist now has the same name
                        suggested_title = key[len("NEW: "):]
                        if suggested_title.lower() not in existing_titles_lower:
                            new_cache_data[key] = val # Keep old NEW suggestion

                playlist_data_cache = new_cache_data
                save_cache(playlist_data_cache, playlist_data_cache_path, "Playlist Data")
            else:
                print_error("Cannot fetch playlists - API service unavailable. Using potentially stale cache.")

        # Load keyword frequency cache (for tag extraction)
        keyword_frequency = load_cache(keywords_cache_file_path, "Keyword Frequency")

        # Define keyword filtering criteria (for tag extraction)
        required_substrings = ["GTA", "Grand Theft Auto"] # Example, adjust to your niche
        required_substrings_lower = [sub.lower() for sub in required_substrings]
        social_media_lower = {"twitch", "youtube", "facebook", "instagram", "tiktok", "x", "reddit", "discord", "kick"}
        generic_terms_lower = {"gaming", "video game", "gameplay", "shorts", "short", "viral", "trending", "funny", "meme", "edit", "clip", "clips"}

        # Load max_keywords config
        _DEFAULT_MAX_KEYWORDS = 200
        try:
            max_keywords = int(config.get("MAX_KEYWORDS", _DEFAULT_MAX_KEYWORDS))
        except:
            max_keywords = _DEFAULT_MAX_KEYWORDS

        # Load seed_niche
        try:
            niche_file_path = constants.NICHE_FILE_PATH
            with open(niche_file_path, "r", encoding="utf-8") as f:
                seed_niche = f.readline().strip()
        except:
            seed_niche = "GTA" # Default niche if file not found

        # --- Correlate Performance Data with Channels ---
        print(f"{Fore.BLUE}--- Correlating Performance Data with Channels ---{Style.RESET_ALL}")
        channel_performance_feedback = {} # Key: channel_url, Value: List of perf dicts
        correlation_cache = load_correlation_cache()
        uploaded_performance_data = {} # Key: YouTube Video ID, Value: Dict

        if uploaded_sheet:
            print_info("Loading performance data from 'Uploaded' sheet...", 1)
            header_perf = [str(cell.value).lower().strip() if cell.value else '' for cell in uploaded_sheet[1]]
            try:
                # Find columns by name (case-insensitive)
                col_map = {name: i + 1 for i, name in enumerate(header_perf)}
                yt_id_col = col_map.get(UPLOADED_YT_ID_COL_NAME)
                views_col = col_map.get(UPLOADED_VIEWS_COL_NAME)
                likes_col = col_map.get(UPLOADED_LIKES_COL_NAME)
                comments_col = col_map.get(UPLOADED_COMMENTS_COL_NAME)

                if not yt_id_col:
                    print_warning(f"Required column '{UPLOADED_YT_ID_COL_NAME}' not found in '{UPLOADED_SHEET_NAME}'. Performance data skipped.", 1)
                else:
                    loaded_perf_count = 0
                    for row_idx in range(2, uploaded_sheet.max_row + 1):
                        uploaded_yt_id = str(uploaded_sheet.cell(row=row_idx, column=yt_id_col).value or "").strip()
                        if uploaded_yt_id and uploaded_yt_id != "N/A":
                            views = likes = comments = 0
                            try:
                                if views_col: views = int(uploaded_sheet.cell(row=row_idx, column=views_col).value or 0)
                                if likes_col: likes = int(uploaded_sheet.cell(row=row_idx, column=likes_col).value or 0)
                                if comments_col: comments = int(uploaded_sheet.cell(row=row_idx, column=comments_col).value or 0)
                            except (ValueError, TypeError): pass # Ignore conversion errors
                            uploaded_performance_data[uploaded_yt_id] = {"views": views, "likes": likes, "comments": comments}
                            loaded_perf_count += 1
                    print_success(f"Loaded performance data for {loaded_perf_count} uploaded videos from Excel.", 1)
            except Exception as e:
                print_error(f"Error parsing '{UPLOADED_SHEET_NAME}' sheet: {e}", 1, include_traceback=True)
        else:
            print_info(f"'{UPLOADED_SHEET_NAME}' sheet not found. Cannot load performance data.", 1)

        if uploaded_performance_data and correlation_cache:
            print_info(f"Processing {len(correlation_cache)} entries from correlation cache for channel scoring.", 1)
            correlated_count = 0
            for cache_entry in correlation_cache:
                youtube_video_id = cache_entry.get("youtube_video_id")
                # The 'discovery_keyword' now holds the channel URL
                discovery_channel_url = cache_entry.get("discovery_keyword")
                if youtube_video_id and discovery_channel_url and youtube_video_id in uploaded_performance_data:
                    # Check if it looks like a URL (basic check)
                    if "youtube.com" in discovery_channel_url.lower() or "youtu.be" in discovery_channel_url.lower():
                        perf_data = uploaded_performance_data[youtube_video_id]
                        if discovery_channel_url not in channel_performance_feedback:
                            channel_performance_feedback[discovery_channel_url] = []
                        channel_performance_feedback[discovery_channel_url].append(perf_data)
                        correlated_count += 1
            print_success(f"Correlated performance data for {len(channel_performance_feedback)} channels using cache ({correlated_count} video links).")
            cleanup_correlation_cache(days_to_keep=7) # Cleanup old entries
        elif not correlation_cache:
            print_info("Correlation cache is empty. Cannot score channels.")
        elif not uploaded_performance_data:
            print_info("No uploaded performance data loaded. Cannot score channels.")
        # --- End Performance Correlation ---

        # --- Update Channel Scores ---
        print(f"{Fore.BLUE}--- Updating Channel Scores with Performance Feedback ---{Style.RESET_ALL}")
        def calculate_performance_score(views, likes, comments=0): # Same score func as keyword downloader
            view_score = math.log10(views + 1) * 1.5
            like_score = math.log10(likes + 1) * 2.0
            comment_score = math.log10(comments + 1) * 1.0
            return view_score + like_score + comment_score

        total_channels_updated = 0
        new_channel_scores = {} # Use a temporary dict for updates
        if channel_performance_feedback:
            for channel_url, performance_list in channel_performance_feedback.items():
                cumulative_perf_score = 0
                num_videos = 0
                for perf_data in performance_list:
                    video_score = calculate_performance_score(
                        perf_data.get("views", 0),
                        perf_data.get("likes", 0),
                        perf_data.get("comments", 0)
                    )
                    cumulative_perf_score += video_score
                    num_videos += 1
                if num_videos > 0:
                    average_perf_score = cumulative_perf_score / num_videos
                    # Simple scoring: Average performance score. Could be blended later.
                    # Ensure we don't overwrite existing scores unintentionally if needed, but for now, use latest avg.
                    new_channel_scores[channel_url] = max(0, average_perf_score) # Use average perf directly, ensure non-negative
                    total_channels_updated += 1
                    print_info(f"Updated score for channel '{channel_url}': AvgPerf={average_perf_score:.2f}", 1)

            if total_channels_updated > 0:
                channel_scores.update(new_channel_scores) # Merge new scores into existing
                print_success(f"Updated scores for {total_channels_updated} channels based on upload performance.")
                save_cache(channel_scores, channel_performance_cache_path, "Channel Performance")
            else:
                print_info("No channel scores were updated based on available performance data.")
        else:
            print_info("No performance feedback available to update channel scores.")
        # --- End Channel Score Update ---

        # --- Check Metadata Prompt Quality ---
        print(f"\n{Fore.CYAN}{Style.BRIGHT}--- Checking Metadata Prompt Quality ---{Style.RESET_ALL}")
        metadata_api_calls_total = metadata_metrics.get("total_api_calls", 0)
        if metadata_api_calls_total > 0:
            # Calculate rates for all relevant error types
            parse_fail_rate = metadata_metrics.get("parse_failures", 0) / metadata_api_calls_total
            timeout_rate = metadata_metrics.get("timeouts", 0) / metadata_api_calls_total
            empty_desc_rate = metadata_metrics.get("empty_description_errors", 0) / metadata_api_calls_total
            empty_tags_rate = metadata_metrics.get("empty_tags_errors", 0) / metadata_api_calls_total

            # Calculate validation warning rates
            title_mismatch_rate = metadata_metrics.get("validation_title_mismatches", 0) / metadata_api_calls_total
            tag_list_error_rate = metadata_metrics.get("validation_tag_list_errors", 0) / metadata_api_calls_total
            stuffing_rate = metadata_metrics.get("validation_keyword_stuffing", 0) / metadata_api_calls_total

            # Calculate total content errors (for backward compatibility)
            total_errors = sum(metadata_metrics.get(err_type, 0) for err_type in ["parse_failures", "empty_description_errors", "empty_tags_errors"])
            error_rate = total_errors / metadata_api_calls_total

            print_info(f"Metadata API calls: {metadata_api_calls_total}, Content Errors: {total_errors} ({error_rate:.1%}), Timeouts: {timeout_rate:.1%}")
            print_info(f"Validation Warnings - Title Mismatches: {title_mismatch_rate:.1%}, Tag List Errors: {tag_list_error_rate:.1%}, Keyword Stuffing: {stuffing_rate:.1%}")

            # Determine if we need to improve the prompt
            trigger_improvement = False
            reason = ""

            if parse_fail_rate >= METADATA_ERROR_THRESHOLD:
                trigger_improvement = True
                reason = "Parse Failures"
            elif timeout_rate >= METADATA_TIMEOUT_THRESHOLD:
                trigger_improvement = True
                reason = "Timeouts"
            elif empty_desc_rate >= METADATA_ERROR_THRESHOLD:
                trigger_improvement = True
                reason = "Empty Descriptions"
            elif empty_tags_rate >= METADATA_ERROR_THRESHOLD:
                trigger_improvement = True
                reason = "Empty Tags"
            # Check validation warning rates
            elif title_mismatch_rate >= VALIDATION_WARNING_THRESHOLD:
                trigger_improvement = True
                reason = "Title Mismatches"
            elif tag_list_error_rate >= VALIDATION_WARNING_THRESHOLD:
                trigger_improvement = True
                reason = "Tag List Errors"
            elif stuffing_rate >= VALIDATION_WARNING_THRESHOLD:
                trigger_improvement = True
                reason = "Keyword Stuffing"

            if trigger_improvement:
                print_warning(f"Metadata quality issue detected (Reason: {reason} rate above threshold). Attempting to improve prompt...", 1)
                current_prompt_text = load_or_get_seo_prompt_template()
                try:
                    # Import excel_utils for the backup function
                    import excel_utils
                    # Create a backup in the excel_backups folder
                    backup_folder = excel_utils.create_backup_folder()
                    backup_file_path = excel_utils.create_file_backup(seo_metadata_prompt_cache_path, backup_folder, suffix="_prompt")
                    if backup_file_path:
                        print_success(f"Backed up current prompt to {backup_file_path}", 2)
                    else:
                        raise Exception("Backup function returned None")
                except Exception as e:
                    # Fallback to original backup method if excel_utils fails
                    backup_file_path = seo_metadata_prompt_cache_path + ".backup"
                    try:
                        with open(backup_file_path, "w", encoding="utf-8") as bf:
                            bf.write(current_prompt_text)
                            print_success(f"Backed up current prompt to {backup_file_path} (fallback method)", 2)
                    except Exception as e2:
                        print_error(f"Error backing up prompt file: {e2}", 2)
                improved_prompt = improve_metadata_prompt(metadata_metrics)
                if improved_prompt:
                    save_seo_prompt_template(improved_prompt)
                else:
                    print_warning("Could not generate an improved prompt. Keeping current version.", 2)
            else:
                print_success("Metadata prompt is performing well based on error rates and validation warnings.")
        else:
            print_info("No metadata API calls recorded yet. Skipping prompt quality check.")

        # --- Generate Tuning Suggestions ---
        print(f"\n{Fore.CYAN}{Style.BRIGHT}--- Generating Tuning Suggestions ---{Style.RESET_ALL}")
        if performance_metrics.get("total_downloads_attempted", 0) > 0:
            # Only generate suggestions if we have some data
            suggestions = generate_tuning_suggestions(performance_metrics, config)
            if suggestions:
                print_info("Generated tuning suggestions based on performance data:")
                print(f"{Fore.YELLOW}{suggestions}{Style.RESET_ALL}")
                print_info("These suggestions have been saved to the tuning_suggestions.txt file.")
            else:
                print_info("No tuning suggestions generated. Not enough performance data yet.")
        else:
            print_info("Skipping tuning suggestions - not enough performance data yet.")

        # --- Dynamic Source Exploration: Related Channels ---
        print(f"\n{Fore.CYAN}{Style.BRIGHT}--- Dynamic Source Exploration: Related Channels ---{Style.RESET_ALL}")
        # Define how many top channels to check (make configurable?)
        num_top_channels_to_check = 3
        # Define threshold for considering a channel "top" (make configurable?)
        min_score_threshold = 5.0  # Example score threshold

        service = None  # Initialize service variable
        suggested_new_channels = []

        if YOUTUBE_API_AVAILABLE:
            top_performing_channels = [
                url for url, score in sorted(channel_scores.items(), key=lambda item: item[1], reverse=True)
                if score >= min_score_threshold
            ][:num_top_channels_to_check]  # Get top N above threshold

            if top_performing_channels:
                print_info(f"Identified {len(top_performing_channels)} top channels for related search.")
                # Authenticate only if needed
                service = get_authenticated_service()
                if service:
                    current_channels_set = set(channels)  # Set for faster lookups
                    suggested_new_channels = find_related_channels(
                        top_performing_channels,
                        channel_listing_cache,
                        current_channels_set,
                        service
                    )
                else:
                    print_warning("Cannot search for related channels, API service unavailable.")
            else:
                print_info("No channels met the criteria for related channel search.")
        else:
            print_warning("YouTube API not available. Install required packages to enable related channel search.")

        if suggested_new_channels:
            # Log suggestions for manual review
            print_success(f"Found {len(suggested_new_channels)} new channel suggestions. Logging to {suggested_channels_file_path}")
            try:
                with open(suggested_channels_file_path, "a", encoding="utf-8") as f:
                    f.write(f"\n=== Suggestions [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ===\n")
                    for url in suggested_new_channels:
                        f.write(f"{url}\n")
            except Exception as e:
                print_error(f"Failed to write suggestions to log file: {e}")
            # Note: We're not automatically adding channels, just logging them for manual review

        # --- Start Download Process ---
        print(f"\n{Fore.CYAN}{Style.BRIGHT}--- Starting Download Process ---{Style.RESET_ALL}")
        video_counter = get_last_video_index(downloaded_sheet) if downloaded_sheet else 1

        # Calculate quotas with channel scoring
        num_channels = len(channels)
        channel_quotas = {}

        if num_channels > 0:
            # Basic quota calculation
            base_quota = max(0, max_downloads // num_channels)
            remainder = max(0, max_downloads % num_channels)
            channel_quotas = {c: base_quota for c in channels}

            # Apply channel scoring if available
            if channel_scores and len(channel_scores) > 0:
                print_info("Adjusting quotas based on channel performance scores...")
                # Get channels with scores
                scored_channels = {c: channel_scores.get(c, 0) for c in channels if c in channel_scores}
                if scored_channels:
                    # Calculate total score
                    total_score = sum(scored_channels.values())
                    if total_score > 0:
                        # Allocate remainder based on scores
                        for channel, score in sorted(scored_channels.items(), key=lambda x: x[1], reverse=True):
                            if remainder > 0:
                                bonus = min(remainder, int(score * 2) + 1)  # Higher scores get more bonus
                                channel_quotas[channel] += bonus
                                remainder -= bonus
                                print_info(f"Channel '{channel}' gets +{bonus} quota due to score {score:.2f}", 1)

                    # If remainder still exists, distribute evenly
                    if remainder > 0:
                        channels_list = list(channels)
                        for i in range(remainder):
                            channel_quotas[channels_list[i % num_channels]] += 1
                else:
                    print_info("No channel scores available. Using even distribution.")
                    channels_list = list(channels)
                    for i in range(remainder):
                        channel_quotas[channels_list[i % num_channels]] += 1
            else:
                print_info("No channel scores available. Using even distribution.")
                channels_list = list(channels)
                for i in range(remainder):
                    channel_quotas[channels_list[i % num_channels]] += 1
        else:
            print_warning("No channels loaded.")

        print_success(f"Channel download quotas: {channel_quotas}")

        # Initialize run metrics
        run_metrics = {
            "date": datetime.now().isoformat(),
            "shorts_found": 0,
            "suitable_shorts": 0,
            "downloads_attempted": 0,
            "downloads_successful": 0,
            "metadata_errors": 0
        }

        # Initialize counters
        total_downloaded_this_run = 0
        channel_download_counts = {channel: 0 for channel in channels}
        downloaded_video_data = []  # Store tuples for final Excel append
        playlist_cache = load_cache(channel_processed_ids_cache_file, "Processed IDs")

        # Check FFmpeg
        if not os.path.exists(ffmpeg_path):
            print_warning(f"FFmpeg not found at: {ffmpeg_path}")
            print_warning("Downloads requiring audio/video merging may fail.")

        # --- Channel Loop ---
        for channel_url in channels:
            if total_downloaded_this_run >= max_downloads: print("\nReached total download limit."); break
            channel_quota = channel_quotas.get(channel_url, 0)
            if channel_quota <= 0 or channel_download_counts[channel_url] >= channel_quota: continue

            # Ensure playlist cache list exists for the channel
            if channel_url not in playlist_cache: playlist_cache[channel_url] = []

            print(f"\nProcessing Channel: {channel_url} (Quota Left: {channel_quota - channel_download_counts[channel_url]})")

            # --- Channel Listing Cache Logic (Permanent) ---
            fetched_entries = []
            use_cache = False
            # Check if channel data exists in the permanent listing cache
            if channel_url in channel_listing_cache and isinstance(channel_listing_cache.get(channel_url), list):
                fetched_entries = channel_listing_cache[channel_url]
                if fetched_entries: # Ensure cached list is not empty
                    use_cache = True
                    print(f"  Using permanently cached video list ({len(fetched_entries)} entries).")
                else:
                    # Cached list is empty, treat as uncached
                    print(f"  Cached list for {channel_url} is empty. Will refetch.")
                    # Optionally remove empty entry: del channel_listing_cache[channel_url]
            # --- End Cache Check ---

            # --- Fetch list ONLY if not found in cache or cache was empty ---
            if not use_cache:
                playlist_limit = YT_PLAYLIST_FETCH_LIMIT
                print(f"Fetching video list (up to {playlist_limit}, this may take time)...")
                # Ensure shorts_url is defined before use
                shorts_url = f"{channel_url}/shorts"
                ydl_opts_fetch = { 'extract_flat': 'discard_in_playlist', 'playlistend': playlist_limit, 'quiet': True, 'no_warnings': True, 'ignoreerrors': True, 'skip_download': True, 'forcejson': True, 'socket_timeout': 60, }
                try:
                    with yt_dlp.YoutubeDL(ydl_opts_fetch) as ydl:
                        result = ydl.extract_info(shorts_url, download=False) # Use shorts_url here

                    if result and 'entries' in result and isinstance(result['entries'], list):
                        fetched_entries = [e for e in result['entries'] if e and isinstance(e, dict)]
                        print(f"  Finished fetching list. Found {len(fetched_entries)} entries.")
                        if fetched_entries: # Save ONLY if entries were found
                            channel_listing_cache[channel_url] = fetched_entries
                            save_cache(channel_listing_cache, channel_listing_cache_file)
                            print(f"  Saved fetched list to permanent cache for {channel_url}.")
                        else: log_error(f"Fetched 0 valid entries for {channel_url}.")
                    else:
                        print(f"  No entries found/extracted for {channel_url}.")
                        log_error(f"No entries/bad format: {channel_url}.")
                        continue
                except Exception as e:
                    print(f"  Error fetching list: {e}")
                    log_error(f"Error fetching list {channel_url}: {e}")
                    continue
            # --- End Fetch list ---

            if not fetched_entries: continue # Skip if still no entries

            # Sort and process entries
            def get_view_count(entry):
                try: return int(entry.get('view_count')) if entry.get('view_count') is not None else 0;
                except: return 0
            sorted_entries = sorted(fetched_entries, key=get_view_count, reverse=True)

            # --- Video Loop ---
            for entry in sorted_entries:
                if total_downloaded_this_run >= max_downloads: break
                if channel_download_counts[channel_url] >= channel_quota: break

                video_id = entry.get("id"); original_title = entry.get('title', '').strip(); uploader = entry.get('uploader', DEFAULT_UPLOADER_NAME).strip()
                if not video_id or not original_title: continue

                if not isinstance(playlist_cache.get(channel_url), list): playlist_cache[channel_url] = [] # Sanity check
                if str(video_id) in playlist_cache[channel_url]: continue # Check Processed ID cache
                if (original_title, uploader) in previously_downloaded_videos: # Check Title/Uploader cache
                    if str(video_id) not in playlist_cache[channel_url]: playlist_cache[channel_url].append(str(video_id))
                    continue

                video_url = entry.get('url')
                if not video_url: continue

                # --- Prepare filenames & Check existence ---
                video_file_name = f"video{video_counter}.mp4"; video_file_path = os.path.join(download_folder, video_file_name)
                metadata_file_name = f"video{video_counter}.json"; metadata_file_path = os.path.join(metadata_folder, metadata_file_name)
                video_exists = os.path.exists(video_file_path); metadata_exists = os.path.exists(metadata_file_path)

                if video_exists and metadata_exists: video_counter += 1; continue # Skip if both exist
                elif video_exists and not metadata_exists: # Regenerate metadata if needed
                    print(f"  Video {video_file_name} exists, metadata missing. Regenerating...")
                    # Generate metadata with category suggestion
                    seo_metadata = generate_metadata_with_timeout_v2(original_title, uploader, original_title)

                    # --- Get Playlist Suggestion ---
                    if seo_metadata and not ("error" in seo_metadata.get("tags",[]) or "timeout" in seo_metadata.get("tags",[])):
                        # Prepare list of existing titles from cache for Gemini
                        current_existing_titles = [
                            title for key, title in playlist_data_cache.items()
                            if not key.startswith("NEW: ") and title is not None
                        ]

                        playlist_suggestion = get_playlist_suggestion(
                            seo_metadata.get("title", original_title),
                            seo_metadata.get("description", ""),
                            seo_metadata.get("tags", []),
                            current_existing_titles
                        )

                        if playlist_suggestion:
                            if playlist_suggestion.startswith("NEW: "):
                                # It's a new suggestion, store the title
                                seo_metadata["target_playlist"] = playlist_suggestion
                                # Add to cache so we don't suggest it again immediately
                                if playlist_suggestion not in playlist_data_cache:
                                    playlist_data_cache[playlist_suggestion] = None # Mark as suggested
                                    save_cache(playlist_data_cache, playlist_data_cache_path, "Playlist Data")
                            else:
                                # It's an existing title, find its ID from the cache
                                found_id = None
                                for p_id, p_title in playlist_data_cache.items():
                                    if p_title == playlist_suggestion and not p_id.startswith("NEW: "):
                                        found_id = p_id
                                        break
                                if found_id:
                                    seo_metadata["target_playlist"] = found_id
                                    print_info(f"Using existing playlist ID: {found_id}", 4)
                                else:
                                    print_warning(f"Matched title '{playlist_suggestion}' but couldn't find its ID in cache. Skipping playlist.", 4)

                    generated_metadata = save_metadata_file(entry, video_counter, seo_metadata, channel_url)
                    if generated_metadata:
                        ts = generated_metadata.get("download_timestamp", datetime.now().isoformat()); views = generated_metadata.get('view_count', 0)
                        downloaded_video_data.append((f"video{video_counter}", generated_metadata.get("optimized_title"), ts, views, generated_metadata.get("uploader"), generated_metadata.get("original_title")))
                        playlist_cache[channel_url].append(str(video_id)); previously_downloaded_videos.add((original_title, uploader))
                        total_downloaded_this_run += 1; channel_download_counts[channel_url] += 1; video_counter += 1
                    else: log_error(f"Failed regenerating metadata for {video_file_name}."); video_counter += 1
                    continue

                # --- Download ---
                print(f"  Attempting download: {video_file_name}...")
                # Format string updated slightly for clarity/preference, functionality same
                ydl_opts_download = {
                    'format': 'bestvideo[height>=1080][ext=mp4]+bestaudio[ext=m4a]/bestvideo[ext=mp4]+bestaudio/bestvideo[height>=720][ext=mp4]+bestaudio[ext=m4a]/bestvideo[ext=mp4]+bestaudio/best',
                    'outtmpl': video_file_path, 'quiet': True, 'no_warnings': True,
                    'ignoreerrors': False, 'ffmpeg_location': ffmpeg_path,
                    'merge_output_format': 'mp4', 'retries': 2,
                    'writeinfojson': True,  # Enable info.json for tag extraction
                 }
                download_success = False
                try:
                    with yt_dlp.YoutubeDL(ydl_opts_download) as ydl_download: ydl_download.download([video_url])
                    if os.path.exists(video_file_path) and os.path.getsize(video_file_path) > 1024: download_success = True
                except Exception as e:
                    log_error(f"Download error {video_id} ({original_title}): {e}")
                if download_success:
                    # --- Post-Download ---
                    print(f"  Download ok. Generating metadata...")
                    # Generate metadata with category suggestion
                    seo_metadata = generate_metadata_with_timeout_v2(original_title, uploader, original_title)

                    # --- Get Playlist Suggestion ---
                    if seo_metadata and not ("error" in seo_metadata.get("tags",[]) or "timeout" in seo_metadata.get("tags",[])):
                        # Prepare list of existing titles from cache for Gemini
                        current_existing_titles = [
                            title for key, title in playlist_data_cache.items()
                            if not key.startswith("NEW: ") and title is not None
                        ]

                        playlist_suggestion = get_playlist_suggestion(
                            seo_metadata.get("title", original_title),
                            seo_metadata.get("description", ""),
                            seo_metadata.get("tags", []),
                            current_existing_titles
                        )

                        if playlist_suggestion:
                            if playlist_suggestion.startswith("NEW: "):
                                # It's a new suggestion, store the title
                                seo_metadata["target_playlist"] = playlist_suggestion
                                # Add to cache so we don't suggest it again immediately
                                if playlist_suggestion not in playlist_data_cache:
                                    playlist_data_cache[playlist_suggestion] = None # Mark as suggested
                                    save_cache(playlist_data_cache, playlist_data_cache_path, "Playlist Data")
                            else:
                                # It's an existing title, find its ID from the cache
                                found_id = None
                                for p_id, p_title in playlist_data_cache.items():
                                    if p_title == playlist_suggestion and not p_id.startswith("NEW: "):
                                        found_id = p_id
                                        break
                                if found_id:
                                    seo_metadata["target_playlist"] = found_id
                                    print_info(f"Using existing playlist ID: {found_id}", 4)
                                else:
                                    print_warning(f"Matched title '{playlist_suggestion}' but couldn't find its ID in cache. Skipping playlist.", 4)

                    generated_metadata = save_metadata_file(entry, video_counter, seo_metadata, channel_url)
                    if generated_metadata:
                        ts = generated_metadata.get("download_timestamp", datetime.now().isoformat()); views = generated_metadata.get('view_count', 0)
                        downloaded_video_data.append((f"video{video_counter}", generated_metadata.get("optimized_title"), ts, views, generated_metadata.get("uploader"), generated_metadata.get("original_title")))
                        playlist_cache[channel_url].append(str(video_id)); previously_downloaded_videos.add((original_title, uploader))

                        # --- Tag Extraction & Keyword Pool Update ---
                        info_json_path = os.path.splitext(video_file_path)[0] + ".info.json"
                        if os.path.exists(info_json_path):
                            try:
                                with open(info_json_path, 'r', encoding='utf-8') as f:
                                    video_info = json.load(f)
                                video_tags = video_info.get('tags', [])
                                if video_tags:
                                    print_info(f"Processing {len(video_tags)} tags from info file...", 3)
                                    new_unique_tags_found = set()
                                    lower_to_original_keyword_map = {kw.lower(): kw for kw in keyword_frequency.keys()}

                                    # Identify potentially new keywords
                                    for tag in video_tags:
                                        tag_strip = tag.strip()
                                        tag_lower = tag_strip.lower()
                                        if not tag_lower:
                                            continue

                                        # Use the filtering criteria loaded earlier
                                        is_relevant = any(sub in tag_lower for sub in required_substrings_lower)
                                        is_not_social = tag_lower not in social_media_lower
                                        is_not_generic = tag_lower not in generic_terms_lower
                                        is_not_seed = tag_lower != seed_niche.lower()
                                        is_new = tag_lower not in lower_to_original_keyword_map
                                        is_long_enough = len(tag_lower.replace(" ", "")) > 3

                                        if is_relevant and is_not_social and is_not_generic and is_not_seed and is_new and is_long_enough:
                                            new_unique_tags_found.add(tag_strip)

                                    # Add new keywords / replace old ones
                                    if new_unique_tags_found:
                                        print_info(f"Found {len(new_unique_tags_found)} relevant new unique tags from channel video.", 4)
                                        tags_to_add_list = list(new_unique_tags_found)
                                        added_count = 0
                                        for tag_to_add in tags_to_add_list:
                                            if len(keyword_frequency) < max_keywords:
                                                keyword_frequency[tag_to_add] = 0  # Add with score 0
                                                added_count += 1
                                            else:
                                                # Find lowest scored keyword to replace
                                                lowest_kw = min(keyword_frequency.items(), key=lambda x: x[1])
                                                if lowest_kw[1] <= 0:  # Only replace if score is 0 or negative
                                                    del keyword_frequency[lowest_kw[0]]
                                                    keyword_frequency[tag_to_add] = 0
                                                    added_count += 1
                                                    print_info(f"Replaced low-scoring keyword '{lowest_kw[0]}' with new tag '{tag_to_add}'", 4)

                                        if added_count > 0:
                                            print_success(f"Added {added_count} new keywords from tags. Total keywords: {len(keyword_frequency)}.", 4)
                                            save_cache(keyword_frequency, keywords_cache_file_path, "Keyword Frequency")  # Save immediately after adding
                                else:
                                    print_info("No tags found in info file.", 3)
                            except json.JSONDecodeError as json_e:
                                print_error(f"Error decoding info.json '{info_json_path}': {json_e}", 3)
                            except Exception as tag_e:
                                print_error(f"Error processing tags from info.json '{info_json_path}': {tag_e}", 3, include_traceback=True)
                            finally:  # Cleanup info.json AFTER processing
                                try:
                                    if os.path.exists(info_json_path):
                                        os.remove(info_json_path)
                                except OSError as e_del:
                                    print_warning(f"Error deleting info.json '{info_json_path}': {e_del}", 4)
                        else:
                            print_warning(f"Info file not found, skipping tag extraction: {info_json_path}", 3)
                        # --- End Tag Extraction ---

                        video_counter += 1; total_downloaded_this_run += 1; channel_download_counts[channel_url] += 1
                        print(f"  Processed video {video_counter-1} successfully.") # Confirmation log
                    else: # Metadata failed, delete video
                        log_error(f"Metadata failed {video_id}. Deleting video.");
                        if os.path.exists(video_file_path):
                            try: os.remove(video_file_path); print("  Deleted orphaned video.")
                            except OSError as e: log_error(f"Could not delete orphaned video: {e}")
                elif os.path.exists(video_file_path): # Cleanup partial download
                        try:
                            os.remove(video_file_path)
                        except OSError:
                            pass # Or log error
            # --- End Video Loop ---

            print(f"  Finished channel {channel_url}. Saving intermediate Processed ID cache.")
            save_cache(playlist_cache, channel_processed_ids_cache_file) # Save only processed IDs cache here

        # --- End Channel Loop ---
        print("\n--- Finished processing channels or reached download limit. ---")

    except KeyboardInterrupt: print("\n--- Interrupted by user (Ctrl+C). ---"); log_error("Interrupted by user.")
    except Exception as e: error_message = f"FATAL error in main loop: {e}"; print(error_message); log_error(error_message); print("\n--- Traceback ---"); traceback.print_exc(); print("-----------------\n"); log_error(f"Traceback:\n{traceback.format_exc()}")
    finally:
        # --- Final Save Operations ---
        print("\n--- Entering final cleanup and save phase. ---")
        # Save Excel Data
        if wb and downloaded_sheet and downloaded_video_data:
            print(f"Attempting final Excel update ({len(downloaded_video_data)} new)...")
            try:
                # Try to import excel_utils module
                try:
                    import excel_utils
                    excel_utils_available = True
                    print_info("Using excel_utils module for robust Excel saving")
                except ImportError:
                    excel_utils_available = False
                    print_warning("excel_utils module not available. Using fallback Excel saving.")

                # Sort the data by view count
                def get_sort_key(item):
                    try: return int(item[3]) if len(item)>3 and item[3] is not None else 0;
                    except: return 0
                print("  Sorting data...")
                downloaded_video_data.sort(key=get_sort_key, reverse=True)

                # Append rows to the sheet
                print("  Appending rows...")
                for row_data in downloaded_video_data:
                    if len(row_data) == len(EXPECTED_DOWNLOADED_HEADERS):
                        downloaded_sheet.append(row_data)
                    else:
                        log_error(f"Skipping Excel row, wrong count: {row_data}")

                # Save the workbook using excel_utils if available
                print("  Saving workbook...")
                if excel_utils_available:
                    # Extract workbook data for backup in case save fails
                    def extract_data(wb):
                        data = {}
                        for sheet_name in wb.sheetnames:
                            sheet = wb[sheet_name]
                            data[sheet_name] = []
                            for row in sheet.iter_rows(values_only=True):
                                data[sheet_name].append(list(row))
                        return data

                    # Use the robust save mechanism
                    try:
                        # Try to close Excel before saving
                        excel_utils.close_excel_processes_with_file(excel_file)

                        # Use the safe_save_workbook function with extract_data as a fallback
                        if excel_utils.safe_save_workbook(wb, excel_file, close_excel=True, create_backup=True):
                            print_success("Excel saved successfully using excel_utils.")
                        else:
                            # If safe_save_workbook fails, try to save to an alternative file
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            alt_file_path = f"{os.path.splitext(excel_file)[0]}_alt_{timestamp}.xlsx"

                            print_warning(f"Attempting to save to alternative file: {alt_file_path}")
                            if excel_utils.safe_save_workbook(wb, alt_file_path, close_excel=False, create_backup=False):
                                print_success(f"Saved to alternative file: {alt_file_path}")
                            else:
                                # If both Excel saves failed, save as JSON
                                data = extract_data(wb)
                                backup_file = os.path.join(constants.EXCEL_BACKUPS_DIR, f"excel_backup_data_{datetime.now():%Y%m%d_%H%M%S}.json")
                                with open(backup_file, "w", encoding='utf-8') as bf:
                                    json.dump(data, bf, indent=4, default=str)
                                    print_success(f"Saved data as JSON backup: {backup_file}")
                    except Exception as e:
                        # If all save methods failed, create a JSON backup
                        backup_file = os.path.join(constants.EXCEL_BACKUPS_DIR, f"excel_backup_data_{datetime.now():%Y%m%d_%H%M%S}.json")
                        print(f"All Excel save methods failed. Creating JSON backup: {backup_file}")
                        try:
                            with open(backup_file, "w", encoding='utf-8') as bf:
                                json.dump(downloaded_video_data, bf, indent=4)
                                log_error(f"Saved backup {backup_file}.")
                                print("Backup saved.")
                        except Exception as be:
                            log_error(f"CRITICAL: Failed backup save: {be}")
                            print(f"CRITICAL: Failed backup save: {be}")
                else:
                    # Fallback to original save mechanism with simple retry
                    max_retries = 3
                    for attempt in range(max_retries):
                        try:
                            wb.save(excel_file)
                            print_success(f"Excel saved successfully (attempt {attempt+1}).")
                            break
                        except PermissionError as pe:
                            if attempt < max_retries - 1:
                                print_warning(f"PermissionError saving Excel (attempt {attempt+1}/{max_retries}): {pe}")
                                print_info(f"Retrying in 2 seconds...")
                                time.sleep(2)
                            else:
                                raise
                    else:
                        # If all retries failed, create a backup file
                        backup_file = os.path.join(constants.EXCEL_BACKUPS_DIR, f"excel_backup_data_{datetime.now():%Y%m%d_%H%M%S}.json")
                        print(f"All Excel save attempts failed. Creating JSON backup: {backup_file}")
                        try:
                            with open(backup_file, "w", encoding='utf-8') as bf:
                                json.dump(downloaded_video_data, bf, indent=4)
                                log_error(f"Saved backup {backup_file}.")
                                print("Backup saved.")
                        except Exception as be:
                            log_error(f"CRITICAL: Failed backup save: {be}")
                            print(f"CRITICAL: Failed backup save: {be}")
            except Exception as e:
                error_message = f"CRITICAL: Error saving Excel: {e}"
                print(error_message)
                log_error(error_message)
                traceback.print_exc()
                log_error(f"Excel save Traceback:\n{traceback.format_exc()}")

                # Create a JSON backup as last resort
                backup_file = os.path.join(constants.EXCEL_BACKUPS_DIR, f"excel_backup_data_{datetime.now():%Y%m%d_%H%M%S}.json")
                print(f"Attempting backup to {backup_file}...")
                try:
                    with open(backup_file, "w", encoding='utf-8') as bf:
                        json.dump(downloaded_video_data, bf, indent=4)
                        log_error(f"Saved backup {backup_file}.")
                        print("Backup saved.")
                except Exception as be:
                    log_error(f"CRITICAL: Failed backup save: {be}")
                    print(f"CRITICAL: Failed backup save: {be}")

        elif not downloaded_video_data:
            print("\nNo new videos processed. No Excel data to save.")

        # Save Caches
        print("\nSaving final caches...")
        if isinstance(playlist_cache, dict):
            save_cache(playlist_cache, channel_processed_ids_cache_file, "Processed IDs")
        else:
            print_warning("Playlist cache invalid type. Skipping save.")

        # Save the channel listing cache one last time
        if isinstance(channel_listing_cache, dict):
            save_cache(channel_listing_cache, channel_listing_cache_file, "Channel Listing")
        else:
            print_warning("Channel listing cache invalid type. Skipping save.")

        # Save the keyword frequency cache
        if isinstance(keyword_frequency, dict):
            save_cache(keyword_frequency, keywords_cache_file_path, "Keyword Frequency")
        else:
            print_warning("Keyword frequency cache invalid type. Skipping save.")

        # Save the playlist data cache
        if isinstance(playlist_data_cache, dict):
            save_cache(playlist_data_cache, playlist_data_cache_path, "Playlist Data")
        else:
            print_warning("Playlist data cache invalid type. Skipping save.")

        # Update and save performance metrics
        if 'run_metrics' in locals() and isinstance(run_metrics, dict):
            # Update run metrics with final counts
            run_metrics["downloads_successful"] = total_downloaded_this_run

            # Update overall metrics
            performance_metrics["runs"].append(run_metrics)
            performance_metrics["total_shorts_found"] += run_metrics.get("shorts_found", 0)
            performance_metrics["total_suitable_shorts"] += run_metrics.get("suitable_shorts", 0)
            performance_metrics["total_downloads_attempted"] += run_metrics.get("downloads_attempted", 0)
            performance_metrics["total_successful_downloads"] += run_metrics.get("downloads_successful", 0)
            performance_metrics["total_metadata_errors"] += run_metrics.get("metadata_errors", 0)
            performance_metrics["total_download_errors"] += run_metrics.get("downloads_attempted", 0) - run_metrics.get("downloads_successful", 0)

            # Update channel performance scores
            if channel_scores:
                performance_metrics["channel_performance"] = channel_scores

            # Save performance metrics
            save_performance_metrics(performance_metrics)
            print_success("Performance metrics updated and saved.")

            # Keep only last N runs to prevent file growth
            if len(performance_metrics["runs"]) > MAX_RUNS_TO_KEEP:
                performance_metrics["runs"] = performance_metrics["runs"][-MAX_RUNS_TO_KEEP:]
                save_performance_metrics(performance_metrics)
                print_info(f"Trimmed performance metrics to last {MAX_RUNS_TO_KEEP} runs.")

        # --- Cleanup Old Excel Backups ---
        try:
            print(f"\n{Fore.CYAN}--- Cleaning Up Old Excel Backups ---{Style.RESET_ALL}")
            # Try to import excel_utils again just in case
            try:
                import excel_utils
                excel_utils_available = True
            except ImportError:
                excel_utils_available = False
                print_warning("excel_utils module not available. Skipping backup cleanup.")

            if excel_utils_available:
                # Call cleanup, keeping backups for the last 7 days
                deleted, remaining = excel_utils.cleanup_old_backups(days_to_keep=7)
                print_info(f"Excel backup cleanup: Deleted {deleted} old backups, Remaining {remaining} backups.")
        except Exception as cleanup_e:
            print_error(f"Error during backup cleanup: {cleanup_e}")

        print(f"\n{Fore.GREEN}{Style.BRIGHT}--- Channel Downloader Script Execution Finished ---{Style.RESET_ALL}")

# --- Script Entry Point ---
# --- Performance/Metrics Functions ---
def load_performance_metrics():
    """Loads overall performance metrics from the JSON file."""
    default_metrics = {
        "runs": [], "total_shorts_found": 0, "total_suitable_shorts": 0,
        "total_downloads_attempted": 0, "total_successful_downloads": 0,
        "total_metadata_api_calls": 0, "total_metadata_errors": 0,
        "total_download_errors": 0, "channel_performance": {},
        "last_parameter_update": ""
    }
    try:
        if os.path.exists(performance_metrics_file_path):
            with open(performance_metrics_file_path, "r", encoding="utf-8") as f:
                metrics = json.load(f)
            for key, value in default_metrics.items():
                metrics.setdefault(key, value) # Ensure all keys exist
            return metrics
        else:
            return default_metrics
    except Exception as e:
        print_warning(f"Error loading performance metrics: {e}. Using default values.")
        return default_metrics

def save_performance_metrics(metrics):
    """Saves overall performance metrics to the JSON file."""
    try:
        with open(performance_metrics_file_path, "w", encoding="utf-8") as f:
            json.dump(metrics, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print_error(f"Error saving performance metrics: {e}", 1)

def load_metadata_metrics():
    """Loads metadata generation metrics from the JSON file."""
    default_metrics = {
        "total_api_calls": 0, "parse_failures": 0, "timeouts": 0,
        "empty_title_errors": 0, "empty_description_errors": 0, "empty_tags_errors": 0,
        "last_run_date": datetime.now().isoformat(), "error_samples": [],
        "total_api_calls_previous": 0, "total_errors_previous": 0
    }
    try:
        if os.path.exists(metadata_metrics_file_path):
            with open(metadata_metrics_file_path, "r", encoding="utf-8") as f:
                metrics = json.load(f)
            for key, value in default_metrics.items():
                metrics.setdefault(key, value)
            return metrics
        else:
            return default_metrics
    except Exception as e:
        print_warning(f"Error loading metadata metrics: {e}. Using default values.")
        return default_metrics

def save_metadata_metrics(metrics):
    """Saves metadata generation metrics to the JSON file."""
    try:
        metrics["last_run_date"] = datetime.now().isoformat()
        with open(metadata_metrics_file_path, "w", encoding="utf-8") as f:
            json.dump(metrics, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print_error(f"Error saving metadata metrics: {e}", 1)

def add_error_sample(metrics, error_type, error_details, video_title):
    """Adds an error sample to the metadata metrics."""
    if "error_samples" not in metrics:
        metrics["error_samples"] = []
    error_sample = {
        "type": error_type,
        "details": error_details,
        "video_title": video_title,
        "timestamp": datetime.now().isoformat()
    }
    metrics["error_samples"].append(error_sample)
    metrics["error_samples"] = metrics["error_samples"][-MAX_ERROR_SAMPLES:] # Keep only last N
# --- End Performance/Metrics Functions ---

# --- Correlation Cache Functions ---
def load_correlation_cache():
    """Loads the upload correlation cache that maps video index to YouTube ID."""
    default_cache = []
    if not os.path.exists(upload_correlation_cache_path):
        return default_cache
    try:
        with open(upload_correlation_cache_path, "r", encoding="utf-8") as f:
            content = f.read()
            if not content:
                return default_cache
            cache = json.loads(content)
            if not isinstance(cache, list):
                print_warning(f"Correlation cache file '{os.path.basename(upload_correlation_cache_path)}' invalid format.")
                return default_cache
            return cache
    except json.JSONDecodeError:
        print_error(f"Error decoding correlation cache '{os.path.basename(upload_correlation_cache_path)}'.")
        return default_cache
    except Exception as e:
        print_error(f"Error loading correlation cache: {e}", include_traceback=True)
        return default_cache

def save_correlation_cache(cache_data):
    """Saves the upload correlation cache."""
    try:
        with open(upload_correlation_cache_path, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print_error(f"Error saving correlation cache: {e}", include_traceback=True)

def cleanup_correlation_cache(days_to_keep=7):
    """Cleans up old entries from the correlation cache."""
    cache = load_correlation_cache()
    if not cache:
        return

    now = datetime.now()
    cutoff_date = now - timedelta(days=days_to_keep)
    original_count = len(cache)
    cleaned_cache = []
    removed_count = 0

    for entry in cache:
        try:
            added_time = datetime.fromisoformat(entry.get("added_timestamp"))
            if added_time >= cutoff_date:
                cleaned_cache.append(entry)
            else:
                removed_count += 1
        except (ValueError, TypeError, KeyError):
            cleaned_cache.append(entry)
            print_warning(f"Invalid timestamp in correlation cache entry: {entry.get('video_index', 'Unknown')}. Keeping.")

    if removed_count > 0:
        save_correlation_cache(cleaned_cache)
        print_info(f"Cleaned up {removed_count} old entries (>{days_to_keep} days) from correlation cache. {len(cleaned_cache)} remaining.")
    else:
        print_info("No old entries found in correlation cache to cleanup.")
# --- End Correlation Cache Functions ---

# --- SEO Prompt Handling ---
def load_or_get_seo_prompt_template():
    """Loads or gets the SEO metadata prompt template."""
    global _current_seo_prompt_template
    if _current_seo_prompt_template:
        return _current_seo_prompt_template

    prompt_loaded = False
    if os.path.exists(seo_metadata_prompt_cache_path):
        try:
            with open(seo_metadata_prompt_cache_path, "r", encoding="utf-8") as f:
                _current_seo_prompt_template = f.read()
            # Check if the prompt has the required placeholders
            if "{video_topic}" in _current_seo_prompt_template and "{uploader_name}" in _current_seo_prompt_template and "{channel_name}" in _current_seo_prompt_template and "{channel_topic}" in _current_seo_prompt_template:
                print_info(f"Loaded SEO metadata prompt from cache: {os.path.basename(seo_metadata_prompt_cache_path)}")
                prompt_loaded = True
            else:
                print_warning(f"Cached SEO prompt invalid (missing placeholders). Using default.")
                _current_seo_prompt_template = None
        except Exception as e:
            print_warning(f"Error loading cached SEO prompt: {e}. Using default.")
            _current_seo_prompt_template = None

    if not prompt_loaded:
        print_info("Using default inline SEO metadata prompt.")
        # Use the prompt definition from the channel script, ensuring placeholders match
        _current_seo_prompt_template = f"""
        do not include any explanation or any other text. just give me the metadata in below format.
        only apply the below format. do not include any other text or explanation.
        Generate SEO-optimized metadata for a YouTube Shorts video in the following structured format:
        You are a YOUTUBE SEO EXPERT A GURU one in million. you have insight knowledge of youtube shorts.
        you know how the ranking algorithm works and how to get more views and engagement.
        you know how creator like mrbeast, tseries, and other top creators get more views and engagement.
        your master of youtube shorts. you have worked with big creator know all secrets of youtube shorts.
        you have worked in google youtube team and you know all secrets of youtube shorts.
        Our Channel Name is "{{channel_name}}" and we are a channel about {{channel_topic}}. <-- Use channel_name/topic placeholders
        include a copyright fair use disclaimer in the description.
        APPLY ALL OF THE ABOVE KNOWLEDGE AND SECRETS TO BELOW metadata.

        <metadata>
            <title>
                Create an engaging, fast-paced, and action-driven title (max 100 chars incl. #Shorts) with a high CTR based on the video topic: '{{video_topic}}'.
                Use keywords for '{{channel_topic}}'. Use relevant emojis (🔥, 💪, 👀), numbers, power words (BEST, HOT, ULTIMATE, SECRET, TRY THIS). Add "#Shorts" at the end.
            </title>
            <description>
                Write an SEO-optimized description (max 4500 chars):
                    * Start with the optimized video title.
                    * 2-3 sentence engaging summary about '{{video_topic}}' and '{{channel_topic}}', using keywords/LSI naturally.
                    * **Include credit: "Credit: {{uploader_name}}"** <-- Use uploader_name placeholder
                    * Include copyright disclaimer:
                      --------------【Copyright Disclaimer】-------------
                      All the videos, songs, images, and graphics used in the video belong to
                      their respective owners and I or this channel "{{channel_name}}" does not claim any right over them.
                      Copyright Disclaimer under section 107 of the Copyright Act of 1976, allowance is made for "fair use" for purposes such as criticism, comment, news reporting, teaching, scholarship, education and research. Fair use is a use permitted by copyright statute that might otherwise be infringing.
                    * After disclaimer, add 10-15 relevant hashtags (inspired by: {seo_example_hashtags_raw}). <-- Insert example hashtags from config
                    * Add heading "Tags Used in Video :-" and list all tags from <tags> section below, comma-separated.
                    * End with a Call to Action (e.g., "👍 Like & Subscribe to {{channel_name}}!").
                    * Add heading "Ignored Hashtags :-" followed by a diverse list of relevant hashtags.
            </description>
            <tags>
                Suggest 15-25 SEO-friendly tags (comma-separated, max 500 chars total).
                * Start with keywords for '{{video_topic}}'. Include tags for '{{channel_topic}}' and channel name '{{channel_name}}'.
                * Use mix of general/specific tags. Inspire from: {seo_example_tags_raw} <-- Insert example tags from config
            </tags>
        </metadata>

        **Video Topic**: {{video_topic}}
        """
    return _current_seo_prompt_template

def save_seo_prompt_template(prompt_text):
    """Saves the SEO metadata prompt template to the cache file."""
    global _current_seo_prompt_template
    try:
        with open(seo_metadata_prompt_cache_path, "w", encoding="utf-8") as f:
            f.write(prompt_text)
        _current_seo_prompt_template = prompt_text
        print_success(f"Saved updated SEO metadata prompt to cache: {os.path.basename(seo_metadata_prompt_cache_path)}")
    except Exception as e:
        print_error(f"Error saving updated SEO metadata prompt to cache: {e}")

def improve_metadata_prompt(error_metrics):
    """Uses Gemini to improve the SEO metadata prompt based on error patterns."""
    current_prompt = load_or_get_seo_prompt_template()
    if not current_prompt:
        print_error("Cannot improve prompt: Failed to load current SEO prompt.")
        return None

    total_calls = error_metrics.get("total_api_calls", 0)
    if total_calls == 0:
        print_warning("No API calls recorded yet. Cannot improve prompt.")
        return None

    error_summary = [f"Total API calls: {total_calls}"]
    for error_type in ["parse_failures", "timeouts", "empty_description_errors", "empty_tags_errors"]:
        count = error_metrics.get(error_type, 0)
        rate = count / total_calls if total_calls > 0 else 0
        error_summary.append(f"{error_type}: {count} ({rate:.1%})")

    error_samples = error_metrics.get("error_samples", [])
    if error_samples:
        error_summary.append("\nRecent error samples:")
        for i, sample in enumerate(error_samples[-5:], 1):
            error_summary.append(f"Sample {i}: {sample.get('type')} - {sample.get('details')}")
    error_summary_text = "\n".join(error_summary)

    meta_prompt = f"""
    Review the following prompt used to generate SEO-optimized YouTube Shorts metadata for a channel-based downloader:

    ```
    {current_prompt}
    ```

    Based on the following issues observed:

    {error_summary_text}

    Provide an improved version. Focus on clarity, robustness (avoiding parsing failures/empty results), and structure (explicit XML).
    Ensure instructions for title, description (keywords, credit, disclaimer, hashtags, tags list, CTA), and tags are clear and reinforce SEO goals.
    The prompt uses placeholders: {{channel_name}}, {{channel_topic}}, {{video_topic}}, {{uploader_name}}. Ensure these remain.
    Perhaps add a step to double-check all elements are present before finishing.

    Provide ONLY the improved prompt text, without explanations.
    """
    try:
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(meta_prompt)
        improved_prompt = response.text.strip()

        # Basic validation
        if "<metadata>" not in improved_prompt or "<title>" not in improved_prompt or "<description>" not in improved_prompt or "<tags>" not in improved_prompt:
            print_error("Generated prompt missing required XML tags. Keeping current.", 1)
            return None

        # Check channel downloader specific placeholders
        if "{video_topic}" not in improved_prompt or "{uploader_name}" not in improved_prompt or "{channel_name}" not in improved_prompt or "{channel_topic}" not in improved_prompt:
             print_error("Generated prompt missing required placeholders. Keeping current.", 1)
             return None

        return improved_prompt
    except Exception as e:
        print_error(f"Error generating improved SEO metadata prompt: {e}", 1, include_traceback=True)
        return None
# --- End SEO Prompt Handling ---

# --- Category Suggestion ---
def get_suggested_category(title: str, description: str):
    """Uses Gemini to suggest the most appropriate YouTube category based on video content."""
    if not title or not description:
        print_warning("Cannot suggest category: Title or Description empty.", 2)
        return None

    KNOWN_CATEGORIES = [
        "Film & Animation", "Autos & Vehicles", "Music", "Pets & Animals",
        "Sports", "Travel & Events", "Gaming", "People & Blogs",
        "Comedy", "Entertainment", "News & Politics", "Howto & Style",
        "Education", "Science & Technology", "Nonprofits & Activism",
        "Movies", "Shows"
    ]
    KNOWN_CATEGORIES_LOWER = {cat.lower() for cat in KNOWN_CATEGORIES}

    # Improved prompt that explicitly lists valid categories
    prompt = f"""
    YouTube Shorts Title/Description:
    Title: {title}
    Description: {description[:1000]}

    Based on the content above, select ONE most appropriate YouTube video category from this EXACT list:
    {', '.join(KNOWN_CATEGORIES)}

    Output ONLY the category name exactly as written in the list above. Do not add any explanation.
    """

    try:
        print_info("Requesting category suggestion...", 3)
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        suggested_cat_raw = response.text.strip()

        if suggested_cat_raw and suggested_cat_raw.lower() in KNOWN_CATEGORIES_LOWER:
            for known_cat in KNOWN_CATEGORIES:
                if known_cat.lower() == suggested_cat_raw.lower():
                    print_success(f"Suggested category: {known_cat}", 3)
                    return known_cat
            print_warning(f"Could not find exact capitalization for '{suggested_cat_raw}'. Using as is.", 3)
            return suggested_cat_raw
        elif suggested_cat_raw:
            print_warning(f"Gemini suggested unknown category: '{suggested_cat_raw}'. Ignoring.", 3)
            return None
        else:
            print_warning("Gemini returned empty category suggestion.", 3)
            return None
    except Exception as e:
        print_error(f"Error getting category suggestion: {e}", 3)
        return None
# --- End Category Suggestion ---

# --- Tuning Suggestions ---
def generate_performance_summary(metrics):
    """Generates a summary of performance metrics for tuning suggestions."""
    summary = ["=== Channel Downloader Performance Summary ==="]
    runs_data = metrics.get('runs', [])
    summary.append(f"Total runs recorded: {len(runs_data)}")

    # Overall metrics
    summary.append(f"Total videos found (all runs): {metrics.get('total_shorts_found', 0)}")
    summary.append(f"Total suitable videos (all runs): {metrics.get('total_suitable_shorts', 0)}")
    summary.append(f"Total downloads attempted (all runs): {metrics.get('total_downloads_attempted', 0)}")
    summary.append(f"Total successful downloads (all runs): {metrics.get('total_successful_downloads', 0)}")
    summary.append(f"Overall Download success rate: {metrics.get('total_successful_downloads', 0) / max(1, metrics.get('total_downloads_attempted', 1)):.1%}")

    # Metadata performance
    summary.append(f"\n=== Metadata Performance (All Runs) ===")
    summary.append(f"Total metadata API calls: {metrics.get('total_metadata_api_calls', 0)}")
    summary.append(f"Total metadata errors: {metrics.get('total_metadata_errors', 0)}")
    summary.append(f"Overall Metadata error rate: {metrics.get('total_metadata_errors', 0) / max(1, metrics.get('total_metadata_api_calls', 1)):.1%}")

    # Recent runs
    if runs_data:
        summary.append(f"\n=== Recent Runs ({min(5, len(runs_data))}) ===")
        for run in runs_data[-5:]:
            summary.append(f" - Run {run.get('date', 'Unknown')[:10]}: Suitable={run.get('suitable_shorts',0)}, Attempted={run.get('downloads_attempted', 0)}, Succeeded={run.get('downloads_successful', 0)}, MetaErrors={run.get('metadata_errors', 0)}")

    # Channel performance
    channel_performance = metrics.get('channel_performance', {})
    if channel_performance:
         sorted_channels = sorted(channel_performance.items(), key=lambda item: item[1], reverse=True)
         summary.append(f"\n=== Top 10 Performing Channels (by Score) ===")
         for chan, score in sorted_channels[:10]:
             summary.append(f" - {chan}: {score:.2f}")

    return "\n".join(summary)

def generate_tuning_suggestions(metrics, config):
    """Generates parameter tuning suggestions using Gemini."""
    performance_summary = generate_performance_summary(metrics)
    config_text = "\n".join([f"{key}={value}" for key, value in config.items() if 'API_KEY' not in key])

    prompt = f"""
    Analyze the following performance report from a YouTube Shorts CHANNEL downloader script.
    Suggest config.txt adjustments and explain *why* based ONLY on the provided data.

    Performance Report:
    {performance_summary}

    Current Config (excluding API keys):
    {config_text}

    Focus on:
    - MAX_DOWNLOADS: Total downloads per run.
    - METADATA_TIMEOUT_SECONDS: Metadata generation timeout.
    - YT_PLAYLIST_FETCH_LIMIT: Initial videos checked per channel. (Is it finding enough suitable ones?)
    - SEO_... parameters: Are metadata errors high? Maybe review prompt indirectly via these.

    Provide specific, data-driven recommendations ONLY IF the data suggests a change. If performance looks good, say so.
    Format clearly.
    """

    try:
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        suggestions = response.text.strip()

        with open(tuning_suggestions_file_path, "w", encoding="utf-8") as f:
            f.write(f"=== Tuning Suggestions ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ===\nConfig:\n{config_text}\n\nSummary:\n{performance_summary}\n\nSuggestions:\n{suggestions}\n")

        return suggestions
    except Exception as e:
        print_error(f"Error generating tuning suggestions: {e}", 1)
        return None
# --- End Tuning Suggestions ---

# --- YouTube API Authentication ---
def get_authenticated_service():
    """Gets an authenticated YouTube Data API service."""
    if not YOUTUBE_API_AVAILABLE:
        print_error("Google API libraries not available.")
        return None

    import pickle
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request

    # Define constants for authentication
    CLIENT_SECRETS_FILE = constants.CLIENT_SECRETS_FILE
    TOKEN_FILE = constants.TOKEN_FILE
    # Include playlist management scope in addition to readonly
    SCOPES = [
        "https://www.googleapis.com/auth/youtube.readonly",
        "https://www.googleapis.com/auth/youtube" # Full access for playlist management
    ]

    creds = None
    if os.path.exists(TOKEN_FILE):
        try:
            with open(TOKEN_FILE, 'rb') as token:
                creds = pickle.load(token)
            print_success("Cached credentials loaded.")
        except Exception as e:
            print_warning(f"Failed to load cached credentials: {e}. Will re-authenticate.")
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                print_success("Credentials refreshed successfully.")
            except Exception as e:
                print_warning(f"Failed to refresh credentials: {e}. Will perform new auth.")
                creds = None
        else:
            print_info("No valid cached credentials. Starting new authentication flow.")
            if not os.path.exists(CLIENT_SECRETS_FILE):
                print_error(f"FATAL: Client secrets file not found: {CLIENT_SECRETS_FILE}")
                return None
            try:
                flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
                creds = flow.run_local_server(port=0)
                print_success("Authentication flow completed.")
            except Exception as e:
                print_error(f"Auth flow error: {e}", include_traceback=True)
                return None

        if creds and creds.valid:
            try:
                with open(TOKEN_FILE, 'wb') as token:
                    pickle.dump(creds, token)
                print_success(f"Credentials saved to: {TOKEN_FILE}")
            except Exception as e:
                print_warning(f"Failed to save credentials: {e}")

    if creds and creds.valid:
        try:
            service = build('youtube', 'v3', credentials=creds)
            print_success("YouTube Data API service built.")
            return service
        except Exception as e:
            print_error(f"API service build error: {e}", include_traceback=True)
            return None
    else:
        print_error("Authentication failed.")
        return None

# --- Playlist Management Functions ---
def load_cache(cache_file_path, cache_name="Cache"):
    """Loads a cache from a JSON file."""
    default_cache = {"timestamp": datetime.now().isoformat()}
    try:
        if os.path.exists(cache_file_path):
            with open(cache_file_path, "r", encoding="utf-8") as f:
                content = f.read()
                if not content:
                    print_info(f"{cache_name} file exists but is empty. Initializing new cache.")
                    return default_cache
                cache = json.loads(content)
                if not isinstance(cache, dict):
                    print_warning(f"{cache_name} file has invalid format (expected dict). Initializing new cache.")
                    return default_cache
                return cache
        else:
            print_info(f"{cache_name} file not found. Creating new cache.")
            return default_cache
    except json.JSONDecodeError:
        print_error(f"Error decoding JSON from {cache_name} file. Initializing new cache.")
        return default_cache
    except Exception as e:
        print_error(f"Error loading {cache_name}: {e}", include_traceback=True)
        return default_cache

def save_cache(cache_data, cache_file_path, cache_name="Cache"):
    """Saves a cache to a JSON file."""
    try:
        with open(cache_file_path, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=4)
        print_info(f"Saved {cache_name} with {len(cache_data) - 1} entries.") # -1 for timestamp
    except Exception as e:
        print_error(f"Error saving {cache_name}: {e}", include_traceback=True)

def fetch_channel_playlists(service: Any) -> Dict[str, str]:
    """Fetches existing public playlists for the authenticated user's channel."""
    if not service:
        print_warning("YouTube API service not available, cannot fetch playlists.")
        return {}

    playlists_map = {}
    next_page_token = None
    print_info("Fetching existing channel playlists via YouTube API...")
    try:
        while True:
            request = service.playlists().list(
                part="snippet",
                mine=True, # Get playlists for the authenticated user
                maxResults=50,
                pageToken=next_page_token
            )
            response = request.execute()

            for item in response.get("items", []):
                playlist_id = item.get("id")
                playlist_title = item.get("snippet", {}).get("title")
                if playlist_id and playlist_title:
                    playlists_map[playlist_id] = playlist_title

            next_page_token = response.get("nextPageToken")
            if not next_page_token:
                break # Exit loop if no more pages
            time.sleep(0.5) # Avoid hitting API limits too quickly

        print_success(f"Fetched {len(playlists_map)} existing playlists.")
        return playlists_map

    except HttpError as e:
        print_error(f"API Error fetching playlists: {e}", include_traceback=True)
        log_error_to_file(f"API Error fetching playlists: {e}", include_traceback=True)
        return playlists_map # Return what we got so far
    except Exception as e:
        print_error(f"Unexpected error fetching playlists: {e}", include_traceback=True)
        log_error_to_file(f"Unexpected error fetching playlists: {e}", include_traceback=True)
        return playlists_map

def get_playlist_suggestion(video_title: str, video_desc: str, video_tags: List[str], existing_playlist_titles: List[str]) -> Optional[str]:
    """Asks Gemini to match video to an existing playlist or suggest a new one."""
    print_info("Getting playlist suggestion from Gemini...", 3)

    existing_playlists_formatted = "\n - ".join(existing_playlist_titles) if existing_playlist_titles else "None"

    prompt = f"""
    Analyze the following YouTube Short video's metadata:
    Title: {video_title}
    Description: {video_desc[:500]} # Limit description for prompt
    Tags: {', '.join(video_tags[:15])} # Limit tags for prompt

    Here is a list of EXISTING YouTube playlists on the channel:
     - {existing_playlists_formatted}

    Your task is to determine the BEST playlist for this video. **Prioritize matching an existing playlist if possible.**

    1.  **Check Existing:** Review the list of EXISTING playlists. Does the video's core topic (derived from title, description, tags) closely match the theme of ANY of the existing playlists? Be flexible with minor wording differences (e.g., "GTA 6 Leaks" vs "GTA VI Leaks").
    2.  **Decision:**
        *   If YES, and you find a strong thematic match, output the EXACT name of the single best matching EXISTING playlist from the list provided. **Do this even if the match isn't perfect but is reasonably close.**
        *   If NO (no reasonable match found, or the existing list is 'None'), ONLY THEN suggest a SHORT, relevant, and SEO-friendly NEW playlist title (max 60 characters) that accurately reflects the video's specific topic. Prefix the suggestion with "NEW: ". Example: "NEW: GTA V Funny Moments" or "NEW: Vice City Secrets".

    Output ONLY the chosen existing playlist title OR the "NEW: " prefixed suggestion. Do not add any other explanation.
    """

    try:
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        suggestion = response.text.strip()

        if suggestion:
            if suggestion.startswith("NEW: "):
                new_title = suggestion[len("NEW: "):].strip()
                if new_title:
                    print_success(f"Gemini suggested NEW playlist: '{new_title}'", 4)
                    return suggestion # Return with "NEW: " prefix
                else:
                    print_warning("Gemini suggested 'NEW: ' but title was empty.", 4)
                    return None
            elif suggestion in existing_playlist_titles:
                print_success(f"Gemini matched to EXISTING playlist: '{suggestion}'", 4)
                return suggestion # Return the exact existing title
            else:
                print_warning(f"Gemini output '{suggestion}' doesn't match existing or NEW format. Treating as NEW suggestion.", 4)
                # Fallback: Treat unexpected output as a new suggestion if it's not empty
                return f"NEW: {suggestion[:60]}" # Add prefix and limit length

        else:
            print_warning("Gemini returned empty playlist suggestion.", 4)
            return None

    except Exception as e:
        print_error(f"Error getting playlist suggestion from Gemini: {e}", 3)
        log_error(f"Error Gemini playlist suggestion: {e}")
        return None
def create_playlist(service: Any, title: str, description: str = "") -> Optional[str]:
    """Creates a new private playlist and returns its ID."""
    if not service or not title:
        print_warning("Missing service or title for creating playlist.")
        return None

    try:
        request = service.playlists().insert(
            part="snippet,status",
            body={
                "snippet": {
                    "title": title,
                    "description": description or f"Auto-created playlist for {title}",
                },
                "status": {
                    "privacyStatus": "private"  # Start as private, can be changed later
                }
            }
        )
        response = request.execute()
        playlist_id = response.get("id")
        if playlist_id:
            print_success(f"Created new playlist: '{title}' (ID: {playlist_id})")
            return playlist_id
        else:
            print_error("Failed to create playlist - no ID returned.")
            return None
    except HttpError as e:
        print_error(f"API Error creating playlist: {e}", include_traceback=True)
        log_error(f"API Error creating playlist: {e}")
        return None
    except Exception as e:
        print_error(f"Unexpected error creating playlist: {e}", include_traceback=True)
        log_error(f"Unexpected error creating playlist: {e}")
        return None

def add_video_to_playlist(service: Any, video_id: str, playlist_id: str) -> bool:
    """Adds a video to a specified playlist using YouTube Data API."""
    if not service or not video_id or not playlist_id:
        print_warning("Missing service, video ID, or playlist ID for adding to playlist.")
        return False

    try:
        request = service.playlistItems().insert(
            part="snippet",
            body={
                "snippet": {
                    "playlistId": playlist_id,
                    "resourceId": {
                        "kind": "youtube#video",
                        "videoId": video_id
                    }
                    # Optional: Set position 'position': 0 # Add to top
                }
            }
        )
        response = request.execute()
        if response and response.get("id"):
            print_success(f"Added video {video_id} to playlist {playlist_id}")
            return True
        else:
            print_warning(f"Unexpected response adding video to playlist: {response}")
            return False
    except HttpError as e:
        print_error(f"API Error adding video to playlist: {e}", include_traceback=True)
        log_error(f"API Error adding video to playlist: {e}")
        return False
    except Exception as e:
        print_error(f"Unexpected error adding video to playlist: {e}", include_traceback=True)
        log_error(f"Unexpected error adding video to playlist: {e}")
        return False
# --- End Playlist Management Functions ---

# --- Related Channels Discovery ---
def find_related_channels(top_channels: List[str], channel_listing_cache: Dict, current_channels_set: Set[str], service: Any, max_to_suggest: int = 5) -> List[str]:
    """Finds suggested related channels using YouTube API."""
    if not service:
        print_warning("YouTube API service not available, skipping related channel search.")
        return []

    suggested_channels = set()
    checked_channels = 0

    print_info("Searching for related channels based on top performers...")

    for channel_url in top_channels:
        if checked_channels >= max_to_suggest:  # Limit API calls per run
            break

        print_info(f"Checking related channels for: {channel_url}", 1)
        video_id_to_check = None
        # Get a video ID from this channel's cache
        cached_entries = channel_listing_cache.get(channel_url, [])
        if cached_entries and isinstance(cached_entries, list) and len(cached_entries) > 0:
            # Try getting the first video's ID (often highest views if sorted)
            video_id_to_check = cached_entries[0].get('id')

        if not video_id_to_check:
            print_warning(f"Could not find a video ID in cache for {channel_url} to find related channels.", 2)
            continue

        try:
            search_response = service.search().list(
                part="snippet",
                relatedToVideoId=video_id_to_check,
                type="channel",
                maxResults=5  # Limit results per query
            ).execute()

            found_count = 0
            for item in search_response.get("items", []):
                related_channel_id = item.get("snippet", {}).get("channelId")
                related_channel_title = item.get("snippet", {}).get("title")
                if related_channel_id and related_channel_title:
                    related_channel_url = f"https://www.youtube.com/channel/{related_channel_id}"
                    # Check if it's new and not already suggested
                    if related_channel_url not in current_channels_set and related_channel_url not in suggested_channels:
                        print_success(f"Found related channel suggestion: '{related_channel_title}' ({related_channel_url})", 2)
                        suggested_channels.add(related_channel_url)
                        found_count += 1

            if found_count == 0:
                print_info("No new related channels found for this source.", 2)

            checked_channels += 1  # Count this channel check towards the limit
            time.sleep(1)  # Avoid hitting API limits too quickly

        except HttpError as e:
            print_error(f"API Error finding related channels for video {video_id_to_check}: {e}", 2)
            # Break if quota exceeded or serious error
            if e.resp.status in [403, 400]:
                print_error("API quota likely exceeded or bad request. Stopping related channel search for this run.", 2)
                break
        except Exception as e:
            print_error(f"Unexpected error finding related channels for video {video_id_to_check}: {e}", 2, include_traceback=True)
            # Optionally break on unexpected errors too

    return list(suggested_channels)
# --- End Related Channels Discovery ---

# --- Enhanced Metadata Generation ---
def generate_metadata_with_timeout_v2(video_title, uploader_name, original_title="Unknown Title", timeout=METADATA_TIMEOUT_SECONDS):
    """Generates metadata with timeout, includes category suggestion."""
    metadata_metrics = load_metadata_metrics()
    metadata_metrics["total_api_calls"] += 1
    error_type = None
    error_details = None
    result_metadata = None

    try:
        with concurrent.futures.ThreadPoolExecutor() as executor:
            print_info("Generating primary SEO metadata (Title/Desc/Tags)...", 2)
            # Use the load_or_get_seo_prompt_template function directly instead of generate_seo_metadata_v2
            prompt_template = load_or_get_seo_prompt_template()
            if not prompt_template:
                print_error("Failed to load SEO prompt template. Using minimal fallback.")
                prompt_template = """<metadata><title>{video_topic} #Shorts</title><description>Credit: {uploader_name}</description><tags>shorts</tags></metadata>"""

            # Format the prompt with actual data
            prompt = prompt_template.replace("{video_topic}", video_title)\
                                    .replace("{uploader_name}", uploader_name)\
                                    .replace("{channel_name}", seo_channel_name)\
                                    .replace("{channel_topic}", seo_channel_topic)\
                                    .replace("{original_title}", original_title)

            # Create a function to generate metadata that can be executed with timeout
            def generate_metadata_internal():
                try:
                    model = genai.GenerativeModel("gemini-2.0-flash")
                    response = model.generate_content(prompt)
                    raw_text = response.text

                    # Default fallback structure
                    metadata = {
                        "title": f"{video_title} #Shorts",
                        "description": f"Default SEO description.\nCredit: {uploader_name}\nOriginal Title: {original_title}",
                        "tags": []
                    }

                    # --- Parsing (Adjusted for SEO prompt structure) ---
                    title_match = re.search(r"<title>(.*?)</title>", raw_text, re.DOTALL | re.IGNORECASE)
                    desc_match = re.search(r"<description>(.*?)</description>", raw_text, re.DOTALL | re.IGNORECASE)
                    tags_match = re.search(r"<tags>(.*?)</tags>", raw_text, re.DOTALL | re.IGNORECASE)

                    parsing_warnings = []
                    if title_match and title_match.group(1).strip():
                        metadata["title"] = title_match.group(1).strip()
                    else:
                        parsing_warnings.append("title")
                        metadata["title"] = f"{video_title} #Shorts"

                    if desc_match and desc_match.group(1).strip():
                        metadata["description"] = desc_match.group(1).strip()
                        # Ensure credit is present (as per prompt structure)
                        if f"Credit: {uploader_name}" not in metadata["description"]:
                            metadata["description"] += f"\n\nCredit: {uploader_name}"
                        if f"Original Title: {original_title}" not in metadata["description"] and original_title != "Unknown Title":
                            metadata["description"] += f"\nOriginal Title: {original_title}"
                    else:
                        parsing_warnings.append("description")
                        metadata["description"] = f"Default desc.\nCredit: {uploader_name}\nOriginal Title: {original_title}"

                    if tags_match and tags_match.group(1).strip():
                        tags_raw = tags_match.group(1).strip()
                        metadata["tags"] = [tag.strip() for tag in re.split(r'[,\n]', tags_raw) if tag.strip()] # Handle comma or newline
                    else:
                        parsing_warnings.append("tags")
                        metadata["tags"] = [seo_channel_topic.lower(), "shorts"]

                    if parsing_warnings:
                        print_warning(f"Could not parse <{'>, <'.join(parsing_warnings)}> for topic: {video_title}. Used fallbacks.", 1)

                    # --- Post-Processing (Title Length/Suffix) ---
                    temp_title = metadata.get("title", video_title)
                    if len(temp_title) > TARGET_TITLE_LENGTH: # Use TARGET_TITLE_LENGTH (90)
                        truncated = temp_title[:TARGET_TITLE_LENGTH]
                        last_space = truncated.rfind(' ')
                        temp_title = truncated[:last_space].strip() if last_space > 0 else truncated.strip()
                    if not temp_title.lower().endswith(SHORTS_SUFFIX.lower()):
                        if len(temp_title) + len(SHORTS_SUFFIX) <= MAX_TITLE_LENGTH:
                            temp_title += SHORTS_SUFFIX
                    metadata["title"] = temp_title

                    # --- Cross-Validation Checks ---
                    validation_warnings = []
                    # Check 1: Title in Description (simplified check)
                    try:
                        title_check = metadata.get("title", "").replace(SHORTS_SUFFIX, "").strip().lower()
                        desc_check = metadata.get("description", "").lower()
                        if title_check and title_check not in desc_check:
                            # Allow for minor variations, check first ~20 chars maybe?
                            if title_check[:20] not in desc_check[:max(200, len(title_check)+50)]:  # Check beginning of desc
                                validation_warnings.append("Title not found near start of description.")
                    except Exception as e:
                        validation_warnings.append(f"Title check failed: {e}")

                    # Check 2: Tags listed in Description
                    try:
                        tags_heading = "Tags Used in Video :-".lower()
                        desc_check = metadata.get("description", "").lower()
                        tags_list = metadata.get("tags", [])
                        heading_index = desc_check.find(tags_heading)
                        if tags_list and heading_index == -1:
                            validation_warnings.append("Tags list heading missing in description.")
                        elif tags_list and heading_index != -1:
                            desc_after_heading = desc_check[heading_index:]
                            # Check if at least one of the first 5 tags is listed
                            if not any(tag.lower() in desc_after_heading for tag in tags_list[:5]):
                                validation_warnings.append("First few tags not found listed in description after heading.")
                    except Exception as e:
                        validation_warnings.append(f"Tag list check failed: {e}")

                    # Check 3: Basic Keyword Stuffing
                    try:
                        text_to_check = metadata.get("description", "") + " " + " ".join(metadata.get("tags", []))
                        words = re.findall(r'\b\w{4,}\b', text_to_check.lower())  # Words 4+ chars
                        if len(words) > 20:  # Only check if there's enough text
                            word_counts = collections.Counter(words)
                            most_common = word_counts.most_common(3)
                            # Check if top words appear excessively (e.g., > 15 times)
                            if most_common and most_common[0][1] > 15:
                                validation_warnings.append(f"Potential keyword stuffing detected for word: '{most_common[0][0]}' ({most_common[0][1]} times).")
                    except Exception as e:
                        validation_warnings.append(f"Stuffing check failed: {e}")

                    if validation_warnings:
                        print_warning(f"Metadata validation warnings for '{video_title}':", 1)
                        for warn in validation_warnings:
                            print_warning(f"- {warn}", 2)
                    # --- End Cross-Validation Checks ---

                    return metadata
                except Exception as e:
                    print_error(f"Error during metadata generation/processing for '{video_title}': {e}", 1, include_traceback=True)
                    return {
                        "title": f"{video_title[:80]} #Shorts",
                        "description": f"Default desc. Error: {e}\nCredit: {uploader_name}\nOriginal Title: {original_title}",
                        "tags": ["error"]
                    }

            # Execute with timeout
            future = executor.submit(generate_metadata_internal)
            result_metadata = future.result(timeout=timeout)

            # Final check
            if not isinstance(result_metadata, dict) or not result_metadata.get("title") or not result_metadata.get("description") or not result_metadata.get("tags"):
                print_error(f"Critical Warning: Metadata invalid/empty for '{video_title}'. Final fallback.", 1)
                result_metadata = {
                    "title": f"{video_title[:80]} #Shorts",
                    "description": f"Final fallback.\nCredit: {uploader_name}\nOriginal Title: {original_title}",
                    "tags": ["fallback"]
                }

            # Add category suggestion
            suggested_category = None
            if result_metadata:
                gen_title = result_metadata.get("title", video_title)
                gen_desc = result_metadata.get("description", "")
                if gen_title and gen_desc:
                    try:
                        suggested_category = get_suggested_category(gen_title, gen_desc)
                        if suggested_category:
                            result_metadata['suggested_category'] = suggested_category
                    except Exception as cat_err:
                        print_error(f"Error during category suggestion call: {cat_err}", 2)
                else:
                    print_warning("Skipping category suggestion due to empty title/description.", 2)

            # Check for errors (parsing/empty)
            if "Default SEO description" in result_metadata.get("description", ""):
                metadata_metrics["empty_description_errors"] += 1
                error_type = "empty_description"
                error_details = f"Failed description: {video_title}"

            if not result_metadata.get("tags") or "error" in result_metadata.get("tags",[]):
                metadata_metrics["empty_tags_errors"] += 1
                error_type = error_type or "empty_tags"
                error_details = error_details or f"Failed tags: {video_title}"

            if error_type and error_details:
                add_error_sample(metadata_metrics, error_type, error_details, video_title)

            # Validate metadata and track stuffing warnings
            validate_generated_metadata(result_metadata, video_title, metadata_metrics)

            save_metadata_metrics(metadata_metrics)
            return result_metadata

    except concurrent.futures.TimeoutError:
        print_warning(f"Primary metadata generation timed out for: {video_title}", 1)
        metadata_metrics["timeouts"] += 1
        add_error_sample(metadata_metrics, "timeout", f"Timeout for: {video_title}", video_title)

        fallback_metadata = {
            "title": f"{video_title[:80]} #Shorts",
            "description": f"Timeout.\nCredit: {uploader_name}\nOriginal Title: {original_title}",
            "tags": ["timeout"]
        }

        # Validate metadata and track stuffing warnings
        validate_generated_metadata(fallback_metadata, video_title, metadata_metrics)

        save_metadata_metrics(metadata_metrics)
        return fallback_metadata
    except Exception as e:
        print_error(f"Error submitting metadata generation job for '{video_title}': {e}", 1, include_traceback=True)
        metadata_metrics["parse_failures"] += 1
        add_error_sample(metadata_metrics, "exception", f"Error: {str(e)} for: {video_title}", video_title)

        fallback_metadata = {
            "title": f"{video_title[:80]} #Shorts",
            "description": f"Error: {e}\nCredit: {uploader_name}\nOriginal Title: {original_title}",
            "tags": ["error"]
        }

        # Validate metadata and track stuffing warnings
        validate_generated_metadata(fallback_metadata, video_title, metadata_metrics)

        save_metadata_metrics(metadata_metrics)
        return fallback_metadata
# --- End Enhanced Metadata Generation ---

if __name__ == "__main__":
    main()
