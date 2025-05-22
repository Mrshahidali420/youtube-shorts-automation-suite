# --- Combined downloader.py ---
import os
import json
import yt_dlp
import google.generativeai as genai
import re
import concurrent.futures
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import random
import math
import time
import traceback
import collections # For word counting in metadata validation
import colorama # Import colorama
from colorama import Fore, Style, Back # Import specific styles
import shutil # For file backups

# --- Print Helper Functions ---
# (Using the slightly more detailed style from downloader - B.py)
def print_info(msg, indent=0): prefix = "  " * indent; print(f"{prefix}{Fore.BLUE}i INFO:{Style.RESET_ALL} {msg}")
def print_success(msg, indent=0): prefix = "  " * indent; print(f"{prefix}{Fore.GREEN}OK SUCCESS:{Style.RESET_ALL} {msg}")
def print_warning(msg, indent=0): prefix = "  " * indent; print(f"{prefix}{Fore.YELLOW}WARN WARNING:{Style.RESET_ALL} {msg}")
def print_error(msg, indent=0, include_traceback=False):
    prefix = "  " * indent
    print(f"{prefix}{Fore.RED}ERR ERROR:{Style.RESET_ALL} {msg}")
    if include_traceback:
        traceback.print_exc()
def print_fatal(msg, indent=0): prefix = "  " * indent; print(f"{prefix}{Back.RED}{Fore.WHITE}{Style.BRIGHT} FATAL: {msg} {Style.RESET_ALL}"); exit(1)

# --- NEW: Import constants from the new location ---
try:
    from .utils import constants # If downloader_keyword.py is a module in youtube_shorts
    from .utils import metadata_generator # Import metadata_generator from utils
    UTILS_AVAILABLE = True
except ImportError:
    # Fallback if run as script directly from youtube_shorts folder
    try:
        from utils import constants
        from utils import metadata_generator # Import metadata_generator from utils
        UTILS_AVAILABLE = True
    except ImportError:
        # Absolute fallback if utils isn't directly importable (e.g. running from project root)
        # This might indicate a PYTHONPATH issue or that the script isn't being run as part of the package.
        # For now, assume it can find 'utils' if it's in the same directory level.
        print("CRITICAL: Could not import constants.py. Ensure it's in utils/ and runnable.")
        # Define a minimal set of constants here to avoid crashing, but this is not ideal.
        class MinimalConstants:
            CONFIG_FILE_PATH = "config/config.txt" # Will be relative to where script is run
            KEYWORDS_FILE_PATH = "config/keywords.txt"
            SHORTS_DOWNLOADS_DIR = "output/shorts_downloads"
            SHORTS_METADATA_DIR = "output/shorts_metadata"
            EXCEL_FILE_PATH = "data/shorts_data.xlsx"
            DOWNLOADER_KEYWORD_LOG_FILE = "logs/downloader_keyword_log.txt"
            GENERATED_KEYWORDS_CACHE_FILE = "data/generated_keywords_cache.json"
            FFMPEG_PATH = "drivers/ffmpeg.exe"
            SEO_METADATA_PROMPT_FILE = "config/seo_metadata_prompt.txt" # Corrected name
            METADATA_METRICS_FILE = "data/metadata_metrics.json" # General
            PERFORMANCE_METRICS_FILE = "data/performance_metrics.json" # General
            TUNING_SUGGESTIONS_LOG_FILE = "logs/tuning_suggestions.log" # General
            UPLOAD_CORRELATION_CACHE = "data/upload_correlation_cache.json"
            PLAYLIST_DATA_CACHE_FILE = "data/playlists_data_cache.json"
            DOWNLOADED_SHEET_NAME = "Downloaded"
            UPLOADED_SHEET_NAME = "Uploaded"
            NICHE_FILE_PATH = "config/niche.txt"
            # For excel_utils fallback
            BACKUP_FOLDER = "backups/excel"
            # YouTube limits
            YOUTUBE_TITLE_LIMIT = 100
            YOUTUBE_DESCRIPTION_LIMIT = 4950
            YOUTUBE_TAG_LIMIT = 100
            YOUTUBE_TOTAL_TAGS_LIMIT = 460
            YOUTUBE_MAX_TAGS_COUNT = 40
            # Metadata settings
            METADATA_TIMEOUT_SECONDS = 60
            METADATA_ERROR_THRESHOLD = 5
            METADATA_TIMEOUT_THRESHOLD = 3
            VALIDATION_WARNING_THRESHOLD = 5
            MAX_ERROR_SAMPLES = 20
            # Performance settings
            MIN_RUNS_BEFORE_TUNING = 5
            RUN_SUMMARIES_LOG_FILE = "logs/run_summaries.log"
        constants = MinimalConstants()
        print("WARNING: Using minimal fallback constants due to import error.")
        UTILS_AVAILABLE = False
        # Create a dummy metadata_generator to avoid errors
        class DummyMetadataGenerator:
            def improve_metadata(self, *args, **kwargs):
                print("WARNING: metadata_generator not available. Skipping metadata improvement.")
                return args[0] if args else {}

            def validate_metadata(self, *args, **kwargs):
                print("WARNING: metadata_generator not available. Skipping metadata validation.")
                return args[0] if args else {}

            def analyze_metadata_quality(self, *args, **kwargs):
                print("WARNING: metadata_generator not available. Skipping metadata quality analysis.")
                return {"overall_score": 0}

        metadata_generator = DummyMetadataGenerator()
# --- End NEW Import ---

# --- Runtime Configuration Constants ---
# These values can be overridden by config.txt
YT_SEARCH_RESULTS_PER_KEYWORD = 50 # Number of search results to fetch per keyword
VIDEOS_TO_DOWNLOAD_PER_KEYWORD = 5  # Max videos to download for a single keyword
KEYWORDS_TO_PROCESS_PER_RUN = 5  # Number of keywords to select for each run
MIN_KEYWORDS_THRESHOLD = 20      # Minimum number of keywords before generating new ones
NEW_KEYWORDS_TO_GENERATE = 10    # Number of new keywords to generate when needed

# Constants for metadata and formatting
MAX_TITLE_LENGTH = getattr(constants, 'YOUTUBE_TITLE_LIMIT', 100)  # Use YOUTUBE_TITLE_LIMIT or default to 100
TARGET_TITLE_LENGTH = 90           # Target length before adding #Shorts
METADATA_TIMEOUT_SECONDS = getattr(constants, 'METADATA_TIMEOUT_SECONDS', 60)  # Default to 60 seconds if not defined
INITIAL_KEYWORDS_COUNT = 15        # Number of keywords to generate initially
SHORTS_SUFFIX = " #Shorts"

# Excel Sheet Names / Headers / Indices
DOWNLOADED_SHEET_NAME = constants.DOWNLOADED_SHEET_NAME
UPLOADED_SHEET_NAME = constants.UPLOADED_SHEET_NAME

# Additional keyword optimization settings
TOP_KEYWORDS_TO_USE = 5          # Number of top-performing keywords to use for new keyword generation

# Metadata prompt improvement settings
METADATA_ERROR_THRESHOLD = getattr(constants, 'METADATA_ERROR_THRESHOLD', 5)
METADATA_TIMEOUT_THRESHOLD = getattr(constants, 'METADATA_TIMEOUT_THRESHOLD', 3)
VALIDATION_WARNING_THRESHOLD = getattr(constants, 'VALIDATION_WARNING_THRESHOLD', 5)
MAX_ERROR_SAMPLES = getattr(constants, 'MAX_ERROR_SAMPLES', 20)

# Parameter tuning settings
MIN_RUNS_BEFORE_TUNING = getattr(constants, 'MIN_RUNS_BEFORE_TUNING', 5)
BASE_WEIGHT_FOR_KEYWORDS = 1     # Base weight to add to all keywords to prevent stagnation

# --- Global Cache for SEO Prompt ---
_current_seo_prompt_template = None # Will be loaded/set later



# --- Function Definitions ---

# --- Cache Functions ---
def load_cache(cache_file_path, cache_name="Cache"):
    """Loads a cache from a JSON file."""
    default_cache = {"timestamp": datetime.now().isoformat()}
    try:
        if os.path.exists(cache_file_path):
            with open(cache_file_path, "r", encoding="utf-8") as f:
                content = f.read()
                if not content:
                    print_info(f"{cache_name} file exists but is empty. Initializing new cache.")
                    return default_cache
                cache = json.loads(content)
                if not isinstance(cache, dict):
                    print_warning(f"{cache_name} file has invalid format (expected dict). Initializing new cache.")
                    return default_cache
                return cache
        else:
            print_info(f"{cache_name} file not found. Creating new cache.")
            return default_cache
    except json.JSONDecodeError:
        print_error(f"Error decoding JSON from {cache_name} file. Initializing new cache.")
        return default_cache
    except Exception as e:
        print_error(f"Error loading {cache_name}: {e}", include_traceback=True)
        return default_cache

# --- Playlist Management Functions ---
def get_playlist_suggestion(video_title: str, video_desc: str, video_tags: list, existing_playlist_titles: list) -> str:
    """Asks Gemini to match video to an existing playlist or suggest a new one."""
    print_info("Getting playlist suggestion from Gemini...", 3)

    existing_playlists_formatted = "\n - ".join(existing_playlist_titles) if existing_playlist_titles else "None"

    prompt = f"""
    Analyze the following YouTube Short video's metadata:
    Title: {video_title}
    Description: {video_desc[:500]} # Limit description for prompt
    Tags: {', '.join(video_tags[:15])} # Limit tags for prompt

    Here is a list of EXISTING YouTube playlists on the channel:
     - {existing_playlists_formatted}

    Your task is to determine the BEST playlist for this video. **Prioritize matching an existing playlist if possible.**

    1.  **Check Existing:** Review the list of EXISTING playlists. Does the video's core topic (derived from title, description, tags) closely match the theme of ANY of the existing playlists? Be flexible with minor wording differences (e.g., "GTA 6 Leaks" vs "GTA VI Leaks").
    2.  **Decision:**
        *   If YES, and you find a strong thematic match, output the EXACT name of the single best matching EXISTING playlist from the list provided. **Do this even if the match isn't perfect but is reasonably close.**
        *   If NO (no reasonable match found, or the existing list is 'None'), ONLY THEN suggest a SHORT, relevant, and SEO-friendly NEW playlist title (max 60 characters) that accurately reflects the video's specific topic. Prefix the suggestion with "NEW: ". Example: "NEW: GTA V Funny Moments" or "NEW: Vice City Secrets".

    Output ONLY the chosen existing playlist title OR the "NEW: " prefixed suggestion. Do not add any other explanation.
    """

    try:
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        suggestion = response.text.strip()

        if suggestion:
            if suggestion.startswith("NEW: "):
                new_title = suggestion[len("NEW: "):].strip()
                if new_title:
                    print_success(f"Gemini suggested NEW playlist: '{new_title}'", 4)
                    return suggestion # Return with "NEW: " prefix
                else:
                    print_warning("Gemini suggested 'NEW: ' but title was empty.", 4)
                    return None
            elif suggestion in existing_playlist_titles:
                print_success(f"Gemini matched to EXISTING playlist: '{suggestion}'", 4)
                return suggestion # Return the exact existing title
            else:
                print_warning(f"Gemini output '{suggestion}' doesn't match existing or NEW format. Treating as NEW suggestion.", 4)
                # Fallback: Treat unexpected output as a new suggestion if it's not empty
                return f"NEW: {suggestion[:60]}" # Add prefix and limit length

        else:
            print_warning("Gemini returned empty playlist suggestion.", 4)
            return None

    except Exception as e:
        print_error(f"Error getting playlist suggestion from Gemini: {e}", 3)
        return None

def get_authenticated_service():
    """Gets an authenticated YouTube Data API service."""
    try:
        import pickle
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
    except ImportError:
        print_error("Google API libraries not available. Install with: pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
        return None

    # Define constants for authentication
    script_dir = os.path.dirname(os.path.abspath(__file__))
    CLIENT_SECRETS_FILE = os.path.join(script_dir, "client_secret.json")
    TOKEN_FILE = os.path.join(script_dir, "token.json")
    # Include playlist management scope in addition to readonly
    SCOPES = [
        "https://www.googleapis.com/auth/youtube.readonly",
        "https://www.googleapis.com/auth/youtube" # Full access for playlist management
    ]

    creds = None
    if os.path.exists(TOKEN_FILE):
        try:
            with open(TOKEN_FILE, 'rb') as token:
                creds = pickle.load(token)
            print_success("Cached credentials loaded.")
        except Exception as e:
            print_warning(f"Failed to load cached credentials: {e}. Will re-authenticate.")
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                print_success("Credentials refreshed successfully.")
            except Exception as e:
                print_warning(f"Failed to refresh credentials: {e}. Will perform new auth.")
                creds = None
        else:
            print_info("No valid cached credentials. Starting new authentication flow.")
            if not os.path.exists(CLIENT_SECRETS_FILE):
                print_error(f"FATAL: Client secrets file not found: {CLIENT_SECRETS_FILE}")
                return None
            try:
                flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
                creds = flow.run_local_server(port=0)
                print_success("Authentication flow completed.")
            except Exception as e:
                print_error(f"Auth flow error: {e}", include_traceback=True)
                return None

        if creds and creds.valid:
            try:
                with open(TOKEN_FILE, 'wb') as token:
                    pickle.dump(creds, token)
                print_success(f"Credentials saved to: {TOKEN_FILE}")
            except Exception as e:
                print_warning(f"Failed to save credentials: {e}")

    if creds and creds.valid:
        try:
            service = build('youtube', 'v3', credentials=creds)
            print_success("YouTube Data API service built.")
            return service
        except Exception as e:
            print_error(f"API service build error: {e}", include_traceback=True)
            return None
    else:
        print_error("Authentication failed.")
        return None

def fetch_channel_playlists(service) -> dict:
    """Fetches existing public playlists for the authenticated user's channel."""
    if not service:
        print_warning("YouTube API service not available, cannot fetch playlists.")
        return {}

    playlists_map = {}
    next_page_token = None
    print_info("Fetching existing channel playlists via YouTube API...")
    try:
        while True:
            request = service.playlists().list(
                part="snippet",
                mine=True, # Get playlists for the authenticated user
                maxResults=50,
                pageToken=next_page_token
            )
            response = request.execute()

            for item in response.get("items", []):
                playlist_id = item.get("id")
                playlist_title = item.get("snippet", {}).get("title")
                if playlist_id and playlist_title:
                    playlists_map[playlist_id] = playlist_title

            next_page_token = response.get("nextPageToken")
            if not next_page_token:
                break # Exit loop if no more pages
            time.sleep(0.5) # Avoid hitting API limits too quickly

        print_success(f"Fetched {len(playlists_map)} existing playlists.")
        return playlists_map

    except Exception as e:
        print_error(f"Error fetching playlists: {e}", include_traceback=True)
        return playlists_map # Return what we got so far

def create_playlist(service, title: str, description: str = "") -> str:
    """Creates a new private playlist and returns its ID."""
    if not service or not title:
        print_warning("Missing service or title for creating playlist.")
        return None

    print_info(f"Attempting to create new playlist via API: '{title}'", 3)
    try:
        request = service.playlists().insert(
            part="snippet,status",
            body={
              "snippet": {
                "title": title,
                "description": description,
                "defaultLanguage": "en" # Optional: set language
              },
              "status": {
                "privacyStatus": "private" # Or "public", "unlisted"
              }
            }
        )
        response = request.execute()
        playlist_id = response.get("id")
        if playlist_id:
            print_success(f"Successfully created playlist '{title}' with ID: {playlist_id}", 4)
            return playlist_id
        else:
            print_error(f"Failed to create playlist '{title}'. No ID returned.", 4)
            return None
    except Exception as e:
         print_error(f"Error creating playlist '{title}': {e}", 4, include_traceback=True)
         return None

def add_video_to_playlist(service, video_id: str, playlist_id: str) -> bool:
    """Adds a video to a specified playlist using YouTube Data API."""
    if not service or not video_id or not playlist_id:
        print_warning("Missing service, video ID, or playlist ID for adding to playlist.")
        return False
    try:
        request = service.playlistItems().insert(
            part="snippet",
            body={
                "snippet": {
                    "playlistId": playlist_id,
                    "resourceId": {
                        "kind": "youtube#video",
                        "videoId": video_id
                    }
                    # Optional: Set position 'position': 0 # Add to top
                }
            }
        )
        response = request.execute()
        print_success(f"Successfully added video {video_id} to playlist {playlist_id}.", 4)
        return True
    except Exception as e:
        print_error(f"Error adding video {video_id} to playlist {playlist_id}: {e}", 4, include_traceback=True)
        return False

# --- Correlation Cache Functions ---
def load_correlation_cache():
    """Loads the upload correlation cache from JSON file."""
    cache_path = constants.UPLOAD_CORRELATION_CACHE
    default_cache = []
    if not os.path.exists(cache_path):
        print_info("Correlation cache file not found. Cannot correlate performance.")
        return default_cache
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            content = f.read()
            if not content: return default_cache
            cache = json.loads(content)
            if not isinstance(cache, list):
                 print_warning(f"Correlation cache file '{os.path.basename(cache_path)}' has invalid format. Returning empty cache.")
                 return default_cache
            return cache
    except json.JSONDecodeError:
        print_error(f"Error decoding JSON from correlation cache file '{os.path.basename(cache_path)}'. Returning empty cache.")
        return default_cache
    except Exception as e:
        print_error(f"Error loading correlation cache: {e}", include_traceback=True)
        return default_cache

def save_correlation_cache(cache_data):
    """Saves the upload correlation cache to JSON file."""
    cache_path = constants.UPLOAD_CORRELATION_CACHE
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print_error(f"Error saving correlation cache: {e}", include_traceback=True)

def cleanup_correlation_cache(days_to_keep=7):
    """Removes entries older than specified days from the correlation cache."""
    cache = load_correlation_cache()
    if not cache:
        return # Nothing to cleanup

    now = datetime.now()
    cutoff_date = now - timedelta(days=days_to_keep)
    original_count = len(cache)
    cleaned_cache = []
    removed_count = 0

    for entry in cache:
        try:
            added_time = datetime.fromisoformat(entry.get("added_timestamp"))
            if added_time >= cutoff_date:
                cleaned_cache.append(entry)
            else:
                removed_count += 1
        except (ValueError, TypeError, KeyError):
            # If timestamp is invalid or missing, keep the entry for now
            # Or decide to remove it if strict cleanup is needed
            print_warning(f"Invalid or missing timestamp in correlation cache entry: {entry.get('video_index', 'Unknown')}. Keeping entry.")
            cleaned_cache.append(entry) # Keep potentially problematic entries for manual review

    if removed_count > 0:
        save_correlation_cache(cleaned_cache)
        print_info(f"Cleaned up {removed_count} old entries (>{days_to_keep} days) from correlation cache. {len(cleaned_cache)} remaining.")
    else:
        print_info("No old entries found in correlation cache to cleanup.")

# --- Keyword Generation (Kept from downloader - B.py - includes top performers) ---
def generate_keywords_from_niche(seed_niche, num_keywords=10, top_performing_keywords=None):
    """Generates keywords related to the seed niche using Gemini API."""
    base_prompt = f"""
    Generate keyword list of {num_keywords} diverse and specific keywords strictly related to the "{seed_niche}" niche for youtube shorts section for ranking.
    Focus on aspects like gameplay, release information, characters, storyline, map, news, leaks, secrets, tips, comparisons with other games, fan theories, vehicle speculation, etc., specifically for "{seed_niche}" .
    Ensure most keywords contain "GTA", or "Grand Theft Auto" .
    Exclude generic gaming terms (like 'gaming', 'gameplay' unless highly specific like 'GTA gameplay reveal'), keywords *only* related to previous Grand Theft Auto games. Comparisons are okay if they mention "{seed_niche}".
    Exclude names of social media platforms or streaming services (YouTube, Twitch, TikTok, etc.).
    Provide each keyword on a new line. Do not include any explanations, introductory phrases, or bullet points/numbering. Just the raw keywords.
    """
    if top_performing_keywords and len(top_performing_keywords) > 0:
        top_keywords_str = ", ".join(top_performing_keywords)
        prompt = f"""
        {base_prompt}

        Additionally, use these successful keywords as inspiration for generating similar high-performing keywords: {top_keywords_str}.
        These keywords have performed well in the past, so generate keywords that are similar in style and content.
        """
    else: prompt = base_prompt

    try:
        model = genai.GenerativeModel("gemini-2.0-flash") # Using 2.0 flash
        response = model.generate_content(prompt)
        keywords = [line.strip() for line in response.text.splitlines() if line.strip()]
        # Basic filter applied here too
        keywords = [kw for kw in keywords if any(sub.lower() in kw.lower() for sub in ["GTA", "Grand Theft Auto"]) and len(kw) > 3]
        return keywords
    except Exception as e:
        print_error(f"Error generating keywords from niche '{seed_niche}': {e}", 1, include_traceback=True)
        return []

# --- Metadata Validation ---
def validate_generated_metadata(metadata, video_title, metadata_metrics=None):
    """Performs cross-validation checks on generated metadata and updates metrics."""
    warnings = []
    title = metadata.get("title", "")
    description = metadata.get("description", "")
    tags = metadata.get("tags", [])

    # Flags to track specific validation issues (to only increment metrics once per validation)
    title_mismatch_detected = False
    tag_list_error_detected = False
    stuffing_detected = False

    if not title: warnings.append("Generated title is empty.")
    if not description: warnings.append("Generated description is empty.")
    if not tags: warnings.append("Generated tags list is empty.")

    # 1. Title in Description (improved check)
    # Remove #Shorts for check, compare lowercase
    title_base = title.replace("#Shorts", "").strip().lower()

    # Check if title is at the beginning of the description (first 200 chars)
    # This is more lenient than requiring exact match
    desc_start = description[:200].lower()

    if title_base and len(title_base) > 5:  # Only check if title is meaningful
        # Try different approaches to find the title in the description
        title_words = set(re.findall(r'\b\w+\b', title_base))
        significant_words = [w for w in title_words if len(w) > 3 and w not in ['this', 'that', 'with', 'from', 'have', 'what', 'when', 'where', 'will', 'your']]

        # Check if at least 70% of significant words from title appear in the first 200 chars of description
        if len(significant_words) > 0:
            words_found = sum(1 for word in significant_words if word in desc_start)
            match_percentage = words_found / len(significant_words)

            if match_percentage < 0.7:  # Less than 70% of significant words found
                warnings.append(f"Generated title base ('{title_base[:50]}...') not found in description.")
                if metadata_metrics and not title_mismatch_detected:
                    metadata_metrics["validation_title_mismatches"] += 1
                    title_mismatch_detected = True

    # 2. Tags mentioned in Description (improved check)
    tags_in_desc_heading = "Tags Used in Video :-"
    tags_listed_incorrectly = False

    if tags_in_desc_heading.lower() in description.lower():
        try:
            # Find text after heading, split by comma/newline, strip whitespace
            desc_parts = description.lower().split(tags_in_desc_heading.lower(), 1)
            if len(desc_parts) > 1:
                # Take text after heading, stop at next potential heading or end
                tags_text_area = desc_parts[1].split("\n\n")[0].split("ignored hashtags :-")[0]

                # Improved tag extraction - handle asterisk prefixes and other formatting
                raw_tags = re.split(r'[\n,]', tags_text_area)
                tags_mentioned_in_desc = set()

                for tag in raw_tags:
                    # Clean up the tag - remove asterisks, hashes, and other formatting
                    cleaned_tag = re.sub(r'^\s*[\*\-•#]\s*', '', tag).strip().lower()
                    if cleaned_tag:
                        tags_mentioned_in_desc.add(cleaned_tag)

                # Clean up the generated tags for comparison
                tags_generated_lower = set()
                for tag in tags:
                    # Remove special characters and normalize
                    cleaned_tag = re.sub(r'[^\w\s]', '', tag.lower()).strip()
                    if cleaned_tag:
                        tags_generated_lower.add(cleaned_tag)

                # Compare the cleaned sets
                # Allow for some flexibility - check if at least 80% of tags are present
                if tags_generated_lower and tags_mentioned_in_desc:
                    common_tags = tags_generated_lower.intersection(tags_mentioned_in_desc)
                    match_percentage = len(common_tags) / len(tags_generated_lower)

                    if match_percentage < 0.8:  # Less than 80% of tags found
                        missing_in_desc = tags_generated_lower - tags_mentioned_in_desc
                        extra_in_desc = tags_mentioned_in_desc - tags_generated_lower

                        if missing_in_desc:
                            warnings.append(f"Tags listed in <tags> but missing under description heading: {missing_in_desc}")
                            tags_listed_incorrectly = True
                        if extra_in_desc:
                            warnings.append(f"Tags mentioned under description heading but not in <tags>: {extra_in_desc}")
                            tags_listed_incorrectly = True

        except Exception as e:
            warnings.append(f"Error parsing tags from description: {e}")
            tags_listed_incorrectly = True
    else:
        warnings.append(f"'{tags_in_desc_heading}' heading not found in description.")
        tags_listed_incorrectly = True

    if tags_listed_incorrectly and metadata_metrics and not tag_list_error_detected:
        metadata_metrics["validation_tag_list_errors"] += 1
        tag_list_error_detected = True

    # 3. Keyword Stuffing (Simple Heuristic)
    words_in_desc = description.lower().split()
    if tags and words_in_desc:
        max_tag_word_occurrences = 7 # Configurable threshold
        # Use Counter for more efficient word counting
        word_counts = collections.Counter(words_in_desc)

        for tag in tags:
            tag_words = tag.lower().split()
            for tag_word in tag_words:
                if len(tag_word) > 3: # Only check meaningful words
                    count = word_counts.get(tag_word, 0)
                    if count > max_tag_word_occurrences:
                        warnings.append(f"Potential keyword stuffing: Word '{tag_word}' (from tag '{tag}') appears {count} times in description.")
                        if metadata_metrics and not stuffing_detected:
                            metadata_metrics["validation_keyword_stuffing"] += 1
                            stuffing_detected = True
                        break # Only report once per tag

    if warnings:
        print_warning(f"Metadata validation warnings for '{video_title}':", 3)
        for warn in warnings:
            print_warning(f"  - {warn}", 4)
        # Optionally add warnings to metadata object?
        # metadata['validation_warnings'] = warnings
    else:
        print_success(f"Metadata validation passed for '{video_title}'.", 3)

# --- Metadata Prompt Improvement (Adjusted for SEO context) ---
def improve_metadata_prompt(error_metrics):
    """Uses Gemini to improve the SEO metadata prompt based on error patterns."""
    current_prompt = load_or_get_seo_prompt_template()
    if not current_prompt:
        print_error("Cannot improve prompt: Failed to load current SEO prompt.")
        return None

    total_calls = error_metrics.get("total_api_calls", 0)
    if total_calls == 0:
        print_warning("No API calls recorded yet. Cannot improve prompt.")
        return None

    error_summary = [f"Total API calls: {total_calls}"]
    # Calculate error rates for relevant metrics
    for error_type in ["parse_failures", "timeouts", "empty_description_errors", "empty_tags_errors"]:
        count = error_metrics.get(error_type, 0)
        rate = count / total_calls if total_calls > 0 else 0
        # Clarify metric name in summary
        metric_name = error_type.replace("_", " ").title()
        error_summary.append(f"{metric_name}: {count} ({rate:.1%})")

    # Add validation warning metrics
    validation_summary = []
    for warning_type in ["validation_title_mismatches", "validation_tag_list_errors", "validation_keyword_stuffing"]:
        count = error_metrics.get(warning_type, 0)
        rate = count / total_calls if total_calls > 0 else 0
        # Clarify metric name in summary
        metric_name = warning_type.replace("validation_", "").replace("_", " ").title()
        validation_summary.append(f"{metric_name}: {count} ({rate:.1%})")

    error_summary.append("\n=== Validation Warnings ===")
    error_summary.extend(validation_summary)

    error_samples = error_metrics.get("error_samples", [])
    if error_samples:
        error_summary.append("\nRecent error samples:")
        for i, sample in enumerate(error_samples[-5:], 1):
            error_summary.append(f"Sample {i}: {sample.get('type')} - {sample.get('details')}")
    error_summary_text = "\n".join(error_summary)

    # *** ADJUSTED META-PROMPT FOR SEO CONTEXT WITH VALIDATION WARNINGS ***
    meta_prompt = f"""
    Review the following prompt used to generate SEO-optimized YouTube Shorts metadata:

    ```
    {current_prompt}
    ```

    Based on the following issues observed (metrics over multiple runs):

    {error_summary_text}

    Please provide an improved version of the SEO metadata generation prompt. Focus on clarity, robustness, and addressing the specific issues indicated by the metrics. The improved prompt should instruct the metadata generation model to:

    1.  Adhere strictly to the requested XML structure (`<metadata>`, `<title>`, `<description>`, `<tags>`) and output *only* this structure.
    2.  Generate SEO-focused titles, descriptions, and tags.
    3.  Use keywords **naturally** within the description. **Explicitly instruct the model to avoid excessive repetition of the same keywords or phrases.** Aim for readability and value over simple repetition.
    4.  Ensure the title is properly reflected in the description content (to prevent title mismatches).
    5.  Provide clear instructions for including all tags in the "Tags Used in Video :-" section of the description, ensuring the tag list in the description matches exactly with the tags in the <tags> section.
    6.  Ensure all specific elements requested within the description (e.g., title inclusion, credit, disclaimer, tags list, CTA) are present and correctly formatted.
    7.  Generate comprehensive and relevant tag lists.
    8.  Minimize parsing failures and the generation of empty/incomplete `<title>`, `<description>`, or `<tags>` elements.
    9.  Add a final self-check step within the prompt for the model to verify all elements are present and constraints (like keyword usage) are met before outputting.

    Provide ONLY the improved prompt text, without any explanations or additional text.
    """

    try:
        model = genai.GenerativeModel("gemini-2.0-flash") # Using 2.0 flash
        response = model.generate_content(meta_prompt)
        improved_prompt = response.text.strip()

        # Basic validation
        if "<metadata>" not in improved_prompt or "<title>" not in improved_prompt or \
           "<description>" not in improved_prompt or "<tags>" not in improved_prompt:
            print_error("Generated prompt is missing required XML tags. Keeping current prompt.", 1)
            return None
        if "{video_topic}" not in improved_prompt or "{uploader_name}" not in improved_prompt:
             print_error("Generated prompt is missing required placeholders. Keeping current prompt.", 1)
             return None

        return improved_prompt
    except Exception as e:
        print_error(f"Error generating improved SEO metadata prompt: {e}", 1, include_traceback=True)
        return None

def load_or_get_seo_prompt_template():
    """Loads the SEO prompt template, trying cache first, then default."""
    global _current_seo_prompt_template
    if _current_seo_prompt_template:
        return _current_seo_prompt_template

    prompt_cache_path = constants.SEO_METADATA_PROMPT_FILE
    prompt_loaded = False

    # Try loading from cache file
    if os.path.exists(prompt_cache_path):
        try:
            with open(prompt_cache_path, "r", encoding="utf-8") as f:
                _current_seo_prompt_template = f.read()
            if "{video_topic}" in _current_seo_prompt_template and "{uploader_name}" in _current_seo_prompt_template:
                print_info(f"Loaded SEO metadata prompt from cache: {os.path.basename(prompt_cache_path)}")
                prompt_loaded = True
            else:
                print_warning(f"Cached SEO prompt in '{os.path.basename(prompt_cache_path)}' seems invalid (missing placeholders). Using default.")
                _current_seo_prompt_template = None # Reset to force using default
        except Exception as e:
            print_warning(f"Error loading cached SEO prompt from '{os.path.basename(prompt_cache_path)}': {e}. Using default.")
            _current_seo_prompt_template = None

    # If not loaded from cache, use the default inline prompt
    if not prompt_loaded:
        print_info("Using default inline SEO metadata prompt.")
        _current_seo_prompt_template = """
        do not include any explanation or any other text. just give me the metadata in below format.
        only apply the below format. do not include any other text or explanation.
        Generate SEO-optimized metadata for a YouTube Shorts video in the following structured format:
        You are a YOUTUBE SEO EXPERT A GURU one in million. you have insight knowlodge of youtube shorts.
        you know how the ranking algorithm works and how to get more views and engagement.
        you know how creator like mrbeast, tseries, and other top creators get more views and engagement.
        your master of youtube shorts. you have worked with big creato know all secrets of youtube shorts.
        you have worked in google youtube team and you know all secrets of youtube shorts.
        Our Channel Name is "Random GTA Shorts" and we are a gaming channel focused on Grand Theft Auto.
        Our channel is about GTA but mostly cover gta vi news, gameplay, trailers, secrets, and more. We are creating youtube shorts on GTA games.
        include a copyright fair use disclaimer in the description.
        APPLY ALL OF THE ABOVE KNOWLEDGE AND SECRETS TO BELOW metadata.

        <metadata>
            <title>
                Create an engaging, fast-paced, and action-driven title with a high CTR (Click-Through Rate).
                Follow these best practices, drawing inspiration from successful short-form content:
                    * Use primary keywords related to {video_topic} (e.g., GTA, GTA VI, GTA gameplay, GTA VI trailer, GTA 6, Grand Theft Auto , GTA news, GTA gameplay ).
                    * Make it engaging by using relevant emojis if applicable (like 🔥, 💥, 🎮), numbers, or power words if applicable (like BEST, HOT, ULTIMATE, NEW).
                    * Use numbers if applicable (e.g., "5 Secrets", "Top 10").
                    * Use power words if applicable.
                    * Use words that trigger curiosity if applicable.
                    * Keep it under 100 characters for better visibility on mobile devices.
                    * Avoid clickbait, but make it intriguing.
                    * Use a mix of uppercase and lowercase letters for emphasis and readability.
                    * Use a question format to spark curiosity if applicable.
                    * Use a dash (-) to create a sense of urgency if applicable.
                    * Add 2-3 relevant hashtags related to the video at the end for extra visibility .
            </title>
            <description>
                Write an SEO-optimized description with the following structure:
                    * The description should be Beautifully formated and seo optimized.
                    * Put the video title at the beginning. Start with the video title.
                    * Keep it under 4500 characters.
                    * Use a mix of primary and secondary keywords related to {video_topic}.
                    * Description should be beautifully written and engaging.
                    * Include main keywords in the first 150 characters.
                    * Use LSI (Latent Semantic Indexing) keywords naturally.
                    * **Include credit to the original uploader in the following format: "Credit to: {uploader_name}".**
                    * include original video title in format: "Original Title: {original_title}
                    * Include a copyright fair use disclaimer similar to this:
                        --------------【Copyright Disclaimer】-------------
                        All the videos, songs, images, and graphics used in the video belong to their respective owners and I or this channel does not claim any rights over them.

                        Copyright Disclaimer under section 107 of the Copyright Act of 1976, allowance is made for “fair use” for purposes such as criticism,comment, news reporting, teaching, scholarship, education and research. Fair use is a use permitted by copyright statute that might otherwise beinfringing.
                    * Use a mix of short and long sentences for better readability.
                    * After the copyright disclaimer, include 10 to 15 relevant hashtags related to the video.
                    * After that use an heading "Tags Used in Video :-" and include all the tags from <tags> </tags> section.
                    * Use bullet points or numbered lists for key points.
                    * Include a brief summary of the video content in a few sentences.
                    * Highlight the benefits of watching the video if applicable).
                    * End with a Call to Action (e.g., "Like and Subscribe for more GTA content!").
                    * Add relevant hashtags at the end for discoverability related to {video_topic}
                    * At the end use an heading "Ignored Hashtags :-" related to {video_topic} and use as many hashtags as possible but it should not go over description limit of 4500 characters.
            </description>
            <tags>
                Suggest SEO-friendly 25-35 tags and keywords it should help with seo ranking for {video_topic} for better discoverability.
                    * Ranking keywords should be at the top of the list .
                    * Start with the main keyword related to {video_topic}.
                    * Use a mix of general & niche tags and keywords related to {video_topic}.
                    * Stay under 500 characters to avoid getting cut off.
            </tags>
        </metadata>

        **Video Topic**: {video_topic} the title should be close to this topic but improve it. and make it more engaging and catchy.
        """
    return _current_seo_prompt_template

# Function to save the potentially improved SEO prompt
def save_seo_prompt_template(prompt_text):
    """Saves the SEO prompt template to the cache file."""
    global _current_seo_prompt_template
    prompt_cache_path = constants.SEO_METADATA_PROMPT_FILE
    try:
        with open(prompt_cache_path, "w", encoding="utf-8") as f:
            f.write(prompt_text)
        _current_seo_prompt_template = prompt_text # Update in-memory cache
        print_success(f"Saved updated SEO metadata prompt to cache: {os.path.basename(prompt_cache_path)}")
    except Exception as e:
        print_error(f"Error saving updated SEO metadata prompt to cache: {e}")


# MODIFIED FOR COMBINED SCRIPT - Uses SEO Prompt
def generate_seo_metadata(video_topic, uploader_name="Unknown Uploader", original_title="Unknown Title"):
    """Generates SEO-optimized metadata using Gemini API (Strategy A)."""

    # Load the current SEO prompt template
    prompt_template = load_or_get_seo_prompt_template()
    if not prompt_template:
        print_error("Failed to load SEO prompt template. Using minimal fallback.")
        # Define a minimal fallback inline if loading fails completely
        prompt_template = """
        <metadata>
            <title>Short video about {video_topic}</title>
            <description>Watch this short video about {video_topic}. Credit to: {uploader_name}. Original Title: {original_title}</description>
            <tags>{video_topic}, shorts, gta</tags>
        </metadata>
        """

    # Format the prompt with actual data
    prompt = prompt_template.replace("{video_topic}", video_topic)\
                            .replace("{uploader_name}", uploader_name)\
                            .replace("{original_title}", original_title) # Add original title replacement

    # Default fallback structure
    metadata = {
        "title": video_topic,
        "description": f"Default SEO description for {video_topic}.\n\nCredit to: {uploader_name}\nOriginal Title: {original_title}",
        "tags": ["gta", "shorts", "gaming", video_topic.lower().replace(" ", "")] # Basic tags
    }

    try:
        # Using 2.0-flash for better quality and performance
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        raw_text = response.text

        # Reset before parsing
        metadata = {"title": video_topic, "description": "", "tags": []}

        # --- Parsing (Adjusted for SEO prompt structure) ---
        title_match = re.search(r"<title>(.*?)</title>", raw_text, re.DOTALL | re.IGNORECASE)
        desc_match = re.search(r"<description>(.*?)</description>", raw_text, re.DOTALL | re.IGNORECASE)
        tags_match = re.search(r"<tags>(.*?)</tags>", raw_text, re.DOTALL | re.IGNORECASE)

        parsing_warnings = []
        if title_match and title_match.group(1).strip():
            metadata["title"] = title_match.group(1).strip()
        else:
            parsing_warnings.append("title")
            metadata["title"] = video_topic # Keep topic as fallback before length check

        if desc_match and desc_match.group(1).strip():
            metadata["description"] = desc_match.group(1).strip()
            # Ensure uploader credit and original title are present (redundant if prompt followed)
            if f"Credit to: {uploader_name}" not in metadata["description"]:
                 metadata["description"] += f"\n\nCredit to: {uploader_name}"
            if f"Original Title: {original_title}" not in metadata["description"]:
                 metadata["description"] += f"\nOriginal Title: {original_title}"
        else:
            parsing_warnings.append("description")
            metadata["description"] = f"Default SEO description for {video_topic}. Parsing error.\n\nCredit to: {uploader_name}\nOriginal Title: {original_title}"

        if tags_match and tags_match.group(1).strip():
            tags_raw = tags_match.group(1).strip()
            # Handle comma or newline separated tags
            metadata["tags"] = [tag.strip() for tag in re.split(r'[,\n]', tags_raw) if tag.strip()]
        else:
            parsing_warnings.append("tags")
            metadata["tags"] = ["gta", "shorts", "gaming", video_topic.lower().replace(" ", "")]

        if parsing_warnings:
            print_warning(f"Could not parse <{'>, <'.join(parsing_warnings)}> for topic: {video_topic}. Used fallbacks.", 1)


        # --- Post-Processing (SEO Strategy) ---

        # Title Length and #Shorts Handling
        temp_title = metadata.get("title", video_topic)
        if len(temp_title) > TARGET_TITLE_LENGTH: # Use TARGET_TITLE_LENGTH (90)
            truncated = temp_title[:TARGET_TITLE_LENGTH]
            last_space = truncated.rfind(' ')
            temp_title = truncated[:last_space].strip() if last_space > 0 else truncated.strip()

        if not temp_title.endswith(SHORTS_SUFFIX):
            if len(temp_title) + len(SHORTS_SUFFIX) <= MAX_TITLE_LENGTH:
                temp_title += SHORTS_SUFFIX
            # else: Title is already too long even without suffix

        metadata["title"] = temp_title

        # Add top 5 tags as hashtags to the description if not already present
        # (The SEO prompt asks for hashtags at the end, but this ensures some are there if the prompt fails)
        tags_list = metadata.get("tags", [])
        num_hashtags_to_add = min(len(tags_list), 5)
        if num_hashtags_to_add > 0:
            hashtags_from_tags = " ".join([f"#{tag.replace(' ', '').replace('-', '').lower()}" for tag in tags_list[:num_hashtags_to_add]])
            current_desc = metadata.get("description", "")
            # Check if these hashtags are already roughly in the description
            if hashtags_from_tags.lower() not in current_desc.lower():
                 if current_desc and not current_desc.endswith(("\n", "\n\n")):
                     current_desc += "\n\n"
                 metadata["description"] = current_desc + hashtags_from_tags
                 print_info(f"Added hashtags from tags to description: {hashtags_from_tags}", 2)

    except Exception as e:
        print_error(f"Error during metadata generation/processing [SEO] for '{video_topic}': {e}", 1, include_traceback=True)
        # Fallback metadata consistent with SEO Style
        fallback_title = video_topic
        if len(fallback_title) > TARGET_TITLE_LENGTH:
             truncated = fallback_title[:TARGET_TITLE_LENGTH]
             last_space = truncated.rfind(' ')
             fallback_title = truncated[:last_space].strip() if last_space > 0 else truncated.strip()
        if not fallback_title.endswith(SHORTS_SUFFIX) and len(fallback_title) + len(SHORTS_SUFFIX) <= MAX_TITLE_LENGTH:
             fallback_title += SHORTS_SUFFIX
        metadata = {
            "title": fallback_title,
            "description": f"Default SEO description for {video_topic}. Error during generation: {e}\n\nCredit to: {uploader_name}\nOriginal Title: {original_title}",
            "tags": ["gta", "shorts", "gaming", "error", video_topic.lower().replace(" ", "")]
        }

    # Final check for basic structure integrity
    if not isinstance(metadata, dict) or not metadata.get("title") or not metadata.get("description") or not metadata.get("tags"):
        print_error(f"Critical Warning [SEO]: Metadata became invalid/empty for '{video_topic}'. Using final minimal fallback.", 1)
        fallback_title = video_topic
        if len(fallback_title) > TARGET_TITLE_LENGTH:
             truncated = fallback_title[:TARGET_TITLE_LENGTH]
             last_space = truncated.rfind(' ')
             fallback_title = truncated[:last_space].strip() if last_space > 0 else truncated.strip()
        if not fallback_title.endswith(SHORTS_SUFFIX) and len(fallback_title) + len(SHORTS_SUFFIX) <= MAX_TITLE_LENGTH:
             fallback_title += SHORTS_SUFFIX
        metadata = {
            "title": fallback_title,
            "description": f"Default SEO description for {video_topic}. Final fallback.\n\nCredit to: {uploader_name}\nOriginal Title: {original_title}",
            "tags": ["gta", "shorts", "gaming", "error", "final_fallback"]
        }
    return metadata


# MODIFIED FOR COMBINED SCRIPT - Uses SEO Style Fallback & Tracks SEO Errors
def generate_metadata_with_timeout(video_title, uploader_name, original_title="Unknown Title", timeout=METADATA_TIMEOUT_SECONDS):
    """Generates metadata with a timeout (SEO Style Fallback) and suggests category."""
    metadata_metrics = load_metadata_metrics()
    metadata_metrics["total_api_calls"] += 1 # Count main metadata call
    error_type = None
    error_details = None
    result_metadata = None # Initialize result

    try:
        with concurrent.futures.ThreadPoolExecutor() as executor:
            # --- Generate Title/Desc/Tags ---
            print_info("Generating primary SEO metadata (Title/Desc/Tags)...", 2)
            future = executor.submit(generate_seo_metadata, video_title, uploader_name, original_title)
            result_metadata = future.result(timeout=timeout) # Get the main metadata first

            # --- Now Suggest Category (if primary metadata succeeded) ---
            suggested_category = None
            if result_metadata: # Only proceed if primary metadata generation didn't fail/timeout critically
                 # Use the *generated* title and description for category suggestion
                gen_title = result_metadata.get("title", video_title)
                gen_desc = result_metadata.get("description", "")
                if gen_title and gen_desc:
                    # We don't need timeout here, make it a quick synchronous call
                    # No separate thread needed for the category suggestion
                    try:
                        suggested_category = get_suggested_category(gen_title, gen_desc)
                        if suggested_category:
                            result_metadata['suggested_category'] = suggested_category
                    except Exception as cat_err:
                         print_error(f"Error occurred during category suggestion call: {cat_err}", 2)
                         # Continue without suggested category if this specific call fails
                else:
                     print_warning("Skipping category suggestion because generated title/description was empty.", 2)

            # Check for parsing errors / empty results specifically for SEO context
            # (Title check is less strict as fallback is the topic itself before length adjustment)
            # Check if description is still the default/error fallback
            if f"Default SEO description for {video_title}" in result_metadata.get("description", ""):
                 metadata_metrics["empty_description_errors"] += 1
                 error_type = "empty_description"
                 error_details = f"Failed to generate description for: {video_title}"
            # Check if tags are still the default/error fallback
            if result_metadata.get("tags") == ["gta", "shorts", "gaming", video_title.lower().replace(" ", "")]:
                 metadata_metrics["empty_tags_errors"] += 1
                 error_type = error_type or "empty_tags" # Keep first error type found
                 error_details = error_details or f"Failed to generate tags for: {video_title}"
            # Add more specific checks based on SEO prompt expectations if needed

            if error_type and error_details:
                add_error_sample(metadata_metrics, error_type, error_details, video_title)

            # Validate metadata and track stuffing warnings
            if result_metadata: # Only validate if we have metadata
                validate_generated_metadata(result_metadata, video_title, metadata_metrics)

            save_metadata_metrics(metadata_metrics) # Save after checking result
            return result_metadata # Return the metadata dict (potentially with suggested_category)

    except concurrent.futures.TimeoutError:
        print_warning(f"Primary metadata generation timed out [SEO] for: {video_title}", 1)
        metadata_metrics["timeouts"] += 1
        add_error_sample(metadata_metrics, "timeout", f"Metadata generation timed out for: {video_title}", video_title)

        # Fallback metadata consistent with SEO Style (NO suggested category here)
        fallback_title = video_title
        if len(fallback_title) > TARGET_TITLE_LENGTH:
             truncated = fallback_title[:TARGET_TITLE_LENGTH]
             last_space = truncated.rfind(' ')
             fallback_title = truncated[:last_space].strip() if last_space > 0 else truncated.strip()
        if not fallback_title.endswith(SHORTS_SUFFIX) and len(fallback_title) + len(SHORTS_SUFFIX) <= MAX_TITLE_LENGTH:
             fallback_title += SHORTS_SUFFIX

        fallback_metadata = {
            "title": fallback_title,
            "description": f"Default SEO description for {video_title}. Generation timed out.\n\nCredit to: {uploader_name}\nOriginal Title: {original_title}",
            "tags": ["gta", "shorts", "gaming", "timeout", video_title.lower().replace(" ", "")]
            # No 'suggested_category' in timeout fallback
        }

        # Validate metadata and track stuffing warnings
        validate_generated_metadata(fallback_metadata, video_title, metadata_metrics)

        save_metadata_metrics(metadata_metrics)
        return fallback_metadata
    except Exception as e:
        print_error(f"Error submitting primary metadata generation job [SEO] for '{video_title}': {e}", 1, include_traceback=True)
        metadata_metrics["parse_failures"] += 1 # Use parse_failures for general exceptions during generation
        add_error_sample(metadata_metrics, "exception", f"Error: {str(e)} for: {video_title}", video_title)

        # Fallback metadata consistent with SEO Style (NO suggested category here)
        fallback_title = video_title
        if len(fallback_title) > TARGET_TITLE_LENGTH:
             truncated = fallback_title[:TARGET_TITLE_LENGTH]
             last_space = truncated.rfind(' ')
             fallback_title = truncated[:last_space].strip() if last_space > 0 else truncated.strip()
        if not fallback_title.endswith(SHORTS_SUFFIX) and len(fallback_title) + len(SHORTS_SUFFIX) <= MAX_TITLE_LENGTH:
             fallback_title += SHORTS_SUFFIX

        fallback_metadata = {
            "title": fallback_title,
            "description": f"Default SEO description for {video_title}. Error during task submission: {e}\n\nCredit to: {uploader_name}\nOriginal Title: {original_title}",
            "tags": ["gta", "shorts", "gaming", "error", "submission"]
            # No 'suggested_category' in exception fallback
        }

        # Validate metadata and track stuffing warnings
        validate_generated_metadata(fallback_metadata, video_title, metadata_metrics)

        save_metadata_metrics(metadata_metrics)
        return fallback_metadata

# --- Keyword Generation (Kept from downloader - B.py - includes top performers) ---
def generate_keywords_from_niche(seed_niche, num_keywords=10, top_performing_keywords=None):
    """Generates keywords related to the seed niche using Gemini API."""
    base_prompt = f"""
    Generate keyword list of {num_keywords} diverse and specific keywords strictly related to the "{seed_niche}" niche for youtube shorts section for ranking.
    Focus on aspects like gameplay, release information, characters, storyline, map, news, leaks, secrets, tips, comparisons with other games, fan theories, vehicle speculation, etc., specifically for "{seed_niche}" .
    Ensure most keywords contain "GTA", or "Grand Theft Auto" .
    Exclude generic gaming terms (like 'gaming', 'gameplay' unless highly specific like 'GTA gameplay reveal'), keywords *only* related to previous Grand Theft Auto games. Comparisons are okay if they mention "{seed_niche}".
    Exclude names of social media platforms or streaming services (YouTube, Twitch, TikTok, etc.).
    Provide each keyword on a new line. Do not include any explanations, introductory phrases, or bullet points/numbering. Just the raw keywords.
    """
    if top_performing_keywords and len(top_performing_keywords) > 0:
        top_keywords_str = ", ".join(top_performing_keywords)
        prompt = f"""
        {base_prompt}

        Additionally, use these successful keywords as inspiration for generating similar high-performing keywords: {top_keywords_str}.
        These keywords have performed well in the past, so generate keywords that are similar in style and content.
        """
    else: prompt = base_prompt

    try:
        model = genai.GenerativeModel("gemini-2.0-flash") # Using 2.0 flash
        response = model.generate_content(prompt)
        keywords = [line.strip() for line in response.text.splitlines() if line.strip()]
        # Basic filter applied here too
        keywords = [kw for kw in keywords if any(sub.lower() in kw.lower() for sub in ["GTA", "Grand Theft Auto"]) and len(kw) > 3]
        return keywords
    except Exception as e:
        print_error(f"Error generating keywords from niche '{seed_niche}': {e}", 1, include_traceback=True)
        return []

# --- Metadata Prompt Improvement (Adjusted for SEO context) ---
def improve_metadata_prompt(error_metrics):
    """Uses Gemini to improve the SEO metadata prompt based on error patterns."""
    current_prompt = load_or_get_seo_prompt_template()
    if not current_prompt:
        print_error("Cannot improve prompt: Failed to load current SEO prompt.")
        return None

    total_calls = error_metrics.get("total_api_calls", 0)
    if total_calls == 0:
        print_warning("No API calls recorded yet. Cannot improve prompt.")
        return None

    error_summary = [f"Total API calls: {total_calls}"]
    # Calculate error rates for relevant metrics
    for error_type in ["parse_failures", "timeouts", "empty_description_errors", "empty_tags_errors", "keyword_stuffing_warnings"]:
        count = error_metrics.get(error_type, 0)
        rate = count / total_calls if total_calls > 0 else 0
        # Clarify metric name in summary
        metric_name = error_type.replace("_", " ").title()
        error_summary.append(f"{metric_name}: {count} ({rate:.1%})")

    error_samples = error_metrics.get("error_samples", [])
    if error_samples:
        error_summary.append("\nRecent error samples:")
        for i, sample in enumerate(error_samples[-5:], 1):
            error_summary.append(f"Sample {i}: {sample.get('type')} - {sample.get('details')}")
    error_summary_text = "\n".join(error_summary)

    # *** ADJUSTED META-PROMPT FOR SEO CONTEXT ***
    meta_prompt = f"""
    Review the following prompt used to generate SEO-optimized YouTube Shorts metadata:

    ```
    {current_prompt}
    ```

    Based on the following issues observed:

    {error_summary_text}

    Please provide an improved version of the SEO metadata generation prompt. Focus on clarity and robustness to avoid these specific problems, especially parsing failures, empty results, and keyword stuffing. The prompt should:

    1.  Be absolutely explicit about the required XML structure (`<metadata>`, `<title>`, `<description>`, `<tags>`) and that *only* this structure should be output.
    2.  Reinforce the instructions for generating SEO-focused titles, detailed descriptions (including keywords used *naturally*, structure, credits, disclaimers, hashtags), and comprehensive tag lists.
    3.  Instruct the model to avoid excessive repetition of the same keywords within the description.
    4.  Ensure the instructions for description elements (like credits, original title placeholders) are crystal clear to minimize omissions.
    5.  Maintain the overall goal of SEO optimization for discoverability.
    6.  Perhaps add an instruction to double-check that all requested elements (title, description, tags) are included within their respective XML tags before finishing.

    Provide ONLY the new prompt text, without any explanations or additional text.
    """

    try:
        model = genai.GenerativeModel("gemini-1.5-flash") # Using flash
        response = model.generate_content(meta_prompt)
        improved_prompt = response.text.strip()

        # Basic validation
        if "<metadata>" not in improved_prompt or "<title>" not in improved_prompt or \
           "<description>" not in improved_prompt or "<tags>" not in improved_prompt:
            print_error("Generated prompt is missing required XML tags. Keeping current prompt.", 1)
            return None
        if "{video_topic}" not in improved_prompt or "{uploader_name}" not in improved_prompt:
             print_error("Generated prompt is missing required placeholders. Keeping current prompt.", 1)
             return None

        return improved_prompt
    except Exception as e:
        print_error(f"Error generating improved SEO metadata prompt: {e}", 1, include_traceback=True)
        return None

# --- Category Suggestion Function ---
def get_suggested_category(title: str, description: str):
    """Asks Gemini for the most appropriate YouTube category from a predefined list."""
    if not title or not description:
        print_warning("Cannot suggest category: Title or Description is empty.", 2)
        return None

    # --- Keep the list here for validation ---
    KNOWN_CATEGORIES = [
        "Film & Animation", "Autos & Vehicles", "Music", "Pets & Animals",
        "Sports", "Travel & Events", "Gaming", "People & Blogs",
        "Comedy", "Entertainment", "News & Politics", "Howto & Style",
        "Education", "Science & Technology", "Nonprofits & Activism"
        # Removed "Movies", "Shows" as they are less common for Shorts uploads via Studio
    ]
    KNOWN_CATEGORIES_LOWER = {cat.lower() for cat in KNOWN_CATEGORIES}

    # --- MODIFY THE PROMPT ---
    valid_categories_string = ", ".join([f'"{cat}"' for cat in KNOWN_CATEGORIES]) # Create quoted list for prompt

    prompt = f"""
    Analyze the following YouTube Shorts video Title and Description:

    Title: {title}
    Description: {description[:1000]} # Limit description length

    Select the single BEST matching official YouTube Video Category for this content.
    You MUST choose EXACTLY ONE category name from this official list:
    {valid_categories_string}

    Output ONLY the chosen category name from the list, with the exact capitalization shown in the list, and nothing else.
    For example, if the content is about gaming, output: Gaming
    If it's about fitness exercises, the best fit from the list might be "Howto & Style" or "Sports". Choose the single most appropriate one FROM THE LIST PROVIDED.
    """

    try:
        print_info("Requesting category suggestion (with explicit list)...", 3)
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        suggested_cat_raw = response.text.strip()

        # --- Keep the validation logic, but it should pass more often now ---
        if suggested_cat_raw and suggested_cat_raw in KNOWN_CATEGORIES: # Check against the original list directly now
            print_success(f"Suggested category: {suggested_cat_raw}", 3)
            return suggested_cat_raw
        elif suggested_cat_raw:
            # This might still happen if Gemini hallucinates or ignores instructions
            print_warning(f"Gemini suggested a category NOT in the provided list: '{suggested_cat_raw}'. Ignoring.", 3)
            # Log the failed suggestion for debugging
            print_error(f"Category Suggestion Mismatch: Gemini suggested '{suggested_cat_raw}' which is not in the allowed list based on Title: '{title}'")
            return None
        else:
            print_warning("Gemini returned an empty category suggestion.", 3)
            return None
    except Exception as e:
        print_error(f"Error getting category suggestion: {e}", 3)
        print_error(f"Gemini Category Suggestion Error: {e} for Title: '{title}'") # Log error
        return None

# --- Performance Metrics & Tuning Suggestions (Kept from downloader - B.py) ---
def get_top_performing_keywords(keyword_frequency, top_n=5):
    """Returns the top N performing keywords based on frequency/score."""
    if not keyword_frequency: return []
    # Scores can be floats now, handle potential errors during sort
    items = []
    for kw, score in keyword_frequency.items():
        try: items.append((kw, float(score)))
        except (ValueError, TypeError): items.append((kw, 0.0)) # Assign 0 if score is invalid
    sorted_keywords = sorted(items, key=lambda x: x[1], reverse=True)
    return [kw for kw, _ in sorted_keywords[:top_n]]

def load_performance_metrics():
    """Loads overall performance metrics from the JSON file."""
    metrics_file_path = constants.PERFORMANCE_METRICS_FILE
    default_metrics = {
        "runs": [], "total_shorts_found": 0, "total_suitable_shorts": 0,
        "total_downloads_attempted": 0, "total_successful_downloads": 0,
        "total_metadata_api_calls": 0, "total_metadata_errors": 0,
        "total_download_errors": 0, "keyword_performance": {},
        "last_parameter_update": ""
    }
    try:
        if os.path.exists(metrics_file_path):
            with open(metrics_file_path, "r", encoding="utf-8") as f: metrics = json.load(f)
            for key, value in default_metrics.items(): metrics.setdefault(key, value) # Ensure all keys exist
            return metrics
        else: return default_metrics
    except Exception as e: print_warning(f"Error loading performance metrics: {e}. Using default values."); return default_metrics

def save_performance_metrics(metrics):
    """Saves overall performance metrics to the JSON file."""
    metrics_file_path = constants.PERFORMANCE_METRICS_FILE
    try:
        with open(metrics_file_path, "w", encoding="utf-8") as f: json.dump(metrics, f, ensure_ascii=False, indent=4)
    except Exception as e: print_error(f"Error saving performance metrics: {e}", 1)

def generate_performance_summary(metrics):
    """Generates a text summary of overall performance metrics."""
    summary = ["=== Performance Summary ==="]
    runs_data = metrics.get('runs', [])
    summary.append(f"Total runs recorded: {len(runs_data)}")
    summary.append(f"Total shorts found (all runs): {metrics.get('total_shorts_found', 0)}")
    summary.append(f"Total suitable shorts (all runs): {metrics.get('total_suitable_shorts', 0)}")
    summary.append(f"Total downloads attempted (all runs): {metrics.get('total_downloads_attempted', 0)}")
    summary.append(f"Total successful downloads (all runs): {metrics.get('total_successful_downloads', 0)}")
    summary.append(f"Overall Download success rate: {metrics.get('total_successful_downloads', 0) / max(1, metrics.get('total_downloads_attempted', 1)):.1%}")
    summary.append(f"\n=== Metadata Performance (All Runs) ===")
    summary.append(f"Total metadata API calls: {metrics.get('total_metadata_api_calls', 0)}")
    summary.append(f"Total metadata errors: {metrics.get('total_metadata_errors', 0)}")
    summary.append(f"Overall Metadata error rate: {metrics.get('total_metadata_errors', 0) / max(1, metrics.get('total_metadata_api_calls', 1)):.1%}")

    # Add specific metadata errors/warnings
    metadata_metrics = load_metadata_metrics() # Load the detailed metrics
    summary.append(f"  Parse Failures: {metadata_metrics.get('parse_failures', 0)}")
    summary.append(f"  Timeouts: {metadata_metrics.get('timeouts', 0)}")
    summary.append(f"  Empty Descriptions: {metadata_metrics.get('empty_description_errors', 0)}")
    summary.append(f"  Empty Tags: {metadata_metrics.get('empty_tags_errors', 0)}")
    # Add specific validation warnings
    summary.append(f"  Validation - Title Mismatches: {metadata_metrics.get('validation_title_mismatches', 0)}")
    summary.append(f"  Validation - Tag List Errors: {metadata_metrics.get('validation_tag_list_errors', 0)}")
    summary.append(f"  Validation - Keyword Stuffing: {metadata_metrics.get('validation_keyword_stuffing', 0)}")

    if runs_data:
        summary.append(f"\n=== Recent Runs ({min(5, len(runs_data))}) ===")
        for run in runs_data[-5:]:
            summary.append(f" - Run {run.get('date', 'Unknown')[:10]}: Found={run.get('shorts_found', 0)}, Suitable={run.get('suitable_shorts',0)}, Attempted={run.get('downloads_attempted', 0)}, Succeeded={run.get('downloads_successful', 0)}, MetaErrors={run.get('metadata_errors', 0)}")
    keyword_performance = metrics.get('keyword_performance', {})
    if keyword_performance:
        top_keywords = get_top_performing_keywords(keyword_performance, 10) # Use helper function
        summary.append(f"\n=== Top 10 Performing Keywords (by Score) ===")
        for kw in top_keywords: summary.append(f" - {kw}: {keyword_performance.get(kw, 0):.2f}") # Format score
    return "\n".join(summary)


def generate_detailed_run_summary(run_metrics, performance_metrics, config):
    """Generates a detailed summary of the current run with insights and statistics."""
    if not run_metrics:
        print_warning("No run metrics available for summary generation.")
        return None

    try:
        # Format timestamp
        run_date = datetime.fromisoformat(run_metrics.get("date", datetime.now().isoformat()))
        formatted_date = run_date.strftime("%Y-%m-%d %H:%M:%S")

        # Calculate success rates
        downloads_attempted = run_metrics.get("downloads_attempted", 0)
        downloads_successful = run_metrics.get("downloads_successful", 0)
        download_success_rate = downloads_successful / max(1, downloads_attempted) * 100

        metadata_api_calls = run_metrics.get("metadata_api_calls", 0)
        metadata_errors = run_metrics.get("metadata_errors", 0)
        metadata_success_rate = (metadata_api_calls - metadata_errors) / max(1, metadata_api_calls) * 100

        shorts_found = run_metrics.get("shorts_found", 0)
        suitable_shorts = run_metrics.get("suitable_shorts", 0)
        shorts_suitability_rate = suitable_shorts / max(1, shorts_found) * 100

        # Get current config values
        current_config = {
            "YT_SEARCH_RESULTS_PER_KEYWORD": int(config.get("YT_SEARCH_RESULTS_PER_KEYWORD", YT_SEARCH_RESULTS_PER_KEYWORD)),
            "VIDEOS_TO_DOWNLOAD_PER_KEYWORD": int(config.get("VIDEOS_TO_DOWNLOAD_PER_KEYWORD", VIDEOS_TO_DOWNLOAD_PER_KEYWORD)),
            "KEYWORDS_TO_PROCESS_PER_RUN": int(config.get("KEYWORDS_TO_PROCESS_PER_RUN", KEYWORDS_TO_PROCESS_PER_RUN)),
            "MAX_DOWNLOADS": int(config.get("MAX_DOWNLOADS", 0)),
        }

        # Build summary
        summary = [
            f"=== DETAILED RUN SUMMARY: {formatted_date} ===",
            "",
            "--- Configuration ---",
            f"Search Results Per Keyword: {current_config['YT_SEARCH_RESULTS_PER_KEYWORD']}",
            f"Videos to Download Per Keyword: {current_config['VIDEOS_TO_DOWNLOAD_PER_KEYWORD']}",
            f"Keywords to Process Per Run: {current_config['KEYWORDS_TO_PROCESS_PER_RUN']}",
            f"Max Downloads: {current_config['MAX_DOWNLOADS']}",
            "",
            "--- Performance Metrics ---",
            f"Downloads Attempted: {downloads_attempted}",
            f"Downloads Successful: {downloads_successful}",
            f"Download Success Rate: {download_success_rate:.1f}%",
            f"Shorts Found: {shorts_found}",
            f"Suitable Shorts: {suitable_shorts}",
            f"Shorts Suitability Rate: {shorts_suitability_rate:.1f}%",
            f"Metadata API Calls: {metadata_api_calls}",
            f"Metadata Errors: {metadata_errors}",
            f"Metadata Success Rate: {metadata_success_rate:.1f}%",
            "",
            "--- Keywords Used ---"
        ]

        # Add keywords used
        keywords_used = run_metrics.get("keywords_used", [])
        if keywords_used:
            for keyword in keywords_used:
                summary.append(f"- {keyword}")
        else:
            summary.append("No keywords were used in this run.")

        # Add historical context
        summary.extend([
            "",
            "--- Historical Context ---"
        ])

        # Calculate averages from previous runs
        previous_runs = performance_metrics.get("runs", [])[:-1]  # Exclude current run
        if previous_runs:
            avg_downloads_successful = sum(run.get("downloads_successful", 0) for run in previous_runs) / len(previous_runs)
            avg_metadata_success_rate = sum((run.get("metadata_api_calls", 0) - run.get("metadata_errors", 0)) / max(1, run.get("metadata_api_calls", 0)) * 100 for run in previous_runs) / len(previous_runs)
            avg_shorts_suitability_rate = sum(run.get("suitable_shorts", 0) / max(1, run.get("shorts_found", 0)) * 100 for run in previous_runs) / len(previous_runs)

            summary.extend([
                f"Average Downloads Per Run: {avg_downloads_successful:.1f}",
                f"Average Metadata Success Rate: {avg_metadata_success_rate:.1f}%",
                f"Average Shorts Suitability Rate: {avg_shorts_suitability_rate:.1f}%",
                "",
                "--- Comparison to Average ---",
                f"Downloads: {'Above average' if downloads_successful > avg_downloads_successful else 'Below average'} ({downloads_successful} vs {avg_downloads_successful:.1f})",
                f"Metadata Success: {'Above average' if metadata_success_rate > avg_metadata_success_rate else 'Below average'} ({metadata_success_rate:.1f}% vs {avg_metadata_success_rate:.1f}%)",
                f"Shorts Suitability: {'Above average' if shorts_suitability_rate > avg_shorts_suitability_rate else 'Below average'} ({shorts_suitability_rate:.1f}% vs {avg_shorts_suitability_rate:.1f}%)"
            ])
        else:
            summary.append("No previous runs available for comparison.")

        # Add timestamp
        summary.extend([
            "",
            f"=== END OF SUMMARY: {formatted_date} ==="
        ])

        # Join and save summary
        summary_text = "\n".join(summary)
        summary_path = constants.RUN_SUMMARIES_LOG_FILE

        with open(summary_path, "a", encoding="utf-8") as f:
            f.write("\n\n" + summary_text)

        return summary_text

    except Exception as e:
        print_error(f"Error generating detailed run summary: {e}", include_traceback=True)
        return None

def generate_tuning_suggestions(metrics, config):
    """Generates parameter tuning suggestions using Gemini."""
    performance_summary = generate_performance_summary(metrics)
    config_text = "\n".join([f"{key}={value}" for key, value in config.items() if key != 'API_KEY' and key != 'GEMINI_API_KEY']) # Exclude API keys

    prompt = f"""
    Analyze the following performance report from a YouTube Shorts automation script (SEO strategy) that downloads videos and generates metadata.
    Suggest specific parameters in the config.txt file that might need adjustment, and explain *why* based ONLY on the provided data.
    Suggest potential new values or ranges for these parameters to potentially improve performance (e.g., download rate, metadata quality).

    Performance Report:
    {performance_summary}

    Current Config (excluding API keys):
    {config_text}

    Focus on suggesting adjustments for these parameters if the data indicates a potential issue:
    - MAX_DOWNLOADS: Number of videos to download per run (Consider overall success rate, time per run if available).
    - MAX_KEYWORDS: Max number of keywords to store (Consider keyword pool health, performance concentration).
    - METADATA_TIMEOUT_SECONDS: Timeout for metadata generation (Consider metadata error rate, specifically timeouts).
    - YT_SEARCH_RESULTS_PER_KEYWORD: Number of search results per keyword (Consider shorts found vs suitable shorts ratio).
    - VIDEOS_TO_DOWNLOAD_PER_KEYWORD: Max videos per keyword (Consider if limits are often hit per keyword vs overall download limit).
    - KEYWORDS_TO_PROCESS_PER_RUN: How many keywords are searched each time (Balance exploration vs exploitation).
    - MIN_KEYWORDS_THRESHOLD: When to generate new keywords (Is the pool stagnating?).
    - NEW_KEYWORDS_TO_GENERATE: How many new ones to add (Impact on pool diversity).

    Provide specific, data-driven recommendations ONLY IF THE DATA STRONGLY SUGGESTS a change is needed. If performance looks good, say so.
    Format suggestions clearly.
    """
    try:
        model = genai.GenerativeModel("gemini-2.0-flash") # Using 2.0 flash
        response = model.generate_content(prompt)
        suggestions = response.text.strip()
        suggestions_file_path = constants.TUNING_SUGGESTIONS_LOG_FILE
        with open(suggestions_file_path, "w", encoding="utf-8") as f:
            f.write(f"=== Tuning Suggestions ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ===\nBased on Config:\n{config_text}\n\nPerformance Summary:\n{performance_summary}\n\nSuggestions:\n{suggestions}\n")
        return suggestions
    except Exception as e: print_error(f"Error generating tuning suggestions: {e}", 1); return None

# --- Metadata Metrics Helpers (Kept from downloader - B.py) ---
def load_metadata_metrics():
    """Loads metadata metrics from the JSON file."""
    metrics_file_path = constants.METADATA_METRICS_FILE
    default_metrics = {
        "total_api_calls": 0, "parse_failures": 0, "timeouts": 0,
        "empty_title_errors": 0, "empty_description_errors": 0, "empty_tags_errors": 0,
        # Validation warning metrics
        "validation_title_mismatches": 0,
        "validation_tag_list_errors": 0,
        "validation_keyword_stuffing": 0,
        "last_run_date": datetime.now().isoformat(), "error_samples": [],
        "total_api_calls_previous": 0, "total_errors_previous": 0
    } # Added history tracking
    try:
        if os.path.exists(metrics_file_path):
            with open(metrics_file_path, "r", encoding="utf-8") as f: metrics = json.load(f)
            for key, value in default_metrics.items(): metrics.setdefault(key, value)
            return metrics
        else: return default_metrics
    except Exception as e: print_warning(f"Error loading metadata metrics: {e}. Using default values."); return default_metrics

def save_metadata_metrics(metrics):
    """Saves metadata metrics to the JSON file."""
    metrics_file_path = constants.METADATA_METRICS_FILE
    try:
        metrics["last_run_date"] = datetime.now().isoformat()
        with open(metrics_file_path, "w", encoding="utf-8") as f: json.dump(metrics, f, ensure_ascii=False, indent=4)
    except Exception as e: print_error(f"Error saving metadata metrics: {e}", 1)

def add_error_sample(metrics, error_type, error_details, video_title):
    """Adds an error sample to the metadata metrics."""
    if "error_samples" not in metrics: metrics["error_samples"] = []
    error_sample = { "type": error_type, "details": error_details, "video_title": video_title, "timestamp": datetime.now().isoformat() }
    metrics["error_samples"].append(error_sample)
    metrics["error_samples"] = metrics["error_samples"][-MAX_ERROR_SAMPLES:] # Keep only last N

# --- save_metadata (Adjusted for combined script) ---
def save_metadata(entry, video_index, seo_metadata, metadata_folder_path, current_keyword=None):
    """Saves detailed metadata for a video to a JSON file (SEO Strategy Marker)."""
    video_id = entry.get("id", f"unknown_id_{video_index}")
    original_title = entry.get("title", f"unknown_title_{video_index}")
    uploader = entry.get("uploader", "Unknown Uploader")
    view_count = entry.get("view_count", 0)
    optimized_title = seo_metadata.get("title", original_title)
    optimized_description = seo_metadata.get("description", f"Default description.\n\nCredit to: {uploader}\nOriginal Title: {original_title}")
    optimized_tags = seo_metadata.get("tags", ["error"])

    metadata = {
        "id": video_id, # Original YouTube ID at time of download
        "original_title": original_title,
        "uploader": uploader,
        "view_count": view_count, # Views at time of download
        "optimized_title": optimized_title, # Generated title for upload
        "optimized_description": optimized_description, # Generated description for upload
        "optimized_tags": optimized_tags, # Generated tags for upload
        "download_timestamp": datetime.now().isoformat(),
        "local_filename": f"video{video_index}.mp4",
        "metadata_strategy": "A: SEO/Discoverability", # Strategy marker
        "video_index": f"video{video_index}", # Internal index for correlation
        "discovery_keyword": current_keyword, # Keyword that found this video
        "original_yt_id_at_download": video_id, # Explicitly store original ID
        "suggested_category": seo_metadata.get("suggested_category"), # Will be None if not generated
        "target_playlist": seo_metadata.get("target_playlist") # Playlist ID or NEW suggestion
    }
    metadata_file_name = f"video{video_index}.json"
    metadata_file_path = os.path.join(metadata_folder_path, metadata_file_name)
    try:
        with open(metadata_file_path, "w", encoding="utf-8") as f: json.dump(metadata, f, ensure_ascii=False, indent=4)
        return metadata
    except Exception as e:
        print_error(f"Error saving metadata file {metadata_file_path}: {e}", 1)
        metadata["optimized_description"] = f"Error saving metadata file: {e}\n\nCredit to: {uploader}\nOriginal Title: {original_title}"
        metadata["optimized_tags"] = ["error"]
        return metadata

# --- get_last_video_index (Unchanged) ---
def get_last_video_index(excel_file_path, sheet_name):
    """Finds the highest video index (videoXXX) in the 'Video ID' column of a sheet."""
    max_index = 0
    try:
        wb = load_workbook(excel_file_path, read_only=True, data_only=True)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                 if row and row[0] and isinstance(row[0], str) and row[0].lower().startswith("video"): # Safer check
                    try: index = int(row[0][5:]); max_index = max(max_index, index)
                    except (ValueError, IndexError): continue
        wb.close()
        return max_index + 1
    except FileNotFoundError: print_warning(f"Excel file '{excel_file_path}' not found. Starting video index at 1."); return 1
    except KeyError: print_warning(f"Sheet '{sheet_name}' not found in '{excel_file_path}'. Starting video index at 1."); return 1
    except Exception as e:
        if 'wb' in locals() and wb:
             try: wb.close()
             except: pass
        print_error(f"Error reading last video index from '{excel_file_path}', sheet '{sheet_name}': {e}. Starting video index at 1.", 1); return 1

# --- save_cache (Updated to handle both list and dict with optional name) ---
def save_cache(cache_data, cache_file_path, cache_name="Cache"):
    """Saves cache data (list or dict) to a JSON file."""
    try:
        with open(cache_file_path, "w", encoding="utf-8") as f: json.dump(cache_data, f, ensure_ascii=False, indent=4)
        if isinstance(cache_data, dict) and "timestamp" in cache_data:
            print_info(f"Saved {cache_name} with {len(cache_data) - 1} entries.") # -1 for timestamp
        elif isinstance(cache_data, list):
            print_info(f"Saved {cache_name} with {len(cache_data)} entries.")
        else:
            print_info(f"Saved {cache_name}.")
    except Exception as e: print_error(f"Error saving {cache_name} to {cache_file_path}: {e}", 1)


# --- Main Execution Block ---
if __name__ == "__main__":
    colorama.init(autoreset=True)

    try:
        script_directory = os.path.dirname(os.path.abspath(__file__))

        # --- Path Definitions (now using constants) ---
        CONFIG_FILENAME = constants.CONFIG_FILE_PATH
        # Use getattr with fallback for NICHE_FILE_PATH
        NICHE_FILENAME_PATH = getattr(constants, 'NICHE_FILE_PATH', os.path.join(constants.CONFIG_DIR, "niche.txt"))
        DOWNLOADS_FOLDER = constants.SHORTS_DOWNLOADS_DIR
        METADATA_FOLDER = constants.SHORTS_METADATA_DIR
        FFMPEG_PATH = constants.FFMPEG_PATH # Default location
        EXCEL_FILE_PATH = constants.EXCEL_FILE_PATH
        PLAYLIST_CACHE_FILE_PATH = os.path.join(constants.DATA_DIR, "playlist_cache.json")
        KEYWORDS_CACHE_FILE_PATH = constants.GENERATED_KEYWORDS_CACHE_FILE
        METADATA_CACHE_FILE_PATH = os.path.join(constants.DATA_DIR, "metadata_cache.json") # Currently unused but path defined
        SEO_METADATA_PROMPT_CACHE_PATH = constants.SEO_METADATA_PROMPT_FILE # Path for cached SEO prompt
        PLAYLIST_DATA_CACHE_PATH = constants.PLAYLIST_DATA_CACHE_FILE # Path for playlist data cache

        print(f"{Fore.CYAN}--- Initializing Combined Downloader (SEO Focus + Self-Improvement) ---{Style.RESET_ALL}")

        # --- Create Folders ---
        os.makedirs(DOWNLOADS_FOLDER, exist_ok=True)
        os.makedirs(METADATA_FOLDER, exist_ok=True)

        # --- Load Configuration ---
        config = {}
        try:
            with open(CONFIG_FILENAME, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and "=" in line:
                        key, value = line.split("=", 1)
                        config[key.strip()] = value.strip()
            print_success("Configuration loaded.")
        except FileNotFoundError: print_fatal(f"Configuration file '{CONFIG_FILENAME}' not found. Please create it.")
        except Exception as e: print_fatal(f"Error reading configuration file '{CONFIG_FILENAME}': {e}")

        # Get API Key (Mandatory)
        API_KEY = config.get("API_KEY") or config.get("GEMINI_API_KEY") # Allow either key name
        if not API_KEY: print_fatal("API_KEY or GEMINI_API_KEY not found or empty in 'config.txt'.")
        try: genai.configure(api_key=API_KEY); print_success("Gemini API configured.")
        except Exception as e: print_fatal(f"Failed to configure Gemini API: {e}")

        # Get FFMPEG Path from config if specified
        ffmpeg_path_config = config.get("FFMPEG_PATH")
        if ffmpeg_path_config and os.path.exists(ffmpeg_path_config):
            FFMPEG_PATH = ffmpeg_path_config
            print_info(f"Using FFmpeg path from config: {FFMPEG_PATH}")
        elif not os.path.exists(FFMPEG_PATH):
             # If default path doesn't exist and config path wasn't valid/provided, warn user
             print_warning(f"FFmpeg not found at default path '{FFMPEG_PATH}' or specified config path. Downloads requiring merging might fail.")
             print_warning("Ensure ffmpeg.exe is in the script directory or set FFMPEG_PATH in config.txt")
             FFMPEG_PATH = "ffmpeg" # Fallback to hoping it's in system PATH

        # Get MAX_KEYWORDS (Optional with Default)
        _DEFAULT_MAX_KEYWORDS = 200
        try: max_keywords = int(config.get("MAX_KEYWORDS", _DEFAULT_MAX_KEYWORDS))
        except (ValueError, TypeError): max_keywords = _DEFAULT_MAX_KEYWORDS; print_warning(f"Invalid MAX_KEYWORDS. Using default: {_DEFAULT_MAX_KEYWORDS}")
        if max_keywords <= 0: max_keywords = _DEFAULT_MAX_KEYWORDS; print_warning(f"MAX_KEYWORDS must be positive. Using default: {_DEFAULT_MAX_KEYWORDS}")

        # Get MAX_DOWNLOADS (Mandatory)
        try: max_downloads = int(config["MAX_DOWNLOADS"])
        except KeyError: print_fatal("MAX_DOWNLOADS setting is missing from 'config.txt'.")
        except (ValueError, TypeError): print_fatal(f"Invalid MAX_DOWNLOADS value in 'config.txt'. Must be an integer.")
        if max_downloads <= 0: print_fatal("MAX_DOWNLOADS must be a positive integer in 'config.txt'.")

        # Load keyword-based downloader settings from config

        # YT_SEARCH_RESULTS_PER_KEYWORD (Optional with Default)
        try: YT_SEARCH_RESULTS_PER_KEYWORD = int(config.get("YT_SEARCH_RESULTS_PER_KEYWORD", YT_SEARCH_RESULTS_PER_KEYWORD))
        except (ValueError, TypeError): print_warning(f"Invalid YT_SEARCH_RESULTS_PER_KEYWORD. Using default: {YT_SEARCH_RESULTS_PER_KEYWORD}")
        if YT_SEARCH_RESULTS_PER_KEYWORD <= 0: YT_SEARCH_RESULTS_PER_KEYWORD = 50; print_warning(f"YT_SEARCH_RESULTS_PER_KEYWORD must be positive. Using default: 50")

        # VIDEOS_TO_DOWNLOAD_PER_KEYWORD (Optional with Default)
        try: VIDEOS_TO_DOWNLOAD_PER_KEYWORD = int(config.get("VIDEOS_TO_DOWNLOAD_PER_KEYWORD", VIDEOS_TO_DOWNLOAD_PER_KEYWORD))
        except (ValueError, TypeError): print_warning(f"Invalid VIDEOS_TO_DOWNLOAD_PER_KEYWORD. Using default: {VIDEOS_TO_DOWNLOAD_PER_KEYWORD}")
        if VIDEOS_TO_DOWNLOAD_PER_KEYWORD <= 0: VIDEOS_TO_DOWNLOAD_PER_KEYWORD = 5; print_warning(f"VIDEOS_TO_DOWNLOAD_PER_KEYWORD must be positive. Using default: 5")

        # KEYWORDS_TO_PROCESS_PER_RUN (Optional with Default)
        try: KEYWORDS_TO_PROCESS_PER_RUN = int(config.get("KEYWORDS_TO_PROCESS_PER_RUN", KEYWORDS_TO_PROCESS_PER_RUN))
        except (ValueError, TypeError): print_warning(f"Invalid KEYWORDS_TO_PROCESS_PER_RUN. Using default: {KEYWORDS_TO_PROCESS_PER_RUN}")
        if KEYWORDS_TO_PROCESS_PER_RUN <= 0: KEYWORDS_TO_PROCESS_PER_RUN = 5; print_warning(f"KEYWORDS_TO_PROCESS_PER_RUN must be positive. Using default: 5")

        # MIN_KEYWORDS_THRESHOLD (Optional with Default)
        try: MIN_KEYWORDS_THRESHOLD = int(config.get("MIN_KEYWORDS_THRESHOLD", MIN_KEYWORDS_THRESHOLD))
        except (ValueError, TypeError): print_warning(f"Invalid MIN_KEYWORDS_THRESHOLD. Using default: {MIN_KEYWORDS_THRESHOLD}")
        if MIN_KEYWORDS_THRESHOLD <= 0: MIN_KEYWORDS_THRESHOLD = 20; print_warning(f"MIN_KEYWORDS_THRESHOLD must be positive. Using default: 20")

        # NEW_KEYWORDS_TO_GENERATE (Optional with Default)
        try: NEW_KEYWORDS_TO_GENERATE = int(config.get("NEW_KEYWORDS_TO_GENERATE", NEW_KEYWORDS_TO_GENERATE))
        except (ValueError, TypeError): print_warning(f"Invalid NEW_KEYWORDS_TO_GENERATE. Using default: {NEW_KEYWORDS_TO_GENERATE}")
        if NEW_KEYWORDS_TO_GENERATE <= 0: NEW_KEYWORDS_TO_GENERATE = 10; print_warning(f"NEW_KEYWORDS_TO_GENERATE must be positive. Using default: 10")

        print(f"{Fore.BLUE}{Style.BRIGHT}Settings: Max Downloads={max_downloads}, Max Keywords={max_keywords}{Style.RESET_ALL}")
        print(f"{Fore.BLUE}Keyword Settings: Search Results={YT_SEARCH_RESULTS_PER_KEYWORD}, Keywords Per Run={KEYWORDS_TO_PROCESS_PER_RUN}, Videos Per Keyword={VIDEOS_TO_DOWNLOAD_PER_KEYWORD}{Style.RESET_ALL}")
        print(f"{Fore.BLUE}Keyword Generation: Min Threshold={MIN_KEYWORDS_THRESHOLD}, New Keywords={NEW_KEYWORDS_TO_GENERATE}{Style.RESET_ALL}")

        # --- Load or Create Excel Workbook ---
        excel_loaded_ok = False

        # Try to import excel_utils module
        try:
            import excel_utils
            excel_utils_available = True
            print_info("Using excel_utils module for robust Excel handling")
        except ImportError:
            excel_utils_available = False
            print_warning("excel_utils module not available. Using fallback Excel handling.")

        # Define sheet configuration
        sheets_config = {
            constants.DOWNLOADED_SHEET_NAME: ["Video Index", "Optimized Title", "Downloaded Date", "Views", "Uploader", "Original Title"],
            constants.UPLOADED_SHEET_NAME: ["Video Index", "Optimized Title", "YouTube Video ID", "Upload Timestamp", "Scheduled Time", "Publish Status"]
        }

        if excel_utils_available:
            # Use the excel_utils module for robust Excel handling
            try:
                wb, sheets, save_needed = excel_utils.load_or_create_excel(EXCEL_FILE_PATH, sheets_config)

                if not wb:
                    print_fatal(f"Failed to load or create Excel file: {EXCEL_FILE_PATH}")
                    exit(1)

                downloaded_sheet = sheets.get(constants.DOWNLOADED_SHEET_NAME)
                uploaded_sheet = sheets.get(constants.UPLOADED_SHEET_NAME)

                # Save if needed using robust save mechanism
                if save_needed:
                    if not excel_utils.safe_save_workbook(wb, EXCEL_FILE_PATH, close_excel=True, create_backup=True):
                        print_warning(f"Could not save structural changes to Excel. Will try again later.")

                print_success(f"Excel loaded successfully using excel_utils.")
                excel_loaded_ok = True
            except Exception as e:
                print_error(f"Error using excel_utils: {e}", include_traceback=True)
                print_warning("Falling back to standard Excel handling.")
                excel_utils_available = False  # Force fallback

        # Fallback to original implementation if excel_utils is not available or failed
        if not excel_utils_available:
            try:
                if not os.path.exists(EXCEL_FILE_PATH):
                    wb = Workbook()
                    downloaded_sheet = wb.active; downloaded_sheet.title = constants.DOWNLOADED_SHEET_NAME
                    downloaded_sheet.append(["Video Index", "Optimized Title", "Downloaded Date", "Views", "Uploader", "Original Title"]) # Corrected header
                    uploaded_sheet = wb.create_sheet(title=constants.UPLOADED_SHEET_NAME)
                    uploaded_sheet.append(["Video Index", "Optimized Title", "YouTube Video ID", "Upload Timestamp", "Scheduled Time", "Publish Status"]) # Corrected header

                    # Try to save with a simple retry mechanism
                    max_retries = 3
                    for attempt in range(max_retries):
                        try:
                            wb.save(EXCEL_FILE_PATH)
                            print_success(f"Created new Excel file: {os.path.basename(EXCEL_FILE_PATH)}")
                            break
                        except PermissionError as pe:
                            if attempt < max_retries - 1:
                                print_warning(f"PermissionError saving Excel (attempt {attempt+1}/{max_retries}): {pe}")
                                print_info(f"Retrying in 2 seconds...")
                                time.sleep(2)
                            else:
                                raise
                    excel_loaded_ok = True
                else:
                    wb = load_workbook(EXCEL_FILE_PATH)
                    # Ensure sheets exist and headers are correct
                    if constants.DOWNLOADED_SHEET_NAME not in wb.sheetnames:
                        downloaded_sheet = wb.create_sheet(title=constants.DOWNLOADED_SHEET_NAME)
                        downloaded_sheet.append(["Video Index", "Optimized Title", "Downloaded Date", "Views", "Uploader", "Original Title"])
                        print_warning("Created missing 'Downloaded' sheet.")
                    else:
                        downloaded_sheet = wb[constants.DOWNLOADED_SHEET_NAME] # Check/fix header if needed

                    if constants.UPLOADED_SHEET_NAME not in wb.sheetnames:
                        uploaded_sheet = wb.create_sheet(title=constants.UPLOADED_SHEET_NAME)
                        uploaded_sheet.append(["Video Index", "Optimized Title", "YouTube Video ID", "Upload Timestamp", "Scheduled Time", "Publish Status"])
                        print_warning("Created missing 'Uploaded' sheet.")
                    else:
                        uploaded_sheet = wb[constants.UPLOADED_SHEET_NAME] # Check/fix header if needed

                    print_success(f"Loaded Excel file: {os.path.basename(EXCEL_FILE_PATH)}")
                    excel_loaded_ok = True
            except Exception as e:
                print_fatal(f"Error handling Excel file {EXCEL_FILE_PATH}: {e}")

        if not excel_loaded_ok:
            exit(1) # Should be caught by fatal, but safety check

        # --- Load Caches ---
        print(f"{Fore.BLUE}--- Loading Caches ---{Style.RESET_ALL}")
        # Playlist Cache (Downloaded Video IDs)
        playlist_cache = set()
        if os.path.exists(PLAYLIST_CACHE_FILE_PATH):
            try:
                with open(PLAYLIST_CACHE_FILE_PATH, "r", encoding="utf-8") as f: content = f.read()
                if content:
                    loaded_data = json.loads(content)
                    if isinstance(loaded_data, list): playlist_cache = set(loaded_data); print_success(f"Loaded {len(playlist_cache)} IDs from playlist cache.")
                    else: print_warning(f"Playlist cache '{os.path.basename(PLAYLIST_CACHE_FILE_PATH)}' has invalid format. Initializing empty.")
                else: print_warning(f"Playlist cache '{os.path.basename(PLAYLIST_CACHE_FILE_PATH)}' is empty. Initializing empty.")
            except json.JSONDecodeError: print_warning(f"Playlist cache '{os.path.basename(PLAYLIST_CACHE_FILE_PATH)}' contains invalid JSON. Initializing empty.")
            except Exception as e: print_warning(f"Error loading playlist cache '{os.path.basename(PLAYLIST_CACHE_FILE_PATH)}': {e}. Initializing empty.")
        else: print_info(f"Playlist cache not found. Will create if needed.")
        downloaded_video_ids = playlist_cache

        # Generated Keywords Cache (with Frequency/Score)
        keyword_frequency = {}
        if os.path.exists(KEYWORDS_CACHE_FILE_PATH):
            try:
                with open(KEYWORDS_CACHE_FILE_PATH, "r", encoding="utf-8") as f: content = f.read()
                if content:
                    cached_data = json.loads(content)
                    if isinstance(cached_data, dict): keyword_frequency = cached_data; print_success(f"Loaded {len(keyword_frequency)} keywords from cache (Score format).")
                    # Remove handling for old list format, assume dict format now
                    else: print_warning(f"Keyword cache '{os.path.basename(KEYWORDS_CACHE_FILE_PATH)}' has invalid format (expected dict). Initializing empty.")
                else: print_warning(f"Keyword cache '{os.path.basename(KEYWORDS_CACHE_FILE_PATH)}' is empty. Initializing empty.")
            except json.JSONDecodeError: print_warning(f"Keyword cache '{os.path.basename(KEYWORDS_CACHE_FILE_PATH)}' contains invalid JSON. Initializing empty.")
            except Exception as e: print_warning(f"Error loading keyword cache '{os.path.basename(KEYWORDS_CACHE_FILE_PATH)}': {e}. Initializing empty.")
        else: print_info(f"Keyword cache not found. Will generate keywords if needed.")
        search_keywords = list(keyword_frequency.keys()) # Get keywords from the loaded dict

        # --- Load Playlist Data Cache ---
        print_info("Loading Playlist Data Cache...")
        playlist_data_cache = load_cache(PLAYLIST_DATA_CACHE_PATH, "Playlist Data")

        # --- Keyword Filtering (Kept from B) ---
        print_info("Applying keyword filters...")
        if search_keywords:
            initial_count = len(search_keywords)
            required_substrings = ["GTA", "Grand Theft Auto"]; required_substrings_lower = [sub.lower() for sub in required_substrings]
            social_media_lower = {"twitch", "youtube", "facebook", "instagram", "tiktok", "x", "reddit", "discord", "kick"}
            generic_terms_lower = {"gaming", "video game", "gameplay", "shorts", "short", "viral", "trending", "funny", "meme", "edit", "clip", "clips"}
            filtered_keywords_temp = []; removed_keywords = []; temp_freq = {}
            for keyword in search_keywords:
                kw_lower = keyword.lower()
                contains_required = any(sub in kw_lower for sub in required_substrings_lower)
                is_not_social = kw_lower not in social_media_lower
                is_not_generic = kw_lower not in generic_terms_lower
                is_long_enough = len(keyword.replace(" ", "")) > 3
                if contains_required and is_not_social and is_not_generic and is_long_enough:
                    filtered_keywords_temp.append(keyword)
                    temp_freq[keyword] = keyword_frequency.get(keyword, 0) # Keep score
                else: removed_keywords.append(keyword)
            search_keywords = filtered_keywords_temp; keyword_frequency = temp_freq
            if removed_keywords: print_warning(f"Filtered keywords: Removed {len(removed_keywords)}. Kept {len(search_keywords)}.")
            else: print_success("Keyword filtering completed. No keywords removed.")
        else: print_info("No keywords loaded from cache to filter.")

        # --- Get Seed Niche ---
        try:
            with open(NICHE_FILENAME_PATH, "r", encoding="utf-8") as f: seed_niche = f.readline().strip()
            if not seed_niche: print_fatal(f"The niche file '{NICHE_FILENAME_PATH}' is empty.")
            print_success(f"Main niche: '{seed_niche}'")
        except FileNotFoundError: print_fatal(f"Niche file '{NICHE_FILENAME_PATH}' not found.")
        except Exception as e: print_fatal(f"Error reading niche from '{NICHE_FILENAME_PATH}': {e}")

        # --- Generate Initial Keywords if Cache is Empty ---
        if not search_keywords:
            print_warning("No cached keywords found/loaded. Generating initial keywords...")
            initial_keywords = generate_keywords_from_niche(seed_niche, num_keywords=INITIAL_KEYWORDS_COUNT)
            if initial_keywords:
                new_unique = [kw for kw in initial_keywords if kw not in keyword_frequency]
                search_keywords.extend(new_unique)
                for kw in new_unique: keyword_frequency[kw] = 0 # Start score at 0
                print_success(f"Generated and added {len(new_unique)} initial keywords.")
                save_cache(keyword_frequency, KEYWORDS_CACHE_FILE_PATH); print_success("Saved initial keywords to cache.")
            else: print_fatal("Failed to generate initial keywords. Check Gemini API/prompt. Cannot proceed without keywords.")

        # --- Initialize Counters and State ---
        video_counter = get_last_video_index(EXCEL_FILE_PATH, constants.DOWNLOADED_SHEET_NAME) # Use Downloaded sheet index
        total_downloaded = 0
        processed_keywords_this_run = set()

        print_info(f"Starting video counter at: video{video_counter}")
        print_info(f"Total keywords available for searching: {len(search_keywords)}")

        # --- Load Uploaded Video Performance Data from Excel ---
        print(f"{Fore.BLUE}--- Loading Uploaded Video Performance Data (Excel) ---{Style.RESET_ALL}")
        uploaded_performance_data = {} # Key: YouTube Video ID, Value: Dict of performance data
        try:
            if os.path.exists(EXCEL_FILE_PATH):
                wb_perf = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
                if constants.UPLOADED_SHEET_NAME in wb_perf.sheetnames:
                    sheet_perf = wb_perf[constants.UPLOADED_SHEET_NAME]
                    header_perf = [cell.value for cell in sheet_perf[1]]
                    try:
                        col_indices = {str(name).lower().strip(): i for i, name in enumerate(header_perf, 1) if name}
                        # Need YT ID to link with cache, other stats are optional
                        uploaded_yt_id_col = col_indices.get('youtube video id')
                        views_col = col_indices.get('views (yt)')
                        likes_col = col_indices.get('likes (yt)')
                        comments_col = col_indices.get('comments (yt)')

                        if not uploaded_yt_id_col:
                            print_warning(f"Required column 'YouTube Video ID' not found in '{constants.UPLOADED_SHEET_NAME}' sheet. Performance data cannot be loaded.", 1)
                        else:
                            for row_idx in range(2, sheet_perf.max_row + 1):
                                uploaded_yt_id = str(sheet_perf.cell(row=row_idx, column=uploaded_yt_id_col).value or "").strip()
                                if uploaded_yt_id and uploaded_yt_id != "N/A":
                                    views = 0; likes = 0; comments = 0
                                    try: # Safely get optional stats
                                        if views_col: views = int(sheet_perf.cell(row=row_idx, column=views_col).value or 0)
                                        if likes_col: likes = int(sheet_perf.cell(row=row_idx, column=likes_col).value or 0)
                                        if comments_col: comments = int(sheet_perf.cell(row=row_idx, column=comments_col).value or 0)
                                    except (ValueError, TypeError): pass # Ignore conversion errors, keep 0
                                    # Use YT ID as the key now
                                    uploaded_performance_data[uploaded_yt_id] = {"views": views, "likes": likes, "comments": comments}
                            print_success(f"Loaded performance data for {len(uploaded_performance_data)} uploaded videos from Excel.")
                    except Exception as e: print_error(f"Error parsing '{constants.UPLOADED_SHEET_NAME}' sheet: {e}", 1, include_traceback=True)
                    wb_perf.close()
                else: print_info(f"'{constants.UPLOADED_SHEET_NAME}' sheet not found. No upload performance data to load.")
        except FileNotFoundError: print_info(f"Excel file '{EXCEL_FILE_PATH}' not found. No upload performance data to load.")
        except Exception as e: print_error(f"Error loading Excel for performance data: {e}", 1, include_traceback=True)


        # --- Correlate Performance Data with Keywords using Correlation Cache ---
        print(f"{Fore.BLUE}--- Correlating Performance Data with Keywords (Cache) ---{Style.RESET_ALL}")
        keyword_performance_feedback = {} # Key: keyword, Value: List of perf dicts
        correlation_cache = load_correlation_cache() # Load the new cache

        if uploaded_performance_data and correlation_cache:
            print_info(f"Processing {len(correlation_cache)} entries from correlation cache.", 1)
            correlated_count = 0
            # Iterate through the cache entries
            for cache_entry in correlation_cache:
                youtube_video_id = cache_entry.get("youtube_video_id")
                discovery_keyword = cache_entry.get("discovery_keyword")
                video_index = cache_entry.get("video_index") # For logging clarity

                # Check if we have performance data for this YT ID
                if youtube_video_id and discovery_keyword and youtube_video_id in uploaded_performance_data:
                    perf_data = uploaded_performance_data[youtube_video_id]

                    # Add to keyword performance feedback
                    if discovery_keyword not in keyword_performance_feedback:
                        keyword_performance_feedback[discovery_keyword] = []
                    keyword_performance_feedback[discovery_keyword].append(perf_data)
                    correlated_count += 1
                    # print_info(f"Correlated {video_index} performance (YT ID: {youtube_video_id}) with keyword '{discovery_keyword}'.", 2) # Verbose log

            print_success(f"Correlated performance data for {len(keyword_performance_feedback)} keywords using cache ({correlated_count} video links).")
            # --- Cleanup old entries from correlation cache ---
            cleanup_correlation_cache(days_to_keep=7) # Adjust days_to_keep if needed

        elif not correlation_cache:
            print_info("Correlation cache is empty. Cannot correlate performance.")
        elif not uploaded_performance_data:
             print_info("No uploaded performance data loaded from Excel. Cannot correlate.")

        # --- Update Keyword Scores ---
        print(f"{Fore.BLUE}--- Updating Keyword Scores with Performance Feedback ---{Style.RESET_ALL}")
        def calculate_performance_score(views, likes, comments=0):
            view_score = math.log10(views + 1) * 1.5; like_score = math.log10(likes + 1) * 2.0; comment_score = math.log10(comments + 1) * 1.0
            return view_score + like_score + comment_score

        total_keywords_updated = 0
        if keyword_performance_feedback:
            for keyword, performance_list in keyword_performance_feedback.items():
                # --- Skip channel URLs (from channel-based downloader) ---
                if "youtube.com" in keyword.lower() or "youtu.be" in keyword.lower():
                    print_info(f"Skipping score update for channel URL source: {keyword}", 1)
                    continue
                # --- End channel URL check ---

                cumulative_perf_score = 0; num_videos = 0
                for perf_data in performance_list:
                    video_score = calculate_performance_score(perf_data.get("views", 0), perf_data.get("likes", 0), perf_data.get("comments", 0))
                    cumulative_perf_score += video_score; num_videos += 1
                    # print_info(f"Video for '{keyword}' score: {video_score:.2f}", 2) # Verbose
                if num_videos > 0:
                    average_perf_score = cumulative_perf_score / num_videos
                    # Update existing score or initialize if keyword somehow missed loading
                    current_discovery_score = keyword_frequency.get(keyword, 0) # Score might already include past performance
                    # Blend performance with discovery count - Adjust PERF_SCORE_WEIGHT as needed
                    PERF_SCORE_WEIGHT = 5.0 # How much impact avg performance has
                    # Simple additive model for now: new_score = discovery_count + avg_perf * weight
                    # A more complex model could average scores over time, etc.
                    new_score = current_discovery_score + (average_perf_score * PERF_SCORE_WEIGHT)
                    keyword_frequency[keyword] = max(0, new_score) # Ensure non-negative
                    total_keywords_updated += 1
                    print_info(f"Updated score for '{keyword}': AvgPerf={average_perf_score:.2f}, NewScore={new_score:.2f}", 1)

            if total_keywords_updated > 0:
                print_success(f"Updated scores for {total_keywords_updated} keywords based on upload performance.")
                save_cache(keyword_frequency, KEYWORDS_CACHE_FILE_PATH); print_success("Saved updated keyword performance cache.")
            else: print_info("No keyword scores were updated based on available performance data.")
        else: print_info("No performance feedback available to update keyword scores.")

        # --- Check if Metadata Prompt Needs Improvement ---
        print(f"\n{Fore.CYAN}{Style.BRIGHT}--- Checking Metadata Prompt Quality ---{Style.RESET_ALL}")
        metadata_metrics = load_metadata_metrics()
        total_api_calls = metadata_metrics.get("total_api_calls", 0)
        if total_api_calls > 0:
            # Calculate rates for all relevant error types
            parse_fail_rate = metadata_metrics.get("parse_failures", 0) / total_api_calls
            timeout_rate = metadata_metrics.get("timeouts", 0) / total_api_calls
            empty_desc_rate = metadata_metrics.get("empty_description_errors", 0) / total_api_calls
            empty_tags_rate = metadata_metrics.get("empty_tags_errors", 0) / total_api_calls

            # Calculate validation warning rates
            title_mismatch_rate = metadata_metrics.get("validation_title_mismatches", 0) / total_api_calls
            tag_list_error_rate = metadata_metrics.get("validation_tag_list_errors", 0) / total_api_calls
            stuffing_rate = metadata_metrics.get("validation_keyword_stuffing", 0) / total_api_calls

            # Calculate total content errors (for backward compatibility)
            total_errors = sum(metadata_metrics.get(err_type, 0) for err_type in ["parse_failures", "empty_description_errors", "empty_tags_errors"])
            error_rate = total_errors / total_api_calls

            print_info(f"Metadata API calls: {total_api_calls}, Content Errors: {total_errors} ({error_rate:.1%}), Timeouts: {timeout_rate:.1%}")
            print_info(f"Validation Warnings - Title Mismatches: {title_mismatch_rate:.1%}, Tag List Errors: {tag_list_error_rate:.1%}, Keyword Stuffing: {stuffing_rate:.1%}")

            # Determine if we need to improve the prompt
            trigger_improvement = False
            reason = ""

            if parse_fail_rate >= METADATA_ERROR_THRESHOLD:
                trigger_improvement = True
                reason = "Parse Failures"
            elif timeout_rate >= METADATA_TIMEOUT_THRESHOLD:
                trigger_improvement = True
                reason = "Timeouts"
            elif empty_desc_rate >= METADATA_ERROR_THRESHOLD:
                trigger_improvement = True
                reason = "Empty Descriptions"
            elif empty_tags_rate >= METADATA_ERROR_THRESHOLD:
                trigger_improvement = True
                reason = "Empty Tags"
            # Check validation warning rates
            elif title_mismatch_rate >= VALIDATION_WARNING_THRESHOLD:
                trigger_improvement = True
                reason = "Title Mismatches"
            elif tag_list_error_rate >= VALIDATION_WARNING_THRESHOLD:
                trigger_improvement = True
                reason = "Tag List Errors"
            elif stuffing_rate >= VALIDATION_WARNING_THRESHOLD:
                trigger_improvement = True
                reason = "Keyword Stuffing"

            if trigger_improvement:
                print_warning(f"Metadata quality issue detected (Reason: {reason} rate above threshold). Attempting to improve prompt...", 1)
                # Backup current prompt (cache or default) using excel_utils
                current_prompt_text = load_or_get_seo_prompt_template()
                try:
                    # Import excel_utils for the backup function
                    import excel_utils
                    # Create a backup in the excel_backups folder
                    backup_folder = excel_utils.create_backup_folder()
                    backup_file_path = excel_utils.create_file_backup(SEO_METADATA_PROMPT_CACHE_PATH, backup_folder, suffix="_prompt")
                    if backup_file_path:
                        print_success(f"Backed up current prompt to {backup_file_path}", 2)
                    else:
                        raise Exception("Backup function returned None")
                except Exception as e:
                    # Fallback to original backup method if excel_utils fails
                    backup_file_path = SEO_METADATA_PROMPT_CACHE_PATH + ".backup"
                    try:
                        with open(backup_file_path, "w", encoding="utf-8") as bf: bf.write(current_prompt_text)
                        print_success(f"Backed up current prompt to {backup_file_path} (fallback method)", 2)
                    except Exception as e2: print_error(f"Error backing up prompt file: {e2}", 2)
                # Generate improved prompt
                improved_prompt = improve_metadata_prompt(metadata_metrics)
                if improved_prompt: save_seo_prompt_template(improved_prompt) # Save to cache file
                else: print_warning("Could not generate an improved prompt. Keeping current version.", 2)
            else: print_success("Metadata prompt is performing well based on error rates and validation warnings.")
        else: print_info("No metadata API calls recorded yet. Skipping prompt quality check.")

        # --- Main Download Loop ---
        print(f"\n{Fore.CYAN}{Style.BRIGHT}--- Starting Main Download Loop ---{Style.RESET_ALL}")
        if not search_keywords: print_warning("No search keywords available. Exiting loop.")
        else:
            performance_metrics = load_performance_metrics() # Load overall performance metrics
            run_metrics = { "date": datetime.now().isoformat(), "downloads_attempted": 0, "downloads_successful": 0, "shorts_found": 0, "suitable_shorts": 0, "metadata_api_calls": 0, "metadata_errors": 0, "download_errors": 0, "keywords_used": [] }
            metadata_api_calls_start = metadata_metrics.get("total_api_calls", 0) # Track calls for this run
            metadata_errors_start = metadata_metrics.get("total_metadata_errors", 0) # Track errors for this run

            # Check keyword threshold
            if len(search_keywords) < MIN_KEYWORDS_THRESHOLD:
                print_warning(f"Keyword count ({len(search_keywords)}) below threshold ({MIN_KEYWORDS_THRESHOLD}). Generating new keywords...", 1)
                top_keywords = get_top_performing_keywords(keyword_frequency, TOP_KEYWORDS_TO_USE)
                if top_keywords: print_info(f"Using top {len(top_keywords)} keywords as inspiration: {', '.join(top_keywords)}", 2)
                new_keywords = generate_keywords_from_niche(seed_niche, num_keywords=NEW_KEYWORDS_TO_GENERATE, top_performing_keywords=top_keywords)
                if new_keywords:
                    new_unique_keywords = [kw for kw in new_keywords if kw not in keyword_frequency]
                    if new_unique_keywords:
                        print_success(f"Generated {len(new_unique_keywords)} new unique keywords.", 2)
                        search_keywords.extend(new_unique_keywords) # Add to list used for selection
                        for kw in new_unique_keywords: keyword_frequency[kw] = 0 # Add to score dict
                        save_cache(keyword_frequency, KEYWORDS_CACHE_FILE_PATH) # Save updated pool
                    else: print_warning("No new unique keywords generated.", 2)
                else: print_error("Failed to generate new keywords.", 2)

            # Determine keywords to process
            keywords_to_process_count = 0
            if max_downloads > 0: keywords_to_process_count = min(KEYWORDS_TO_PROCESS_PER_RUN, min(math.ceil(max_downloads / VIDEOS_TO_DOWNLOAD_PER_KEYWORD), len(search_keywords)))
            else: print_warning("MAX_DOWNLOADS is <= 0. No videos will be downloaded.")

            if keywords_to_process_count > 0:
                # Weighted random selection
                print_info(f"Selecting {keywords_to_process_count} keywords using weighted random sampling...")
                keywords_list = list(keyword_frequency.keys()); weights = [float(keyword_frequency.get(kw, 0) + BASE_WEIGHT_FOR_KEYWORDS) for kw in keywords_list] # Ensure float weights, add base
                # Handle case where all weights might be zero or negative after base adjustment
                if all(w <= 0 for w in weights): weights = [1.0] * len(weights) # Equal weight if all are non-positive
                elif any(w < 0 for w in weights): weights = [max(0.01, w) for w in weights] # Ensure non-negative, small positive floor

                try:
                    # Ensure k is not larger than population if weights are involved and list isn't empty
                    k_val = min(len(keywords_list), keywords_to_process_count) if keywords_list else 0
                    if k_val > 0:
                         keywords_to_search = random.choices(keywords_list, weights=weights, k=k_val)
                    else:
                         keywords_to_search = [] # No keywords to select
                except ValueError as e:
                    print_error(f"Error during weighted keyword selection: {e}. Falling back to simple random sampling.", 1)
                    k_val = min(len(keywords_list), keywords_to_process_count) if keywords_list else 0
                    keywords_to_search = random.sample(keywords_list, k=k_val) if k_val > 0 else []

                print_info(f"Will search using {len(keywords_to_search)} selected keywords:", 1)
                for i, kw in enumerate(keywords_to_search): print_info(f"  {i+1}. {kw} (Score: {keyword_frequency.get(kw, 0):.2f})", 2) # Format score

                # --- The actual downloading loop ---
                for keyword in keywords_to_search:
                    if total_downloaded >= max_downloads: print_warning(f"Reached total download limit ({max_downloads}). Stopping."); break
                    if keyword in processed_keywords_this_run: print_warning(f"Skipping already processed keyword in this run: {keyword}"); continue
                    processed_keywords_this_run.add(keyword)

                    search_terms = f"{keyword} shorts"; ydl_search_query = f"ytsearch{YT_SEARCH_RESULTS_PER_KEYWORD}:{search_terms}"
                    print(f"\n{Style.BRIGHT}{Fore.CYAN}--- Processing Keyword: '{keyword}' (Score: {keyword_frequency.get(keyword, 0):.2f}) ---{Style.RESET_ALL}")
                    print_info(f"Searching YouTube with yt-dlp query: {ydl_search_query}", 1)

                    videos_downloaded_for_keyword = 0
                    ydl_opts_search = { 'extract_flat': True, 'playlist_items': f'1-{YT_SEARCH_RESULTS_PER_KEYWORD}', 'quiet': True, 'no_warnings': True, 'nocheckcertificate': True, 'source_address': '0.0.0.0', 'default_search': 'ytsearch', }

                    try:
                        with yt_dlp.YoutubeDL(ydl_opts_search) as ydl: result = ydl.extract_info(ydl_search_query, download=False)
                        if not result or 'entries' not in result or not result.get('entries'): print_warning(f"No videos found or error in search results for keyword: {keyword}"); continue

                        found_count = len(result['entries']); run_metrics["shorts_found"] += found_count; performance_metrics["total_shorts_found"] += found_count
                        print_info(f"Found {found_count} potential videos for '{keyword}'. Filtering...", 1)

                        for entry in result['entries']:
                            if not entry or not isinstance(entry, dict): print_warning("Skipping invalid entry.", 2); continue
                            video_id = entry.get("id"); original_title = entry.get("title", "unknown_title").strip(); uploader = entry.get("uploader", "Unknown Uploader").strip(); duration = entry.get('duration')

                            # Filtering
                            if not video_id: print_warning("Skipping entry with missing ID.", 2); continue
                            if video_id in downloaded_video_ids: print_info(f"Skipping already downloaded ID: {video_id}", 2); continue # Use info level for already downloaded
                            if duration is None: print_warning(f"Skipping {video_id} - Duration unknown.", 2); continue
                            try:
                                duration_seconds = float(duration)
                                if duration_seconds > constants.MAX_SHORT_DURATION: print_info(f"Skipping {video_id} - Too long ({duration_seconds:.1f}s).", 2); continue
                                if duration_seconds <= 0: print_warning(f"Skipping {video_id} - Invalid duration ({duration_seconds:.1f}s).", 2); continue
                            except (ValueError, TypeError): print_warning(f"Skipping {video_id} - Invalid duration value: {duration}", 2); continue

                            # Process Valid Short
                            run_metrics["suitable_shorts"] += 1; performance_metrics["total_suitable_shorts"] += 1
                            if total_downloaded >= max_downloads: print_warning("Reached total download limit.", 1); break
                            if videos_downloaded_for_keyword >= VIDEOS_TO_DOWNLOAD_PER_KEYWORD: print_info(f"Reached download limit ({VIDEOS_TO_DOWNLOAD_PER_KEYWORD}) for keyword '{keyword}'.", 1); break

                            print_info(f"Processing valid Short (ID: {video_id}, {duration_seconds:.1f}s): {Style.BRIGHT}{original_title[:70]}...{Style.RESET_ALL}", 1)

                            # Generate Metadata
                            print_info("Generating SEO metadata...", 2)
                            seo_metadata = generate_metadata_with_timeout(original_title, uploader, original_title)
                            if not seo_metadata or "error" in seo_metadata.get("tags",[]):
                                print_warning("Metadata generation failed or resulted in error fallback.", 3)
                                run_metrics["metadata_errors"] += 1
                                performance_metrics["total_metadata_errors"] += 1
                            else:
                                print_success("Metadata generated.", 3)

                                # Get category suggestion if not already present
                                if "suggested_category" not in seo_metadata:
                                    print_info("Getting category suggestion...", 3)
                                    suggested_category = get_suggested_category(seo_metadata.get("title", original_title), seo_metadata.get("description", ""))
                                    if suggested_category:
                                        seo_metadata["suggested_category"] = suggested_category
                                        print_success(f"Category suggestion: {suggested_category}", 3)
                                    else:
                                        print_warning("No category suggestion available.", 3)

                                # Optimize metadata using utils.metadata_generator if available
                                if UTILS_AVAILABLE:
                                    print_info("Optimizing generated metadata using utils.metadata_generator...", 3)
                                    # Prepare metadata dict for improvement
                                    metadata_to_improve = {
                                        "title": seo_metadata.get("title", original_title),
                                        "description": seo_metadata.get("description", ""),
                                        "tags": seo_metadata.get("tags", [])
                                    }

                                    # Get niche/keyword for optimization
                                    optimization_keyword = keyword if keyword else seed_niche

                                    # Improve metadata
                                    print_info(f"Improving metadata with keyword: {optimization_keyword}", 3)
                                    optimized_meta_content = metadata_generator.improve_metadata(
                                        metadata_to_improve,
                                        keyword=optimization_keyword,
                                        metrics=metadata_metrics
                                    )

                                    # Validate the newly optimized metadata
                                    final_validated_meta = metadata_generator.validate_metadata(
                                        optimized_meta_content,
                                        original_title=metadata_to_improve.get("title"),
                                        keyword=optimization_keyword,
                                        metrics=metadata_metrics
                                    )

                                    # Analyze quality of the final optimized metadata
                                    quality_scores = metadata_generator.analyze_metadata_quality(
                                        final_validated_meta,
                                        keyword=optimization_keyword
                                    )

                                    # Update seo_metadata with the optimized and validated content
                                    seo_metadata["title"] = final_validated_meta.get("title", seo_metadata.get("title"))
                                    seo_metadata["description"] = final_validated_meta.get("description", seo_metadata.get("description"))
                                    seo_metadata["tags"] = final_validated_meta.get("tags", seo_metadata.get("tags"))
                                    seo_metadata["quality_scores"] = quality_scores
                                    seo_metadata["optimization_details"] = {
                                        "optimized_by_downloader": True,
                                        "optimization_timestamp": datetime.now().isoformat(),
                                        "method": "utils.metadata_generator.improve_metadata",
                                        "overall_quality_score": quality_scores.get("overall_score", 0)
                                    }
                                    print_success(f"Metadata optimized. Overall Quality: {quality_scores.get('overall_score', 0):.2f}", 3)
                                    save_metadata_metrics(metadata_metrics) # Save metrics after optimization attempt
                                else:
                                    print_warning("Metadata optimization skipped: utils.metadata_generator not available.", 3)

                                # Get playlist suggestion
                                print_info("Getting playlist suggestion...", 3)
                                # Extract existing playlist titles from cache
                                existing_playlist_titles = []
                                for key, value in playlist_data_cache.items():
                                    if key != "timestamp" and not key.startswith("NEW: "):
                                        existing_playlist_titles.append(value)

                                # Get suggestion from Gemini
                                playlist_suggestion = get_playlist_suggestion(
                                    seo_metadata.get("title", original_title),
                                    seo_metadata.get("description", ""),
                                    seo_metadata.get("tags", []),
                                    existing_playlist_titles
                                )

                                if playlist_suggestion:
                                    seo_metadata["target_playlist"] = playlist_suggestion
                                    print_success(f"Playlist suggestion: {playlist_suggestion}", 3)

                                    # Add to cache if it's a NEW suggestion
                                    if playlist_suggestion.startswith("NEW: ") and playlist_suggestion not in playlist_data_cache:
                                        playlist_data_cache[playlist_suggestion] = None  # None indicates not created yet
                                        save_cache(playlist_data_cache, PLAYLIST_DATA_CACHE_PATH, "Playlist Data")
                                        print_info(f"Added new playlist suggestion to cache: {playlist_suggestion}", 4)
                                else:
                                    print_warning("No playlist suggestion available.", 3)

                            # Metadata API calls are tracked inside generate_metadata_with_timeout

                            # Prepare filenames
                            video_file_name = f"video{video_counter}.mp4"; video_file_path = os.path.join(DOWNLOADS_FOLDER, video_file_name); info_json_path = os.path.splitext(video_file_path)[0] + ".info.json"
                            if os.path.exists(video_file_path): print_warning(f"Target file exists: {video_file_name}. Skipping download, incrementing counter.", 2); video_counter += 1; continue

                            print_info(f"Downloading video{video_counter}: {video_file_name} (Views: {entry.get('view_count', 'N/A')})", 2)
                            ydl_opts_download = { 'format': 'bestvideo[height>=1080][ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best', 'outtmpl': {'default': video_file_path}, 'merge_output_format': 'mp4', 'writeinfojson': True, 'quiet': True, 'no_warnings': True, 'nocheckcertificate': True, 'source_address': '0.0.0.0', 'ffmpeg_location': FFMPEG_PATH, 'retries': 3, 'fragment_retries': 3, }
                            download_success = False; run_metrics["downloads_attempted"] += 1; performance_metrics["total_downloads_attempted"] += 1
                            try:
                                with yt_dlp.YoutubeDL(ydl_opts_download) as ydl_download: ydl_download.download([f"https://www.youtube.com/watch?v={video_id}"])
                                download_success = True
                            except yt_dlp.utils.DownloadError as dl_error: print_error(f"Download Error (ID {video_id}): {dl_error}", 3); run_metrics["download_errors"] += 1; performance_metrics["total_download_errors"] += 1; # Cleanup partials...
                            except Exception as general_e: print_error(f"Unexpected Download Error (ID {video_id}): {general_e}", 3, include_traceback=True); run_metrics["download_errors"] += 1; performance_metrics["total_download_errors"] += 1; continue # Skip to next video on unexpected error

                            # Post-Download
                            if download_success:
                                print_success(f"Download successful: {video_file_name}", 3)
                                print_info("Saving metadata JSON...", 3); saved_metadata = save_metadata(entry, video_counter, seo_metadata, METADATA_FOLDER, current_keyword=keyword); print_success("Metadata saved.", 4)
                                print_info("Adding entry to Excel (in memory)...", 3); downloaded_sheet.append([ f"video{video_counter}", saved_metadata["optimized_title"], datetime.now().strftime("%Y-%m-%d %H:%M:%S"), saved_metadata["view_count"], saved_metadata["uploader"], saved_metadata["original_title"] ]); print_success("Entry added.", 4)

                                # NOTE: We no longer delete info.json here - it will be deleted after tag extraction

                                downloaded_video_ids.add(video_id) # Add to skip list
                                keyword_frequency[keyword] = keyword_frequency.get(keyword, 0) + 1 # Increment discovery score
                                run_metrics["downloads_successful"] += 1; performance_metrics["total_successful_downloads"] += 1
                                if keyword not in run_metrics["keywords_used"]: run_metrics["keywords_used"].append(keyword)
                                performance_metrics.setdefault("keyword_performance", {})[keyword] = performance_metrics["keyword_performance"].get(keyword, 0) + 1 # Increment overall keyword score

                                # --- Tag Extraction & Keyword Pool Update ---
                                if os.path.exists(info_json_path):
                                    try:
                                        with open(info_json_path, 'r', encoding='utf-8') as f: video_info = json.load(f)
                                        video_tags = video_info.get('tags', [])
                                        if video_tags:
                                            print_info(f"Processing {len(video_tags)} tags from info file...", 3)
                                            new_unique_tags_found = set(); lower_to_original_keyword_map = {kw.lower(): kw for kw in keyword_frequency.keys()}

                                            # Update scores for existing keywords found in tags
                                            for tag in video_tags:
                                                tag_lower = tag.strip().lower()
                                                if tag_lower in lower_to_original_keyword_map:
                                                    original_kw = lower_to_original_keyword_map[tag_lower]
                                                    keyword_frequency[original_kw] = keyword_frequency.get(original_kw, 0) + 0.1 # Small bump for being a tag

                                            # Identify potentially new keywords
                                            for tag in video_tags:
                                                tag_strip = tag.strip(); tag_lower = tag_strip.lower()
                                                if not tag_lower: continue
                                                is_relevant = any(sub in tag_lower for sub in required_substrings_lower)
                                                is_not_social = tag_lower not in social_media_lower; is_not_generic = tag_lower not in generic_terms_lower
                                                is_not_seed = tag_lower != seed_niche.lower(); is_new = tag_lower not in lower_to_original_keyword_map
                                                is_long_enough = len(tag_lower.replace(" ", "")) > 3
                                                if is_relevant and is_not_social and is_not_generic and is_not_seed and is_new and is_long_enough:
                                                    new_unique_tags_found.add(tag_strip)

                                            # Add new keywords / replace old ones
                                            if new_unique_tags_found:
                                                print_info(f"Found {len(new_unique_tags_found)} relevant new unique tags to potentially add.", 4)
                                                tags_to_add_list = list(new_unique_tags_found); added_count = 0
                                                for tag_to_add in tags_to_add_list:
                                                    if len(keyword_frequency) < max_keywords: # Check current size of dict
                                                        keyword_frequency[tag_to_add] = 0 # Add with score 0
                                                        added_count += 1
                                                    else: # Keyword limit reached - Tiered Removal
                                                        kw_to_remove = None; min_score = float('inf')
                                                        # Try removing score <= 1 first
                                                        candidates = [kw for kw, score in keyword_frequency.items() if kw.lower() != seed_niche.lower() and score <= 1]
                                                        if candidates: candidates.sort(key=lambda kw: (keyword_frequency.get(kw, 0), kw)); kw_to_remove = candidates[0]
                                                        else: # Find absolute lowest score >= 1
                                                            eligible = [(kw, score) for kw, score in keyword_frequency.items() if kw.lower() != seed_niche.lower()]
                                                            if eligible: kw_to_remove, min_score = min(eligible, key=lambda item: item[1])

                                                        if kw_to_remove:
                                                            print_warning(f"Removing '{kw_to_remove}' (Score: {keyword_frequency.get(kw_to_remove, 0):.2f}) to add '{tag_to_add}'.", 5)
                                                            keyword_frequency.pop(kw_to_remove, None)
                                                            keyword_frequency[tag_to_add] = 0; added_count += 1
                                                        else: print_warning(f"Keyword limit reached, no suitable keyword found for removal. Cannot add '{tag_to_add}'.", 5); break # Stop trying to add tags from this video
                                                if added_count > 0: print_success(f"Added {added_count} new keywords from tags. Total keywords: {len(keyword_frequency)}.", 4)
                                        else: print_info("No tags found in info file.", 3)
                                    except json.JSONDecodeError as json_e: print_error(f"Error decoding info.json '{info_json_path}': {json_e}", 3)
                                    except Exception as tag_e: print_error(f"Error processing tags from info.json '{info_json_path}': {tag_e}", 3, include_traceback=True)
                                    finally: # Cleanup info.json
                                        try: os.remove(info_json_path); # print_info(f"Deleted info file: {info_json_path}", 4) # Less verbose
                                        except OSError as e_del: print_warning(f"Error deleting info.json '{info_json_path}': {e_del}", 4)
                                else: print_warning(f"Info file not found, skipping tag extraction: {info_json_path}", 3)

                                video_counter += 1; total_downloaded += 1; videos_downloaded_for_keyword += 1

                    except yt_dlp.utils.DownloadError as search_err: print_error(f"Error during YouTube search for keyword '{keyword}': {search_err}", 1)
                    except Exception as loop_err: print_error(f"Unexpected error processing keyword '{keyword}': {loop_err}", 1, include_traceback=True)

                    print(f"{Fore.CYAN}--- Finished Keyword: '{keyword}'. Videos downloaded: {videos_downloaded_for_keyword} ---")
                    save_cache(keyword_frequency, KEYWORDS_CACHE_FILE_PATH); print_success("Saved keyword cache.", 1) # Save freq cache after each keyword
            else: print_info("No keywords selected for searching (MAX_DOWNLOADS likely 0 or limit reached).")

            # --- End of Run Updates ---
            # Update run metrics with metadata stats for the run
            run_metrics["metadata_api_calls"] = metadata_metrics.get("total_api_calls", 0) - metadata_api_calls_start
            run_metrics["metadata_errors"] = metadata_metrics.get("total_metadata_errors", 0) - metadata_errors_start
            performance_metrics.setdefault("runs", []).append(run_metrics)
            if len(performance_metrics["runs"]) > 20: performance_metrics["runs"] = performance_metrics["runs"][-20:] # Keep last 20 runs
            save_performance_metrics(performance_metrics) # Save overall metrics

            # Generate detailed run summary
            print_info("Generating detailed run summary...")
            run_summary = generate_detailed_run_summary(run_metrics, performance_metrics, config)
            if run_summary:
                print_success(f"Run summary generated and saved to {os.path.basename(constants.RUN_SUMMARIES_LOG_FILE)}")
                print_info(f"Summary preview:\n{run_summary[:400]}...", 1)
            else:
                print_warning("Could not generate detailed run summary.", 1)

            # Parameter Tuning Check
            if len(performance_metrics["runs"]) >= MIN_RUNS_BEFORE_TUNING:
                print_info("Generating parameter tuning suggestions...")
                suggestions = generate_tuning_suggestions(performance_metrics, config)
                if suggestions: print_success(f"Tuning suggestions generated and saved to {os.path.basename(constants.TUNING_SUGGESTIONS_LOG_FILE)}"); print_info(f"Suggestions preview:\n{suggestions[:400]}...", 1)
                else: print_warning("Could not generate parameter tuning suggestions.", 1)

        # --- Refresh Playlist Data if Needed ---
        print(f"\n{Fore.CYAN}{Style.BRIGHT}--- Refreshing Playlist Data if Needed ---{Style.RESET_ALL}")

        # Simple cache expiry - refetch all if cache is old (could be more granular)
        refetch_playlists = True
        if "timestamp" in playlist_data_cache:
            try:
                cache_time = datetime.fromisoformat(playlist_data_cache["timestamp"])
                if datetime.now() - cache_time < timedelta(hours=24): # 24hr cache
                    refetch_playlists = False
                    print_info("Using cached playlist data (less than 24 hours old).")
            except:
                print_warning("Error reading playlist cache timestamp. Will refetch.")

        # Initialize service variable for potential API calls
        service = None
        existing_playlists_api = {} # Will store API results if fetched

        if refetch_playlists:
            print_info("Fetching fresh playlist data from YouTube API...")
            service = get_authenticated_service() # Get service if needed
            if service:
                existing_playlists_api = fetch_channel_playlists(service)
                # Update cache: Keep existing suggestions, update API ones
                new_cache_data = {"timestamp": datetime.now().isoformat()}
                for p_id, p_title in existing_playlists_api.items():
                    new_cache_data[p_id] = p_title # Add/update fetched playlists
                # Keep previous NEW suggestions if they don't conflict with fetched IDs/Titles
                existing_titles_lower = {t.lower() for t in existing_playlists_api.values()}
                for key, val in playlist_data_cache.items():
                    if key.startswith("NEW: ") and key not in new_cache_data:
                        # Avoid adding if an existing playlist now has the same name
                        suggested_title = key[len("NEW: "):]
                        if suggested_title.lower() not in existing_titles_lower:
                            new_cache_data[key] = val # Keep old NEW suggestion

                playlist_data_cache = new_cache_data
                save_cache(playlist_data_cache, PLAYLIST_DATA_CACHE_PATH, "Playlist Data")
            else:
                print_error("Cannot fetch playlists - API service unavailable. Using potentially stale cache.")

        # --- Final Operations ---
        print(f"\n{Fore.CYAN}{Style.BRIGHT}--- Final Operations ---{Style.RESET_ALL}")

        # Sort Downloaded Sheet by Views
        print_info("Sorting downloaded data by views (descending)...", 1)
        if downloaded_sheet.max_row > 1:
            header = [cell.value for cell in downloaded_sheet[1]]
            try:
                view_col_idx = header.index("Views") if "Views" in header else -1
                if view_col_idx != -1:
                    data_rows = list(downloaded_sheet.iter_rows(min_row=2, values_only=True))
                    def get_view_count_safe(row):
                        try: return int(row[view_col_idx]) if row[view_col_idx] is not None else 0
                        except (ValueError, TypeError, IndexError): return 0
                    data_rows.sort(key=get_view_count_safe, reverse=True)
                    downloaded_sheet.delete_rows(2, downloaded_sheet.max_row -1)
                    for row_data in data_rows: downloaded_sheet.append(row_data)
                    print_success("Sorting complete.", 2)
                else: print_warning("'Views' column not found. Skipping sorting.", 2)
            except Exception as sort_e: print_warning(f"Error during sorting: {sort_e}. Skipping.", 2)
        else: print_info("No data rows in 'Downloaded' sheet to sort.", 2)

        # Save Excel
        print_info("Saving final Excel file...", 1)

        # Try to import excel_utils module
        try:
            import excel_utils
            excel_utils_available = True
            print_info("Using excel_utils module for robust Excel saving", 2)
        except ImportError:
            excel_utils_available = False
            print_warning("excel_utils module not available. Using fallback Excel saving.", 2)

        if excel_utils_available:
            # Extract workbook data for backup in case save fails
            def extract_data(wb):
                data = {}
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    data[sheet_name] = []
                    for row in sheet.iter_rows(values_only=True):
                        data[sheet_name].append(list(row))
                return data

            # Use the robust save mechanism
            if excel_utils.save_workbook_with_fallback(wb, EXCEL_FILE_PATH, extract_data):
                print_success(f"Saved final updates to {os.path.basename(EXCEL_FILE_PATH)} using excel_utils.", 2)
            else:
                # If all save methods failed, create a JSON backup
                backup_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"excel_backup_data_{datetime.now():%Y%m%d_%H%M%S}.json")
                print_warning(f"All Excel save methods failed. Creating JSON backup: {backup_file}", 2)
                try:
                    with open(backup_file, "w", encoding='utf-8') as bf:
                        json.dump(extract_data(wb), bf, indent=4, default=str)
                        print_success(f"Saved backup to {backup_file}", 2)
                except Exception as be:
                    print_error(f"CRITICAL: Failed backup save: {be}", 2)
        else:
            # Fallback to original save mechanism with simple retry
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    wb.save(EXCEL_FILE_PATH)
                    print_success(f"Saved final updates to {os.path.basename(EXCEL_FILE_PATH)} (attempt {attempt+1}).", 2)
                    break
                except PermissionError as pe:
                    if attempt < max_retries - 1:
                        print_warning(f"PermissionError saving Excel (attempt {attempt+1}/{max_retries}): {pe}", 2)
                        print_info(f"Retrying in 2 seconds...", 2)
                        time.sleep(2)
                    else:
                        # On last attempt failure, try to save to a backup file
                        print_error(f"Failed to save Excel after {max_retries} attempts: {pe}", 2)
                        backup_path = f"{EXCEL_FILE_PATH}.backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
                        try:
                            wb.save(backup_path)
                            print_success(f"Saved backup Excel file: {backup_path}", 2)
                        except Exception as be:
                            print_error(f"Failed to save backup Excel file: {be}", 2)
                except Exception as e:
                    print_error(f"Error saving final Excel file '{os.path.basename(EXCEL_FILE_PATH)}': {e}", 2)
                    # Try to save to a backup file
                    backup_path = f"{EXCEL_FILE_PATH}.backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
                    try:
                        wb.save(backup_path)
                        print_success(f"Saved backup Excel file: {backup_path}", 2)
                    except Exception as be:
                        print_error(f"Failed to save backup Excel file: {be}", 2)
                    break

        # Final Cache Saves
        print_info("Performing final cache saves...", 1)
        save_cache(list(downloaded_video_ids), PLAYLIST_CACHE_FILE_PATH, "Downloaded Video IDs")
        save_cache(keyword_frequency, KEYWORDS_CACHE_FILE_PATH, "Keyword Frequency")
        # Safely save playlist data cache if it exists
        if 'playlist_data_cache' in locals() and isinstance(playlist_data_cache, dict):
            save_cache(playlist_data_cache, PLAYLIST_DATA_CACHE_PATH, "Playlist Data")
        else:
            print_warning("Playlist data cache invalid type or not initialized. Skipping save.")
        # Metadata metrics were saved after prompt check and during timeout/error handling
        print_success("Caches saved.")

        print(f"\n{Style.BRIGHT}{Fore.GREEN}----- Script Execution Finished -----{Style.RESET_ALL}")
        print(f"{Fore.GREEN}Total videos downloaded in this run: {total_downloaded}")
        print(f"{Fore.GREEN}Total unique video IDs tracked (incl. previous): {len(downloaded_video_ids)}")
        print(f"{Fore.GREEN}Final keyword count: {len(keyword_frequency)}")
        print(f"{Fore.GREEN}Next video index will be: video{video_counter}")
        print(f"{Style.BRIGHT}{Fore.GREEN}-------------------------------------\n{Style.RESET_ALL}")

    # --- Error Handling ---
    except FileNotFoundError as fnf_err: print_fatal(f"Required file not found: {fnf_err}")
    except ValueError as val_err: print_fatal(f"Configuration issue: {val_err}")
    except KeyError as key_err: print_fatal(f"Missing required configuration key: {key_err}")
    except Exception as main_err:
        print(f"\n{Back.RED}{Fore.WHITE}{Style.BRIGHT}!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        print(f" FATAL ERROR during script execution: {main_err}")
        print(f"!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! {Style.RESET_ALL}\n")
        traceback.print_exc()
    finally:
        # --- Cleanup Old Excel Backups ---
        try:
            print(f"\n{Fore.CYAN}--- Cleaning Up Old Excel Backups ---{Style.RESET_ALL}")
            # Try to import excel_utils again just in case
            try:
                import excel_utils
                excel_utils_available = True
            except ImportError:
                excel_utils_available = False
                print_warning("excel_utils module not available. Skipping backup cleanup.")

            if excel_utils_available:
                # Call cleanup, keeping backups for the last 7 days
                deleted, remaining = excel_utils.cleanup_old_backups(days_to_keep=7)
                print_info(f"Excel backup cleanup: Deleted {deleted} old backups, Remaining {remaining} backups.")
        except Exception as cleanup_e:
            print_error(f"Error during backup cleanup: {cleanup_e}")

        print(f"{Fore.CYAN}--- Script execution attempt finished. ---{Style.RESET_ALL}")