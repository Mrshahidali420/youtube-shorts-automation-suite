#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Utilities Module

This module provides robust Excel file handling with automatic process closing,
retry mechanisms, and backup functionality to prevent permission errors.
"""

import os
import time
import json
import shutil
import subprocess
import traceback
import platform # For OS detection in process management
import logging # Use project's logging if available, else basic
from datetime import datetime, timedelta
from typing import Optional, Tuple, List, Dict, Any, Callable, Union

# Try to import openpyxl, but provide fallback mechanisms if not available
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # This print will be replaced by proper logging if logger is configured
    print("WARNING: openpyxl not available. Excel functionality will be limited.")

# Try to import psutil for process management
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False
    print("WARNING: psutil not available. Excel process management will be limited.")

# Import constants module for paths
try:
    # Assuming excel_utils.py is in youtube_shorts/ and constants.py is in youtube_shorts/utils/
    from .utils import constants
    CONSTANTS_IMPORTED = True
except ImportError:
    try: # Fallback if run from project root or utils is directly in PYTHONPATH
        from utils import constants
        CONSTANTS_IMPORTED = True
    except ImportError:
        try:
            from youtube_shorts import constants
            CONSTANTS_IMPORTED = True
        except ImportError:
            CONSTANTS_IMPORTED = False
            print("CRITICAL WARNING: constants.py not found. Using fallback paths for excel_utils.")
            # Define fallback paths if constants cannot be imported
            class FallbackConstants:
                PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__)) # Guesses project root if not in package
                EXCEL_BACKUPS_DIR = os.path.join(PROJECT_ROOT, "backups", "excel")
                JSON_BACKUPS_DIR = os.path.join(PROJECT_ROOT, "backups", "json_data") # Assuming JSON backups also go here
                LOGS_DIR = os.path.join(PROJECT_ROOT, "logs")
                BACKUPS_DIR = os.path.join(PROJECT_ROOT, "backups")
                DATA_DIR = os.path.join(PROJECT_ROOT, "data")
            constants = FallbackConstants()

# --- Setup Logger (can be overridden by the main script's logger) ---
logger = logging.getLogger("excel_utils")
if not logger.handlers: # Configure only if no handlers are already set
    logs_dir = constants.LOGS_DIR if CONSTANTS_IMPORTED else "."
    os.makedirs(logs_dir, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(os.path.join(logs_dir, "excel_utils.log")), # Use logs_dir from constants
            logging.StreamHandler()
        ]
    )
# --- End Logger Setup ---

# Constants
MAX_SAVE_RETRIES = 3
RETRY_DELAY_SECONDS = 2
EXCEL_PROCESS_NAMES_WINDOWS = ["excel.exe", "EXCEL.EXE"]
EXCEL_PROCESS_NAMES_MACOS = ["Microsoft Excel", "Excel"]
EXCEL_PROCESS_NAMES_LINUX = ["excel.exe", "libreoffice", "openoffice", "soffice"] # Common Linux office suites
SYSTEM_PLATFORM = platform.system()

# Backup folder now comes from constants
BACKUP_FOLDER_BASE = constants.EXCEL_BACKUPS_DIR # For .xlsx files
JSON_BACKUP_FOLDER_BASE = constants.JSON_BACKUPS_DIR # For .json backups


# --- Logging Functions ---
# These can be replaced with your project's logging functions
def log_info(msg: str, indent: int = 0) -> None:
    """Log an informational message."""
    prefix = "  " * indent
    print(f"{prefix}INFO: {msg}")


def log_success(msg: str, indent: int = 0) -> None:
    """Log a success message."""
    prefix = "  " * indent
    print(f"{prefix}SUCCESS: {msg}")


def log_warning(msg: str, indent: int = 0) -> None:
    """Log a warning message."""
    prefix = "  " * indent
    print(f"{prefix}WARNING: {msg}")


def log_error(msg: str, indent: int = 0, include_traceback: bool = False) -> None:
    """Log an error message with optional traceback."""
    prefix = "  " * indent
    print(f"{prefix}ERROR: {msg}")
    if include_traceback:
        traceback.print_exc()


def log_error_to_file(msg: str, log_file: str = "excel_error.log", include_traceback: bool = False) -> None:
    """Log an error message to a file."""
    try:
        with open(log_file, "a", encoding="utf-8") as f:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{timestamp}] ERROR: {msg}\n")
            if include_traceback:
                f.write(f"Traceback:\n{traceback.format_exc()}\n")
    except Exception as e:
        print(f"ERROR: Failed to write to error log: {e}")


# --- Excel Process Management ---
def find_excel_processes() -> List[Any]:
    """Find all running Excel processes."""
    excel_processes = []
    if not PSUTIL_AVAILABLE:
        log_warning("Cannot find Excel processes: psutil not available")
        return excel_processes

    try:
        process_names_to_check = []
        if SYSTEM_PLATFORM == "Windows": process_names_to_check = EXCEL_PROCESS_NAMES_WINDOWS
        elif SYSTEM_PLATFORM == "Darwin": process_names_to_check = EXCEL_PROCESS_NAMES_MACOS
        elif SYSTEM_PLATFORM == "Linux": process_names_to_check = EXCEL_PROCESS_NAMES_LINUX

        for proc in psutil.process_iter(['pid', 'name']):
            if proc.info['name'] and any(name.lower() in proc.info['name'].lower() for name in process_names_to_check):
                excel_processes.append(proc)
    except Exception as e:
        log_error(f"Error finding Excel processes: {e}")
    return excel_processes


def find_excel_processes_with_file(file_path: str) -> List[psutil.Process]:
    """Find Excel processes that have the specified file open."""
    file_path = os.path.abspath(file_path)
    excel_processes = find_excel_processes()
    processes_with_file = []

    try:
        for proc in excel_processes:
            try:
                for file in proc.open_files():
                    if file.path.lower() == file_path.lower():
                        processes_with_file.append(proc)
                        break
            except (psutil.AccessDenied, psutil.NoSuchProcess):
                continue
    except Exception as e:
        log_error(f"Error checking Excel processes for file: {e}")

    return processes_with_file


def close_excel_processes_with_file(file_path: str) -> bool:
    """Close Excel processes that have the specified file open."""
    processes = find_excel_processes_with_file(file_path)
    if not processes:
        return True  # No processes to close

    success = True
    for proc in processes:
        try:
            proc_name = proc.name()
            proc_pid = proc.pid
            log_warning(f"Attempting to close Excel process (PID: {proc_pid}) with file: {os.path.basename(file_path)}")
            proc.terminate()
            proc.wait(timeout=5)  # Wait up to 5 seconds for graceful termination
            log_success(f"Successfully closed Excel process (PID: {proc_pid})")
        except Exception as e:
            log_error(f"Failed to close Excel process (PID: {proc.pid}): {e}")
            success = False

    return success


def force_close_all_excel_processes() -> bool:
    """Force close all Excel processes (use with caution)."""
    processes = find_excel_processes()
    if not processes:
        return True  # No processes to close

    success = True
    for proc in processes:
        try:
            proc_name = proc.name()
            proc_pid = proc.pid
            log_warning(f"Force closing Excel process: {proc_name} (PID: {proc_pid})")
            proc.kill()  # Force kill
            log_success(f"Killed Excel process: {proc_name} (PID: {proc_pid})")
        except Exception as e:
            log_error(f"Failed to kill Excel process (PID: {proc.pid}): {e}")
            success = False

    return success


def close_excel_via_taskkill() -> bool:
    """Close Excel using Windows taskkill command (fallback method)."""
    try:
        log_warning("Attempting to close Excel using taskkill command")
        subprocess.run(["taskkill", "/F", "/IM", "excel.exe"],
                      stdout=subprocess.PIPE,
                      stderr=subprocess.PIPE,
                      check=False)
        return True
    except Exception as e:
        log_error(f"Failed to close Excel using taskkill: {e}")
        return False


# --- Backup Functions ---
def create_backup_folder(base_backup_dir: str = BACKUP_FOLDER_BASE) -> str:
    """
    Ensure a backup folder exists. Uses the path from constants.
    Args:
        base_backup_dir: The base directory for backups (e.g., backups/excel or backups/json_data)
    Returns:
        str: Path to the backup folder
    """
    # BACKUP_FOLDER_BASE is now a full path from constants.py
    try:
        os.makedirs(base_backup_dir, exist_ok=True)
        if not os.access(base_backup_dir, os.W_OK):
            # Fallback if default backup dir is not writable
            log_warning(f"Default backup folder '{base_backup_dir}' not writable. Trying project root.")
            fallback_dir = os.path.join(constants.PROJECT_ROOT, "backups", os.path.basename(base_backup_dir))
            os.makedirs(fallback_dir, exist_ok=True)
            if os.access(fallback_dir, os.W_OK):
                log_info(f"Using fallback backup folder: {fallback_dir}")
                return fallback_dir
            else:
                log_error(f"Fallback backup folder '{fallback_dir}' also not writable. Backup might fail.")
                return base_backup_dir # Return original path, let it fail later if truly unwritable
        log_info(f"Ensured backup folder exists: {base_backup_dir}")
        return base_backup_dir
    except Exception as e:
        log_error(f"Error ensuring backup folder '{base_backup_dir}': {e}. Backup might fail.")
        return base_backup_dir # Return original path


def create_file_backup(file_path: str, backup_base_dir: Optional[str] = None, suffix: str = "") -> Optional[str]:
    """
    Create a backup of any file into the specified base backup directory.
    """
    if not os.path.exists(file_path):
        log_warning(f"Cannot backup non-existent file: {file_path}")
        return None

    try:
        # Determine the correct backup directory (excel or json)
        if backup_base_dir is None:
            if file_path.endswith(".xlsx"):
                backup_base_dir = BACKUP_FOLDER_BASE # backups/excel/
            elif file_path.endswith(".json"):
                backup_base_dir = JSON_BACKUP_FOLDER_BASE # backups/json_data/
            else: # Default to a generic subdir in main backups if type unknown
                backup_base_dir = os.path.join(constants.BACKUPS_DIR, "other")

        # Ensure this specific backup sub-folder exists
        target_backup_folder = create_backup_folder(backup_base_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f") # Added microseconds for uniqueness
        filename = os.path.basename(file_path)
        file_name_part, file_ext_part = os.path.splitext(filename)
        backup_filename = f"{file_name_part}{suffix}_backup_{timestamp}{file_ext_part}" # Suffix before timestamp
        backup_path = os.path.join(target_backup_folder, backup_filename)

        shutil.copy2(file_path, backup_path)
        log_success(f"Created file backup: {backup_path}")
        return backup_path
    except Exception as e:
        log_error(f"Failed to create file backup for {file_path}: {e}")
        return None


def create_excel_backup(file_path: str) -> Optional[str]:
    """Create a backup of an Excel file into the excel backups directory."""
    return create_file_backup(file_path, backup_base_dir=BACKUP_FOLDER_BASE, suffix="") # No extra suffix for excel


def save_data_as_json_backup(data: Any, original_file_path: str) -> Optional[str]:
    """Save data as a JSON backup into the json backups directory."""
    try:
        target_backup_folder = create_backup_folder(JSON_BACKUP_FOLDER_BASE)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
        filename = os.path.basename(original_file_path)
        backup_filename = f"{os.path.splitext(filename)[0]}_data_backup_{timestamp}.json"
        backup_path = os.path.join(target_backup_folder, backup_filename)

        with open(backup_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4, default=str) # Default str for non-serializable

        log_success(f"Created JSON data backup: {backup_path}")
        return backup_path
    except Exception as e:
        log_error(f"Failed to create JSON data backup for {original_file_path}: {e}")
        return None


def cleanup_old_backups(days_to_keep: int = 7) -> Tuple[int, int]:
    """
    Deletes backup files older than a specified number of days from ALL backup subdirectories.
    """
    total_deleted = 0
    total_remaining = 0
    backup_sub_dirs = [constants.EXCEL_BACKUPS_DIR, constants.JSON_BACKUPS_DIR,
                       os.path.join(constants.BACKUPS_DIR, "prompts"),
                       os.path.join(constants.BACKUPS_DIR, "other")]

    for backup_folder_path in backup_sub_dirs:
        if not os.path.isdir(backup_folder_path):
            log_info(f"Backup sub-folder '{backup_folder_path}' does not exist. Skipping cleanup for this folder.")
            continue

        log_info(f"Starting backup cleanup in '{backup_folder_path}'. Keeping backups from the last {days_to_keep} days.")
        deleted_in_subdir, remaining_in_subdir = _cleanup_specific_backup_folder(backup_folder_path, days_to_keep)
        total_deleted += deleted_in_subdir
        total_remaining += remaining_in_subdir

    log_info(f"Total backup cleanup finished. Deleted: {total_deleted}, Remaining: {total_remaining} across all backup types.")
    return total_deleted, total_remaining

def _cleanup_specific_backup_folder(backup_folder_path: str, days_to_keep: int) -> Tuple[int, int]:
    """Helper function to clean up a specific backup folder."""
    deleted_count = 0
    remaining_count = 0
    error_count = 0
    now = datetime.now()
    cutoff_time = now - timedelta(days=days_to_keep)

    try:
        for filename in os.listdir(backup_folder_path):
            file_path = os.path.join(backup_folder_path, filename)
            if not os.path.isfile(file_path): continue

            try:
                file_mod_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                if file_mod_time < cutoff_time:
                    log_info(f"Deleting old backup: {filename} (Modified: {file_mod_time.strftime('%Y-%m-%d')})", indent=1)
                    try: os.remove(file_path); deleted_count += 1
                    except Exception as e: log_error(f"Error deleting {filename}: {e}", 1); error_count += 1; remaining_count +=1
                else:
                    remaining_count += 1
            except Exception as e: log_error(f"Error processing {filename}: {e}", 1); error_count +=1; remaining_count +=1

        log_info(f"Cleanup for '{backup_folder_path}': Deleted: {deleted_count}, Remaining: {remaining_count}, Errors: {error_count}")
        return deleted_count, remaining_count
    except Exception as e:
        log_error(f"Error during cleanup of '{backup_folder_path}': {e}")
        return 0, len(os.listdir(backup_folder_path)) if os.path.exists(backup_folder_path) else 0


# --- Excel Loading Functions ---
def safe_load_workbook(file_path: str, read_only: bool = False, data_only: bool = False) -> Tuple[Optional[Any], Optional[str]]:
    """
    Safely load an Excel workbook with error handling.

    Returns:
        Tuple of (workbook object or None, error message or None)
    """
    if not OPENPYXL_AVAILABLE:
        return None, "openpyxl not available"

    if not os.path.exists(file_path):
        return None, f"File not found: {file_path}"

    try:
        wb = load_workbook(file_path, read_only=read_only, data_only=data_only)
        return wb, None
    except Exception as e:
        error_msg = f"Error loading workbook {file_path}: {str(e)}"
        log_error(error_msg)
        return None, error_msg


def create_new_workbook(file_path: str, sheets_config: Dict[str, List[str]]) -> Tuple[Optional[Any], Optional[str]]:
    """
    Create a new Excel workbook with specified sheets and headers.

    Args:
        file_path: Path to save the new workbook
        sheets_config: Dict mapping sheet names to header lists

    Returns:
        Tuple of (workbook object or None, error message or None)
    """
    if not OPENPYXL_AVAILABLE:
        return None, "openpyxl not available"

    try:
        wb = Workbook()

        # Configure first sheet (active sheet)
        first_sheet_name = next(iter(sheets_config))
        first_sheet = wb.active
        first_sheet.title = first_sheet_name
        if sheets_config[first_sheet_name]:
            first_sheet.append(sheets_config[first_sheet_name])

        # Create additional sheets
        for sheet_name, headers in sheets_config.items():
            if sheet_name == first_sheet_name:
                continue  # Skip the first sheet as it's already created

            sheet = wb.create_sheet(title=sheet_name)
            if headers:
                sheet.append(headers)

        # Save the workbook
        wb.save(file_path)
        log_success(f"Created new Excel file: {file_path}")
        return wb, None
    except Exception as e:
        error_msg = f"Error creating workbook {file_path}: {str(e)}"
        log_error(error_msg)
        return None, error_msg


# --- Excel Saving Functions ---
def safe_save_workbook(wb: Any, file_path: str, close_excel: bool = True,
                      create_backup: bool = True, max_retries: int = MAX_SAVE_RETRIES) -> bool:
    """
    Safely save an Excel workbook with retry mechanism and Excel process handling.

    Args:
        wb: The workbook object to save
        file_path: Path to save the workbook
        close_excel: Whether to attempt to close Excel processes with the file open
        create_backup: Whether to create a backup before saving
        max_retries: Maximum number of save attempts

    Returns:
        bool: True if save was successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        log_error("Cannot save workbook: openpyxl not available")
        return False

    # Create a backup before attempting to save
    if create_backup:
        create_excel_backup(file_path)

    # Try to save with retries
    for attempt in range(max_retries):
        try:
            wb.save(file_path)
            log_success(f"Excel file saved successfully: {file_path}")
            return True
        except PermissionError as pe:
            log_warning(f"PermissionError on attempt {attempt+1}/{max_retries}: {pe}")

            if close_excel:
                # Try to close Excel processes that have the file open
                if close_excel_processes_with_file(file_path):
                    log_info(f"Closed Excel processes with file open. Retrying save...")
                else:
                    log_warning("Failed to close some Excel processes. Retrying anyway...")

            # Wait before retrying
            if attempt < max_retries - 1:
                time.sleep(RETRY_DELAY_SECONDS)
        except Exception as e:
            log_error(f"Error saving workbook on attempt {attempt+1}/{max_retries}: {e}")
            if attempt < max_retries - 1:
                time.sleep(RETRY_DELAY_SECONDS)

    # If we get here, all save attempts failed
    log_error(f"Failed to save Excel file after {max_retries} attempts: {file_path}")
    return False


def save_workbook_with_fallback(wb: Any, file_path: str, data_extractor: Optional[Callable] = None) -> bool:
    """
    Save a workbook with fallback to alternative file and JSON backup.

    Args:
        wb: The workbook object to save
        file_path: Path to save the workbook
        data_extractor: Optional function to extract data from workbook for JSON backup

    Returns:
        bool: True if any save method was successful, False if all failed
    """
    # Try to save normally with Excel process handling
    if safe_save_workbook(wb, file_path, close_excel=True, create_backup=True):
        return True

    # If normal save failed, try saving to an alternative filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    alt_file_path = f"{os.path.splitext(file_path)[0]}_alt_{timestamp}.xlsx"

    log_warning(f"Attempting to save to alternative file: {alt_file_path}")
    if safe_save_workbook(wb, alt_file_path, close_excel=False, create_backup=False):
        log_success(f"Saved to alternative file: {alt_file_path}")
        return True

    # If both Excel saves failed and we have a data extractor, save as JSON
    if data_extractor:
        try:
            data = data_extractor(wb)
            json_backup_path = save_data_as_json_backup(data, file_path)
            if json_backup_path:
                log_success(f"Saved data as JSON backup: {json_backup_path}")
                return True
        except Exception as e:
            log_error(f"Failed to extract and save data as JSON: {e}")

    return False


# --- Data Extraction Helpers ---
def extract_sheet_data(sheet: Any) -> List[List[Any]]:
    """Extract all data from a worksheet as a list of rows."""
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


def extract_workbook_data(wb: Any) -> Dict[str, List[List[Any]]]:
    """Extract all data from a workbook as a dictionary of sheet data."""
    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data[sheet_name] = extract_sheet_data(sheet)
    return data


# --- High-Level Functions ---
def load_or_create_excel(file_path: str, sheets_config: Dict[str, List[str]]) -> Tuple[Optional[Any], Dict[str, Any], bool]:
    """
    Load an existing Excel file or create a new one if it doesn't exist.

    Args:
        file_path: Path to the Excel file
        sheets_config: Dict mapping sheet names to header lists

    Returns:
        Tuple of (workbook object or None, dict of sheets by name, whether save is needed)
    """
    save_needed = False
    sheets = {}

    if not OPENPYXL_AVAILABLE:
        return None, sheets, False

    if not os.path.exists(file_path):
        # Create new workbook
        wb, error = create_new_workbook(file_path, sheets_config)
        if wb:
            # Get all sheets
            for sheet_name in wb.sheetnames:
                sheets[sheet_name] = wb[sheet_name]
            return wb, sheets, False  # No save needed as we just created and saved it
        else:
            log_error(f"Failed to create Excel file: {error}")
            return None, sheets, False

    # Load existing workbook
    wb, error = safe_load_workbook(file_path)
    if not wb:
        log_error(f"Failed to load Excel file: {error}")
        return None, sheets, False

    # Check and create missing sheets
    for sheet_name, headers in sheets_config.items():
        if sheet_name not in wb.sheetnames:
            log_warning(f"Sheet '{sheet_name}' not found. Creating...")
            sheet = wb.create_sheet(title=sheet_name)
            if headers:
                sheet.append(headers)
            sheets[sheet_name] = sheet
            save_needed = True
        else:
            sheets[sheet_name] = wb[sheet_name]

    return wb, sheets, save_needed


def append_rows_to_sheet(sheet: Any, rows: List[List[Any]], expected_column_count: Optional[int] = None) -> int:
    """
    Append rows to a worksheet with validation.

    Args:
        sheet: The worksheet to append to
        rows: List of rows to append
        expected_column_count: Optional validation for column count

    Returns:
        int: Number of rows successfully appended
    """
    if not rows:
        return 0

    appended_count = 0
    for row in rows:
        if expected_column_count is not None and len(row) != expected_column_count:
            log_warning(f"Skipping row with incorrect column count. Expected {expected_column_count}, got {len(row)}")
            continue

        try:
            sheet.append(row)
            appended_count += 1
        except Exception as e:
            log_error(f"Error appending row: {e}")

    return appended_count


def get_last_row_index(sheet: Any) -> int:
    """Get the index of the last row in a worksheet."""
    return sheet.max_row


def get_last_video_index(sheet: Any, index_column: int = 1, prefix: str = "video") -> int:
    """
    Find the highest video index (e.g., video123) in a column.

    Args:
        sheet: The worksheet to search
        index_column: The column index (1-based) to search
        prefix: The prefix of the video index (e.g., "video")

    Returns:
        int: The highest video index found + 1, or 1 if none found
    """
    max_index = 0

    # Skip header row
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=index_column).value
        if cell_value and isinstance(cell_value, str) and cell_value.lower().startswith(prefix.lower()):
            try:
                index = int(cell_value[len(prefix):])
                max_index = max(max_index, index)
            except (ValueError, IndexError):
                continue

    return max_index + 1


def find_column_index(sheet: Any, column_name: str, case_sensitive: bool = False) -> Optional[int]:
    """
    Find the index of a column by name in the header row.

    Args:
        sheet: The worksheet to search
        column_name: The name of the column to find
        case_sensitive: Whether to perform case-sensitive matching

    Returns:
        int: The 1-based column index, or None if not found
    """
    if sheet.max_row == 0:
        return None

    search_name = column_name if case_sensitive else column_name.lower()

    for cell in sheet[1]:
        cell_value = str(cell.value) if cell.value is not None else ""
        compare_value = cell_value if case_sensitive else cell_value.lower()

        if compare_value == search_name:
            return cell.column

    return None


def archive_old_excel_entries(wb: Any, sheet_name: str, date_col_name: str, days_to_keep: int) -> bool:
    """
    Moves old entries from a source sheet to an archive sheet.

    Args:
        wb: The workbook object
        sheet_name: Name of the source sheet
        date_col_name: Name of the column containing dates
        days_to_keep: Number of days to keep entries before archiving

    Returns:
        bool: True if any entries were archived, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        log_error("Cannot archive entries: openpyxl not available")
        return False

    if sheet_name not in wb.sheetnames:
        log_warning(f"Cannot archive: Source sheet '{sheet_name}' not found.")
        return False

    source_sheet = wb[sheet_name]
    archive_sheet_name = f"{sheet_name}_Archive"

    # Ensure archive sheet exists and has headers
    if archive_sheet_name not in wb.sheetnames:
        log_info(f"Creating archive sheet: '{archive_sheet_name}'")
        archive_sheet = wb.create_sheet(archive_sheet_name)
        # Copy headers from source
        headers = [cell.value for cell in source_sheet[1]]
        archive_sheet.append(headers)
    else:
        archive_sheet = wb[archive_sheet_name]

    # Find date column index (case-insensitive)
    date_col_idx = find_column_index(source_sheet, date_col_name, case_sensitive=False)
    if not date_col_idx:
        log_error(f"Cannot archive sheet '{sheet_name}': Date column '{date_col_name}' not found.")
        return False

    cutoff_date = datetime.now() - timedelta(days=days_to_keep)
    moved_count = 0
    rows_to_delete = []

    log_info(f"Archiving entries older than {days_to_keep} days from '{sheet_name}'...")

    # Iterate backwards to safely delete rows
    for row_idx in range(source_sheet.max_row, 1, -1):  # Skip header row
        date_cell = source_sheet.cell(row=row_idx, column=date_col_idx)
        entry_date = None

        try:
            if isinstance(date_cell.value, datetime):
                entry_date = date_cell.value
            elif isinstance(date_cell.value, float):  # Handle Excel date numbers
                entry_date = datetime.fromtimestamp(time.mktime(time.gmtime((date_cell.value - 25569) * 86400.0)))
            elif isinstance(date_cell.value, str) and date_cell.value.strip() and date_cell.value.strip().upper() != "N/A":
                # Try common date formats
                try:
                    entry_date = datetime.strptime(date_cell.value.strip(), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        entry_date = datetime.strptime(date_cell.value.strip(), "%Y-%m-%d")
                    except ValueError:
                        try:
                            entry_date = datetime.strptime(date_cell.value.strip(), "%m/%d/%Y %H:%M:%S")
                        except ValueError:
                            try:
                                entry_date = datetime.strptime(date_cell.value.strip(), "%m/%d/%Y")
                            except ValueError:
                                log_warning(f"Could not parse date '{date_cell.value}' in row {row_idx}. Skipping.")
                                continue

            if entry_date and entry_date < cutoff_date:
                # Copy row to archive sheet
                row_values = [cell.value for cell in source_sheet[row_idx]]
                archive_sheet.append(row_values)
                rows_to_delete.append(row_idx)
                moved_count += 1

        except (ValueError, TypeError) as parse_err:
            log_warning(f"Skipping row {row_idx} in '{sheet_name}' due to date parse error: {parse_err} (Value: '{date_cell.value}')")
        except Exception as e:
            log_error(f"Unexpected error processing row {row_idx} in '{sheet_name}': {e}")

    # Delete rows from source sheet after iteration
    if rows_to_delete:
        log_info(f"Deleting {len(rows_to_delete)} archived rows from '{sheet_name}'...")
        # Sort indices descending to delete from bottom up
        for row_idx in sorted(rows_to_delete, reverse=True):
            source_sheet.delete_rows(row_idx)
        log_success(f"Archived {moved_count} entries from '{sheet_name}'.")
        return True  # Indicate changes were made
    else:
        log_info(f"No entries old enough to archive found in '{sheet_name}'.")
        return False


# --- Main Function for Testing (Ensure test_file uses constants.DATA_DIR) ---
def test_excel_utils():
    """Test the Excel utilities module."""
    if not OPENPYXL_AVAILABLE:
        log_error("Skipping excel_utils tests: openpyxl library not installed.")
        return

    test_file_name = "test_excel_utils.xlsx"
    test_file_path = os.path.join(constants.DATA_DIR, test_file_name) # Test file in data dir
    log_info(f"Starting Excel utilities test using file: {test_file_path}")

    sheets_config = {
        "Data": ["ID", "Name", "Value", "Date", "Status"],
        "Config": ["Setting", "Value", "Description"]
    }

    # Test loading or creating Excel
    wb, sheets, _ = load_or_create_excel(test_file_path, sheets_config)
    if wb:
        log_success("Excel file created/loaded successfully for test")

        # Test appending rows with dates
        current_date = datetime.now()
        old_date = current_date - timedelta(days=10)
        very_old_date = current_date - timedelta(days=200)

        rows = [
            ["ID1", "Current", 100, current_date, "Active"],
            ["ID2", "Old", 200, old_date, "Pending"],
            ["ID3", "Very Old", 300, very_old_date, "Archived"]
        ]
        append_rows_to_sheet(sheets["Data"], rows)

        if save_workbook_with_fallback(wb, test_file_path, extract_workbook_data):
            log_success(f"Saved test workbook to {test_file_path}")

            # Test archiving
            archive_result = archive_old_excel_entries(wb, "Data", "Date", 30) # Test archive
            if archive_result:
                log_success("Archiving test successful")
            else:
                log_warning("No entries were archived (this is normal if running test multiple times)")

            if save_workbook_with_fallback(wb, test_file_path, extract_workbook_data):
                log_success("Final save after archive successful")
            else:
                log_error("Final save after archive failed")
        else:
            log_error("Test failed: Could not save test workbook")

        # Cleanup test file
        try:
            os.remove(test_file_path)
            log_info(f"Cleaned up test file: {test_file_path}")
        except Exception as e:
            log_warning(f"Could not cleanup test file {test_file_path}: {e}")
    else:
        log_error("Test failed: Could not load or create test workbook")

    log_info("Excel utilities test complete.")


if __name__ == "__main__":
    test_excel_utils()
