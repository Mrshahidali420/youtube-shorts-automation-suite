# --- START OF FILE uploader.py ---

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YouTube Shorts Uploader

This script uploads videos from the metadata folder to YouTube,
handling metadata, scheduling (including dynamic based on analytics),
playlist management, and error reporting. It prioritizes using the
Page Object Model (POM) uploader if available.

Copyright (c) 2023-2025 Shahid Ali
License: MIT License
GitHub: https://github.com/Mrshahidali420/youtube-shorts-automation
Version: 1.7.0 # Updated version to reflect recent changes
"""

import os
import time
import json
import re
import random
import csv # Keep import for potential future use
from datetime import datetime, timedelta, time as dt_time # Added time import
from typing import Optional, Tuple, List, Dict, Any # Import Any and others

import traceback # For detailed error logging to file
import platform # For OS detection
import subprocess # For running FFmpeg
import signal # For sending signals (like SIGINT equivalent) on non-Windows
import sys # For command-line arguments

# Import Google's Generative AI library for self-improvement features
try:
    import google.generativeai as genai
except ImportError:
    print("Warning: Google Generative AI library not found. Self-improvement features will be disabled.")
    print("To enable, install with: pip install google-generativeai")
    genai = None

# Google API imports for YouTube Data/Analytics & Auth
try:
    import pickle
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    GOOGLE_API_AVAILABLE = True
except ImportError:
    print("ERROR: Google API libraries not found. Install with:")
    print("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
    GOOGLE_API_AVAILABLE = False

# --- Colorama Setup ---
try:
    import colorama
    from colorama import Fore, Style, init
    init(autoreset=True) # Automatically reset style after each print
    COLOR_ENABLED = True
    # print(f"{Fore.GREEN}Colorama loaded successfully. Colored output enabled.{Style.RESET_ALL}") # Optional startup message
except ImportError:
    print("Warning: 'colorama' not found. Install it for colored output (`pip install colorama`). Output will be monochrome.")
    # Define dummy color objects if colorama is not available
    class DummyColor:
        def __getattr__(self, name): return ""
    Fore = DummyColor(); Style = DummyColor() # Assign instances
    COLOR_ENABLED = False
# --- End Colorama Setup ---

# Selenium Imports
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    TimeoutException,
    WebDriverException,
    JavascriptException,
    NoSuchWindowException,
    InvalidSessionIdException,
    NoSuchElementException
)
from webdriver_manager.firefox import GeckoDriverManager

# Date parsing imports
try:
    import dateutil.parser as date_parser
    DATEUTIL_AVAILABLE = True
except ImportError:
    print("Warning: dateutil not available. Date parsing will use fallback methods.")
    DATEUTIL_AVAILABLE = False

# Excel Imports (with fallback)
try:
    # Assume it might be in utils or root based on previous fix attempt
    script_dir_excel = os.path.dirname(os.path.abspath(__file__))
    utils_path_excel = os.path.join(script_dir_excel, 'utils')
    excel_utils_path_root = os.path.join(script_dir_excel, 'excel_utils.py')

    if os.path.isdir(utils_path_excel) and 'excel_utils.py' in os.listdir(utils_path_excel):
         from utils.excel_utils import (
             load_workbook_safely, save_workbook_safely, create_excel_backup,
             find_column_indices, get_cell_value, parse_date_value,
             load_or_create_excel, safe_save_workbook, save_workbook_with_fallback,
             extract_workbook_data, append_rows_to_sheet, archive_old_excel_entries,
             close_excel_processes_with_file # Import process closer
         )
         EXCEL_UTILS_AVAILABLE = True
    elif os.path.exists(excel_utils_path_root):
         import excel_utils
         # Define a fallback function for load_workbook_safely if it doesn't exist
         if hasattr(excel_utils, 'load_workbook_safely'):
             load_workbook_safely = excel_utils.load_workbook_safely
         else:
             # Fallback implementation
             def load_workbook_safely(file_path, read_only=False, data_only=False):
                 """Fallback implementation of load_workbook_safely"""
                 print("WARNING: Using fallback load_workbook_safely function")
                 try:
                     from openpyxl import load_workbook
                     return load_workbook(file_path, read_only=read_only, data_only=data_only)
                 except Exception as e:
                     print(f"ERROR: Failed to load workbook: {e}")
                     return None

         # Define a fallback function for save_workbook_safely if it doesn't exist
         if hasattr(excel_utils, 'save_workbook_safely'):
             save_workbook_safely = excel_utils.save_workbook_safely
         else:
             # Fallback implementation
             def save_workbook_safely(workbook, file_path):
                 """Fallback implementation of save_workbook_safely"""
                 print("WARNING: Using fallback save_workbook_safely function")
                 try:
                     workbook.save(file_path)
                     return True
                 except Exception as e:
                     print(f"ERROR: Failed to save workbook: {e}")
                     return False

         # Get other functions with fallbacks
         create_excel_backup = getattr(excel_utils, 'create_excel_backup', lambda *args, **kwargs: None)
         find_column_indices = getattr(excel_utils, 'find_column_indices', lambda *args, **kwargs: {})
         get_cell_value = getattr(excel_utils, 'get_cell_value', lambda *args, **kwargs: None)
         parse_date_value = getattr(excel_utils, 'parse_date_value', lambda *args, **kwargs: None)
         load_or_create_excel = getattr(excel_utils, 'load_or_create_excel', None)
         safe_save_workbook = getattr(excel_utils, 'safe_save_workbook', None)
         save_workbook_with_fallback = getattr(excel_utils, 'save_workbook_with_fallback', None)
         extract_workbook_data = getattr(excel_utils, 'extract_workbook_data', lambda *args, **kwargs: {})
         append_rows_to_sheet = getattr(excel_utils, 'append_rows_to_sheet', lambda *args, **kwargs: False)
         archive_old_excel_entries = getattr(excel_utils, 'archive_old_excel_entries', lambda *args, **kwargs: (0, 0))
         close_excel_processes_with_file = getattr(excel_utils, 'close_excel_processes_with_file', lambda *args, **kwargs: False)
         EXCEL_UTILS_AVAILABLE = True
         print("INFO: Using excel_utils found in the root directory.")
    else:
        raise ImportError("excel_utils not found")

except ImportError:
    print("WARNING: excel_utils module not found or failed to import. Using basic openpyxl.")
    from openpyxl import load_workbook, Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    EXCEL_UTILS_AVAILABLE = False

# --- Import YouTube Limits Module ---
try:
    # Assume it might be in utils or root
    script_dir_limits = os.path.dirname(os.path.abspath(__file__))
    utils_path_limits = os.path.join(script_dir_limits, 'utils')
    limits_path_root = os.path.join(script_dir_limits, 'youtube_limits.py')

    if os.path.isdir(utils_path_limits) and 'youtube_limits.py' in os.listdir(utils_path_limits):
        from utils.youtube_limits import (
            validate_description, validate_tags, validate_title, # Add validate_title
            DEFAULT_YOUTUBE_DESCRIPTION_LIMIT, DEFAULT_YOUTUBE_TAG_LIMIT,
            DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT, DEFAULT_YOUTUBE_MAX_TAGS_COUNT,
            DEFAULT_YOUTUBE_TITLE_LIMIT # Add title limit constant
        )
    elif os.path.exists(limits_path_root):
        import youtube_limits
        validate_description = youtube_limits.validate_description
        validate_tags = youtube_limits.validate_tags
        validate_title = youtube_limits.validate_title # Add validate_title
        DEFAULT_YOUTUBE_DESCRIPTION_LIMIT = youtube_limits.DEFAULT_YOUTUBE_DESCRIPTION_LIMIT
        DEFAULT_YOUTUBE_TAG_LIMIT = youtube_limits.DEFAULT_YOUTUBE_TAG_LIMIT
        DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT = youtube_limits.DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT
        DEFAULT_YOUTUBE_MAX_TAGS_COUNT = youtube_limits.DEFAULT_YOUTUBE_MAX_TAGS_COUNT
        DEFAULT_YOUTUBE_TITLE_LIMIT = youtube_limits.DEFAULT_YOUTUBE_TITLE_LIMIT # Add title limit constant
        print("INFO: Using youtube_limits found in the root directory.")
    else:
        raise ImportError("youtube_limits.py not found")

except ImportError:
    print(f"{Fore.RED}{Style.BRIGHT}ERROR:{Style.RESET_ALL}{Fore.YELLOW} Could not import from youtube_limits.py. Using fallback validation.")
    print(f"{Fore.YELLOW}       Ensure youtube_limits.py is available for accurate validation.")
    # Dummy functions/constants if module is missing
    DEFAULT_YOUTUBE_DESCRIPTION_LIMIT = 4900
    DEFAULT_YOUTUBE_TAG_LIMIT = 100
    DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT = 450
    DEFAULT_YOUTUBE_MAX_TAGS_COUNT = 40
    DEFAULT_YOUTUBE_TITLE_LIMIT = 100
    def validate_description(desc: str, limit: int = DEFAULT_YOUTUBE_DESCRIPTION_LIMIT) -> Tuple[str, List[str]]:
        warnings = []; validated_desc = desc
        if len(desc) > limit: warnings.append(f"Desc truncated: {len(desc)}->{limit}"); validated_desc = desc[:limit]
        return validated_desc, warnings
    def validate_tags(tags: List[str], tag_char_limit: int = DEFAULT_YOUTUBE_TAG_LIMIT, total_char_limit: int = DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT, max_count_limit: int = DEFAULT_YOUTUBE_MAX_TAGS_COUNT) -> Tuple[List[str], List[str]]:
        warnings = []; validated_tags = []; current_total_chars = 0
        for tag in tags:
            if not isinstance(tag, str) or not tag.strip(): warnings.append(f"Skip invalid tag:'{tag}'"); continue
            cleaned_tag = tag.strip()
            if len(validated_tags) >= max_count_limit: warnings.append(f"Tag count limit {max_count_limit}. Skip:'{cleaned_tag[:20]}...'"); continue
            if len(cleaned_tag) > tag_char_limit: warnings.append(f"Tag truncated:{len(cleaned_tag)}->{tag_char_limit}:'{cleaned_tag[:20]}...'"); cleaned_tag = cleaned_tag[:tag_char_limit]
            potential_len = current_total_chars + len(cleaned_tag) + (1 if validated_tags else 0)
            if potential_len > total_char_limit: warnings.append(f"Total tag limit {total_char_limit}. Skip:'{cleaned_tag[:20]}...'"); continue
            validated_tags.append(cleaned_tag); current_total_chars += len(cleaned_tag) + (1 if len(validated_tags) > 1 else 0)
        return validated_tags, warnings
    def validate_title(title: str, limit: int = DEFAULT_YOUTUBE_TITLE_LIMIT, add_ellipsis: bool = True) -> tuple[str, list[str]]:
        warnings = []; validated_title = title
        if len(title) > limit: warnings.append(f"Title truncated: {len(title)}->{limit}"); validated_title = title[:limit]
        return validated_title, warnings
# --- End YouTube Limits Import ---

# --- NEW: Import constants from the new location ---
try:
    from .utils import constants # If uploader.py is a module in youtube_shorts
except ImportError:
    # Fallback if run as script directly from youtube_shorts folder
    try:
        from utils import constants
    except ImportError:
        # Absolute fallback if utils isn't directly importable (e.g. running from project root)
        # This might indicate a PYTHONPATH issue or that the script isn't being run as part of the package.
        # For now, assume it can find 'utils' if it's in the same directory level.
        print("CRITICAL: Could not import constants.py. Ensure it's in utils/ and runnable.")
        # Define a minimal set of constants here to avoid crashing, but this is not ideal.
        class MinimalConstants:
            # Base directories
            PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            CONFIG_DIR = os.path.join(PROJECT_ROOT, "config")
            DATA_DIR = os.path.join(PROJECT_ROOT, "data")
            LOGS_DIR = os.path.join(PROJECT_ROOT, "logs")
            OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")

            # File paths
            CONFIG_FILE_PATH = os.path.join(CONFIG_DIR, "config.txt")
            EXCEL_FILE_PATH = os.path.join(DATA_DIR, "shorts_data.xlsx")
            UPLOADER_LOG_FILE = os.path.join(LOGS_DIR, "uploader_log.txt")
            PERFORMANCE_METRICS_FILE = os.path.join(DATA_DIR, "performance_metrics.json")
            UPLOAD_CORRELATION_CACHE = os.path.join(DATA_DIR, "upload_correlation_cache.json")
            ANALYTICS_PEAK_TIMES_CACHE_FILE = os.path.join(DATA_DIR, "analytics_peak_times_cache.json")
            PLAYLIST_DATA_CACHE_FILE = os.path.join(DATA_DIR, "playlists_data_cache.json")

            # Output directories
            SHORTS_METADATA_DIR = os.path.join(OUTPUT_DIR, "shorts_metadata")
            SHORTS_DOWNLOADS_DIR = os.path.join(OUTPUT_DIR, "shorts_downloads")
            DEBUG_RECORDINGS_DIR = os.path.join(OUTPUT_DIR, "debug_recordings")

            # Authentication files
            CLIENT_SECRETS_FILE = os.path.join(DATA_DIR, "client_secret.json")
            TOKEN_FILE = os.path.join(DATA_DIR, "token.json")

            # OAuth scopes
            SCOPES = [
                "https://www.googleapis.com/auth/youtube.readonly",
                "https://www.googleapis.com/auth/yt-analytics.readonly",
                "https://www.googleapis.com/auth/youtube"
            ]

            # YouTube limits
            YOUTUBE_DESCRIPTION_LIMIT = 4950
            YOUTUBE_TAG_LIMIT = 100
            YOUTUBE_TOTAL_TAGS_LIMIT = 460
            YOUTUBE_MAX_TAGS_COUNT = 40

        constants = MinimalConstants()
        print("WARNING: Using minimal fallback constants due to import error.")
# --- End NEW Import ---

# --- Global Path Definitions (now derived from constants) ---
# Using constants directly instead of building paths from script_directory
script_directory = os.path.dirname(os.path.abspath(__file__))  # Keep for backward compatibility

# Input folders for manual workflow
INPUT_METADATA_FOLDER = constants.SHORTS_METADATA_DIR  # Use metadata folder directly
INPUT_VIDEO_FOLDER = constants.SHORTS_DOWNLOADS_DIR  # Use downloads folder directly

# File paths
EXCEL_FILE_PATH = constants.EXCEL_FILE_PATH
CONFIG_FILE_PATH = constants.CONFIG_FILE_PATH
ERROR_LOG_FILE = constants.UPLOADER_LOG_FILE
DEBUG_RECORDING_FOLDER = constants.DEBUG_RECORDINGS_DIR
PERFORMANCE_METRICS_FILE = constants.PERFORMANCE_METRICS_FILE
UPLOADER_ANALYSIS_LOG = os.path.join(constants.LOGS_DIR, "uploader_analysis_log.txt")

# Cache files
UPLOAD_CORRELATION_CACHE_PATH = constants.UPLOAD_CORRELATION_CACHE
# Use getattr with fallback for ANALYTICS_PEAK_TIMES_CACHE_FILE
ANALYTICS_PEAK_TIMES_CACHE_PATH = getattr(constants, 'ANALYTICS_PEAK_TIMES_CACHE_FILE',
                                         os.path.join(constants.DATA_DIR, "analytics_peak_times_cache.json"))
PLAYLIST_DATA_CACHE_PATH = constants.PLAYLIST_DATA_CACHE_FILE
# --- End Path Definitions ---

# --- YouTube API Authentication Constants ---
CLIENT_SECRETS_FILE = constants.CLIENT_SECRETS_FILE
TOKEN_FILE = constants.TOKEN_FILE
# Scopes for YouTube API access - includes Data API, Analytics API, and Playlist Management
SCOPES = constants.SCOPES
# --- End YouTube API Authentication Constants ---

# --- Error Types and Analysis Constants ---
ERROR_TYPES = {
    "title_input": "Finding/setting title input",
    "description_input": "Finding/setting description input",
    "show_more": "Finding/clicking 'Show more' button",
    "tags_input": "Finding/setting tags input",
    "next_button": "Finding/clicking 'Next' button",
    "public_radio": "Finding/selecting 'Public' radio button",
    "schedule_radio": "Finding/selecting 'Schedule' radio button",
    "date_picker": "Finding/setting date in calendar",
    "time_input": "Finding/setting time input",
    "publish_button": "Finding/clicking 'Publish' button",
    "schedule_button": "Finding/clicking 'Schedule' button",
    "confirmation": "Confirming upload completion / Getting ID", # Clarified
    "browser_session": "Browser session issues (crash, closed, unresponsive)", # Clarified
    "file_selection": "Selecting video file in OS dialog", # Added
    "file_upload": "Video file upload process", # Added
    "playlist_api": "Adding video to playlist via API", # Added
    "other": "Other/unclassified errors"
}
MAX_ERROR_SAMPLES = 50
MIN_ERRORS_FOR_ANALYSIS = 10
MIN_ERROR_RATE_FOR_ANALYSIS = 0.15
# --- End Error Types and Analysis Constants ---

# --- Global Variable for Active Recording Process (for cleanup) ---
_current_recording_process: Optional[subprocess.Popen] = None
_current_recording_filename: Optional[str] = None
# --- End Global Variable ---

# --- Logging Helper Functions (Copied from previous analysis) ---
def log_error_to_file(message: str, error_type: str = "other", step: str = "unknown", video_index: str = "UNKNOWN", xpath: str = "", include_traceback: bool = False):
    """Logs a detailed error message to the error log file (plain text) with additional context."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    full_message = f"[{timestamp}] {message}\n"
    if error_type != "other" or step != "unknown" or xpath:
        full_message += f"CONTEXT: Type={error_type}, Step={step}, VideoIndex={video_index}, XPath={xpath}\n"
    if include_traceback:
        try:
            exc_info = traceback.format_exc();
            if exc_info and exc_info.strip() != 'NoneType: None': full_message += exc_info + "\n"
        except Exception as e: full_message += f"[Error getting traceback: {e}]\n"
    try:
        # Ensure logs directory exists
        os.makedirs(os.path.dirname(ERROR_LOG_FILE), exist_ok=True)
        with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f: f.write(full_message)
        update_error_metrics(error_type, step, video_index, message, xpath)
    except (IOError, PermissionError) as e: print(f"CRITICAL: Failed to write to error log file '{ERROR_LOG_FILE}': {e}")
    except Exception as e: print(f"CRITICAL: Unexpected error writing to error log file '{ERROR_LOG_FILE}': {e}")

def load_performance_metrics():
    """Loads performance metrics from the JSON file."""
    default_metrics = {
        "total_uploads_attempted": 0, "total_uploads_successful": 0, "total_errors": 0,
        "error_counts": {error_type: 0 for error_type in ERROR_TYPES.keys()},
        "error_samples": [], "runs": [], "last_analysis_date": ""
    }
    try:
        if os.path.exists(PERFORMANCE_METRICS_FILE):
            with open(PERFORMANCE_METRICS_FILE, "r", encoding="utf-8") as f: metrics = json.load(f)
            for key, value in default_metrics.items(): metrics.setdefault(key, value)
            for error_type in ERROR_TYPES.keys(): metrics["error_counts"].setdefault(error_type, 0)
            return metrics
        else: return default_metrics
    except Exception as e: print(f"{Fore.YELLOW}Error loading performance metrics: {e}. Using default values."); return default_metrics

def save_performance_metrics(metrics):
    """Saves performance metrics to the JSON file."""
    try:
        # Prune runs if too long
        if len(metrics.get("runs", [])) > 50: # Keep last 50 runs max
            metrics["runs"] = metrics["runs"][-50:]
        # Ensure data directory exists
        os.makedirs(os.path.dirname(PERFORMANCE_METRICS_FILE), exist_ok=True)
        with open(PERFORMANCE_METRICS_FILE, "w", encoding="utf-8") as f: json.dump(metrics, f, ensure_ascii=False, indent=4)
    except Exception as e: print(f"{Fore.RED}Error saving performance metrics: {e}")

def update_error_metrics(error_type, step, video_index, error_message, xpath=""):
    """Updates the error metrics in the performance metrics file."""
    try:
        metrics = load_performance_metrics()
        metrics["total_errors"] += 1
        metrics["error_counts"][error_type] = metrics["error_counts"].get(error_type, 0) + 1
        error_sample = { "type": error_type, "step": step, "video_index": video_index, "message": error_message, "xpath": xpath, "timestamp": datetime.now().isoformat() }
        metrics["error_samples"].append(error_sample)
        metrics["error_samples"] = metrics["error_samples"][-MAX_ERROR_SAMPLES:] # Keep only last N samples
        save_performance_metrics(metrics)
    except Exception as e: print(f"{Fore.RED}Error updating error metrics: {e}")

def print_section_header(title: str): print(f"\n{Style.BRIGHT}{Fore.CYAN}--- {title} ---{Style.RESET_ALL}")
def print_info(message: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.DIM}{Fore.BLUE}i INFO:{Style.RESET_ALL} {message}")
def print_success(message: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.GREEN}OK SUCCESS:{Style.RESET_ALL} {Fore.GREEN}{message}{Style.RESET_ALL}")
def print_warning(message: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.YELLOW}WARN WARNING:{Style.RESET_ALL} {Fore.YELLOW}{message}{Style.RESET_ALL}")
def print_error(message: str, indent: int = 0, log_to_file: bool = True, include_traceback: bool = False, error_type: str = "other", step: str = "unknown", video_index: str = "UNKNOWN", xpath: str = ""):
    prefix = "  " * indent
    print(f"{prefix}{Style.BRIGHT}{Fore.RED}ERR ERROR:{Style.RESET_ALL} {Fore.RED}{message}{Style.RESET_ALL}")
    if log_to_file: log_error_to_file(f"ERROR: {message}", error_type=error_type, step=step, video_index=video_index, xpath=xpath, include_traceback=include_traceback)
def print_fatal(message: str, indent: int = 0, log_to_file: bool = True, include_traceback: bool = True):
    prefix = "  " * indent
    print(f"{prefix}{Style.BRIGHT}{Fore.RED}FATAL ERROR:{Style.RESET_ALL} {Fore.RED}{message}{Style.RESET_ALL}")
    if log_to_file: log_error_to_file(f"FATAL: {message}", include_traceback=include_traceback)
def print_config(key: str, value: any): print(f"  {Fore.MAGENTA}{key:<28}:{Style.RESET_ALL} {Style.BRIGHT}{value}{Style.RESET_ALL}")
# --- End Logging Helper Functions ---

# --- YouTube Analytics API Functions ---
def load_peak_times_cache() -> Optional[Dict]:
    """Loads peak times data from cache."""
    if not os.path.exists(ANALYTICS_PEAK_TIMES_CACHE_PATH): return None
    try:
        with open(ANALYTICS_PEAK_TIMES_CACHE_PATH, "r", encoding="utf-8") as f: data = json.load(f)
        if isinstance(data, dict) and "timestamp" in data and "peak_hours" in data: return data
        else: print_warning("Invalid format in peak times cache file."); return None
    except Exception as e: print_error(f"Error loading peak times cache: {e}"); return None

def save_peak_times_cache(peak_hours: List[int]):
    """Saves peak times data and timestamp to cache."""
    data = { "timestamp": datetime.now().isoformat(), "peak_hours": peak_hours }
    try:
        # Ensure data directory exists
        os.makedirs(os.path.dirname(ANALYTICS_PEAK_TIMES_CACHE_PATH), exist_ok=True)
        with open(ANALYTICS_PEAK_TIMES_CACHE_PATH, "w", encoding="utf-8") as f: json.dump(data, f, indent=4)
        print_success(f"Saved peak times to cache: {ANALYTICS_PEAK_TIMES_CACHE_PATH}")
    except Exception as e: print_error(f"Error saving peak times cache: {e}")

def get_peak_viewer_hours_from_api(service: Any, days_back: int, num_peak_hours: int) -> Optional[List[int]]:
    """Queries YouTube Analytics API to find the top N peak viewer hours."""
    if not service: print_error("Analytics API service not available for peak time query."); return None
    print_info(f"Querying YouTube Analytics for peak hours over last {days_back} days...")
    end_date = datetime.now().date() - timedelta(days=1); start_date = end_date - timedelta(days=days_back - 1)
    start_date_str = start_date.strftime('%Y-%m-%d'); end_date_str = end_date.strftime('%Y-%m-%d')
    try:
        analytics = build('youtubeAnalytics', 'v2', credentials=service._credentials) # Build analytics service
        response = analytics.reports().query(
            ids='channel==MINE', startDate=start_date_str, endDate=end_date_str,
            metrics='views', dimensions='hour', sort='-views', maxResults=24
        ).execute()
        if 'rows' not in response: print_warning("No data rows returned from Analytics API."); return None
        hourly_views = {}
        for row in response['rows']:
            try: hour = int(row[0]); views = int(row[1]); hourly_views[hour] = views
            except (IndexError, ValueError, TypeError): print_warning(f"Skipping invalid row: {row}"); continue
        if not hourly_views: print_warning("Could not parse valid hourly view data."); return None
        sorted_hours = sorted(hourly_views.items(), key=lambda item: item[1], reverse=True)
        peak_hours = [hour for hour, views in sorted_hours[:num_peak_hours]]; peak_hours.sort()
        print_success(f"Identified top {len(peak_hours)} peak hours: {peak_hours}")
        return peak_hours
    except HttpError as e: print_error(f"YouTube Analytics API error: {e}", include_traceback=True); log_error_to_file(f"Analytics API HttpError: {e}", include_traceback=True); return None
    except Exception as e: print_error(f"Unexpected error querying Analytics API: {e}", include_traceback=True); log_error_to_file(f"Unexpected error querying Analytics API: {e}", include_traceback=True); return None
# --- End YouTube Analytics API Functions ---

# --- Playlist Management Functions ---
def load_playlist_cache() -> Dict:
    """Loads the playlist data cache from JSON file."""
    default_cache = {"timestamp": datetime.now().isoformat()}
    try:
        if os.path.exists(PLAYLIST_DATA_CACHE_PATH):
            with open(PLAYLIST_DATA_CACHE_PATH, "r", encoding="utf-8") as f: content = f.read()
            if not content: print_info("Playlist cache file empty."); return default_cache
            cache = json.loads(content)
            if not isinstance(cache, dict): print_warning(f"Playlist cache file invalid format."); return default_cache
            return cache
        else: print_info(f"Playlist cache file not found."); return default_cache
    except json.JSONDecodeError: print_error(f"Error decoding playlist cache."); return default_cache
    except Exception as e: print_error(f"Error loading playlist cache: {e}", include_traceback=True); return default_cache

def save_playlist_cache(cache_data: Dict):
    """Saves the playlist data cache to JSON file."""
    try:
        # Ensure data directory exists
        os.makedirs(os.path.dirname(PLAYLIST_DATA_CACHE_PATH), exist_ok=True)
        with open(PLAYLIST_DATA_CACHE_PATH, "w", encoding="utf-8") as f: json.dump(cache_data, f, ensure_ascii=False, indent=4)
        print_info(f"Saved playlist cache with {len(cache_data) - 1} entries.") # -1 for timestamp
    except Exception as e: print_error(f"Error saving playlist cache: {e}", include_traceback=True)

def fetch_channel_playlists(service: Any) -> Dict[str, str]:
    """Fetches existing playlists for the authenticated user's channel."""
    if not service: print_warning("YouTube API service not available, cannot fetch playlists."); return {}
    playlists_map = {}; next_page_token = None
    print_info("Fetching existing channel playlists via YouTube API...")
    try:
        while True:
            request = service.playlists().list(part="snippet", mine=True, maxResults=50, pageToken=next_page_token); response = request.execute()
            for item in response.get("items", []):
                playlist_id = item.get("id"); playlist_title = item.get("snippet", {}).get("title")
                if playlist_id and playlist_title: playlists_map[playlist_id] = playlist_title
            next_page_token = response.get("nextPageToken")
            if not next_page_token: break; time.sleep(0.5)
        print_success(f"Fetched {len(playlists_map)} existing playlists.")
        return playlists_map
    except HttpError as e: print_error(f"API Error fetching playlists: {e}", include_traceback=True); log_error_to_file(f"API Error fetching playlists: {e}", include_traceback=True); return playlists_map
    except Exception as e: print_error(f"Unexpected error fetching playlists: {e}", include_traceback=True); log_error_to_file(f"Unexpected error fetching playlists: {e}", include_traceback=True); return playlists_map

def create_playlist(service: Any, title: str, description: str = "") -> Optional[str]:
    """Creates a new private playlist and returns its ID."""
    if not service or not title: print_warning("Missing service or title for creating playlist."); return None
    print_info(f"Attempting to create new playlist via API: '{title}'", 3)
    try:
        request = service.playlists().insert(part="snippet,status", body={"snippet": {"title": title, "description": description, "defaultLanguage": "en"}, "status": {"privacyStatus": "private"}})
        response = request.execute(); playlist_id = response.get("id")
        if playlist_id: print_success(f"Successfully created playlist '{title}' with ID: {playlist_id}", 4); return playlist_id
        else: print_error(f"Failed to create playlist '{title}'. No ID returned.", 4); return None
    except HttpError as e: print_error(f"API Error creating playlist '{title}': {e}", 4); log_error_to_file(f"API Error creating playlist '{title}': {e}", include_traceback=True); return None
    except Exception as e: print_error(f"Unexpected error creating playlist '{title}': {e}", 4, include_traceback=True); log_error_to_file(f"Unexpected error creating playlist '{title}': {e}", include_traceback=True); return None

def add_video_to_playlist(service: Any, video_id: str, playlist_id: str) -> bool:
    """Adds a video to a specified playlist using YouTube Data API."""
    if not service or not video_id or not playlist_id: print_warning("Missing service, video ID, or playlist ID for adding to playlist."); return False
    try:
        request = service.playlistItems().insert(part="snippet", body={"snippet": {"playlistId": playlist_id, "resourceId": {"kind": "youtube#video", "videoId": video_id}}})
        response = request.execute()
        print_success(f"Successfully added video {video_id} to playlist {playlist_id}.", 4)
        return True
    except HttpError as e:
        error_details = e.resp.get('content', b'').decode('utf-8')
        if e.resp.status == 404:
             if "playlistNotFound" in error_details: print_warning(f"Playlist ID '{playlist_id}' not found.", 4)
             elif "videoNotFound" in error_details: print_warning(f"Video ID '{video_id}' not found.", 4)
             else: print_error(f"API Error 404 adding video {video_id} to playlist {playlist_id}: {e}", 4)
        elif e.resp.status == 403: print_error(f"Permission denied adding video {video_id} to playlist {playlist_id}. Check API scopes/permissions. Error: {e}", 4)
        else: print_error(f"API Error adding video {video_id} to playlist {playlist_id}: {e}", 4)
        log_error_to_file(f"API Error adding video {video_id} to playlist {playlist_id}: {e}", include_traceback=True)
        return False
    except Exception as e: print_error(f"Unexpected error adding video {video_id} to playlist {playlist_id}: {e}", 4, include_traceback=True); log_error_to_file(f"Unexpected error adding video {video_id} to playlist {playlist_id}: {e}", include_traceback=True); return False
# --- End Playlist Management Functions ---

# --- YouTube API Authentication Function ---
def get_authenticated_service() -> Optional[Any]:
    """Authenticates with YouTube API and returns a service object."""
    if not GOOGLE_API_AVAILABLE:
        print_warning("Google API libraries not available. Cannot authenticate.")
        return None

    try:
        # Ensure data directory exists for token
        os.makedirs(os.path.dirname(TOKEN_FILE), exist_ok=True)

        creds = None
        # The file token.json stores the user's access and refresh tokens
        if os.path.exists(TOKEN_FILE):
            try:
                with open(TOKEN_FILE, 'rb') as token:
                    creds = pickle.load(token)
            except Exception as e:
                print_warning(f"Error loading token file: {e}")

        # If there are no valid credentials available, let the user log in
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())
                except Exception as e:
                    print_error(f"Error refreshing credentials: {e}")
                    # If refresh fails, force re-authentication
                    creds = None

            if not creds:
                if not os.path.exists(CLIENT_SECRETS_FILE):
                    print_error(f"Client secrets file not found at: {CLIENT_SECRETS_FILE}")
                    return None

                try:
                    flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
                    creds = flow.run_local_server(port=0)
                except Exception as e:
                    print_error(f"Error during authentication flow: {e}", include_traceback=True)
                    return None

            # Save the credentials for the next run
            try:
                with open(TOKEN_FILE, 'wb') as token:
                    pickle.dump(creds, token)
                print_success("Saved authentication credentials to token file.")
            except Exception as e:
                print_warning(f"Error saving credentials: {e}")

        # Build and return the YouTube service
        try:
            service = build('youtube', 'v3', credentials=creds)
            print_success("Successfully authenticated with YouTube API.")
            return service
        except Exception as e:
            print_error(f"Error building YouTube service: {e}", include_traceback=True)
            return None

    except Exception as e:
        print_error(f"Unexpected error in authentication: {e}", include_traceback=True)
        return None
# --- End YouTube API Authentication Function ---

# --- Configuration Loading (Moved earlier for global use) ---
config = {}
try:
    print_info(f"Loading configuration from: {CONFIG_FILE_PATH}")
    with open(CONFIG_FILE_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and "=" in line:
                key, value = line.split("=", 1)
                config[key.strip()] = value.strip()
    print_success("Configuration loaded.")
except FileNotFoundError: print_fatal(f"Configuration file '{CONFIG_FILE_PATH}' not found. Cannot continue.", log_to_file=False); raise
except Exception as e: print_fatal(f"Error reading configuration file '{CONFIG_FILE_PATH}': {e}. Cannot continue.", log_to_file=False); raise
# --- End Configuration Loading ---

# --- Get Configurable Settings (Moved earlier for global use) ---
# General settings
_DEFAULT_MAX_UPLOADS = 25
_DEFAULT_CATEGORY = "Gaming"
try: max_uploads = int(config.get("MAX_UPLOADS", _DEFAULT_MAX_UPLOADS)); assert max_uploads > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid MAX_UPLOADS in config. Using default: {_DEFAULT_MAX_UPLOADS}"); max_uploads = _DEFAULT_MAX_UPLOADS
upload_category = config.get("UPLOAD_CATEGORY", _DEFAULT_CATEGORY).strip()
profile_path_config = config.get("PROFILE_PATH")
# Scheduling Mode Settings
_DEFAULT_SCHEDULING_MODE = "analytics_priority"
_DEFAULT_SCHEDULE_INTERVAL = 120
_DEFAULT_MIN_SCHEDULE_AHEAD = 20
_DEFAULT_CUSTOM_TIMES_STR = "9:00 AM, 3:00 PM"
scheduling_mode = config.get("SCHEDULING_MODE", _DEFAULT_SCHEDULING_MODE).strip().lower()
if scheduling_mode not in ["default_interval", "custom_tomorrow", "analytics_priority"]: print_warning(f"Invalid SCHEDULING_MODE '{scheduling_mode}'. Using default: '{_DEFAULT_SCHEDULING_MODE}'"); scheduling_mode = _DEFAULT_SCHEDULING_MODE
try: schedule_interval_minutes = int(config.get("SCHEDULE_INTERVAL_MINUTES", _DEFAULT_SCHEDULE_INTERVAL)); assert schedule_interval_minutes > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid SCHEDULE_INTERVAL_MINUTES. Using default: {_DEFAULT_SCHEDULE_INTERVAL}"); schedule_interval_minutes = _DEFAULT_SCHEDULE_INTERVAL
custom_schedule_times_str = config.get("CUSTOM_SCHEDULE_TIMES", _DEFAULT_CUSTOM_TIMES_STR).strip(); parsed_config_times: List[dt_time] = []
if scheduling_mode == 'custom_tomorrow' and custom_schedule_times_str:
    for time_str in [t.strip() for t in custom_schedule_times_str.split(',') if t.strip()]:
        try: parsed_config_times.append(datetime.strptime(time_str, "%I:%M %p").time())
        except ValueError:
            try: parsed_config_times.append(datetime.strptime(time_str, "%H:%M").time()); print_warning(f"Parsed custom time '{time_str}' using 24h fmt.", 1)
            except ValueError: print_warning(f"Invalid format '{time_str}'. Skip.", 1)
    if parsed_config_times: print_success(f"Parsed {len(parsed_config_times)} custom schedule times.", 1); parsed_config_times.sort()
    else: print_warning("No valid custom times. 'custom_tomorrow' will use fallback interval.", 1)
elif scheduling_mode == 'custom_tomorrow': print_info("'CUSTOM_SCHEDULE_TIMES' empty. 'custom_tomorrow' will use fallback.", 1)
try: min_schedule_ahead_minutes = int(config.get("MIN_SCHEDULE_AHEAD_MINUTES", _DEFAULT_MIN_SCHEDULE_AHEAD)); assert min_schedule_ahead_minutes >= 5
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid MIN_SCHEDULE_AHEAD_MINUTES. Using default: {_DEFAULT_MIN_SCHEDULE_AHEAD}"); min_schedule_ahead_minutes = _DEFAULT_MIN_SCHEDULE_AHEAD
# Analytics-Based Scheduling Settings
_DEFAULT_ANALYTICS_DAYS = 7; _DEFAULT_ANALYTICS_PEAK_HOURS = 5; _DEFAULT_ANALYTICS_CACHE_HOURS = 24
enable_analytics_scheduling = True if scheduling_mode == 'analytics_priority' else config.get("ENABLE_ANALYTICS_SCHEDULING", "false").strip().lower() == 'true'
try: analytics_days_to_analyze = int(config.get("ANALYTICS_DAYS_TO_ANALYZE", _DEFAULT_ANALYTICS_DAYS)); assert analytics_days_to_analyze > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid ANALYTICS_DAYS_TO_ANALYZE. Using default: {_DEFAULT_ANALYTICS_DAYS}"); analytics_days_to_analyze = _DEFAULT_ANALYTICS_DAYS
try: analytics_peak_hours_count = int(config.get("ANALYTICS_PEAK_HOURS_COUNT", _DEFAULT_ANALYTICS_PEAK_HOURS)); assert 0 < analytics_peak_hours_count <= 24
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid ANALYTICS_PEAK_HOURS_COUNT. Using default: {_DEFAULT_ANALYTICS_PEAK_HOURS}"); analytics_peak_hours_count = _DEFAULT_ANALYTICS_PEAK_HOURS
try: analytics_cache_expiry_hours = int(config.get("ANALYTICS_CACHE_EXPIRY_HOURS", _DEFAULT_ANALYTICS_CACHE_HOURS)); assert analytics_cache_expiry_hours > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid ANALYTICS_CACHE_EXPIRY_HOURS. Using default: {_DEFAULT_ANALYTICS_CACHE_HOURS}"); analytics_cache_expiry_hours = _DEFAULT_ANALYTICS_CACHE_HOURS
# Debug Recording Settings
enable_debug_recording = config.get("ENABLE_DEBUG_RECORDING", "False").strip().lower() == 'true'
ffmpeg_path_config = config.get("FFMPEG_PATH", "ffmpeg").strip() # Default to 'ffmpeg'
# Gemini API Configuration
gemini_api_key = config.get("GEMINI_API_KEY", "").strip()
if genai and gemini_api_key:
    try: genai.configure(api_key=gemini_api_key); print_success("Gemini API configured successfully.")
    except Exception as e: print_warning(f"Failed to configure Gemini API: {e}. Self-improvement features disabled.")
elif genai: print_warning("GEMINI_API_KEY not found. Self-improvement disabled."); print_info("Add GEMINI_API_KEY=... to config.txt")
# Read YouTube Limits from Config
try: cfg_desc_limit = int(config.get("YOUTUBE_DESCRIPTION_LIMIT", DEFAULT_YOUTUBE_DESCRIPTION_LIMIT)); assert cfg_desc_limit > 0
except: print_warning(f"Invalid YOUTUBE_DESCRIPTION_LIMIT. Using default: {DEFAULT_YOUTUBE_DESCRIPTION_LIMIT}"); cfg_desc_limit = DEFAULT_YOUTUBE_DESCRIPTION_LIMIT
try: cfg_tag_limit = int(config.get("YOUTUBE_TAG_LIMIT", DEFAULT_YOUTUBE_TAG_LIMIT)); assert cfg_tag_limit > 0
except: print_warning(f"Invalid YOUTUBE_TAG_LIMIT. Using default: {DEFAULT_YOUTUBE_TAG_LIMIT}"); cfg_tag_limit = DEFAULT_YOUTUBE_TAG_LIMIT
try: cfg_total_tags_limit = int(config.get("YOUTUBE_TOTAL_TAGS_LIMIT", DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT)); assert cfg_total_tags_limit > 0
except: print_warning(f"Invalid YOUTUBE_TOTAL_TAGS_LIMIT. Using default: {DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT}"); cfg_total_tags_limit = DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT
try: cfg_max_tags_count = int(config.get("YOUTUBE_MAX_TAGS_COUNT", DEFAULT_YOUTUBE_MAX_TAGS_COUNT)); assert cfg_max_tags_count > 0
except: print_warning(f"Invalid YOUTUBE_MAX_TAGS_COUNT. Using default: {DEFAULT_YOUTUBE_MAX_TAGS_COUNT}"); cfg_max_tags_count = DEFAULT_YOUTUBE_MAX_TAGS_COUNT
# --- End Configurable Settings ---


# --- Helper Functions ---
def mimic_human_action_delay(min_sec=0.3, max_sec=0.8): time.sleep(random.uniform(min_sec, max_sec))
# --- End Helper Functions ---


# --- WebDriver Configuration ---
def configure_driver() -> Optional[webdriver.Firefox]:
    print_info("Configuring Firefox WebDriver...")
    firefox_options = FirefoxOptions()
    firefox_options.add_argument("--start-maximized")
    firefox_options.add_argument("--disable-notifications")
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0" # Example
    firefox_options.set_preference("general.useragent.override", user_agent)
    firefox_options.set_preference("media.peerconnection.enabled", False); firefox_options.set_preference("geo.enabled", False)
    firefox_options.add_argument('--disable-blink-features=AutomationControlled')

    if profile_path_config:
        profile_path = profile_path_config.strip()
        if os.path.isdir(profile_path):
            try: firefox_options.profile = profile_path; print_info(f"Using Firefox profile: {profile_path}", 1)
            except Exception as e: print_warning(f"Error setting profile '{profile_path}'. Default used. Error: {e}", 1); log_error_to_file(f"Warning: {msg}")
        else: print_warning(f"PROFILE_PATH '{profile_path}' not found. Default used.", 1); log_error_to_file(f"Warning: {msg}")
    else: print_info("Using default Firefox profile (PROFILE_PATH not set).", 1)

    driver = None
    try:
        print_info("Setting up GeckoDriver using webdriver-manager...", 1)
        geckodriver_log_path = os.path.join(script_directory, "geckodriver.log")
        try: service = FirefoxService(executable_path=GeckoDriverManager().install(), log_path=geckodriver_log_path)
        except Exception as e: print_warning(f"Could not set geckodriver log path '{geckodriver_log_path}': {e}. Default path used.", 2); service = FirefoxService(executable_path=GeckoDriverManager().install())
        driver = webdriver.Firefox(service=service, options=firefox_options)
        print_success("WebDriver setup complete.", 1); print_info(f"GeckoDriver log: {geckodriver_log_path}", 2)
        return driver
    except WebDriverException as e:
        if "connection refused" in str(e).lower(): msg = f"WebDriver setup failed: Connection Refused. Check Firefox/firewall. Error: {e}"
        elif "binary" in str(e).lower(): msg = f"WebDriver setup failed: Firefox Binary issue. Check install. Error: {e}"
        else: msg = f"WebDriver setup failed: {e}\n{Fore.YELLOW}       Check Firefox install, network, compatibility.{Style.RESET_ALL}"
        print_error(msg, 1); log_error_to_file(f"WebDriver setup failed: {e}", include_traceback=True); raise
    except Exception as e: msg = f"Unexpected error during WebDriver setup: {e}"; print_error(msg, 1); log_error_to_file(f"Unexpected error during WebDriver setup: {e}", include_traceback=True); raise
# --- End WebDriver Configuration ---


# --- Excel Update Function ---
def update_excel_data(downloaded_sheet: Any, uploaded_sheet: Any, video_index_str: str, optimized_title: str, upload_time: datetime, schedule_time: Optional[datetime] = None, publish_status: str = "Published", youtube_video_id: Optional[str] = None):
    """Moves entry from Downloaded to Uploaded sheet and adds YT ID/timestamps."""
    print_info(f"Updating Excel data in memory for video index: {video_index_str}", indent=1)
    try:
        # --- Find and remove from Downloaded sheet (handle case where sheet might be None) ---
        rows_to_delete = []; row_found_in_downloaded = False
        if downloaded_sheet: # Check if the sheet object is valid
            for row_idx in range(downloaded_sheet.max_row, 1, -1): # Iterate backwards from max_row
                cell = downloaded_sheet.cell(row=row_idx, column=1) # Assume Video Index is column 1
                # Safer comparison: handle None values and convert to string for comparison
                if cell.value is not None and str(cell.value).strip().lower() == f"video{video_index_str}".lower():
                     rows_to_delete.append(row_idx); row_found_in_downloaded = True
            if row_found_in_downloaded:
                for row_idx in sorted(rows_to_delete, reverse=True): # Sort to delete from bottom up
                    downloaded_sheet.delete_rows(row_idx)
                print_success(f"Removed entry for video{video_index_str} from 'Downloaded' sheet.", indent=2)
            else: print_warning(f"Could not find video{video_index_str} in 'Downloaded' sheet to remove.", indent=2)
        else: print_warning("'Downloaded' sheet object is invalid/None. Cannot remove entry.", indent=2)

        # --- Add to Uploaded sheet (handle case where sheet might be None) ---
        if uploaded_sheet: # Check if the sheet object is valid
            # Ensure consistent formatting for Excel
            upload_time_str = upload_time.strftime("%Y-%m-%d %H:%M:%S")
            # --- FIX APPLIED HERE ---
            # If schedule_time is None (published immediately), use upload_time_str for the schedule column
            schedule_time_for_excel_str = schedule_time.strftime("%Y-%m-%d %H:%M:%S") if schedule_time else upload_time_str
            # --- END FIX ---
            # Include YouTube Video ID in the row data if available
            row_data = [ f"video{video_index_str}", optimized_title, youtube_video_id if youtube_video_id else "N/A", upload_time_str, schedule_time_for_excel_str, publish_status ]
            uploaded_sheet.append(row_data)
            print_success(f"Appended entry to 'Uploaded' sheet (Status: {publish_status}, Schedule: {schedule_time_for_excel_str}, YT ID: {youtube_video_id if youtube_video_id else 'N/A'}).", indent=2)
        else: print_error(f"Cannot append to 'Uploaded' sheet - sheet object is invalid/None.", indent=2)
    except AttributeError as ae: msg = f"Excel sheet object missing expected methods (maybe None or wrong type?): {ae}"; print_error(msg, indent=1, include_traceback=True)
    except Exception as e: msg = f"Error updating Excel data in memory for video{video_index_str}: {e}"; print_error(msg, indent=1, include_traceback=True)
# --- End Excel Update Function ---


# --- Correlation Cache Functions (Copied from downloader) ---
# (load_correlation_cache, save_correlation_cache, add_to_correlation_cache - kept same)
def load_correlation_cache():
    default_cache = []
    if not os.path.exists(UPLOAD_CORRELATION_CACHE_PATH): return default_cache
    try:
        with open(UPLOAD_CORRELATION_CACHE_PATH, "r", encoding="utf-8") as f: content = f.read()
        if not content: return default_cache
        cache = json.loads(content)
        if not isinstance(cache, list): print_warning(f"Correlation cache invalid format."); return default_cache
        return cache
    except json.JSONDecodeError: print_error(f"Error decoding correlation cache."); return default_cache
    except Exception as e: print_error(f"Error loading correlation cache: {e}"); return default_cache

def save_correlation_cache(cache_data):
    try:
        # Ensure data directory exists
        os.makedirs(os.path.dirname(UPLOAD_CORRELATION_CACHE_PATH), exist_ok=True)
        with open(UPLOAD_CORRELATION_CACHE_PATH, "w", encoding="utf-8") as f: json.dump(cache_data, f, ensure_ascii=False, indent=4)
    except Exception as e: print_error(f"Error saving correlation cache: {e}")

def add_to_correlation_cache(video_index_str: str, discovery_keyword: Optional[str], youtube_video_id: str):
    if not youtube_video_id: return # Don't add if upload failed
    cache = load_correlation_cache()
    new_entry = { "video_index": video_index_str, "discovery_keyword": discovery_keyword if discovery_keyword else "Unknown", "youtube_video_id": youtube_video_id, "added_timestamp": datetime.now().isoformat() }
    cache.append(new_entry)
    save_correlation_cache(cache)
    print_info(f"Added {video_index_str} (YT ID: {youtube_video_id}) to correlation cache.", 2)
# --- End Correlation Cache Functions ---


# --- Check and Update Scheduled Status Function ---
def check_and_update_scheduled(uploaded_sheet: Any) -> bool: # Use Any type hint for sheet
    """Checks 'Uploaded' sheet and updates 'Scheduled' to 'Published' if time has passed."""
    print_info("Checking status of previously scheduled videos...")
    updated_count = 0; changes_made = False; now = datetime.now()
    if not uploaded_sheet or uploaded_sheet.max_row < 2: print_info("Uploaded sheet empty or only headers. No check needed.", 1); return False

    # --- Column Mapping (More Robust) ---
    header = [str(cell.value).lower().strip() if cell.value else '' for cell in uploaded_sheet[1]]
    col_indices = {name: i+1 for i, name in enumerate(header) if name} # Map lower header to 1-based index

    schedule_time_col_idx = col_indices.get('schedule time') # Correct target
    publish_status_col_idx = col_indices.get('publish status') # Correct target
    video_id_col_idx = col_indices.get('video index') # Use Video Index for logging

    # Validate that required columns were found
    if not schedule_time_col_idx: print_warning("'Schedule Time' column not found. Cannot check scheduled status.", 1); return False
    if not publish_status_col_idx: print_warning("'Publish Status' column not found. Cannot check/update scheduled status.", 1); return False
    if not video_id_col_idx: print_warning("'Video Index' column not found. Status check logging will use row number.", 1)
    # --- End Column Mapping ---

    for row_idx in range(2, uploaded_sheet.max_row + 1):
        try:
            schedule_time_cell = uploaded_sheet.cell(row=row_idx, column=schedule_time_col_idx)
            publish_status_cell = uploaded_sheet.cell(row=row_idx, column=publish_status_col_idx)
            video_id_str = str(uploaded_sheet.cell(row=row_idx, column=video_id_col_idx).value) if video_id_col_idx else f"Row {row_idx}"

            # Check if the status is exactly "Scheduled" (case-insensitive check)
            if isinstance(publish_status_cell.value, str) and publish_status_cell.value.strip().lower() == "scheduled" and schedule_time_cell.value:
                schedule_time_value = schedule_time_cell.value
                schedule_time = None # Reset for each row

                try:
                    # --- Use Robust Date Parsing ---
                    if EXCEL_UTILS_AVAILABLE:
                        schedule_time = parse_date_value(schedule_time_value)
                    elif DATEUTIL_AVAILABLE:
                        if isinstance(schedule_time_value, datetime): schedule_time = schedule_time_value
                        else: schedule_time = date_parser.parse(str(schedule_time_value), fuzzy=True)
                    else: # Fallback
                        if isinstance(schedule_time_value, datetime): schedule_time = schedule_time_value
                        elif isinstance(schedule_time_value, float): schedule_time = datetime.fromtimestamp(time.mktime(time.gmtime((schedule_time_value - 25569) * 86400.0)))
                        else: schedule_time = datetime.strptime(str(schedule_time_value), "%Y-%m-%d %H:%M:%S")
                    # --- End Robust Date Parsing ---

                    if schedule_time and now >= schedule_time:
                        publish_status_cell.value = "Published" # Update the status cell
                        print_success(f"Updated status for {video_id_str} to 'Published'.", 2)
                        updated_count += 1
                        changes_made = True
                except (ValueError, TypeError) as e: print_warning(f"Could not parse schedule time '{schedule_time_value}' for {video_id_str}: {e}", 2); log_error_to_file(f"Warning: Parse schedule time {video_id_str}: {e}")
        except Exception as e: video_id_str = str(uploaded_sheet.cell(row=row_idx, column=video_id_col_idx).value) if video_id_col_idx else f"Row {row_idx}"; msg = f"Unexpected error checking schedule status for {video_id_str}: {e}"; print_error(msg, 2); log_error_to_file(msg, include_traceback=True)
    if updated_count > 0: print_info(f"Checked scheduled videos. Updated status for {updated_count} videos.", 1)
    else: print_info("Checked scheduled videos. No status updates needed based on current time.", 1)
    return changes_made
# --- End Check and Update Scheduled Status Function ---


# --- select_date_in_calendar (Legacy Fallback) ---
def select_date_in_calendar(driver, schedule_time):
    """Selects the date in the calendar using Backspace clearing + Fallback XPaths."""
    print_info("Selecting date in calendar (using old Backspace method + Fallbacks)...", 3)
    try:
        time.sleep(1) # Allow calendar to potentially render
        date_input_container_xpaths = [ "//tp-yt-paper-dialog[@id='dialog']//tp-yt-paper-input[@id='textbox']", ]
        date_input_xpaths = [ ".//input", "//tp-yt-paper-dialog[@id='dialog']//input[contains(@class,'tp-yt-paper-input')]", ]
        date_input = None; container_found = False
        # Try finding within container first
        for container_xpath in date_input_container_xpaths:
             try:
                 date_input_container = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, container_xpath)))
                 for input_xpath_rel in [".//input", ".//input[@id='input']"]:
                     try: date_input = date_input_container.find_element(By.XPATH, input_xpath_rel); container_found = True; break
                     except NoSuchElementException: continue
                 if date_input: break
             except Exception: continue
        # If not found in container, try direct XPaths
        if not date_input:
             print_info("Could not find date input via container, trying direct XPaths...", 4)
             for xpath in date_input_xpaths:
                 try: date_input = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath))); break
                 except Exception: continue
        # If still not found after all attempts
        if not date_input: raise TimeoutException("Failed to find Date input field with any provided XPath.")

        # Format date (OLD STYLE - might be incompatible with newer UI)
        date_string = schedule_time.strftime("%b %d, %Y")
        print_info(f"Formatted date string (old format): '{date_string}'", 4)
        print_info("Clearing date input using Backspace x 15...", 4)
        # Clicking before clearing might help focus
        try: date_input.click(); time.sleep(0.1)
        except: print_warning("Could not click date input before clearing.", 5)
        for _ in range(15): date_input.send_keys(Keys.BACKSPACE); time.sleep(0.05) # Shorter delay
        time.sleep(0.2); print_info("Backspace clearing complete.", 4)
        print_info(f"Sending date keys: '{date_string}'", 4); date_input.send_keys(date_string); time.sleep(0.5)
        print_info("Sending Enter key for date...", 4); date_input.send_keys(Keys.ENTER); time.sleep(1) # Wait after Enter
        print_success("Date selected using old method.", 4); return True
    except TimeoutException as e: msg = f"Timeout finding date input element using old XPaths + Fallbacks: {e}"; print_error(msg, 3, error_type="date_picker", step="select_date"); return False
    except Exception as e: msg = f"Error selecting date in calendar (using old method + Fallbacks): {e}"; print_error(msg, 3, include_traceback=True, error_type="date_picker", step="select_date"); return False
# --- End select_date_in_calendar ---


# --- upload_video (Legacy Fallback - kept for compatibility) ---
def upload_video(
    driver: webdriver.Firefox,
    video_file: str,
    metadata: dict,
    publish_now: bool = True,
    schedule_time: Optional[datetime] = None,
    desc_limit: int = DEFAULT_YOUTUBE_DESCRIPTION_LIMIT,
    tag_char_limit: int = DEFAULT_YOUTUBE_TAG_LIMIT,
    total_char_limit: int = DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT,
    max_count_limit: int = DEFAULT_YOUTUBE_MAX_TAGS_COUNT,
    service: Optional[Any] = None # Added service parameter for playlist management
) -> Optional[str]:
    """Handles the video upload process on YouTube Studio (LEGACY IMPLEMENTATION)."""
    video_index = metadata.get('video_index', 'UNKNOWN')
    print_section_header(f"Uploading Video Index: {video_index} (Using Legacy Function)")
    print_info(f"File: {os.path.basename(video_file)}", 1)
    print_info(f"Title: {metadata.get('optimized_title', 'N/A')}", 1)

    # Validate metadata (uses imported or fallback functions)
    print_info("Validating metadata...", 1)
    validated_desc, desc_warnings = validate_description(metadata.get('optimized_description', ''), limit=desc_limit)
    validated_tags, tag_warnings = validate_tags(metadata.get('optimized_tags', []), tag_char_limit, total_char_limit, max_count_limit)
    metadata['optimized_description'] = validated_desc # Update metadata dict
    metadata['optimized_tags'] = validated_tags # Update metadata dict
    if desc_warnings or tag_warnings:
        print_warning("Metadata validation finished with warnings:", 2);
        for w in desc_warnings: print_warning(w, 3)
        for w in tag_warnings: print_warning(w, 3)
    else: print_success("Metadata validated successfully.", 2)

    upload_successful = False
    youtube_video_id: Optional[str] = None

    try:
        print_info("Navigating to YouTube Studio...", 1); studio_url = "https://studio.youtube.com/"; driver.get(studio_url)
        wait_long = WebDriverWait(driver, 60); wait_medium = WebDriverWait(driver, 30); wait_short = WebDriverWait(driver, 15)

        create_button_selector = "ytcp-button#create-icon, yt-icon-button#create-icon-button"; create_button = wait_long.until(EC.element_to_be_clickable((By.CSS_SELECTOR, create_button_selector))); create_button.click(); mimic_human_action_delay(0.2, 0.5)
        upload_videos_selector = "tp-yt-paper-item#text-item-0"; upload_button = wait_short.until(EC.element_to_be_clickable((By.CSS_SELECTOR, upload_videos_selector))); upload_button.click(); mimic_human_action_delay(0.5, 1.0)

        file_input_xpath = "//input[@type='file']"; file_input = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, file_input_xpath))); abs_video_path = os.path.abspath(video_file); print_info(f"Selecting file: {abs_video_path}", 1)
        if not os.path.exists(abs_video_path): print_error(f"Video file missing: {abs_video_path}", 1); return None
        file_input.send_keys(abs_video_path)

        # --- Simplified Legacy Element Interaction (More Prone to Breaking) ---
        # --- Title ---
        print_info("Setting title (legacy)...", 2)
        title_xpath = "//ytcp-mention-textbox[@label='Title (required)']//div[@id='textbox']" # Common legacy path
        try: title_input = wait_long.until(EC.element_to_be_clickable((By.XPATH, title_xpath))); title_input.click(); title_input.send_keys(Keys.CONTROL + "a"); title_input.send_keys(Keys.DELETE); title_input.send_keys(metadata.get('optimized_title', f'Video {video_index}')); print_success("Title set.", 3)
        except Exception as e: print_error(f"Failed legacy title set: {e}", 3); raise # Fail fast on legacy errors

        # --- Description ---
        print_info("Setting description (legacy - JS)...", 2)
        desc_xpath = "//ytcp-mention-textbox[@label='Description']//div[@id='textbox']" # Common legacy path
        try: desc_input = wait_medium.until(EC.presence_of_element_located((By.XPATH, desc_xpath))); driver.execute_script("arguments[0].textContent = arguments[1];", desc_input, metadata.get('optimized_description', '')); driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", desc_input); print_success("Description set (JS).", 3)
        except Exception as e: print_error(f"Failed legacy description set: {e}", 3); raise

        # --- Show More ---
        print_info("Clicking 'Show more' (legacy)...", 2)
        show_more_xpath = "//ytcp-button[@id='toggle-button']//div[contains(text(), 'Show more')]"
        try: show_more_button = wait_medium.until(EC.element_to_be_clickable((By.XPATH, show_more_xpath))); show_more_button.click(); print_success("'Show more' clicked.", 3); time.sleep(1)
        except Exception as e: print_warning(f"Could not click 'Show more' (might be expanded): {e}", 3) # Don't fail if 'Show more' isn't needed

        # --- Tags ---
        print_info("Adding tags (legacy)...", 2)
        tags_xpath = "//input[contains(@aria-label, 'Tags')]"
        try: tags_input = wait_medium.until(EC.element_to_be_clickable((By.XPATH, tags_xpath)))
        except Exception as e: print_error(f"Tags input not found: {e}", 3); raise
        try:
            for tag in metadata.get('optimized_tags', []): tags_input.send_keys(tag + ","); mimic_human_action_delay(0.05, 0.15)
            print_success(f"Added {len(metadata.get('optimized_tags', []))} tags.", 3)
        except Exception as e: print_error(f"Error adding tags: {e}", 3); raise

        # --- Category ---
        # Legacy category selection is complex and prone to breaking, skipping explicit selection here.
        # Relies on the default category being correct or user setting it manually before running.
        print_info("Skipping explicit category selection (rely on default).", 2)

        # --- Wait for Upload Completion (Legacy - Simple Check) ---
        print_info("Waiting for upload complete check (legacy)...", 2)
        try: WebDriverWait(driver, 1800).until(EC.text_to_be_present_in_element((By.XPATH, "//span[contains(@class, 'progress-label')]"), "Checks complete")); print_success("Upload/Checks complete indicator found.", 3)
        except TimeoutException: print_error("Timeout waiting for upload/checks complete indicator.", 3); raise

        # --- Next Buttons (Legacy - assumes 3 clicks) ---
        print_info("Clicking 'Next' buttons (legacy)...", 2)
        next_button_xpath = "//ytcp-button[@id='next-button']"
        for i in range(3):
            try: print_info(f"Clicking Next ({i+1}/3)...", 3); next_button = wait_medium.until(EC.element_to_be_clickable((By.XPATH, next_button_xpath))); next_button.click(); print_success("Next clicked.", 3); time.sleep(1.5) # Slightly longer wait
            except Exception as e: print_error(f"Error clicking Next ({i+1}/3): {e}", 3); raise

        # --- Visibility (Legacy) ---
        print_info("Setting visibility (legacy)...", 2)
        if publish_now:
            try: public_radio_xpath = "//tp-yt-paper-radio-button[@name='PUBLIC']"; public_radio = wait_medium.until(EC.element_to_be_clickable((By.XPATH, public_radio_xpath))); public_radio.click(); print_success("Selected 'Public'.", 3)
            except Exception as e: print_error(f"Failed to select 'Public': {e}", 3); raise
            try: publish_button_xpath = "//ytcp-button[@id='done-button']"; publish_button = wait_medium.until(EC.element_to_be_clickable((By.XPATH, publish_button_xpath))); publish_button.click(); print_success("Clicked 'Publish'.", 3)
            except Exception as e: print_error(f"Failed to click 'Publish': {e}", 3); raise
        else: # Schedule
            if not schedule_time: print_error("Schedule time required.", 3); raise ValueError("Schedule time missing")
            try: schedule_radio_xpath = "//tp-yt-paper-radio-button[@name='SCHEDULE']"; schedule_radio = wait_medium.until(EC.element_to_be_clickable((By.XPATH, schedule_radio_xpath))); schedule_radio.click(); print_success("Selected 'Schedule'.", 3); time.sleep(1)
            except Exception as e: print_error(f"Failed to select 'Schedule': {e}", 3); raise
            # Date/Time selection using the old helper function
            if not select_date_in_calendar(driver, schedule_time): raise Exception("Failed legacy date selection")
            time_str = schedule_time.strftime("%H:%M")
            try: time_input_xpath = "//ytcp-form-input-container[@id='time-of-day-container']//input"; time_input = wait_medium.until(EC.presence_of_element_located((By.XPATH, time_input_xpath))); time_input.clear(); time_input.send_keys(time_str); time_input.send_keys(Keys.ENTER); print_success(f"Time set to {time_str}.", 3); time.sleep(1)
            except Exception as e: print_error(f"Failed to set time: {e}", 3); raise
            try: schedule_button_xpath = "//ytcp-button[@id='done-button']"; schedule_button = wait_medium.until(EC.element_to_be_clickable((By.XPATH, schedule_button_xpath))); schedule_button.click(); print_success("Clicked 'Schedule'.", 3)
            except Exception as e: print_error(f"Failed to click 'Schedule': {e}", 3); raise

        # --- Confirmation & ID Capture (Legacy - less reliable) ---
        print_info("Waiting for confirmation (legacy)...", 3)
        try:
            # Wait for the 'Video processed' or similar message
            WebDriverWait(driver, 300).until(lambda d: "processing abandoned" not in d.page_source.lower() and ("processing complete" in d.page_source.lower() or "video published" in d.page_source.lower() or "video scheduled" in d.page_source.lower() or d.find_elements(By.XPATH, "//span[contains(text(),'Video link')]"))) # Add video link check
            print_success("Upload confirmed (text check).", 4)
            # Try to find the Share URL to extract the ID
            share_url_element_xpath = "//a[contains(@href, 'youtu.be/') or contains(@href, 'youtube.com/shorts/')]" # Simplified
            share_url_element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, share_url_element_xpath)))
            share_url = share_url_element.get_attribute('href'); print_info(f"Captured Share URL: {share_url}", 4)
            match = re.search(r"(?:youtu\.be/|youtube\.com/(?:shorts/|watch\?v=))([^?&]+)", share_url)
            if match: youtube_video_id = match.group(1); print_success(f"Parsed YouTube Video ID: {youtube_video_id}", 4); upload_successful = True
            else: print_warning("Could not parse YT ID from Share URL.", 4); upload_successful = False # Fail if ID missing
            # Close the confirmation dialog (best effort)
            try: close_button = driver.find_element(By.XPATH, "//ytcp-button[@id='close-button']"); close_button.click(); print_success("Confirmation dialog closed.", 4)
            except Exception as close_e: print_warning(f"Could not close confirmation dialog: {close_e}", 4)
        except TimeoutException: print_error("Timeout waiting for upload confirmation.", 4); upload_successful = False
        # --- End Confirmation ---

    except (NoSuchWindowException, InvalidSessionIdException) as sess_err: msg = f"Browser session lost (legacy): {sess_err}"; print_error(msg, 1); raise
    except TimeoutException as e: msg = f"Timeout (legacy): {e}"; print_error(msg, 1); return None
    except Exception as e: msg = f"Error during legacy upload: {e}"; print_error(msg, 1, include_traceback=True); return None
    finally: print_info(f"--- Finished Legacy Upload Attempt for {video_index} ---", 1)

    return youtube_video_id if upload_successful else None
# --- End upload_video (Legacy Fallback) ---


# --- POM Adapter Function ---
def use_pom_uploader(
    video_file: str,
    metadata: Dict[str, Any],
    publish_now: bool = True,
    schedule_time: Optional[datetime] = None,
    ffmpeg_path: Optional[str] = None,
    enable_recording: bool = False,
    profile_path: Optional[str] = None, # Add profile_path
    service: Optional[Any] = None # Add service for playlists
) -> Optional[str]:
    """
    Adapter function to use the new Page Object Model uploader.
    Maintains backward compatibility while using the new implementation.
    """
    driver = None # Initialize driver to None
    try:
        # Import the POM uploader components dynamically
        from uploader_pom import setup_browser, upload_video_pom

        print_info("Using new Page Object Model uploader implementation")

        # Set up browser
        driver = setup_browser(profile_path)
        if not driver:
            print_error("Failed to set up browser for POM uploader. Exiting attempt.")
            return None # Return None if browser setup fails

        # Upload video using POM implementation
        video_id = upload_video_pom(
            driver=driver,
            video_file=video_file,
            metadata=metadata,
            publish_now=publish_now,
            schedule_time=schedule_time,
            ffmpeg_path=ffmpeg_path,
            enable_recording=enable_recording
        )

        # --- Handle Playlist Addition AFTER successful upload ---
        if video_id: # Only if upload succeeded and we got an ID
             target_playlist = metadata.get("target_playlist")
             if target_playlist:
                 print_info(f"Attempting playlist management for YT ID: {video_id}, Target: '{target_playlist}'", 1)
                 playlist_id_to_add = None
                 service_needed = False

                 if target_playlist.startswith("NEW: "):
                     new_playlist_title = target_playlist[len("NEW: "):].strip()
                     if new_playlist_title:
                         print_info(f"Need to create NEW playlist: '{new_playlist_title}'", 2)
                         service_needed = True
                         if not service: service = get_authenticated_service() # Get service if needed
                         if service:
                             created_id = create_playlist(service, new_playlist_title)
                             if created_id:
                                 playlist_id_to_add = created_id
                                 # Update cache
                                 try:
                                     pl_cache = load_playlist_cache()
                                     pl_cache[created_id] = new_playlist_title
                                     if target_playlist in pl_cache: del pl_cache[target_playlist]
                                     pl_cache["timestamp"] = datetime.now().isoformat()
                                     save_playlist_cache(pl_cache)
                                 except Exception as cache_e: print_warning(f"Could not update playlist cache after creation: {cache_e}", 3)
                             else: print_warning(f"Failed to create new playlist '{new_playlist_title}'.", 2)
                         else: print_warning("Cannot create new playlist - API service unavailable.", 2)
                     else: print_warning("Gemini suggested 'NEW:' but title was empty. Skipping playlist.", 2)
                 else:
                     # Assume target_playlist is an existing ID
                     playlist_id_to_add = target_playlist
                     print_info(f"Targeting existing playlist ID: {playlist_id_to_add}", 2)
                     service_needed = True

                 # Add to playlist if ID found/created
                 if playlist_id_to_add:
                     if service_needed and not service: service = get_authenticated_service()
                     if service: add_video_to_playlist(service, video_id, playlist_id_to_add)
                     else: print_warning(f"Cannot add to playlist {playlist_id_to_add} - API service unavailable.", 2)
             else: print_info("No target playlist specified in metadata.", 1)
        # --- End Playlist Management ---

        return video_id

    except ImportError:
        print_warning("Page Object Model uploader not available (ImportError). Falling back to legacy uploader.")
        # Fall back to the original upload_video function within this file
        # Need to configure driver separately for legacy function
        driver_legacy = configure_driver() # Setup driver for legacy
        if not driver_legacy:
             print_error("Failed to setup driver for legacy fallback.")
             return None
        try:
            # Call legacy function
            legacy_video_id = upload_video(
                driver=driver_legacy,
                video_file=video_file,
                metadata=metadata,
                publish_now=publish_now,
                schedule_time=schedule_time,
                desc_limit=cfg_desc_limit, # Pass limits
                tag_char_limit=cfg_tag_limit,
                total_char_limit=cfg_total_tags_limit,
                max_count_limit=cfg_max_tags_count,
                service=service # Pass service for playlists
            )
            return legacy_video_id
        finally:
             if driver_legacy:
                 try: driver_legacy.quit()
                 except: pass # Ignore errors during quit
    except Exception as e:
        print_error(f"Unexpected error in use_pom_uploader adapter: {e}", include_traceback=True)
        return None
    finally:
        # Ensure the driver for POM is closed if it was created
        if driver:
            try: driver.quit()
            except Exception as dq_e: print_warning(f"Minor error closing POM driver: {dq_e}")
# --- End POM Adapter Function ---


# --- delete_uploaded_files function ---
def delete_uploaded_files(video_file: str, metadata_file_path: str) -> bool:
    """Deletes the video and its corresponding metadata JSON file."""
    print_info(f"Attempting cleanup for: {os.path.basename(video_file)}", 1)
    deleted_video = False; deleted_metadata = False
    try:
        if os.path.exists(video_file): os.remove(video_file); print_success(f"Deleted video file: {os.path.basename(video_file)}", 2); deleted_video = True
        else: print_info(f"Video file not found (already deleted?): {os.path.basename(video_file)}", 2); deleted_video = True # Treat missing as success
    except Exception as e: print_error(f"Error deleting video file {os.path.basename(video_file)}: {e}", 2)
    try:
        if os.path.exists(metadata_file_path): os.remove(metadata_file_path); print_success(f"Deleted metadata file: {os.path.basename(metadata_file_path)}", 2); deleted_metadata = True
        else: print_info(f"Metadata file not found (already deleted?): {os.path.basename(metadata_file_path)}", 2); deleted_metadata = True # Treat missing as success
    except Exception as e: print_error(f"Error deleting metadata file {os.path.basename(metadata_file_path)}: {e}", 2)
    success = deleted_video and deleted_metadata
    if not success: print_warning("Cleanup incomplete.", 1)
    return success
# --- End delete_uploaded_files function ---


# --- FFmpeg Recording Helper Functions ---
def start_recording(video_index: str, ffmpeg_cmd_path: str, driver: webdriver.Firefox) -> Optional[Tuple[subprocess.Popen, str]]:
    """Starts the FFmpeg screen recording."""
    global _current_recording_process, _current_recording_filename
    if not driver: print_error("Cannot start recording: Selenium driver invalid.", 3); return None
    print_info(f"Attempting start debug recording for video index: {video_index}", 2)
    try:
        os.makedirs(DEBUG_RECORDING_FOLDER, exist_ok=True); timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = os.path.join(DEBUG_RECORDING_FOLDER, f"recording_video_{video_index}_{timestamp}.mp4")
        cmd = [ffmpeg_cmd_path, '-y', '-loglevel', 'error', '-f']
        system = platform.system()
        if system == "Windows":
             cmd.extend(['gdigrab']);
             try: window_title = driver.title;
             except WebDriverException: window_title = "Mozilla Firefox" # Fallback title
             if not window_title or len(window_title) > 100: print_warning("Firefox window title problematic. Falling back to desktop capture.", 4); cmd.extend(['-i', 'desktop'])
             else: print_info(f"Targeting Firefox window title: '{window_title}'", 4); cmd.extend(['-i', f'title={window_title}']); time.sleep(0.3)
        elif system == "Linux": print_warning("Linux: Using full desktop capture (x11grab).", 4); display = os.environ.get('DISPLAY');
        if not display: print_error("Linux: DISPLAY env var not set.", 3); return None; cmd.extend(['x11grab', '-r', '15', '-i', display]) # Framerate before input for x11grab
        elif system == "Darwin": print_warning("macOS: Using full desktop capture (avfoundation). Check permissions.", 4); cmd.extend(['avfoundation', '-r', '15', '-i', '1:none']) # Framerate before input
        else: print_error(f"Unsupported OS for recording: {system}", 3); return None
        if len(cmd) <= 5: print_error("Failed determine input method for FFmpeg.", 3); return None
        cmd.extend(['-c:v', 'libx264', '-preset', 'ultrafast', '-crf', '28', '-pix_fmt', 'yuv420p', output_filename])
        print_info(f"FFmpeg command list: {cmd}", 3)
        creationflags = 0;
        if system == "Windows": creationflags = subprocess.CREATE_NO_WINDOW
        process = subprocess.Popen(cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, creationflags=creationflags)
        time.sleep(2.0); # Allow init
        if process.poll() is not None: stderr_output = "";
        try: stderr_output = process.stderr.read().decode('utf-8', errors='ignore');
        except: pass; print_error(f"FFmpeg failed start (code: {process.returncode}). Error:", 3);
        if stderr_output: print(f"{Fore.RED}{stderr_output}{Style.RESET_ALL}"); log_error_to_file(f"FFmpeg failed start {video_index}. Cmd: {' '.join(cmd)}\nRC: {process.returncode}\nStderr: {stderr_output}");
        else: print_info("No stderr captured.", 4); return None
        print_success("FFmpeg recording started.", 3); _current_recording_process = process; _current_recording_filename = output_filename; return process, output_filename
    except FileNotFoundError: print_error(f"FFmpeg not found: '{ffmpeg_cmd_path}'. Check FFMPEG_PATH config.", 3); log_error_to_file(f"FFmpeg not found at path: {ffmpeg_cmd_path}"); return None
    except WebDriverException as e: print_error(f"Selenium error during recording setup: {e}", 3); log_error_to_file(f"Selenium error recording setup {video_index}: {e}", include_traceback=True); return None
    except OSError as e: print_error(f"OS Error starting FFmpeg: {e}", 3); log_error_to_file(f"OS Error starting FFmpeg {video_index}: {e}", include_traceback=True); return None
    except Exception as e: print_error(f"Unexpected error starting FFmpeg: {e}", 3); log_error_to_file(f"Unexpected error starting FFmpeg {video_index}: {e}", include_traceback=True); return None

def stop_recording(process: Optional[subprocess.Popen], filename: Optional[str], keep_file: bool):
    """Stops the FFmpeg recording process and handles the output file."""
    global _current_recording_process, _current_recording_filename
    if not process or not filename: print_info("No active recording process/filename to stop.", 2); return
    print_info(f"Stopping recording process (PID: {process.pid}). Keep file: {keep_file}", 2)
    if process.poll() is None:
        print_info("Attempting graceful shutdown (sending 'q')...", 3); graceful_success = False; shutdown_timeout = 5
        try:
            if process.stdin:
                if platform.system() == "Windows": process.stdin.write(b'q')
                else: process.stdin.write(b'q\n')
                process.stdin.flush(); process.stdin.close(); print_info(f"Waiting up to {shutdown_timeout}s...", 4); process.wait(timeout=shutdown_timeout); graceful_success = True; print_success("FFmpeg exited gracefully.", 4)
            else: print_warning("Process stdin not available.", 4)
        except (subprocess.TimeoutExpired, OSError, BrokenPipeError, ValueError) as e: print_warning(f"Graceful shutdown failed/timed out: {e}. Terminating/Killing...", 4); graceful_success = False
        except Exception as e: print_warning(f"Unexpected error during graceful shutdown: {e}. Terminating/Killing...", 4); graceful_success = False
        if not graceful_success and process.poll() is None:
            try:
                print_info("Terminating/Killing FFmpeg process...", 4)
                if platform.system() == "Windows": subprocess.run(['taskkill', '/F', '/PID', str(process.pid)], check=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                else: os.kill(process.pid, signal.SIGINT); time.sleep(1); os.kill(process.pid, signal.SIGTERM); time.sleep(1); os.kill(process.pid, signal.SIGKILL)
                process.wait(timeout=3); print_success("FFmpeg terminated/killed.", 4)
            except (subprocess.TimeoutExpired, ProcessLookupError): print_warning("Process likely already exited.", 4)
            except Exception as term_e: print_error(f"Error during terminate/kill: {term_e}", 4)
    else: print_info(f"Recording process already terminated (Code: {process.poll()}).", 3)
    if not keep_file:
        print_info(f"Deleting recording file: {filename}", 3)
        try:
            if os.path.exists(filename): os.remove(filename); print_success("Recording file deleted.", 4)
            else: print_info("Recording file not found (already deleted?).", 4)
        except Exception as e: print_error(f"Error deleting recording file {filename}: {e}", 4); log_error_to_file(f"Error deleting recording {filename}: {e}")
    else:
        if os.path.exists(filename): print_success(f"Keeping recording file: {filename}", 3)
        else: print_warning(f"Intended keep recording, but file not found: {filename}", 3)
    _current_recording_process = None; _current_recording_filename = None
# --- End FFmpeg Recording Helper Functions ---


# --- analyze_upload_errors_with_gemini Function ---
def analyze_upload_errors_with_gemini():
    """Analyzes upload errors using Gemini AI and generates suggestions."""
    if not genai: print_warning("Gemini AI not available. Skipping error analysis."); return None
    try:
        metrics = load_performance_metrics()
        if metrics["total_errors"] < MIN_ERRORS_FOR_ANALYSIS: print(f"{Fore.YELLOW}Not enough errors ({metrics['total_errors']}) for analysis. Need {MIN_ERRORS_FOR_ANALYSIS}."); return None
        total_attempts = metrics["total_uploads_attempted"]; error_rate = metrics["total_errors"] / max(1, total_attempts)
        if error_rate < MIN_ERROR_RATE_FOR_ANALYSIS: print(f"{Fore.YELLOW}Error rate ({error_rate:.1%}) below threshold ({MIN_ERROR_RATE_FOR_ANALYSIS:.1%}). Analysis skipped."); return None
        error_log_content = "";
        if os.path.exists(ERROR_LOG_FILE):
            try:
                 with open(ERROR_LOG_FILE, "r", encoding="utf-8") as f: error_log_content = f.read()
            except Exception as log_read_e: print_warning(f"Could not read error log file {ERROR_LOG_FILE}: {log_read_e}")
        summary = ["=== Performance Summary ==="]
        summary.append(f"Total upload attempts: {metrics['total_uploads_attempted']}")
        summary.append(f"Total successful uploads: {metrics['total_uploads_successful']}")
        summary.append(f"Total errors: {metrics['total_errors']}")
        summary.append(f"Overall success rate: {metrics['total_uploads_successful'] / max(1, metrics['total_uploads_attempted']):.1%}")
        summary.append("\n=== Error Type Breakdown ===")
        for error_type, count in sorted(metrics["error_counts"].items(), key=lambda item: item[1], reverse=True):
            if count > 0: summary.append(f"{ERROR_TYPES.get(error_type, error_type)}: {count} ({count / max(1, metrics['total_errors']):.1%})")
        if metrics["error_samples"]:
            summary.append("\n=== Recent Error Samples ===")
            for i, sample in enumerate(metrics["error_samples"][-5:], 1): summary.append(f"Sample {i}: Type={sample['type']}, Step={sample['step']}, Msg={sample['message'][:100]}..., XPath={sample.get('xpath', 'N/A')}")
        performance_summary = "\n".join(summary)
        prompt = f"""Analyze the following performance data and error logs from a YouTube Shorts uploader script using Selenium WebDriver.
        Performance Summary:
        {performance_summary}
        Recent Error Log Entries:
        {error_log_content[-5000:] if len(error_log_content) > 5000 else error_log_content}
        Based on this data, provide:
        1. Analysis of the most common/critical failure points.
        2. Potential reasons (e.g., outdated XPaths, network issues, UI changes).
        3. Specific technical suggestions: Alternative XPaths, timeout adjustments, retry logic, error handling improvements.
        4. Configuration parameter suggestions (if applicable). Format clearly."""
        model = genai.GenerativeModel("gemini-2.0-flash"); response = model.generate_content(prompt); analysis = response.text.strip()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(UPLOADER_ANALYSIS_LOG, "a", encoding="utf-8") as f: f.write(f"\n\n=== Analysis Generated on {timestamp} ===\n\n{analysis}\n\n=== End of Analysis ===\n")
        metrics["last_analysis_date"] = timestamp; save_performance_metrics(metrics)
        return analysis
    except Exception as e: print(f"{Fore.RED}Error analyzing upload errors: {e}"); traceback.print_exc(); return None
# --- End analyze_upload_errors_with_gemini Function ---


# --- Main Execution Logic ---
def main():
    """Main function to drive the uploader script."""
    global _current_recording_process, _current_recording_filename

    # --- Argument Parsing ---
    analyze_mode = False; publish_mode_override = None; max_uploads_override = None
    if len(sys.argv) > 1:
        if sys.argv[1].lower() in ["--analyze", "-a"]: analyze_mode = True
        if "--publish" in sys.argv: publish_mode_override = True
        if "--schedule" in sys.argv: publish_mode_override = False
        try: max_arg_idx = sys.argv.index("--max"); max_uploads_override = int(sys.argv[max_arg_idx + 1])
        except (ValueError, IndexError): pass
    # --- End Argument Parsing ---

    if analyze_mode:
        # --- Run Analysis Mode ---
        print_section_header("Starting YouTube Uploader in Analysis Mode")
        if genai is None: print_error("Google Generative AI library not found. Cannot run analysis."); return
        if not gemini_api_key: print_error("GEMINI_API_KEY not found in config. Cannot run analysis."); return
        print_info("Analyzing upload errors...")
        analysis = analyze_upload_errors_with_gemini()
        if analysis: print_success("Analysis completed successfully."); print_section_header("Analysis Results"); print(analysis); print_info(f"\nFull analysis saved to {UPLOADER_ANALYSIS_LOG}")
        else: print_warning("Analysis could not be completed. Check logs.")
        return
        # --- End Analysis Mode ---

    # --- Normal Upload Mode ---
    print_section_header("Starting YouTube Uploader Script"); start_time = time.time()
    driver = None; wb = None; downloaded_sheet = None; uploaded_sheet = None; excel_save_required = False
    metrics = load_performance_metrics() # Load metrics
    run_upload_attempts = 0; run_upload_successes = 0 # Track for this run

    try:
        # --- Print Configuration ---
        print_section_header("Configuration Settings")
        final_max_uploads = max_uploads_override if max_uploads_override is not None else max_uploads # Apply override
        print_config("Max Uploads per Run", final_max_uploads)
        print_config("Video Category", upload_category)
        print_config("Profile Path", profile_path_config if profile_path_config else f"{Style.DIM}Default{Style.RESET_ALL}")
        print_config("Input Metadata Folder", INPUT_METADATA_FOLDER)
        print_config("Input Video Folder", INPUT_VIDEO_FOLDER)
        print_config("Excel Data File", EXCEL_FILE_PATH)
        print_config("Error Log File", ERROR_LOG_FILE)
        print_config("Desc Char Limit", cfg_desc_limit); print_config("Tag Char Limit", cfg_tag_limit); print_config("Total Tag Chars Limit", cfg_total_tags_limit); print_config("Max Tags Count", cfg_max_tags_count)
        final_scheduling_mode = scheduling_mode if publish_mode_override is None else ("publish_now" if publish_mode_override else "force_schedule")
        print_config("Scheduling Mode", f"{final_scheduling_mode} {'(Overridden by CLI)' if publish_mode_override is not None else ''}")
        if final_scheduling_mode == 'analytics_priority': print_config("-> Analytics Enabled", f"{Fore.GREEN}True{Style.RESET_ALL}"); # Print analytics details...
        elif final_scheduling_mode == 'custom_tomorrow': print_config("Custom Schedule Times", ", ".join([t.strftime('%I:%M %p') for t in parsed_config_times]) if parsed_config_times else "None")
        print_config("Fallback Interval (mins)", schedule_interval_minutes); print_config("Min Schedule Ahead (mins)", min_schedule_ahead_minutes)
        if enable_analytics_scheduling and final_scheduling_mode != 'analytics_priority': print_config("Analytics Scheduling", f"{Fore.GREEN}Enabled{Style.RESET_ALL}") # Print analytics details...
        print_config("Enable Debug Recording", f"{Fore.GREEN}{enable_debug_recording}{Style.RESET_ALL}" if enable_debug_recording else f"{Fore.YELLOW}{enable_debug_recording}{Style.RESET_ALL}")
        if enable_debug_recording: print_config("FFmpeg Path", ffmpeg_path_config); print_config("Recording Output Dir", DEBUG_RECORDING_FOLDER)
        # --- End Print Config ---

        # --- Setup ---
        print_section_header("Initializing WebDriver"); driver = configure_driver()
        if not driver: raise Exception("WebDriver initialization failed. Check logs.")

        print_section_header("Loading Excel Data")
        excel_headers = { "Downloaded": ["Video Index", "Optimized Title", "Downloaded Date", "Views", "Uploader", "Original Title"], "Uploaded": ["Video Index", "Optimized Title", "YouTube Video ID", "Upload Timestamp", "Schedule Time", "Publish Status", "Views (YT)", "Likes (YT)", "Comments (YT)", "Last Updated"] }
        if EXCEL_UTILS_AVAILABLE:
            try:
                wb, sheets, save_needed = load_or_create_excel(EXCEL_FILE_PATH, excel_headers)
                assert wb is not None, "Workbook failed load/create (excel_utils)"
                downloaded_sheet = sheets.get("Downloaded")
                uploaded_sheet = sheets.get("Uploaded")
                print_success("Excel loaded (excel_utils).", 1)
                if save_needed:
                    excel_save_required = True # Mark save needed if headers corrected
            except Exception as e:
                print_error(f"Error using excel_utils: {e}", 1, include_traceback=True)
                print_fatal("Cannot proceed without valid Excel handling.")
                raise
        else: # Fallback
            try:
                if not os.path.exists(EXCEL_FILE_PATH): wb = Workbook(); downloaded_sheet = wb.active; downloaded_sheet.title = "Downloaded"; downloaded_sheet.append(excel_headers["Downloaded"]); uploaded_sheet = wb.create_sheet("Uploaded"); uploaded_sheet.append(excel_headers["Uploaded"]); wb.save(EXCEL_FILE_PATH); print_success("Created new Excel.", 1)
                else: wb = load_workbook(EXCEL_FILE_PATH); downloaded_sheet = wb["Downloaded"] if "Downloaded" in wb.sheetnames else None; uploaded_sheet = wb["Uploaded"] if "Uploaded" in wb.sheetnames else None # Handle potentially missing sheets
                if downloaded_sheet is None: downloaded_sheet = wb.create_sheet("Downloaded", 0); downloaded_sheet.append(excel_headers["Downloaded"]); excel_save_required=True; print_warning("Created missing 'Downloaded' sheet.", 1)
                if uploaded_sheet is None: uploaded_sheet = wb.create_sheet("Uploaded"); uploaded_sheet.append(excel_headers["Uploaded"]); excel_save_required=True; print_warning("Created missing 'Uploaded' sheet.", 1)
                print_success("Excel loaded (openpyxl).", 1)
            except Exception as e: print_fatal(f"Error handling Excel file {EXCEL_FILE_PATH}: {e}"); raise

        print_section_header("Checking Scheduled Video Status")
        if check_and_update_scheduled(uploaded_sheet): excel_save_required = True

        # --- Excel Archiving ---
        print_section_header("Archiving Old Excel Entries")
        if EXCEL_UTILS_AVAILABLE:
            try:
                ARCHIVE_DAYS = int(config.get("EXCEL_ARCHIVE_DAYS", "180")); print_info(f"Archiving entries older than {ARCHIVE_DAYS} days", 1)
                archived_dl = archive_old_excel_entries(wb, "Downloaded", "Downloaded Date", ARCHIVE_DAYS) if downloaded_sheet else False
                archived_ul = archive_old_excel_entries(wb, "Uploaded", "Upload Timestamp", ARCHIVE_DAYS) if uploaded_sheet else False
                if archived_dl or archived_ul: print_success("Archiving complete.", 1); excel_save_required = True
                else: print_info("No entries were old enough to archive.", 1)
            except Exception as e: print_error(f"Error during Excel archiving: {e}", 1); log_error_to_file(f"ERROR: Excel archiving failed: {e}", include_traceback=True)
        else: print_warning("Excel archiving skipped: excel_utils module not available", 1)
        # --- End Excel Archiving ---

        # --- YouTube Analytics & Playlist API ---
        service = None; peak_hours_from_analytics = []; playlist_cache = load_playlist_cache()
        if enable_analytics_scheduling or any(metadata.get("target_playlist") for metadata in [...] ): # Need to check actual metadata later
            print_section_header("Fetching/Loading API Data (Analytics/Playlists)")
            cached_data = load_peak_times_cache(); cache_is_fresh = False
            if cached_data:
                try: cache_timestamp = datetime.fromisoformat(cached_data["timestamp"]);
                except: cache_timestamp = datetime.min # Handle invalid timestamp
                if datetime.now() - cache_timestamp < timedelta(hours=analytics_cache_expiry_hours): peak_hours_from_analytics = cached_data["peak_hours"]; cache_is_fresh = True; print_info(f"Using fresh peak hours data from cache: {peak_hours_from_analytics}", 1)
                else: print_info("Peak times cache expired. Will fetch fresh data.")
            if not cache_is_fresh or not playlist_cache.get("timestamp") or datetime.now() - datetime.fromisoformat(playlist_cache.get("timestamp", datetime.min.isoformat())) > timedelta(hours=24):
                print_info("Attempting to fetch fresh API data (Analytics Peak Times / Playlists)...")
                if GOOGLE_API_AVAILABLE:
                    service = get_authenticated_service() # Authenticate (includes necessary scopes)
                    if service:
                        if not cache_is_fresh and enable_analytics_scheduling:
                            fetched_hours = get_peak_viewer_hours_from_api(service, analytics_days_to_analyze, analytics_peak_hours_count)
                            if fetched_hours is not None: peak_hours_from_analytics = fetched_hours; save_peak_times_cache(peak_hours_from_analytics)
                            else: print_warning("Failed fetch peak hours. Dynamic scheduling might use fallback.", 1)
                        # Fetch playlists regardless of cache age if service is active
                        existing_playlists_api = fetch_channel_playlists(service)
                        if existing_playlists_api:
                            new_cache_data = {"timestamp": datetime.now().isoformat()}
                            for p_id, p_title in existing_playlists_api.items(): new_cache_data[p_id] = p_title
                            existing_titles_lower = {t.lower() for t in existing_playlists_api.values()}
                            for key, val in playlist_cache.items():
                                if key.startswith("NEW: ") and key not in new_cache_data:
                                    suggested_title = key[len("NEW: "):]
                                    if suggested_title.lower() not in existing_titles_lower: new_cache_data[key] = val
                            playlist_cache = new_cache_data; save_playlist_cache(playlist_cache)
                        else: print_warning("No playlists fetched from API.", 1)
                    else: print_error("YouTube API service unavailable. Cannot fetch peak times or playlists.", 1)
                else: print_warning("Google API libraries not available. Cannot fetch API data.", 1)
        # --- End API Integration ---

        uploaded_count = 0
        # --- Scheduling Tracking ---
        last_schedule_time: Optional[datetime] = None; first_video_this_run = True
        remaining_config_times: List[dt_time] = list(parsed_config_times) if final_scheduling_mode == 'custom_tomorrow' else []
        if final_scheduling_mode == 'custom_tomorrow': print_info(f"Starting with {len(remaining_config_times)} custom time slots.", 1)
        # --- End Scheduling Tracking ---

        print_section_header(f"Scanning for Videos to Upload (Max: {final_max_uploads})"); print_info(f"Scanning metadata folder: {INPUT_METADATA_FOLDER}", 1); all_metadata_files = []
        try:
            if not os.path.isdir(INPUT_METADATA_FOLDER): print_warning(f"Input metadata folder '{INPUT_METADATA_FOLDER}' not found.", 1)
            else:
                metadata_files_raw = [f for f in os.listdir(INPUT_METADATA_FOLDER) if f.lower().endswith('.json') and re.match(r'video\d+\.json', f, re.IGNORECASE)]
                def get_video_index_from_filename(filename): match = re.search(r'video(\d+)\.json', filename, re.IGNORECASE); return int(match.group(1)) if match else float('inf')
                all_metadata_files = sorted(metadata_files_raw, key=get_video_index_from_filename)
                if all_metadata_files: print_success(f"Found {len(all_metadata_files)} potential metadata files to process.", 1)
                else: print_info("No metadata files matching 'video*.json' found.", 1)
        except Exception as e: print_fatal(f"Error scanning metadata folder '{INPUT_METADATA_FOLDER}': {e}"); raise

        if all_metadata_files: print_section_header(f"Starting Upload Loop (Max: {final_max_uploads})")
        else: print_info("No videos to upload based on scan results.")

        # --- Main Upload Loop ---
        for metadata_file in all_metadata_files:
            if uploaded_count >= final_max_uploads: print_info(f"Reached maximum upload limit ({final_max_uploads}). Stopping."); break

            video_index_match = re.search(r'video(\d+)\.json', metadata_file, re.IGNORECASE)
            if not video_index_match: print_warning(f"Skipping file with unexpected format: {metadata_file}", 1); continue
            video_index = video_index_match.group(1); metadata_path = os.path.join(INPUT_METADATA_FOLDER, metadata_file); video_file_name = f"video{video_index}.mp4"; video_file_path = os.path.join(INPUT_VIDEO_FOLDER, video_file_name)
            print_info(f"--- Processing Video Index: {video_index} ({metadata_file}) ---", 1)

            if not os.path.exists(video_file_path): print_error(f"Video file not found: '{video_file_path}'. Skipping.", 2); continue
            metadata: Dict = {};
            try:
                with open(metadata_path, "r", encoding="utf-8") as f: metadata = json.load(f)
                metadata['video_index'] = video_index; print_success("Metadata loaded.", 2)
                if 'optimized_title' not in metadata or not metadata['optimized_title']: print_warning("Metadata missing 'optimized_title'.", 3)
            except Exception as e: print_error(f"Error reading {metadata_file}: {e}. Skipping.", 2); log_error_to_file(f"ERROR: Reading metadata {metadata_file}: {e}", include_traceback=True); continue

            # --- Scheduling Logic ---
            publish_this_video_now = False; target_schedule_time: Optional[datetime] = None; action_desc = ""
            now = datetime.now(); min_future_time = now + timedelta(minutes=min_schedule_ahead_minutes); tomorrow_date = now.date() + timedelta(days=1); base_time = last_schedule_time if last_schedule_time else now
            print_info(f"Determining schedule (Mode: '{final_scheduling_mode}')...", 2)

            if final_scheduling_mode == "publish_now": publish_this_video_now = True; action_desc = "Publish Now (CLI Override)"
            elif final_scheduling_mode == "force_schedule": publish_this_video_now = False; action_desc = "Schedule (CLI Override)" # Will use fallback interval logic
            elif final_scheduling_mode == 'analytics_priority' and enable_analytics_scheduling and peak_hours_from_analytics:
                print_info(f"Attempting peak hour schedule: {peak_hours_from_analytics}", 3); found_peak_slot = False; check_time = base_time; max_search_time = now + timedelta(hours=48)
                while check_time < max_search_time:
                    check_time += timedelta(minutes=15)
                    if check_time.hour in peak_hours_from_analytics and check_time >= min_future_time:
                         target_schedule_time = check_time; action_desc = f"Schedule (Peak Hour) for {target_schedule_time:%Y-%m-%d %H:%M:%S}"; found_peak_slot = True; print_info(f"Found peak slot: {target_schedule_time:%Y-%m-%d %H:%M:%S}", 4); break
                if not found_peak_slot: print_warning("No suitable peak slot found. Using fallback interval.", 3)
            elif final_scheduling_mode == 'default_interval':
                if first_video_this_run: publish_this_video_now = True; action_desc = "Publish Now (First Video)"
                # Else falls through to fallback
            elif final_scheduling_mode == 'custom_tomorrow':
                if remaining_config_times:
                    slot_time = remaining_config_times[0]; potential_schedule = datetime.combine(tomorrow_date, slot_time)
                    if potential_schedule >= min_future_time and (last_schedule_time is None or potential_schedule > last_schedule_time):
                        target_schedule_time = potential_schedule; action_desc = f"Schedule (Custom Slot) for {target_schedule_time:%Y-%m-%d %H:%M:%S}"; remaining_config_times.pop(0); print_info(f"Using custom slot: {target_schedule_time:%Y-%m-%d %H:%M:%S}. {len(remaining_config_times)} left.", 4)
                    else: print_warning(f"Custom slot {slot_time.strftime('%I:%M %p')} invalid for tomorrow. Fallback.", 4)
                # Else falls through to fallback

            # --- Fallback Interval Logic ---
            if not publish_this_video_now and target_schedule_time is None:
                print_info("Using fallback interval scheduling.", 3)
                calculated_time = base_time + timedelta(minutes=schedule_interval_minutes); target_schedule_time = max(calculated_time, min_future_time)
                action_desc = f"Schedule (Fallback Interval) for {target_schedule_time:%Y-%m-%d %H:%M:%S}"
                print_info(f"Using fallback interval. Next slot: {target_schedule_time:%Y-%m-%d %H:%M:%S}", 3)

            if not publish_this_video_now and target_schedule_time is None: print_error(f"Could not determine schedule time for {video_index}. Skipping.", 2); continue
            print_info(f"Action: {action_desc}", 2)
            if target_schedule_time: last_schedule_time = target_schedule_time # Update for next iteration
            if first_video_this_run and publish_this_video_now: first_video_this_run = False
            # --- End Scheduling Logic ---

            # --- Upload Attempt Loop ---
            max_total_attempts = 3; final_upload_successful = False; captured_youtube_video_id = None; attempted_this_video_metric = False
            for current_attempt in range(1, max_total_attempts + 1):
                print_info(f"\n--- Upload Attempt {current_attempt}/{max_total_attempts} for Video Index: {video_index} ---", 1)
                if current_attempt > 1: print_info(f"Waiting before retry...", 2); time.sleep(random.uniform(5, 10))
                local_recording_process: Optional[subprocess.Popen] = None; local_recording_filename: Optional[str] = None; error_during_this_attempt = False
                if enable_debug_recording:
                    start_result = start_recording(video_index, ffmpeg_path_config, driver);
                    if start_result: local_recording_process, local_recording_filename = start_result
                    else: print_warning(f"Could not start recording for attempt {current_attempt}.", 2)
                try:
                    if current_attempt == 1 and not attempted_this_video_metric: metrics["total_uploads_attempted"] += 1; attempted_this_video_metric = True # Increment global attempts once per video
                    # --- Call Uploader (use_pom_uploader preferred) ---
                    attempt_youtube_video_id = use_pom_uploader( # Use the adapter
                        video_file=video_file_path, metadata=metadata, publish_now=publish_this_video_now, schedule_time=target_schedule_time,
                        ffmpeg_path=ffmpeg_path_config, enable_recording=enable_debug_recording, profile_path=profile_path_config,
                        service=service # Pass service for playlist management
                    )
                    if attempt_youtube_video_id:
                        print_success(f"Upload Attempt {current_attempt} SUCCEEDED (YT ID: {attempt_youtube_video_id}).", 2)
                        final_upload_successful = True
                        captured_youtube_video_id = attempt_youtube_video_id
                        break # Break loop on success
                    else:
                        print_error(f"Upload Attempt {current_attempt} FAILED (No YT ID returned).", 2)
                        error_during_this_attempt = True
                except (NoSuchWindowException, InvalidSessionIdException) as critical_wd_error:
                    error_during_this_attempt = True
                    final_upload_successful = False
                    print_error(f"CRITICAL BROWSER SESSION ERROR attempt {current_attempt}: {critical_wd_error}", 1)
                    log_error_to_file(f"CRITICAL WebDriver error upload {current_attempt} for {video_index}: {critical_wd_error}", include_traceback=True)
                    if local_recording_process:
                        stop_recording(local_recording_process, local_recording_filename, keep_file=True)
                    raise critical_wd_error
                except Exception as upload_err:
                    error_during_this_attempt = True
                    final_upload_successful = False
                    print_error(f"Exception upload attempt {current_attempt}: {upload_err}", 1, include_traceback=True)
                    log_error_to_file(f"Exception upload attempt {current_attempt} for {video_index}: {upload_err}", include_traceback=True)
                finally:
                    if local_recording_process:
                        keep_the_recording = error_during_this_attempt or not final_upload_successful
                        stop_recording(local_recording_process, local_recording_filename, keep_the_recording)
                if not final_upload_successful and current_attempt < max_total_attempts:
                    print_info(f"Will retry. {max_total_attempts - current_attempt} attempts remaining.", 2)
            # --- End Upload Attempt Loop ---

            if not final_upload_successful:
                print_error(f"All {max_total_attempts} upload attempts FAILED for {video_index}.", 1)
                log_error_to_file(f"ERROR: All {max_total_attempts} upload attempts failed for {video_index}.", step="retry_handler")
                continue # Skip post-upload if all attempts failed

            # --- Post-Upload Actions (only if final_upload_successful is True and ID captured) ---
            if final_upload_successful and captured_youtube_video_id:
                print_success(f"Upload confirmed for {video_index} (YT ID: {captured_youtube_video_id})", 1)
                uploaded_count += 1; metrics["total_uploads_successful"] += 1; excel_save_required = True
                status = "Published" if publish_this_video_now else "Scheduled"; actual_schedule_time_for_excel = target_schedule_time if not publish_this_video_now else None
                # Add correlation data BEFORE deleting files
                discovery_keyword_for_cache = metadata.get("discovery_keyword"); add_to_correlation_cache(f"video{video_index}", discovery_keyword_for_cache, captured_youtube_video_id)
                # Update Excel data
                update_excel_data(downloaded_sheet, uploaded_sheet, video_index, metadata.get('optimized_title', f'Video {video_index}'), datetime.now(), actual_schedule_time_for_excel, status, captured_youtube_video_id)
                # Delete local files
                delete_uploaded_files(video_file_path, metadata_path)
                print_success(f"Successfully processed {video_index}. Run count: {uploaded_count}/{final_max_uploads}", 1)
            else: print_error(f"Upload FAILED (Final State) for {video_index}. Files NOT deleted.", 1); log_error_to_file(f"Upload FAILED (final state) for {video_index}. Files kept.", step="post_upload_check")

            # Save metrics after each video processing attempt (success or failure)
            save_performance_metrics(metrics)
            print_info(f"--- End Processing Video Index: {video_index} ---", 1); mimic_human_action_delay(2, 5) # Wait between uploads
        # --- End Main Upload Loop ---

        print_section_header("Finished Processing All Found Videos")
        if uploaded_count == 0 and all_metadata_files: print_info("No videos successfully uploaded.", 1)
        elif uploaded_count > 0: print_success(f"Successfully uploaded {uploaded_count} video(s).", 1)

    except (WebDriverException, NoSuchWindowException, InvalidSessionIdException) as wd_e: print_fatal(f"Critical WebDriver error: {wd_e}", include_traceback=False); log_error_to_file(f"FATAL WebDriver error: {wd_e}", include_traceback=True)
    except KeyboardInterrupt: print_fatal("\nScript interrupted by user (Ctrl+C). Exiting.", log_to_file=False); log_error_to_file("Script interrupted by user (Ctrl+C).")
    except Exception as e: print_fatal(f"Unexpected error in main loop: {e}", include_traceback=False); log_error_to_file(f"FATAL error in main loop: {e}", include_traceback=True)

    finally:
        if _current_recording_process: print_warning("Script exiting with active recording. Emergency stop.", 1); stop_recording(_current_recording_process, _current_recording_filename, keep_file=True)
        if wb and excel_save_required:
            print_section_header("Saving Excel Data")
            if EXCEL_UTILS_AVAILABLE:
                if save_workbook_with_fallback(wb, EXCEL_FILE_PATH, extract_workbook_data): print_success(f"Excel saved (excel_utils): {EXCEL_FILE_PATH}", 1)
                else: print_error("All Excel save methods failed (excel_utils). Backup possibly created.", 1)
            else: # Fallback save
                try: wb.save(EXCEL_FILE_PATH); print_success(f"Excel saved (openpyxl): {EXCEL_FILE_PATH}", 1)
                except Exception as e: print_error(f"Final Excel save failed (openpyxl): {e}", 1)
        elif wb: print_info("No Excel changes to save.", 1)
        if driver:
            print_section_header("Shutting Down WebDriver")
            try:
                driver.quit()
                print_success("WebDriver closed.", 1)
            except Exception as qe:
                print_warning(f"Error quitting WebDriver: {qe}", 1)
                log_error_to_file(f"Warning: Error quitting WebDriver: {qe}")
        end_time = time.time()
        duration = timedelta(seconds=end_time - start_time)
        print_section_header("Script Execution Finished")
        print_info(f"Total execution time: {str(duration).split('.')[0]}", 1)
        if os.path.exists(ERROR_LOG_FILE) and os.path.getsize(ERROR_LOG_FILE) > 100:
            print_warning(f"Errors/warnings logged. Check: {ERROR_LOG_FILE}", 1) # Check size > 100 to avoid warning for rotation header
        else:
            print_success("Script completed without significant errors logged.", 1)
        # Check if analysis should be suggested
        if genai is not None and not analyze_mode:
            try:
                metrics = load_performance_metrics()
                if metrics["total_uploads_attempted"] > 0:
                    error_rate = (metrics["total_uploads_attempted"] - metrics["total_uploads_successful"]) / metrics["total_uploads_attempted"]
                    if error_rate >= MIN_ERROR_RATE_FOR_ANALYSIS and metrics["total_errors"] >= MIN_ERRORS_FOR_ANALYSIS:
                        print_info(f"Error rate is {error_rate:.1%}. Consider running with --analyze flag for AI suggestions.")
                        print_info(f"Command: python \"{os.path.basename(__file__)}\" --analyze")
            except Exception as e: print_warning(f"Could not check error metrics: {e}")

# --- Script Entry Point ---
if __name__ == "__main__":
    try:
        analyze_mode = len(sys.argv) > 1 and sys.argv[1].lower() in ["--analyze", "-a"]
        if os.path.exists(ERROR_LOG_FILE) and not analyze_mode:
            try:
                if os.path.getsize(ERROR_LOG_FILE) > 5 * 1024 * 1024: # Rotate if > 5MB
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S"); archive_path = f"{ERROR_LOG_FILE}.{timestamp}.bak"
                    with open(ERROR_LOG_FILE, "r", encoding="utf-8") as src, open(archive_path, "w", encoding="utf-8") as dst: dst.write(src.read())
                    with open(ERROR_LOG_FILE, "w", encoding="utf-8") as f: f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Log rotated. Archived: {archive_path}\n")
                    print_info(f"Error log rotated. Archived: {archive_path}")
            except Exception as e: print(f"Warning: Error managing log file '{ERROR_LOG_FILE}': {e}")
    except Exception as e: print(f"Warning: Error checking log file: {e}")
    main()

# --- End of the script ---